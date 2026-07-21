#!/usr/bin/env python3
"""Normalize 2024 SAMR nonconforming automotive gasoline detergent rows."""

from __future__ import annotations

import hashlib
import html
import io
import json
import re
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data/samr-china-2024-nonconforming-fuel-additives.jsonl"
REPORT = ROOT / "data/samr-china-2024-nonconforming-fuel-additives-report.json"
SOURCE_ID = "SAMR_CHINA_2024_NONCONFORMING_FUEL_ADDITIVES"
SOURCE_URL = "https://www.samr.gov.cn/zw/zfxxgk/fdzdgknr/zljds/art/2025/art_d75bb5a15c794893bca6cf8baf9f5155.html"
ATTACHMENT_URL = "https://www.samr.gov.cn/cms_files/filemanager/1647978232/attach/20254/e9b731884e6848659c3f7c2b8a98f826.xlsx"
RIGHTS_URL = "https://www.samr.gov.cn/jg/wzsm/art/2021/art_c30de52ec3264bd29886479e5471dc72.html"
EXPECTED_XLSX_SHA256 = "a5c76c5d5834f1d33512e4177b79cc0601f13c4ee0fefa2a8e8c45db8caf1ca5"
SNAPSHOT_DATE = "2026-07-21"
REPORT_DATE = "2025-04-07"


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def empty_marker(value: object) -> str:
    value = clean(value)
    return "" if value in {"/", "-", "—"} else value


def fetch() -> bytes:
    request = urllib.request.Request(ATTACHMENT_URL, headers={
        "User-Agent": "Mozilla/5.0 MFClassifierResearch/1.0",
        "Referer": SOURCE_URL,
    })
    with urllib.request.urlopen(request, timeout=120) as response:
        payload = response.read()
    digest = hashlib.sha256(payload).hexdigest()
    if digest != EXPECTED_XLSX_SHA256:
        raise RuntimeError(f"SAMR 2024 XLSX changed: expected {EXPECTED_XLSX_SHA256}, received {digest}")
    return payload


def split_failures(value: str) -> list[str]:
    return [clean(item).rstrip("*") for item in re.split(r"[,，]", value) if clean(item)]


def main() -> None:
    payload = fetch()
    sheet = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)["Sheet1"]
    header = tuple(clean(value) for value in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[:11])
    expected = ("序号", "产品种类", "受检单位", "标称生产单位", "标称生产单位所在地", "产品名称", "规格型号", "生产日期/批号", "主要不合格项目", "承检机构", "备注")
    if header != expected:
        raise RuntimeError(f"Unexpected SAMR 2024 columns: {header!r}")

    source_rows = []
    for values in sheet.iter_rows(min_row=4, values_only=True):
        if values[1] != "车用汽油清净剂":
            continue
        raw = {
            "source_row": int(values[0]),
            "producer": empty_marker(values[3]),
            "product_name": clean(values[5]),
            "model": clean(values[6]),
            "batch": empty_marker(values[7]),
            "failures": clean(values[8]),
            "note": clean(values[10]),
        }
        if not raw["producer"] or not raw["product_name"] or not raw["model"]:
            raise RuntimeError(f"Incomplete relevant SAMR 2024 row: {raw!r}")
        source_rows.append(raw)
    if len(source_rows) != 23:
        raise RuntimeError(f"Expected 23 gasoline-detergent rows, received {len(source_rows)}")

    records = []
    for row in source_rows:
        source_facts = {key: row[key] for key in sorted(row)}
        records.append({
            "source_id": SOURCE_ID,
            "source_record_id": f"SAMR-CN-2024-{row['source_row']:03d}",
            "source_url": SOURCE_URL,
            "attachment_url": ATTACHMENT_URL,
            "rights_url": RIGHTS_URL,
            "snapshot_date": SNAPSHOT_DATE,
            "report_date": REPORT_DATE,
            "market": "China",
            "manufacturer": row["producer"],
            "brand": row["producer"],
            "brand_basis": "source_reported_nominal_producer_fallback_no_separate_brand_column",
            "product_name": row["product_name"],
            "product_name_basis": "source_reported_product_name_from_national_quality_inspection",
            "model_specification_source_reported": row["model"],
            "product_kind_source_reported": "车用汽油清净剂",
            "product_kind_english": "automotive gasoline detergent additive",
            "family_code": "S",
            "technical": {
                "api_source_reported": [], "sae_source_reported": [],
                "brake_fluid_dot_source_reported": [], "brake_fluid_hzy_source_reported": [],
                "coolant_class_source_reported": [], "washer_fluid_class_source_reported": [],
                "urea_class_source_reported": [],
            },
            "production_date_or_batch_source_reported": row["batch"],
            "inspection_outcome": "nonconforming",
            "inspection_retest_confirmed_nonconforming": "复检仍不合格" in row["note"],
            "nonconforming_items": split_failures(row["failures"]),
            "nonconforming_items_source_reported": row["failures"],
            "source_note": row["note"],
            "lifecycle_status": "official_2024_national_inspection_nonconforming_current_market_status_unverified",
            "source_quality_flags": [
                "historical_regulatory_observation_not_current_catalog_listing",
                "nonconforming_product_do_not_treat_as_approved_or_recommended",
                "brand_falls_back_to_nominal_producer_because_source_has_no_brand_column",
            ],
            "source_facts_sha256": hashlib.sha256(json.dumps(source_facts, ensure_ascii=False, sort_keys=True).encode()).hexdigest(),
            "evidence_status": "official_government_nonconforming_product_inspection_observation",
        })

    output_text = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records)
    OUTPUT.write_text(output_text, encoding="utf-8")
    report = {
        "status": "official_china_2024_nonconforming_gasoline_detergent_observations_normalized",
        "source_id": SOURCE_ID,
        "source_url": SOURCE_URL,
        "attachment_url": ATTACHMENT_URL,
        "snapshot_date": SNAPSHOT_DATE,
        "report_date": REPORT_DATE,
        "source_xlsx_sha256": hashlib.sha256(payload).hexdigest(),
        "source_relevant_rows": len(records),
        "normalized_product_observations": len(records),
        "families": dict(Counter(row["family_code"] for row in records)),
        "retest_confirmed_nonconforming_rows": sum(row["inspection_retest_confirmed_nonconforming"] for row in records),
        "normalized_output_sha256": hashlib.sha256(output_text.encode()).hexdigest(),
        "grain_note": "One row is one nominal producer + product name + model/package identity observed as nonconforming in the 2024 national inspection.",
        "lifecycle_note": "Historical failed-inspection evidence only; never approval, recommendation or evidence of current sale.",
        "rights_note": "The SAMR workbook is not redistributed. Only attributed non-expressive product and inspection facts are normalized.",
        "privacy_note": "Retailers, testing laboratories and producer locations are excluded.",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
