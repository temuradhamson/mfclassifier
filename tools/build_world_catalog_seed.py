#!/usr/bin/env python3
"""Build the normalized, provenance-aware seed for the worldwide product catalog."""

from __future__ import annotations

import hashlib
import gzip
import json
import lzma
import re
import shutil
import sqlite3
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import date
from itertools import combinations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CATALOG = ROOT / "data" / "catalog-v3.json"
POLICY = ROOT / "data" / "global-source-policy.json"
JSONL_OUT = ROOT / "data" / "world-catalog-products.jsonl"
JSONL_GZ_OUT = ROOT / "data" / "world-catalog-products.jsonl.gz"
REPORT_OUT = ROOT / "data" / "world-catalog-report.json"
SQLITE_OUT = ROOT / "data" / "world-catalog.sqlite3"
SQLITE_XZ_OUT = ROOT / "data" / "world-catalog.sqlite3.xz"
XLSX_OUT = ROOT / "deliverables" / "World_lubricants_catalog_seed.xlsx"
AICHILON_DB = Path("/workspace/chilon/aichilon/var/chilon_seed.sqlite")
JASO_JSONL = ROOT / "data" / "jaso-filed-oils.jsonl"
LICENSED_JSONL = ROOT / "data" / "official-licensed-products.jsonl"
USDA_BIOPREFERRED_JSONL = ROOT / "data" / "usda-biopreferred-products.jsonl"
ZF_TE_ML_JSONL = ROOT / "data" / "zf-te-ml-approved-products.jsonl"
ALLISON_JSONL = ROOT / "data" / "allison-approved-fluids.jsonl"
DRIVENTIC_DIWA_JSONL = ROOT / "data" / "driventic-diwa-approved-oils.jsonl"
MERCEDES_DTFR_JSONL = ROOT / "data" / "mercedes-dtfr-approved-fluids.jsonl"
MERCEDES_BEVO_JSONL = ROOT / "data" / "mercedes-bevo-approved-fluids.jsonl"
VOLVO_GENUINE_JSONL = ROOT / "data" / "volvo-genuine-fluids.jsonl"
MACK_GENUINE_JSONL = ROOT / "data" / "mack-genuine-fluids.jsonl"
MACK_2014_APPROVED_JSONL = ROOT / "data" / "mack-2014-approved-oils.jsonl"
CUMMINS_VALVOLINE_2022_JSONL = ROOT / "data" / "cummins-valvoline-2022-products.jsonl"
TAIWAN_CPC_JSONL = ROOT / "data" / "taiwan-cpc-lubricant-products.jsonl"
VOLVO_WEBSDS_JSONL = ROOT / "data" / "volvo-group-websds-lubricant-products.jsonl"
PARKER_DENISON_JSONL = ROOT / "data" / "parker-denison-fluid-ratings.jsonl"
SCANIA_GENUINE_JSONL = ROOT / "data" / "scania-genuine-oils.jsonl"
BRAVA_OFFICIAL_JSONL = ROOT / "data" / "brava-official-products.jsonl"
CEYPETCO_JSONL = ROOT / "data" / "ceypetco-lubricant-products.jsonl"
PSO_OFFICIAL_JSONL = ROOT / "data" / "pso-official-lubricant-products.jsonl"
PERTAMINA_OFFICIAL_JSONL = ROOT / "data" / "pertamina-official-lubricant-products.jsonl"
MAN_SERVICE_JSONL = ROOT / "data" / "man-service-products.jsonl"
LIQUI_MOLY_2020_JSONL = ROOT / "data" / "liqui-moly-2020-products.jsonl"
LIQUI_MOLY_CURRENT_JSONL = ROOT / "data" / "liqui-moly-current-products.jsonl"
ANP_BRAZIL_JSONL = ROOT / "data" / "anp-brazil-lubricant-products.jsonl"
ANP_BRAZIL_MONITORING_JSONL = ROOT / "data" / "anp-brazil-monitoring-observations.jsonl"
ANP_BRAZIL_MONITORING_PDF_JSONL = ROOT / "data" / "anp-brazil-monitoring-pdf-observations.jsonl"
ANP_BRAZIL_MONITORING_PDF_EXCEPTIONS_JSONL = ROOT / "data" / "anp-brazil-monitoring-pdf-exceptions.jsonl"
INDONESIA_NPT_JSONL = ROOT / "data" / "indonesia-npt-lubricant-products.jsonl"
THAILAND_DOEB_JSONL = ROOT / "data" / "thailand-doeb-lubricant-products.jsonl"
DLA_QPD_JSONL = ROOT / "data" / "dla-qpd-lubricant-products.jsonl"
BLUE_ANGEL_JSONL = ROOT / "data" / "blue-angel-de-uz-178-products.jsonl"
AUSTRIAN_ECOLABEL_UZ14_JSONL = ROOT / "data" / "austrian-ecolabel-uz14-products.jsonl"
KOREA_ECOLABEL_JSONL = ROOT / "data" / "korea-ecolabel-el611-lubricants.jsonl"
KOREA_ECOLABEL_EL509_JSONL = ROOT / "data" / "korea-ecolabel-el509-washer-fluids.jsonl"
GREEN_CHOICE_PHILIPPINES_JSONL = ROOT / "data" / "green-choice-philippines-lubricants.jsonl"
UAE_MOIAT_JSONL = ROOT / "data" / "uae-moiat-conformity-products.jsonl"
EAEU_CONFORMITY_JSONL = ROOT / "data" / "eaeu-conformity-lubricant-products.jsonl"
EPA_SAFER_CHOICE_JSONL = ROOT / "data" / "epa-safer-choice-lubricants.jsonl"
EPA_CHEMEXPO_JSONL = ROOT / "data" / "epa-chemexpo-lubricants.jsonl"
PSQCA_ENGINE_OIL_JSONL = ROOT / "data" / "psqca-engine-oil-licences.jsonl"
TISI_TWO_STROKE_OIL_JSONL = ROOT / "data" / "tisi-two-stroke-oil-licences.jsonl"
SAMR_CHINA_2025_JSONL = ROOT / "data" / "samr-china-2025-nonconforming-fluids.jsonl"
SAMR_CHINA_2023_JSONL = ROOT / "data" / "samr-china-2023-nonconforming-fluids.jsonl"
SAMR_CHINA_2024_JSONL = ROOT / "data" / "samr-china-2024-nonconforming-fuel-additives.jsonl"
SHENZHEN_CHINA_2021_JSONL = ROOT / "data" / "shenzhen-2021-nonconforming-automotive-fluids.jsonl"
SHENZHEN_CHINA_2020_JSONL = ROOT / "data" / "shenzhen-2020-automotive-fluid-inspection.jsonl"
SHENZHEN_CHINA_2019_JSONL = ROOT / "data" / "shenzhen-2019-automotive-fluid-inspection.jsonl"
SHENZHEN_CHINA_2025_JSONL = ROOT / "data" / "shenzhen-2025-automotive-fluid-inspection.jsonl"
SHANGHAI_CHINA_2023_2025_JSONL = ROOT / "data" / "shanghai-2023-2025-lubricant-inspections.jsonl"
BEIJING_CHINA_2018_JSONL = ROOT / "data" / "beijing-2018-automotive-fluid-inspections.jsonl"
SHENZHEN_CHINA_2016_2017_JSONL = ROOT / "data" / "shenzhen-2016-2017-lubricant-inspections.jsonl"
QINGDAO_CHINA_2021_2025_JSONL = ROOT / "data" / "qingdao-2021-2025-automotive-fluid-inspections.jsonl"
JILIN_CHINA_2024_2025_JSONL = ROOT / "data" / "jilin-2024-2025-automotive-fluid-inspections.jsonl"
PHILIPPINES_BPS_BRAKE_FLUID_JSONL = ROOT / "data" / "philippines-bps-brake-fluid-products.jsonl"
GHANA_GSA_CERTIFIED_JSONL = ROOT / "data" / "ghana-gsa-certified-lubricant-products.jsonl"
KEBS_SMARK_JSONL = ROOT / "data" / "kebs-smark-lubricant-products.jsonl"
EAST_AFRICA_CERTIFIED_JSONL = ROOT / "data" / "east-africa-certified-lubricant-products.jsonl"
SON_MANCAP_JSONL = ROOT / "data" / "son-mancap-chemical-lubricant-products.jsonl"
RSB_SMARK_JSONL = ROOT / "data" / "rsb-smark-lubricant-products.jsonl"
FUCHS_INDIA_JSONL = ROOT / "data" / "fuchs-india-products.jsonl"
FUCHS_US_JSONL = ROOT / "data" / "fuchs-us-products.jsonl"
FUCHS_GERMANY_JSONL = ROOT / "data" / "fuchs-germany-products.jsonl"
FUCHS_POLAND_JSONL = ROOT / "data" / "fuchs-poland-products.jsonl"
FUCHS_ITALY_JSONL = ROOT / "data" / "fuchs-italy-products.jsonl"
FUCHS_SWEDEN_JSONL = ROOT / "data" / "fuchs-sweden-products.jsonl"
FUCHS_SPAIN_JSONL = ROOT / "data" / "fuchs-spain-products.jsonl"
FUCHS_FRANCE_JSONL = ROOT / "data" / "fuchs-france-products.jsonl"
FUCHS_TURKEY_JSONL = ROOT / "data" / "fuchs-turkey-products.jsonl"
FUCHS_CANADA_JSONL = ROOT / "data" / "fuchs-canada-products.jsonl"
FUCHS_CHINA_JSONL = ROOT / "data" / "fuchs-china-products.jsonl"
FUCHS_CZECH_JSONL = ROOT / "data" / "fuchs-czech-products.jsonl"
FUCHS_MEXICO_JSONL = ROOT / "data" / "fuchs-mexico-products.jsonl"
FUCHS_SOUTH_AFRICA_JSONL = ROOT / "data" / "fuchs-south-africa-products.jsonl"
FUCHS_BRAZIL_JSONL = ROOT / "data" / "fuchs-brazil-products.jsonl"
FUCHS_NORWAY_JSONL = ROOT / "data" / "fuchs-norway-products.jsonl"
FUCHS_HUNGARY_JSONL = ROOT / "data" / "fuchs-hungary-products.jsonl"
FUCHS_ADDITIONAL_MARKETS = [
    ("denmark", ROOT / "data" / "fuchs-denmark-products.jsonl", "FUCHS_DENMARK_PRODUCT_FINDER", "Denmark"),
    ("finland", ROOT / "data" / "fuchs-finland-products.jsonl", "FUCHS_FINLAND_PRODUCT_FINDER", "Finland"),
    ("portugal", ROOT / "data" / "fuchs-portugal-products.jsonl", "FUCHS_PORTUGAL_PRODUCT_FINDER", "Portugal"),
    ("romania", ROOT / "data" / "fuchs-romania-products.jsonl", "FUCHS_ROMANIA_PRODUCT_FINDER", "Romania"),
    ("austria", ROOT / "data" / "fuchs-austria-products.jsonl", "FUCHS_AUSTRIA_PRODUCT_FINDER", "Austria"),
    ("greece", ROOT / "data" / "fuchs-greece-products.jsonl", "FUCHS_GREECE_PRODUCT_FINDER", "Greece"),
    ("switzerland", ROOT / "data" / "fuchs-switzerland-products.jsonl", "FUCHS_SWITZERLAND_PRODUCT_FINDER", "Switzerland"),
    ("korea", ROOT / "data" / "fuchs-korea-products.jsonl", "FUCHS_KOREA_PRODUCT_FINDER", "Korea"),
    ("uae", ROOT / "data" / "fuchs-uae-products.jsonl", "FUCHS_UAE_PRODUCT_FINDER", "United Arab Emirates"),
    ("argentina", ROOT / "data" / "fuchs-argentina-products.jsonl", "FUCHS_ARGENTINA_PRODUCT_FINDER", "Argentina"),
    ("chile", ROOT / "data" / "fuchs-chile-products.jsonl", "FUCHS_CHILE_PRODUCT_FINDER", "Chile"),
    ("ukraine", ROOT / "data" / "fuchs-ukraine-products.jsonl", "FUCHS_UKRAINE_PRODUCT_FINDER", "Ukraine"),
    ("slovakia", ROOT / "data" / "fuchs-slovakia-products.jsonl", "FUCHS_SLOVAKIA_PRODUCT_FINDER", "Slovakia"),
    ("slovenia", ROOT / "data" / "fuchs-slovenia-products.jsonl", "FUCHS_SLOVENIA_PRODUCT_FINDER", "Slovenia"),
    ("croatia", ROOT / "data" / "fuchs-croatia-products.jsonl", "FUCHS_CROATIA_PRODUCT_FINDER", "Croatia"),
    ("saudi-arabia", ROOT / "data" / "fuchs-saudi-arabia-products.jsonl", "FUCHS_SAUDI_ARABIA_PRODUCT_FINDER", "Saudi Arabia"),
    ("macedonia", ROOT / "data" / "fuchs-macedonia-products.jsonl", "FUCHS_MACEDONIA_PRODUCT_FINDER", "Macedonia"),
]
SCHEMA_VERSION = 1
SNAPSHOT_DATE = "2026-07-22"

FAMILY_NAMES = {
    "M": "Моторные масла",
    "T": "Трансмиссионные масла",
    "H": "Гидравлические масла",
    "I": "Индустриальные масла",
    "C": "Компрессорные масла",
    "U": "Турбинные масла",
    "E": "Электроизоляционные масла",
    "G": "Пластичные смазки",
    "TF": "Охлаждающие и технические жидкости",
    "S": "Специальные продукты",
}

EXPECTED_ENKT_BASE = {
    "M": "19.20.29.110",
    "T": "19.20.29.120",
    "H": "19.20.29.130",
    "I": "19.20.29.140",
    "C": "19.20.29.150",
    "U": "19.20.29.160",
    "E": "19.20.29.172",
    "G": "19.20.29.210",
}


def text(value) -> str:
    return str(value or "").strip()


def normalize(value) -> str:
    value = unicodedata.normalize("NFKC", text(value)).casefold().replace("ё", "е")
    # Preserve letters, numbers and combining marks from every script.  The old
    # Latin/Cyrillic-only expression collapsed Thai, Chinese and Korean company
    # names to an empty identity, generating thousands of false cross-source
    # duplicate candidates.
    normalized = "".join(
        character if unicodedata.category(character)[0] in {"L", "M", "N"} else " "
        for character in value
    )
    return re.sub(r"\s+", " ", normalized).strip()


def source_for(row: dict) -> str:
    source = text(row.get("source"))
    if source == "products_classified_2026":
        return "user-supplied-chilon-2026"
    if source == "legacy_mfclassifier":
        return "legacy-mfclassifier"
    if source == "aichilon_internal":
        return "aichilon-internal"
    if source.startswith("JASO_"):
        return source
    if "+" in source:
        return "user-supplied-chilon-2026"
    return "legacy-mfclassifier"


def extract_sae(row: dict) -> tuple[str, str]:
    raw = " ".join([text(row.get("sae_class")), text(row.get("name"))]).upper()
    engine = re.search(r"(?<![0-9])((?:0|5|10|15|20|25)W)[-\s]?([0-9]{2})(?![0-9])", raw)
    if engine:
        return f"{engine.group(1)}-{engine.group(2)}", ""
    gear = re.search(r"(?<![0-9])(70W|75W|80W|85W)(?:[-\s]?([0-9]{2,3}))?(?![0-9])", raw)
    if gear:
        return "", gear.group(1) + (f"-{gear.group(2)}" if gear.group(2) else "")
    mono = re.fullmatch(r"(?:SAE\s*)?([0-9]{2,3})", text(row.get("sae_class")).upper())
    if mono:
        return (f"SAE {mono.group(1)}", "") if row.get("category_code") == "M" else ("", f"SAE {mono.group(1)}")
    return "", ""


def extract_performance(row: dict) -> dict:
    raw = " ".join([text(row.get("api_class")), text(row.get("technical_document"))]).upper()
    api_patterns = [
        r"\bAPI\s+(SP|SN\s+PLUS|SN|SM|SL|SJ|SH|SG|CK-4|CJ-4|CI-4\s+PLUS|CI-4|CH-4|CG-4|CF-4|CF|FA-4)(?:\s*[/,]\s*(CF|SN|SM|SL|SJ))?",
        r"\b(GL-[1-6])\b",
    ]
    api = []
    api_gl = []
    for index, pattern in enumerate(api_patterns):
        for match in re.finditer(pattern, raw):
            values = [v.replace("  ", " ") for v in match.groups() if v]
            (api_gl if index else api).extend(values)
    acea = []
    for match in re.finditer(r"\b(?:ACEA\s*)?([ACE][0-9](?:/[BCE][0-9])?(?:-[0-9]{2})?|A[0-9]/B[0-9])\b", raw):
        acea.append(match.group(1))
    ilsac = []
    for match in re.finditer(r"\b(?:ILSAC\s*)?(GF-[1-7])\b", raw):
        ilsac.append(match.group(1))
    return {
        "api": sorted(set(api)),
        "api_gl": sorted(set(api_gl)),
        "acea": sorted(set(acea)),
        "ilsac": sorted(set(ilsac)),
        "performance_raw": text(row.get("api_class")) or text(row.get("technical_document")),
    }


def canonical_record(row: dict) -> dict:
    sae_engine, sae_gear = extract_sae(row)
    performance = extract_performance(row)
    family_code = text(row.get("category_code"))
    iso_vg = ""
    if family_code in {"H", "I", "C", "U", "E"} and row.get("viscosity") not in (None, ""):
        iso_vg = text(row.get("viscosity"))
    nlgi = text(row.get("grease_class"))
    if family_code == "G" and not nlgi:
        match = re.search(r"\bNLGI\s*(00|0|[1-6])\b|\bEP\s*(00|0|[1-6])\b", text(row.get("name")), re.I)
        if match:
            nlgi = next(value for value in match.groups() if value is not None)
    class_match = row.get("class_match") or {}
    source_id = source_for(row)
    source_record_id = text(row.get("legacy_id") or row.get("source_number") or row.get("id"))
    specs = {
        **performance,
        "sae_engine": sae_engine,
        "sae_gear": sae_gear,
        "iso_vg": iso_vg,
        "din_gost_class": text(row.get("din_gost_class")),
        "gost_name": text(row.get("gost_name")),
        "coolant_class": text(row.get("coolant_class")),
        "grease_class": text(row.get("grease_class")),
        "nlgi": nlgi,
    }
    signature_parts = [
        normalize(row.get("brand")), normalize(row.get("name")), normalize(sae_engine),
        normalize(sae_gear), normalize(iso_vg), normalize("|".join(performance["api"])),
        normalize("|".join(performance["api_gl"])), normalize("|".join(performance["acea"])),
        normalize("|".join(performance["ilsac"])), normalize(row.get("din_gost_class")),
        normalize(row.get("coolant_class")), normalize(row.get("grease_class")),
        normalize(nlgi),
    ]
    canonical_key = "|".join(signature_parts)
    canonical_hash = hashlib.sha256(canonical_key.encode()).hexdigest()[:20]
    record = {
        "product_id": f"WC-{canonical_hash}",
        "canonical_key": canonical_key,
        "manufacturer": text(row.get("brand")),
        "brand": text(row.get("brand")),
        "product_name_raw": text(row.get("name")),
        "product_name_normalized": normalize(row.get("name")),
        "market": "UZ",
        "family_code": family_code,
        "family": FAMILY_NAMES.get(family_code, text(row.get("family"))),
        "category": text(row.get("category")),
        "specifications": specs,
        "candidate_technical_profile_id": text(class_match.get("class_id")),
        "profile_match_confidence": class_match.get("confidence"),
        "profile_match_status": text(class_match.get("status")),
        "profile_match_basis": class_match.get("basis") or [],
        "source_id": source_id,
        "source_record_id": source_record_id,
        "source_row": row.get("source_row"),
        "evidence_status": (
            "project_source_row" if source_id == "user-supplied-chilon-2026"
            else "internal_product_master" if source_id == "aichilon-internal"
            else "legacy_needs_official_tds"
        ),
        "lifecycle_status": "active_or_unknown",
        "certificate_status": text(row.get("certificate_status")),
        "codes": {
            k: {
                "value": text(row.get(k)),
                "source_id": (
                    "legacy-mfclassifier" if k in {"ikpu", "enkt", "skp"}
                    else source_id
                ),
                "status": "source_reported_unverified",
            }
            for k in ["tnved_code", "ikpu", "enkt", "skp"] if text(row.get(k))
        },
        "certificate": {
            "number": text(row.get("certificate_number")),
            "issued_at": text(row.get("certificate_issued_at")),
            "expires_at": text(row.get("certificate_expires_at")),
            "local_producer_certificate": text(row.get("local_producer_certificate")),
            "technical_document": text(row.get("technical_document")),
        },
        "snapshot_date": SNAPSHOT_DATE,
    }
    return record


def aichilon_family(category: str, name: str) -> str:
    value = normalize(name)
    if "смазк" in value or "grease" in value:
        return "G"
    if any(token in value for token in ["сож", "антифриз", "coolant", "тормозн", "теплоносител"]):
        return "TF"
    if any(token in value for token in ["трансмис", "редуктор", "gear", " atf "]):
        return "T"
    if "гидрав" in value:
        return "H"
    if "компресс" in value:
        return "C"
    if "турбин" in value:
        return "U"
    if "трансформ" in value or "электроизоля" in value:
        return "E"
    if "мотор" in value or re.search(r"\b(?:0|5|10|15|20|25)w[ -]?[0-9]{2}\b", value):
        return "M"
    if normalize(category) == "масла":
        return "I"
    return "S"


def aichilon_seed() -> tuple[list[dict], list[dict], list[dict]]:
    db = sqlite3.connect(f"file:{AICHILON_DB}?mode=ro", uri=True)
    products = []
    exclusions = []
    for product_id, brand, category, name, is_active, archive_type, archive_reason in db.execute("""
        SELECT p.id, b.code, p.category, p.name, p.is_active, p.archive_type, p.archive_reason
        FROM products p JOIN brands b ON b.id=p.brand_id ORDER BY p.id
    """):
        if normalize(name) in {"в литрах", "мой тестовый товар"}:
            exclusions.append({"source_record_id": product_id, "name": name, "reason": "non_product_service_or_test_row"})
            continue
        family_code = aichilon_family(category or "", name)
        viscosity = ""
        match = re.search(r"\b(?:ISO\s*)?VG\s*([0-9]{1,4})\b", name, re.I)
        if not match and family_code in {"H", "I", "C", "U"}:
            match = re.search(r"\b([0-9]{1,4})\s*$", name)
        if match:
            viscosity = match.group(1)
        api_match = re.search(r"\bAPI\s+(.+)$", name, re.I)
        products.append({
            "id": f"AICHILON-{product_id}",
            "source_number": product_id,
            "source_row": None,
            "brand": brand,
            "name": name,
            "category": category or FAMILY_NAMES.get(family_code, ""),
            "category_code": family_code,
            "family": FAMILY_NAMES.get(family_code, ""),
            "viscosity": viscosity,
            "api_class": api_match.group(1) if api_match else "",
            "source": "aichilon_internal",
            "is_active": bool(is_active),
            "archive_type": archive_type,
            "archive_reason": archive_reason,
        })
    packages = [dict(zip(
        ["package_id", "source_product_id", "package_name", "unit", "quantity_per_package", "is_active", "weight_kg", "density_kg_per_l", "archive_type", "archive_reason"], row
    )) for row in db.execute("""
        SELECT id, product_id, package_name, unit, quantity_per_package, is_active,
               weight_kg, density_kg_per_l, archive_type, archive_reason
        FROM product_packages ORDER BY id
    """)]
    db.close()
    return products, packages, exclusions


def jaso_record(row: dict) -> dict:
    generic = {
        "id": f"{row['source_id']}-{row['source_row_number']}",
        "source_number": row["source_row_number"],
        "brand": row["submitter"],
        "name": row["product_name"],
        "category": {
            "JASO_4T": "Моторные масла для четырёхтактных мотоциклов",
            "JASO_DEO": "Дизельные моторные масла",
            "JASO_2T": "Масла для двухтактных двигателей",
        }[row["source_id"]],
        "category_code": "M",
        "family": "Моторные масла",
        "sae_class": row.get("sae_viscosity", ""),
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["submitter"]
    record["brand"] = row["submitter"]
    record["market"] = "GLOBAL_JASO_FILED"
    record["evidence_status"] = "official_filed_registry"
    record["lifecycle_status"] = "filed_as_of_list_date"
    record["snapshot_date"] = row["list_date"]
    record["specifications"]["jaso"] = row["jaso_classification"].split(",")
    record["specifications"]["jaso_family_detail"] = row["family_detail"]
    record["canonical_key"] += f"|jaso_oil_code:{normalize(row['oil_code'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    record["codes"]["jaso_oil_code"] = {
        "value": row["oil_code"],
        "source_id": row["source_id"],
        "status": "official_filed_registry",
    }
    return record


def licensed_record(row: dict) -> dict:
    generic = {
        "id": f"{row['source_id']}-{row['source_record_id']}",
        "source_number": row["source_record_id"],
        "brand": row["manufacturer"],
        "name": row["product_name"],
        "category": row["category"],
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": row.get("viscosity", ""),
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["manufacturer"]
    record["brand"] = row["manufacturer"]
    record["market"] = "EU_ECOLABEL" if row["source_id"] == "EU_ECOLABEL_LUBRICANTS" else "GLOBAL_OFFICIAL_LICENSED"
    record["source_id"] = row["source_id"]
    record["source_record_id"] = row["source_record_id"]
    record["source_row"] = row["source_row_number"]
    record["evidence_status"] = "official_licensed_registry"
    record["lifecycle_status"] = "listed_as_of_snapshot"
    record["snapshot_date"] = row["snapshot_date"]
    record["specifications"]["licensed_standard"] = row["specification"]
    record["specifications"]["certification_tags"] = row.get("certification_tags", [])
    if row.get("available_in"):
        record["specifications"]["available_in"] = row["available_in"]
    record["canonical_key"] += f"|official_registry:{normalize(row['source_id'])}:{normalize(row['source_record_id'])}:{normalize(row['product_name'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    if row.get("license_number"):
        system = (
            "GM_DEXOS_LICENSE" if row["source_id"].startswith("GM_")
            else "EU_ECOLABEL_LICENSE" if row["source_id"] == "EU_ECOLABEL_LUBRICANTS"
            else "NMMA_REGISTRATION"
        )
        record["codes"][system.lower()] = {
            "value": row["license_number"],
            "source_id": row["source_id"],
            "status": "official_licensed_registry",
        }
        record["certificate"]["number"] = row["license_number"]
        record["certificate"]["expires_at"] = row.get("expiration_date", "")
    for system, value in row.get("external_codes", {}).items():
        if value:
            record["codes"][system.lower()] = {
                "value": value,
                "source_id": row["source_id"],
                "status": "official_licensed_registry",
            }
    return record


def biopreferred_record(row: dict) -> dict:
    categories = row.get("categories", [])
    generic = {
        "id": f"USDA-BIOPREFERRED-{row['product_id']}",
        "source_number": row["product_id"],
        "brand": row["company"],
        "name": row["product_name"],
        "category": "; ".join(categories),
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "source": "usda-biopreferred",
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["company"]
    record["brand"] = row["company"]
    record["market"] = "US"
    record["source_id"] = "usda-biopreferred"
    record["source_record_id"] = row["product_id"]
    record["source_row"] = None
    record["evidence_status"] = "official_government_program_catalog"
    record["lifecycle_status"] = "listed_as_of_snapshot"
    record["snapshot_date"] = row["snapshot_date"]
    record["specifications"]["usda_biopreferred_categories"] = categories
    statuses = []
    if row.get("usda_certified_biobased"):
        statuses.append("USDA Certified Biobased")
    if row.get("mandatory_federal_purchasing"):
        statuses.append("USDA Mandatory Federal Purchasing category")
    record["specifications"]["usda_biopreferred_status"] = statuses
    record["canonical_key"] += f"|usda_biopreferred_product_id:{normalize(row['product_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    record["codes"]["usda_biopreferred_product_id"] = {
        "value": row["product_id"],
        "source_id": "usda-biopreferred",
        "status": "official_government_program_catalog",
    }
    return record


def zf_te_ml_record(row: dict) -> dict:
    viscosity = ""
    if row["family_code"] == "H":
        match = re.search(r"\b(?:ISO\s*)?(?:VG\s*)?([0-9]{2,3})\b", row["product_name"], re.I)
        if match:
            viscosity = match.group(1)
    generic = {
        "id": f"ZF-TE-ML-{row['approval_number']}",
        "source_number": row["approval_number"],
        "brand": row["manufacturer"],
        "name": row["product_name"],
        "category": "Смазочные материалы, одобренные ZF TE-ML",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "viscosity": viscosity,
        "source": "ZF_TE_ML",
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["manufacturer"]
    record["brand"] = row["manufacturer"]
    record["market"] = "GLOBAL_ZF_APPROVED"
    record["source_id"] = "ZF_TE_ML"
    record["source_record_id"] = row["approval_number"]
    record["source_row"] = None
    record["evidence_status"] = "official_oem_approval_registry"
    record["lifecycle_status"] = "approved_as_of_list_date"
    record["snapshot_date"] = row["list_date"]
    record["specifications"]["oem_approvals"] = [
        f"ZF {item['te_ml_sheet']} class {item['lubricant_class']}"
        for item in row["approval_occurrences"]
    ]
    record["specifications"]["zf_te_ml_sheets"] = row["te_ml_sheets"]
    record["specifications"]["zf_lubricant_classes"] = row["lubricant_classes"]
    record["specifications"]["product_name_aliases"] = row["product_name_aliases"]
    record["specifications"]["manufacturer_country"] = row["manufacturer_country"]
    record["specifications"]["licensed_standard"] = "ZF TE-ML"
    record["canonical_key"] += f"|zf_approval_number:{normalize(row['approval_number'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    record["codes"]["zf_approval_number"] = {
        "value": row["approval_number"],
        "source_id": "ZF_TE_ML",
        "status": "official_oem_approval_registry",
    }
    return record


def allison_record(row: dict) -> dict:
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["marketer_brand"],
        "name": row["product_name"],
        "category": "Жидкости, одобренные Allison Transmission",
        "category_code": "T",
        "family": FAMILY_NAMES["T"],
        "sae_class": row["product_name"],
        "source": "ALLISON_APPROVED_FLUIDS",
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["marketer_brand"]
    record["brand"] = row["marketer_brand"]
    record["market"] = "GLOBAL_ALLISON_APPROVED"
    record["source_id"] = "ALLISON_APPROVED_FLUIDS"
    record["source_record_id"] = row["source_record_id"]
    record["source_row"] = None
    record["evidence_status"] = "official_oem_approval_registry"
    record["lifecycle_status"] = "approved_as_of_list_date"
    record["snapshot_date"] = row["list_date"]
    record["specifications"]["oem_approvals"] = row["specifications"]
    record["specifications"]["allison_approval_numbers"] = row["approval_numbers"]
    record["specifications"]["licensed_standard"] = "Allison TES"
    record["canonical_key"] += f"|allison_registry_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, approval_number in enumerate(row["approval_numbers"], 1):
        record["codes"][f"allison_approval_number_{index}"] = {
            "system": "ALLISON_APPROVAL_NUMBER",
            "value": approval_number,
            "source_id": "ALLISON_APPROVED_FLUIDS",
            "status": "official_oem_approval_registry",
        }
    return record


def driventic_diwa_record(row: dict) -> dict:
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["marketer_brand"],
        "name": row["product_name"],
        "category": "Масла, одобренные для автоматических трансмиссий DIWA",
        "category_code": "T",
        "family": FAMILY_NAMES["T"],
        "source": "DRIVENTIC_DIWA_APPROVED_OILS",
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["marketer_brand"]
    record["brand"] = row["marketer_brand"]
    record["market"] = "GLOBAL_DRIVENTIC_DIWA_APPROVED"
    record["source_id"] = "DRIVENTIC_DIWA_APPROVED_OILS"
    record["source_record_id"] = row["source_record_id"]
    record["source_row"] = None
    record["evidence_status"] = "official_oem_approval_registry"
    record["lifecycle_status"] = "approved_as_of_current_published_list"
    record["snapshot_date"] = row["snapshot_date"]
    record["specifications"]["oem_approvals"] = ["Driventic DIWA approved oil"]
    record["specifications"]["diwa_oil_change_intervals_km"] = [str(value) for value in row["oil_change_intervals_km"]]
    record["specifications"]["driventic_publication_numbers"] = sorted({
        item["publication_number"] for item in row["approval_occurrences"]
    })
    record["specifications"]["licensed_standard"] = "Driventic DIWA"
    record["canonical_key"] += f"|driventic_diwa_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def mercedes_dtfr_record(row: dict) -> dict:
    sae_class = row["sae_grades"][0] if row["sae_grades"] else ""
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["company"],
        "name": row["product_name"],
        "category": "Эксплуатационные жидкости с допуском Mercedes-Benz Trucks DTFR",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": sae_class,
        "source": "MERCEDES_DTFR_APPROVED_FLUIDS",
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["company"]
    record["brand"] = row["company"]
    record["market"] = "GLOBAL_MERCEDES_DTFR_APPROVED"
    record["source_id"] = "MERCEDES_DTFR_APPROVED_FLUIDS"
    record["source_record_id"] = row["source_record_id"]
    record["source_row"] = None
    record["evidence_status"] = "official_oem_approval_registry"
    record["lifecycle_status"] = "historical_approval" if row["historical_only"] else "approved_as_of_current_registry"
    record["snapshot_date"] = row["snapshot_date"]
    record["specifications"]["oem_approvals"] = [
        f"Mercedes-Benz Trucks {sheet}" for sheet in row["dtfr_sheets"]
    ]
    record["specifications"]["mercedes_dtfr_sheets"] = row["dtfr_sheets"]
    record["specifications"]["sae_grades_source_reported"] = row["sae_grades"]
    record["specifications"]["licensed_standard"] = "Mercedes-Benz Trucks DTFR"
    record["canonical_key"] += f"|mercedes_dtfr_product_id:{normalize(row['dtfr_product_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    record["codes"]["mercedes_dtfr_product_id"] = {
        "system": "MERCEDES_DTFR_PRODUCT_ID",
        "value": row["dtfr_product_id"],
        "source_id": "MERCEDES_DTFR_APPROVED_FLUIDS",
        "status": "official_oem_approval_registry",
    }
    return record


def mercedes_bevo_record(row: dict) -> dict:
    sae_class = row["sae_grades"][0] if row["sae_grades"] else ""
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["company"],
        "name": row["product_name"],
        "category": "Эксплуатационные жидкости с допуском Mercedes-Benz BeVo",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": sae_class,
        "source": "MERCEDES_BENZ_BEVO_APPROVED_FLUIDS",
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["company"]
    record["brand"] = row["company"]
    record["market"] = "GLOBAL_MERCEDES_BEVO_APPROVED"
    record["source_id"] = "MERCEDES_BENZ_BEVO_APPROVED_FLUIDS"
    record["source_record_id"] = row["source_record_id"]
    record["source_row"] = None
    record["evidence_status"] = "official_oem_approval_registry"
    record["lifecycle_status"] = "historical_approval" if row["historical_only"] else "approved_as_of_current_registry"
    record["snapshot_date"] = row["snapshot_date"]
    record["specifications"]["oem_approvals"] = [
        f"Mercedes-Benz BeVo {sheet}" for sheet in row["bevo_sheets"]
    ]
    record["specifications"]["mercedes_bevo_sheets"] = row["bevo_sheets"]
    record["specifications"]["sae_grades_source_reported"] = row["sae_grades"]
    record["specifications"]["licensed_standard"] = "Mercedes-Benz BeVo"
    record["canonical_key"] += f"|mercedes_bevo_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, product_id in enumerate(row["bevo_product_ids"], 1):
        record["codes"][f"mercedes_bevo_product_id_{index}"] = {
            "system": "MERCEDES_BEVO_PRODUCT_ID",
            "value": product_id,
            "source_id": "MERCEDES_BENZ_BEVO_APPROVED_FLUIDS",
            "status": "official_oem_approval_registry",
        }
    return record


def merge_mercedes_bevo_evidence(target: dict, source_record: dict, raw: dict) -> None:
    specs = target["specifications"]
    specs["oem_approvals"] = sorted(set(specs.get("oem_approvals", [])) | {
        f"Mercedes-Benz BeVo {sheet}" for sheet in raw["bevo_sheets"]
    })
    specs["mercedes_bevo_sheets"] = sorted(
        set(specs.get("mercedes_bevo_sheets", [])) | set(raw["bevo_sheets"])
    )
    specs["sae_grades_source_reported"] = sorted(
        set(specs.get("sae_grades_source_reported", [])) | set(raw["sae_grades"])
    )
    for key, code in source_record["codes"].items():
        target["codes"][f"bevo_{raw['source_record_id']}_{key}"] = code
    if not raw["historical_only"]:
        target["lifecycle_status"] = "approved_as_of_current_registry"


def volvo_genuine_record(row: dict) -> dict:
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Оригинальные эксплуатационные жидкости Volvo",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": f"API {specs['api_gl']}" if specs.get("api_gl") else "",
        "grease_class": specs.get("nlgi", ""),
        "source": "VOLVO_GENUINE_FLUIDS",
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["manufacturer"]
    record["brand"] = row["brand"]
    record["market"] = row["market"]
    record["source_id"] = "VOLVO_GENUINE_FLUIDS"
    record["source_record_id"] = row["source_record_id"]
    record["source_row"] = None
    record["evidence_status"] = "official_manufacturer_product_catalog"
    record["lifecycle_status"] = "current_official_catalog"
    record["snapshot_date"] = row["snapshot_date"]
    record["specifications"].update(specs)
    record["specifications"]["source_url"] = row["source_url"]
    record["canonical_key"] += f"|volvo_genuine_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, part_number in enumerate(specs.get("part_numbers", []), 1):
        record["codes"][f"volvo_part_number_{index}"] = {
            "system": "VOLVO_PART_NUMBER",
            "value": part_number,
            "source_id": "VOLVO_GENUINE_FLUIDS",
            "status": "official_manufacturer_product_catalog",
        }
    return record


def mack_genuine_record(row: dict) -> dict:
    specs = row["specifications"]
    strict_mack_standards = [
        value for value in specs.get("mack_standards", []) if value != "37319"
    ]
    performance = []
    if specs.get("api"):
        performance.append(f"API {'/'.join(specs['api'])}")
    if specs.get("mack_standard"):
        performance.append(f"Mack {specs['mack_standard']}")
    if strict_mack_standards:
        performance.append(f"Mack {'/'.join(strict_mack_standards)}")
    if specs.get("mack_product_class"):
        performance.append(f"Mack {specs['mack_product_class']}")
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Оригинальные эксплуатационные жидкости Mack",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": " ".join(performance),
        "viscosity": specs.get("iso_vg", ""),
        "source": "MACK_GENUINE_FLUIDS",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "MACK_GENUINE_FLUIDS",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_url": row["source_url"],
        "product_overview_url": row["product_overview_url"],
        "engine_oil_standard_evidence_url": row["engine_oil_standard_evidence_url"],
        "safety_document_url_source_reported": row["safety_document_url_source_reported"],
        "source_quality_flags": row["source_quality_flags"],
        "product_sheet_part_numbers": row["product_sheet_part_numbers"],
        "product_sheet_url": row["product_sheet_url"],
    })
    if performance:
        record["canonical_key"] += f"|mack_professional_performance:{normalize('|'.join(performance))}"
    record["canonical_key"] += f"|mack_genuine_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    all_part_numbers = [package["part_number"] for package in row["packages"]]
    all_part_numbers.extend(
        number for number in row["product_sheet_part_numbers"] if number not in all_part_numbers
    )
    for index, part_number in enumerate(all_part_numbers, 1):
        record["codes"][f"mack_part_number_{index}"] = {
            "system": "MACK_PART_NUMBER",
            "value": part_number,
            "source_id": "MACK_GENUINE_FLUIDS",
            "status": "current_official_catalog",
        }
    return record


def mack_2014_approved_record(row: dict) -> dict:
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Historical Mack approved oils (Service Bulletin 175-61-08)",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": "API CJ-4" if specs.get("api") == ["CJ-4"] else "",
        "source": "MACK_2014_APPROVED_OILS",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "MACK_2014_APPROVED_OILS",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_oem_approval_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_url": row["source_url"],
        "source_document": row["source_document"],
        "source_document_date": row["source_document_date"],
        "product_name_source": row["product_name_source"],
        "source_occurrences": row["source_occurrences"],
        "source_quality_flags": row["source_quality_flags"],
    })
    sections = "|".join(specs["mack_approval_sections"])
    record["canonical_key"] += f"|mack_historical_approval:{normalize(sections)}"
    record["canonical_key"] += f"|mack_2014_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def cummins_valvoline_2022_record(row: dict) -> dict:
    specs = row["specifications"]
    performance = []
    if specs.get("api"):
        performance.append(f"API {'/'.join(specs['api'])}")
    if specs.get("acea"):
        performance.append(f"ACEA {'/'.join(specs['acea'])}")
    if specs.get("cummins_ces_source_reported"):
        performance.append(f"Cummins CES {'/'.join(specs['cummins_ces_source_reported'])}")
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Исторический официальный каталог Valvoline–Cummins Europe 2022",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": "; ".join(performance),
        "viscosity": specs.get("iso_vg", ""),
        "grease_class": specs.get("nlgi", ""),
        "source": "CUMMINS_VALVOLINE_EU_2022_CATALOG",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "CUMMINS_VALVOLINE_EU_2022_CATALOG",
        "source_record_id": row["source_record_id"],
        "source_row": row["source_page"],
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_url": row["source_url"],
        "source_document": row["source_document"],
        "source_document_date": row["source_document_date"],
        "source_page": row["source_page"],
        "source_quality_flags": row["source_quality_flags"],
    })
    record["canonical_key"] += f"|cummins_valvoline_2022_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, package in enumerate(row["packages"], 1):
        record["codes"][f"valvoline_article_number_{index}"] = {
            "system": "VALVOLINE_ARTICLE_NUMBER",
            "value": package["article_number"],
            "source_id": "CUMMINS_VALVOLINE_EU_2022_CATALOG",
            "status": "historical_official_catalog_2022_current_status_unverified",
        }
    return record


def taiwan_cpc_record(row: dict) -> dict:
    """Convert one current official CPC Taiwan product-sheet card."""
    specs = row["specifications"]
    performance = []
    for system, values in (("API", specs.get("api", [])), ("ACEA", specs.get("acea", [])), ("JASO", specs.get("jaso", []))):
        if values:
            performance.append(f"{system} {'/'.join(values)}")
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Current CPC Taiwan lubricant and technical-fluid catalog",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": "; ".join(performance),
        "viscosity": specs.get("iso_vg", ""),
        "grease_class": specs.get("nlgi", ""),
        "coolant_class": "/".join(specs.get("brake_fluid_classes", [])),
        "source": "TAIWAN_CPC_CURRENT_LUBRICANT_CATALOG",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "TAIWAN_CPC_CURRENT_LUBRICANT_CATALOG",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_url": row["source_url"],
        "product_sheet_url": row["product_sheet_url"],
        "product_sheet_data_url": row["product_sheet_data_url"],
        "product_sheet_update_date": row["product_sheet_update_date"],
        "product_sheet_data_sha256": row["product_sheet_data_sha256"],
        "packages_source_reported": row["packages_source_reported"],
        "source_quality_flags": row["source_quality_flags"],
    })
    record["canonical_key"] += f"|taiwan_cpc_product_code:{normalize(row['manufacturer_product_code'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    record["codes"]["cpc_taiwan_product_code"] = {
        "system": "CPC_TAIWAN_PRODUCT_CODE",
        "value": row["manufacturer_product_code"],
        "source_id": "TAIWAN_CPC_CURRENT_LUBRICANT_CATALOG",
        "status": "current_official_catalog",
    }
    return record


def volvo_websds_record(row: dict) -> dict:
    """Convert one conservative product/part identity from Volvo Group WebSDS news."""
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Official Volvo Group SDS change-list observation",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": "; ".join(
            specs.get("volvo_standards_source_reported", [])
            + specs.get("renault_trucks_standards_source_reported", [])
        ),
        "viscosity": specs.get("iso_vg", ""),
        "source": "VOLVO_GROUP_WEBSDS_CHANGE_ARCHIVE",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_sds_change_observation",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_url": row["source_url"],
        "first_observed_sds_date": row["first_observed_sds_date"],
        "last_observed_sds_date": row["last_observed_sds_date"],
        "change_types_source_reported": row["change_types_source_reported"],
        "product_name_source_aliases": row["product_name_source_aliases"],
        "supplier_organizations": row["supplier_organizations"],
        "source_occurrences": row["source_occurrences"],
        "source_quality_flags": row["source_quality_flags"],
    })
    record["canonical_key"] += f"|volvo_websds_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, part_number in enumerate(row["part_numbers"], 1):
        record["codes"][f"volvo_group_websds_part_number_{index}"] = {
            "system": "VOLVO_GROUP_WEBSDS_PART_NUMBER",
            "value": part_number,
            "source_id": row["source_id"],
            "status": "official_sds_change_observation_current_availability_unverified",
        }
    return record


def parker_denison_record(row: dict) -> dict:
    """Convert one Parker-Denison hydraulic-fluid rating row."""
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Parker-Denison current hydraulic-fluid rating list",
        "category_code": "H",
        "family": FAMILY_NAMES["H"],
        "api_class": "/".join(specs["parker_denison_hf_classes"]),
        # A rating row may cover several grades; keep the array in structured
        # evidence instead of asserting one strict-equivalence ISO VG value.
        "viscosity": "",
        "source": "PARKER_DENISON_CURRENT_FLUID_RATINGS",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": row["source_table_row"],
        "evidence_status": "official_oem_approval_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "rating_validity_until_source_reported": row["rating_validity_until_source_reported"],
        "rating_list_review_date": row["rating_list_review_date"],
        "source_url": row["source_url"],
        "source_document_url": row["source_document_url"],
        "source_page": row["source_page"],
        "source_quality_flags": row["source_quality_flags"],
    })
    record["canonical_key"] += f"|parker_denison_rating:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def scania_genuine_record(row: dict) -> dict:
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Оригинальные моторные масла Scania",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine", ""),
        "api_class": f"ACEA {'/'.join(specs.get('acea', []))}" if specs.get("acea") else "",
        "source": "SCANIA_GENUINE_ENGINE_OILS",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "SCANIA_GENUINE_ENGINE_OILS",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"]["source_url"] = row["source_url"]
    record["canonical_key"] += f"|scania_genuine_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def brava_official_record(row: dict) -> dict:
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Официальный каталог Brava Lubricants",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": " ".join(
            [f"API {'/'.join(specs.get('api', []))}" if specs.get("api") else ""]
            + [f"API {'/'.join(specs.get('api_gl', []))}" if specs.get("api_gl") else ""]
            + [f"ACEA {'/'.join(specs.get('acea', []))}" if specs.get("acea") else ""]
            + [f"ILSAC {'/'.join(specs.get('ilsac', []))}" if specs.get("ilsac") else ""]
        ).strip(),
        "viscosity": specs.get("iso_vg", ""),
        "coolant_class": specs.get("brake_fluid_class", ""),
        "source": "BRAVA_LUBRICANTS_OFFICIAL_CATALOG",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "BRAVA_LUBRICANTS_OFFICIAL_CATALOG",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_url": row["source_url"],
        "technical_document_url": row["technical_document_url"],
        "safety_document_url": row["safety_document_url"],
        "source_page_modified_at": row["source_page_modified_at"],
        "source_series": row["source_series"],
        "source_type": row["source_type"],
        "source_variant": row["source_variant"],
        "source_quality_flags": row["source_quality_flags"],
    })
    record["canonical_key"] += f"|brava_official_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, package in enumerate(row["packages"], 1):
        record["codes"][f"brava_part_number_{index}"] = {
            "system": "BRAVA_PART_NUMBER",
            "value": package["part_number"],
            "source_id": "BRAVA_LUBRICANTS_OFFICIAL_CATALOG",
            "status": "current_official_catalog",
        }
    return record


def ceypetco_record(row: dict) -> dict:
    """Convert one current official Ceypetco product-grade row."""
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Official Ceypetco lubricant and technical-fluid catalog",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": " ".join(
            [f"API {'/'.join(specs.get('api', []))}" if specs.get("api") else ""]
            + [f"API {'/'.join(specs.get('api_gl', []))}" if specs.get("api_gl") else ""]
            + [f"ACEA {'/'.join(specs.get('acea', []))}" if specs.get("acea") else ""]
            + [f"ILSAC {'/'.join(specs.get('ilsac', []))}" if specs.get("ilsac") else ""]
        ).strip(),
        "viscosity": specs.get("iso_vg", ""),
        "grease_class": specs.get("nlgi", ""),
        "coolant_class": specs.get("brake_fluid_class", ""),
        "source": "CEYPETCO_OFFICIAL_LUBRICANT_CATALOG",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "CEYPETCO_OFFICIAL_LUBRICANT_CATALOG",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_url": row["source_url"],
        "technical_document_url": row["technical_document_url"],
        "technical_document_sha256": row["technical_document_sha256"],
        "technical_document_issuer": row["technical_document_issuer"],
    })
    # These source conflicts must never leak into strict equivalence through a
    # guessed SAE or color value.
    if "conflicting_sae_within_current_tds" in specs.get("source_quality_flags", []):
        record["specifications"]["sae_engine"] = ""
        record["specifications"]["sae_gear"] = ""
    record["canonical_key"] += f"|ceypetco_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def pso_official_record(row: dict) -> dict:
    """Convert one current PSO series-grade fact row without inventing SKUs."""
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Official Pakistan State Oil lubricant and technical-fluid catalog",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": " ".join(
            [f"API {'/'.join(specs.get('api', []))}" if specs.get("api") else ""]
            + [f"API {'/'.join(specs.get('api_gl', []))}" if specs.get("api_gl") else ""]
            + [f"ACEA {'/'.join(specs.get('acea', []))}" if specs.get("acea") else ""]
            + [f"JASO {'/'.join(specs.get('jaso', []))}" if specs.get("jaso") else ""]
            + [f"ILSAC {'/'.join(specs.get('ilsac', []))}" if specs.get("ilsac") else ""]
        ).strip(),
        "viscosity": specs.get("iso_vg", ""),
        "grease_class": specs.get("nlgi", ""),
        "coolant_class": specs.get("brake_fluid_class", ""),
        "source": "PAKISTAN_STATE_OIL_OFFICIAL_CATALOG",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "PAKISTAN_STATE_OIL_OFFICIAL_CATALOG",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_url": row["source_url"],
        "technical_document_url": row["technical_document_url"],
        "technical_document_sha256": row["technical_document_sha256"],
        "safety_document_url": row["safety_document_url"],
        "source_series": row["source_series"],
        "source_grade": row["source_grade"],
        "source_grade_evidence": row["source_grade_evidence"],
        "source_grade_kind": row["source_grade_kind"],
        "source_quality_flags": row["source_quality_flags"],
        "packages_source_reported_at_series_level": row["packages_source_reported_at_series_level"],
        "publication_restriction": row["publication_restriction"],
    })
    record["canonical_key"] += f"|pso_official_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def merge_pso_catalog_evidence(target: dict, source_record: dict) -> None:
    """Attach exact manufacturer facts to the pre-existing DTFR identity."""
    target_specs = target["specifications"]
    source_specs = source_record["specifications"]
    for key in ("api", "api_gl", "acea", "jaso", "ilsac", "dexron", "oem_approvals"):
        target_specs[key] = sorted(set(target_specs.get(key, [])) | set(source_specs.get(key, [])))
    for key in (
        "source_url", "technical_document_url", "technical_document_sha256",
        "safety_document_url", "source_series", "source_grade",
        "source_grade_evidence", "source_grade_kind", "source_quality_flags",
        "packages_source_reported_at_series_level", "publication_restriction",
        "standards_and_approvals_source_reported",
    ):
        if key in source_specs:
            target_specs[f"pso_{key}"] = source_specs[key]


def pertamina_official_record(row: dict) -> dict:
    """Convert one current Pertamina product-grade fact without inventing a SKU."""
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Official Pertamina Lubricants current product catalog",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "api_class": " ".join(
            [f"API {'/'.join(specs.get('api', []))}" if specs.get("api") else ""]
            + [f"API {'/'.join(specs.get('api_gl', []))}" if specs.get("api_gl") else ""]
            + [f"ACEA {'/'.join(specs.get('acea', []))}" if specs.get("acea") else ""]
            + [f"JASO {'/'.join(specs.get('jaso', []))}" if specs.get("jaso") else ""]
            + [f"ILSAC {'/'.join(specs.get('ilsac', []))}" if specs.get("ilsac") else ""]
        ).strip(),
        "viscosity": specs.get("iso_vg", ""),
        "grease_class": specs.get("nlgi", ""),
        "coolant_class": specs.get("brake_fluid_class", ""),
        "source": "PERTAMINA_LUBRICANTS_OFFICIAL_CATALOG",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "PERTAMINA_LUBRICANTS_OFFICIAL_CATALOG",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update(specs)
    record["specifications"].update({
        "source_grade": row["source_grade"],
        "source_grade_kind": row["source_grade_kind"],
        "source_product_names": row["source_product_names"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_occurrences": row["source_occurrences"],
        "source_quality_flags": row["source_quality_flags"],
        "access_status": row["access_status"],
    })
    record["canonical_key"] += f"|pertamina_official_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def pertamina_grade_value(value) -> str:
    return re.sub(r"^(?:sae|iso vg|nlgi)\s+", "", normalize(value))


def pertamina_grade_signature(record: dict) -> tuple[str, ...]:
    specs = record["specifications"]
    return tuple(pertamina_grade_value(specs.get(key, "")) for key in (
        "sae_engine", "sae_gear", "iso_vg", "nlgi",
    )) + (normalize(specs.get("brake_fluid_class") or specs.get("coolant_class", "")),)


def pertamina_identity_surface(product_name: str) -> str:
    """Normalize labels while retaining the grade encoded in the source name."""
    value = normalize(product_name)
    value = re.split(r"\b(?:api|acea|jaso|ilsac)\b", value, maxsplit=1)[0]
    value = re.sub(r"\b(?:sae|iso vg|nlgi)\b", " ", value)
    # NPT occasionally prints the same grade on both sides of `ISO VG`.
    value = re.sub(r"\b(\d{1,4})\s+\1\b", r"\1", value)
    return re.sub(r"\s+", " ", value).strip()


def pertamina_identity_core(record: dict) -> str:
    value = pertamina_identity_surface(record["product_name_raw"])
    for grade in pertamina_grade_signature(record)[:4]:
        if grade and (value == grade or value.endswith(f" {grade}")):
            value = value[:-len(grade)].strip()
    return value


def is_pertamina_affiliated(record: dict) -> bool:
    return "pertamina" in normalize(record.get("manufacturer")) or "pertamina" in normalize(record.get("brand"))


def merge_pertamina_catalog_evidence(target: dict, source_record: dict, match_basis: str) -> None:
    """Attach current manufacturer facts to one unambiguous affiliated identity."""
    target_specs = target["specifications"]
    source_specs = source_record["specifications"]
    for key in ("api", "api_gl", "acea", "jaso", "ilsac", "oem_approvals"):
        target_specs[key] = sorted(set(target_specs.get(key, [])) | set(source_specs.get(key, [])))
    for key in ("sae_engine", "sae_gear", "iso_vg", "nlgi", "coolant_class"):
        if not target_specs.get(key) and source_specs.get(key):
            target_specs[key] = source_specs[key]
    for key in (
        "source_grade", "source_grade_kind", "source_product_names",
        "source_occurrence_count", "source_occurrences", "source_quality_flags",
        "standards_and_approvals_source_reported", "base_oil_source_reported",
        "access_status", "brake_fluid_class",
    ):
        if key in source_specs:
            target_specs[f"pertamina_{key}"] = source_specs[key]
    target_specs["pertamina_cross_source_match_basis"] = match_basis
    target["lifecycle_status"] = "listed_on_current_official_catalog"
    target["snapshot_date"] = SNAPSHOT_DATE


def man_service_record(row: dict) -> dict:
    specs = row["specifications"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Продукты из действующих рекомендаций MAN по эксплуатационным материалам",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": specs.get("sae_engine") or specs.get("sae_gear") or "",
        "viscosity": specs.get("iso_vg", ""),
        "grease_class": specs.get("nlgi", ""),
        "source": "MAN_CURRENT_SERVICE_PRODUCTS",
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["manufacturer"]
    record["brand"] = row["brand"]
    record["market"] = row["market"]
    record["source_id"] = "MAN_CURRENT_SERVICE_PRODUCTS"
    record["source_record_id"] = row["source_record_id"]
    record["source_row"] = None
    record["evidence_status"] = "official_oem_service_recommendation"
    record["lifecycle_status"] = "recommended_as_of_current_document"
    record["snapshot_date"] = row["snapshot_date"]
    record["specifications"].update(specs)
    record["specifications"]["application"] = row["application"]
    record["specifications"]["man_recommendation_document_date"] = row["document_date"]
    record["specifications"]["man_recommendation_pages"] = row["source_pages"]
    record["specifications"]["source_url"] = row["source_url"]
    record["canonical_key"] += f"|man_service_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def liqui_moly_catalog_record(row: dict) -> dict:
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"], "source_number": row["source_record_id"],
        "brand": row["brand"], "name": row["product_name"],
        "category": "Исторический официальный каталог LIQUI MOLY 2020",
        "category_code": row["family_code"], "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": (technical.get("sae_grades") or [""])[0],
        "api_class": "; ".join(technical.get("api", []) + technical.get("acea", [])),
        "viscosity": (technical.get("iso_vg") or [""])[0],
        "source": "LIQUI_MOLY_2020_PRODUCT_CATALOG",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"], "brand": row["brand"], "market": row["market"],
        "source_id": "LIQUI_MOLY_2020_PRODUCT_CATALOG", "source_record_id": row["source_record_id"],
        "source_row": None, "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"], "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "technical": technical, "document_date": row["document_date"], "source_pages": row["source_pages"],
        "part_numbers": row["part_numbers"], "package_rows": row["package_rows"],
        "source_url": row["source_url"], "catalog_page_url": row["catalog_page_url"],
    })
    record["canonical_key"] += f"|liqui_moly_2020_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, part_number in enumerate(row["part_numbers"], 1):
        record["codes"][f"liqui_moly_part_number_{index}"] = {"system": "LIQUI_MOLY_PART_NUMBER", "value": part_number, "source_id": "LIQUI_MOLY_2020_PRODUCT_CATALOG", "status": "historical_official_catalog"}
    return record


def merge_liqui_moly_evidence(target: dict, source_record: dict, raw: dict) -> None:
    target["specifications"].setdefault("liqui_moly_historical_catalog_entries", []).append({
        "document_date": raw["document_date"], "pages": raw["source_pages"], "technical": raw["technical"],
        "part_numbers": raw["part_numbers"], "package_rows": raw["package_rows"], "source_url": raw["source_url"],
    })
    for key, code in source_record["codes"].items():
        target["codes"][f"liqui_moly_{raw['source_record_id']}_{key}"] = code


def liqui_moly_current_record(row: dict) -> dict:
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Текущий официальный каталог LIQUI MOLY 2026",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": (technical.get("sae_grades") or [""])[0],
        "api_class": "; ".join(row["specifications"] + row["approvals"] + row["recommendations"]),
        "viscosity": (technical.get("iso_vg") or [""])[0],
        "source": "LIQUI_MOLY_CURRENT_OPENAPI",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": "LIQUI_MOLY_CURRENT_OPENAPI",
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_manufacturer_product_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "technical": technical,
        "source_specifications": row["specifications"],
        "source_approvals": row["approvals"],
        "source_recommendations": row["recommendations"],
        "liqui_moly_master_sku": row["master_sku"],
        "liqui_moly_articles": row["articles"],
        "classification_basis": row["classification_basis"],
        "lifecycle_assessment": row["lifecycle_assessment"],
        "historical_candidates": row["historical_candidates"],
        "source_url": row["source_url"],
        "sitemap_url": row["sitemap_url"],
        "api_base_url": row["api_base_url"],
    })
    record["canonical_key"] += f"|liqui_moly_current_master:{normalize(row['master_sku'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    record["codes"]["liqui_moly_master_sku"] = {
        "system": "LIQUI_MOLY_MASTER_SKU",
        "value": row["master_sku"],
        "source_id": "LIQUI_MOLY_CURRENT_OPENAPI",
        "status": "current_official_catalog",
    }
    for index, article in enumerate(row["articles"], 1):
        record["codes"][f"liqui_moly_article_sku_{index}"] = {
            "system": "LIQUI_MOLY_ARTICLE_SKU",
            "value": article["sku"],
            "source_id": "LIQUI_MOLY_CURRENT_OPENAPI",
            "status": "current_official_catalog",
        }
    return record


def merge_liqui_moly_current_evidence(target: dict, source_record: dict, raw: dict) -> None:
    target["specifications"].setdefault("liqui_moly_current_catalog_entries", []).append({
        "master_sku": raw["master_sku"],
        "product_name": raw["product_name"],
        "market": raw["market"],
        "technical": raw["technical"],
        "specifications": raw["specifications"],
        "approvals": raw["approvals"],
        "recommendations": raw["recommendations"],
        "articles": raw["articles"],
        "lifecycle_assessment": raw["lifecycle_assessment"],
        "source_url": raw["source_url"],
    })
    target["lifecycle_status"] = "listed_as_of_current_official_catalog"
    target["snapshot_date"] = raw["snapshot_date"]
    existing_codes = {
        (code.get("system", key).upper(), code["value"])
        for key, code in target["codes"].items()
    }
    for key, code in source_record["codes"].items():
        if (code["system"], code["value"]) not in existing_codes:
            target["codes"][f"liqui_moly_current_{raw['master_sku']}_{key}"] = code
            existing_codes.add((code["system"], code["value"]))


def anp_brazil_record(row: dict) -> dict:
    """Convert one normalized ANP regulatory product-grade row to catalog form."""
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["registration_holder"],
        "name": row["product_name"],
        "category": "Действующий государственный реестр масел и смазок ANP Бразилии",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": row["sae"],
        "api_class": row["performance_level"],
        "viscosity": row["iso_vg"],
        "grease_class": row["nlgi"],
        "source": "ANP_BRAZIL_LUBRICANT_REGISTRY",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["producer"] or row["registration_holder"],
        "brand": row["registration_holder"],
        "market": row["market"],
        "source_id": "ANP_BRAZIL_LUBRICANT_REGISTRY",
        "source_record_id": row["source_record_id"],
        "source_row": row["source_row_numbers"][0],
        "evidence_status": "official_government_regulatory_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "anp_registration_status": row["registration_status"],
        "anp_registration_holder": row["registration_holder"],
        "anp_registration_holder_cnpj": row["registration_holder_cnpj"],
        "anp_company_type": row["company_type"],
        "anp_producer": row["producer"],
        "anp_origin": row["origin"],
        "anp_product_type": row["product_type"],
        "anp_regulatory_purpose": row["regulatory_purpose"],
        "application": row["application"],
        "base_oil_composition": row["base_oil_composition"],
        "performance_raw": row["performance_level"],
        "anp_process_numbers": row["process_numbers"],
        "anp_registration_years": row["registration_years"],
        "packages": row["packages"],
        "classification_basis": row["classification_basis"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_row_numbers": row["source_row_numbers"],
        "source_url": row["source_url"],
        "metadata_url": row["metadata_url"],
    })
    record["canonical_key"] += f"|anp_brazil_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    record["codes"]["anp_registration_number"] = {
        "system": "ANP_BRAZIL_REGISTRATION_NUMBER",
        "value": row["registration_number"],
        "source_id": "ANP_BRAZIL_LUBRICANT_REGISTRY",
        "status": "active_in_current_anp_snapshot",
    }
    return record


def anp_monitoring_performance_key(value: str) -> str:
    """Return the comparable performance token printed in an ANP PML annex."""
    normalized = normalize(value)
    if normalized in {"", "n a", "na", "acea"}:
        return ""
    words = normalized.split()
    if words and words[0] in {"api", "acea", "jaso", "ilsac"}:
        words = words[1:]
    return " ".join(words)


def anp_monitoring_current_performance_keys(value: str) -> set[str]:
    return {
        key
        for part in text(value).split(";")
        if (key := anp_monitoring_performance_key(part))
    }


def anp_monitoring_identity_key(row: dict) -> tuple[str, ...]:
    registration = normalize(row["registration_number"])
    base = (
        registration,
        normalize(row["product_name"]),
        normalize(row["sae"]),
        anp_monitoring_performance_key(row["performance_level"]),
    )
    # A registration number is the stable public identity.  For unregistered
    # samples the holder is required to avoid merging unrelated same-name oils.
    return base if registration else base + (normalize(row["registration_holder"]),)


def anp_monitoring_history_record(identity: tuple[str, ...], occurrences: list[dict]) -> dict:
    row = occurrences[0]
    family_code = row.get("family_hint") or ("T" if anp_monitoring_performance_key(row["performance_level"]) == "gl 4" else "M")
    source_id = row["source_id"]
    holder = row["registration_holder"] or "ANP sample — holder not reported"
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": holder,
        "name": row["product_name"],
        "category": "Исторический мониторинг масел ANP Бразилии",
        "category_code": family_code,
        "family": FAMILY_NAMES[family_code],
        "sae_class": row["sae"],
        "api_class": row["performance_level"],
        "source": source_id,
    }
    record = canonical_record(generic)
    fingerprint = hashlib.sha256("|".join(identity).encode()).hexdigest()[:18]
    record.update({
        "manufacturer": holder,
        "brand": holder,
        "market": "Brazil",
        "source_id": source_id,
        "source_record_id": f"ANP-PML-GRADE-{fingerprint}",
        "source_row": row.get("source_row", row.get("source_page")),
        "evidence_status": "official_government_historical_market_monitoring",
        "lifecycle_status": "historical_sample_not_matched_to_current_anp_registry_grade",
        "snapshot_date": row["snapshot_date"],
    })
    record["canonical_key"] += f"|anp_monitoring_history:{fingerprint}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    if row["registration_number"]:
        record["codes"]["anp_historical_registration_number"] = {
            "system": "ANP_BRAZIL_REGISTRATION_NUMBER",
            "value": row["registration_number"],
            "source_id": source_id,
            "status": "historically_observed_not_currently_matched_grade",
        }
    return record


def merge_anp_monitoring_evidence(target: dict, occurrences: list[dict], match_basis: str, candidate_count: int) -> None:
    years = sorted({row["issue_year"] for row in occurrences if row["issue_year"]})
    sample_ids = sorted({row["sample_id"] for row in occurrences if row["sample_id"]})
    flags = sorted({flag for row in occurrences for flag in row["quality_flags"]})
    holders = sorted(
        {row["registration_holder"] for row in occurrences if row["registration_holder"]},
        key=lambda value: (value.casefold(), value),
    )
    files = sorted({row["source_url"] for row in occurrences})
    source_ids = sorted({row["source_id"] for row in occurrences})
    published_scopes = sorted({row["published_scope"] for row in occurrences if row.get("published_scope")})
    target["specifications"].setdefault("anp_monitoring_history", []).append({
        "reported_product_name": occurrences[0]["product_name"],
        "reported_registration_number": occurrences[0]["registration_number"],
        "reported_sae": occurrences[0]["sae"],
        "reported_performance_level": occurrences[0]["performance_level"],
        "reported_holders": holders,
        "issue_years": years,
        "first_issue_year": years[0] if years else None,
        "last_issue_year": years[-1] if years else None,
        "observation_count": len(occurrences),
        "sample_ids": sample_ids,
        "sample_ids_omitted_count": len(occurrences) - len(sample_ids),
        "quality_flags": flags,
        "source_files": files,
        "source_ids": source_ids,
        "published_scopes": published_scopes,
        "match_basis": match_basis,
        "current_registry_candidate_count": candidate_count,
        "source_page_url": occurrences[0]["source_page_url"],
    })


def indonesia_npt_record(row: dict) -> dict:
    """Convert one published Indonesian NPT table row to catalog form."""
    technical = row["technical"]
    performance = "; ".join(
        f"{system.upper()} {value}"
        for system in ("api", "acea", "jaso")
        for value in technical[system]
    )
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["company"],
        "name": row["product_name"],
        "category": "Государственный перечень смазочных материалов NPT Индонезии 2021–2025",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae"][0] if technical["sae"] else "",
        "api_class": performance,
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "source": "INDONESIA_NPT_LUBRICANT_REGISTRY",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["company"],
        "brand": row["company"],
        "market": row["market"],
        "source_id": "INDONESIA_NPT_LUBRICANT_REGISTRY",
        "source_record_id": row["source_record_id"],
        "source_row": row["source_pdf_page"],
        "evidence_status": (
            "official_government_regulatory_registry"
            if row["registration_number"]
            else "official_government_registry_source_data_issue"
        ),
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "indonesia_npt_class_code": row["npt_class_code"],
        "indonesia_npt_registration_number_raw": row["registration_number_raw"],
        "indonesia_npt_registration_number_status": row["registration_number_status"],
        "indonesia_npt_expiry_raw": row["expiry_raw"],
        "valid_through": row["valid_through"],
        "lifecycle_basis": row["lifecycle_basis"],
        "classification_basis": row["classification_basis"],
        "sae_source_reported": technical["sae"],
        "iso_vg_source_reported": technical["iso_vg"],
        "nlgi_source_reported": technical["nlgi"],
        "api_source_reported": technical["api"],
        "acea_source_reported": technical["acea"],
        "jaso_source_reported": technical["jaso"],
        "source_document_date": row["source_document_date"],
        "source_pdf_page": row["source_pdf_page"],
        "source_url": row["source_url"],
    })
    record["canonical_key"] += f"|indonesia_npt_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    if row["registration_number"]:
        record["codes"]["indonesia_npt_registration_number"] = {
            "system": "INDONESIA_NPT_REGISTRATION_NUMBER",
            "value": row["registration_number"],
            "source_id": "INDONESIA_NPT_LUBRICANT_REGISTRY",
            "status": row["lifecycle_status"],
        }
    return record


def thailand_doeb_record(row: dict) -> dict:
    """Convert one Thai DOEB motor-lubricant registration occurrence."""
    technical = row["technical"]
    performance = "; ".join(
        f"{system.upper()} {value}"
        for system in ("api", "acea", "jaso", "ilsac", "oem", "nmma")
        for value in technical[system]
    )
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["registration_holder"],
        "name": row["product_name"],
        "category": "Государственный реестр моторных масел DOEB Таиланда",
        "category_code": "M",
        "family": FAMILY_NAMES["M"],
        "sae_class": technical["sae"][0],
        "api_class": performance,
        "viscosity": "",
        "grease_class": "",
        "source": "THAILAND_DOEB_LUBRICANT_REGISTRY",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["registration_holder"],
        "brand": row["registration_holder"],
        "market": row["market"],
        "source_id": "THAILAND_DOEB_LUBRICANT_REGISTRY",
        "source_record_id": row["source_record_id"],
        "source_row": row["source_row"],
        "evidence_status": "official_government_regulatory_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "thailand_doeb_registration_holder": row["registration_holder"],
        "thailand_doeb_registration_number_raw_values": row["registration_number_raw_values"],
        "sae_source_raw_values": row["sae_source_raw_values"],
        "sae_source_reported": technical["sae"],
        "api_source_reported": technical["api"],
        "acea_source_reported": technical["acea"],
        "jaso_source_reported": technical["jaso"],
        "ilsac_source_reported": technical["ilsac"],
        "oem_source_reported": technical["oem"],
        "nmma_source_reported": technical["nmma"],
        "standards_source_reported": row["standards"],
        "valid_through": row["valid_through"],
        "lifecycle_basis": row["lifecycle_basis"],
        "classification_basis": row["classification_basis"],
        "source_quality_flags": row["source_quality_flags"],
        "source_occurrences": row["source_occurrences"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_rows": row["source_rows"],
        "source_snapshot_month": row["source_snapshot_month"],
        "source_dataset_url": row["source_dataset_url"],
        "source_mirror_url": row["source_mirror_url"],
        "source_resource_id": row["source_resource_id"],
    })
    record["canonical_key"] += f"|thailand_doeb_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, registration_number in enumerate(row["registration_numbers"], 1):
        record["codes"][f"thailand_doeb_registration_number_{index}"] = {
            "system": "THAILAND_DOEB_REGISTRATION_NUMBER",
            "value": registration_number,
            "source_id": "THAILAND_DOEB_LUBRICANT_REGISTRY",
            "status": row["lifecycle_status"],
        }
    return record


def dla_qpd_record(row: dict) -> dict:
    """Convert one normalized DLA qualification identity to catalog form."""
    source_id = row["source_id"]
    technical = row["technical"]
    qualification_names = sorted({
        value
        for approval in row["qualifications"]
        for value in (approval["qpl_number"], approval["document_id"], approval["government_designation"])
        if value
    })
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["company"],
        "name": row["product_name"],
        "category": f"Официальный реестр квалифицированной продукции DLA, {source_id.replace('DLA_QPD_FSC_', 'FSC ').replace('_LUBRICANT_SCOPE', ' (отобранный профильный сегмент)')}",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae"][0] if technical["sae"] else "",
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "source": source_id,
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["company"],
        "brand": row["company"],
        "market": row["market"],
        "source_id": source_id,
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_government_qualified_product_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "dla_qpd_qualifications": qualification_names,
        "dla_qpd_qualification_records": row["qualifications"],
        "dla_qpd_certified_statuses": row["certified_statuses"],
        "dla_qpd_sam_statuses": row["sam_statuses"],
        "dla_qpd_stop_ship_values": row["stop_ship_values"],
        "dla_qpd_source_types": row["source_types"],
        "dla_qpd_manufacturer_designations_raw": row.get("manufacturer_designations_raw", [row["product_name"]]),
        "dla_qpd_designation_qualifiers": row.get("designation_qualifiers", []),
        "cage_codes": row["cage_codes"],
        "nato_codes": technical["nato_codes"],
        "sae_source_reported": technical["sae"],
        "iso_vg_source_reported": technical["iso_vg"],
        "nlgi_source_reported": technical["nlgi"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_url": row["source_url"],
        "help_url": row["help_url"],
    })
    record["canonical_key"] += f"|dla_qpd_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, qpl_number in enumerate(sorted({qualification["qpl_number"] for qualification in row["qualifications"]})):
        record["codes"][f"dla_qpl_{index}"] = {
            "system": "DLA_QPL_NUMBER",
            "value": qpl_number,
            "source_id": source_id,
            "status": row["lifecycle_status"],
        }
    return record


def blue_angel_record(row: dict) -> dict:
    """Convert one current Blue Angel DE-UZ 178 product to catalog form."""
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["manufacturer"],
        "name": row["product_name"],
        "category": "Blue Angel DE-UZ 178: " + "; ".join(row["official_categories"]),
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae"][0] if technical["sae"] else "",
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "source": "BLUE_ANGEL_DE_UZ_178",
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["manufacturer"],
        "market": row["market"],
        "source_id": "BLUE_ANGEL_DE_UZ_178",
        "source_record_id": row["source_record_id"],
        "source_row": row["source_row_numbers"][0],
        "evidence_status": "official_ecolabel_product_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "blue_angel_standard": row["certification_standard"],
        "blue_angel_categories": row["official_categories"],
        "blue_angel_product_page_urls": row["product_page_urls"],
        "product_external_urls": row["product_external_urls"],
        "classification_basis": row["classification_basis"],
        "iso_vg_source_reported": technical["iso_vg"],
        "nlgi_source_reported": technical["nlgi"],
        "sae_source_reported": technical["sae"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_row_numbers": row["source_row_numbers"],
        "source_export_facts_sha256": row["source_export_facts_sha256"],
        "source_page_facts_sha256": row["source_page_facts_sha256"],
        "category_page_sha256": row["category_page_sha256"],
        "criteria_url": row["criteria_url"],
    })
    record["canonical_key"] += f"|blue_angel_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, url in enumerate(row["product_page_urls"]):
        record["codes"][f"blue_angel_product_{index}"] = {
            "system": "BLUE_ANGEL_PRODUCT_PAGE",
            "value": url,
            "source_id": "BLUE_ANGEL_DE_UZ_178",
            "status": row["lifecycle_status"],
        }
    return record


def merge_blue_angel_evidence(target: dict, source_record: dict, raw: dict) -> None:
    target["specifications"].setdefault("blue_angel_de_uz_178", []).append({
        "source_record_id": raw["source_record_id"],
        "categories": raw["official_categories"],
        "product_page_urls": raw["product_page_urls"],
        "snapshot_date": raw["snapshot_date"],
    })
    existing_codes = {(code.get("system", key).upper(), code["value"]) for key, code in target["codes"].items()}
    for key, code in source_record["codes"].items():
        identity = (code.get("system", key).upper(), code["value"])
        if identity not in existing_codes:
            target["codes"][f"blue_angel_{raw['source_record_id']}_{key}"] = code
            existing_codes.add(identity)


def austrian_ecolabel_uz14_record(row: dict) -> dict:
    """Convert one current Austrian Ecolabel UZ 14 product to catalog form."""
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["licence_number"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": row["certification_standard"],
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae"][0] if technical["sae"] else "",
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": row["source_row"],
        "evidence_status": "official_government_ecolabel_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "austrian_ecolabel_standard": row["certification_standard"],
        "austrian_ecolabel_licence_number": row["licence_number"],
        "austrian_ecolabel_licence_holder": row["licence_holder"],
        "austrian_ecolabel_holder_country_source": row["holder_country_source"],
        "austrian_ecolabel_brand_basis": row["brand_basis"],
        "austrian_ecolabel_classification_basis": row["classification_basis"],
        "source_url": row["source_url"],
        "directory_url": row["directory_url"],
        "licensee_export_url": row["licensee_export_url"],
    })
    record["codes"]["austrian_ecolabel_uz14_certificate"] = {
        "system": "AUSTRIAN_ECOLABEL_UZ14_CERTIFICATE",
        "value": row["licence_number"],
        "source_id": row["source_id"],
        "status": row["lifecycle_status"],
    }
    record["canonical_key"] += f"|austrian_uz14:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def merge_austrian_ecolabel_uz14_evidence(target: dict, source_record: dict, raw: dict) -> None:
    target["specifications"].setdefault("austrian_ecolabel_uz14", []).append({
        "source_record_id": raw["source_record_id"],
        "licence_number": raw["licence_number"],
        "licence_holder": raw["licence_holder"],
        "asserted_family_code": raw["family_code"],
        "classification_basis": raw["classification_basis"],
        "source_url": raw["source_url"],
        "snapshot_date": raw["snapshot_date"],
    })
    existing_codes = {(code.get("system", key).upper(), code["value"]) for key, code in target["codes"].items()}
    for key, code in source_record["codes"].items():
        identity = (code.get("system", key).upper(), code["value"])
        if identity not in existing_codes:
            target["codes"][f"austrian_uz14_{raw['source_record_id']}_{key}"] = code
            existing_codes.add(identity)


def korea_ecolabel_record(row: dict) -> dict:
    """Convert one official Korea Eco-Label product to catalog form."""
    source_id = row["source_id"]
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["certificate_number"],
        "brand": row["manufacturer"],
        "name": row["product_name"],
        "category": f"Korea Eco-Label {row['official_category_code']}: {row['official_use']}",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae"][0] if technical["sae"] else "",
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "source": source_id,
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["manufacturer"],
        "market": row["market"],
        "source_id": source_id,
        "source_record_id": row["source_record_id"],
        "source_row": row["source_row_numbers"][0],
        "evidence_status": "official_government_ecolabel_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
    })
    record["specifications"].update({
        "korea_ecolabel_category_code": row["official_category_code"],
        "korea_ecolabel_category_name": row["official_category_name"],
        "korea_ecolabel_official_use": row["official_use"],
        "korea_ecolabel_initial_certification_date": row["initial_certification_date"],
        "korea_ecolabel_certification_start_date": row["certification_start_date"],
        "korea_ecolabel_certification_end_date": row["certification_end_date"],
        "source_product_names": row.get("source_product_names", [row["product_name"]]),
        "source_model_names": row.get("source_model_names", []),
        "classification_basis": row["classification_basis"],
        "iso_vg_source_reported_or_name_explicit": technical["iso_vg"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_row_numbers": row["source_row_numbers"],
        "packages": row["packages"],
        "source_url": row["source_url"],
    })
    record["codes"]["korea_ecolabel_certificate"] = {
        "system": "KOREA_ECOLABEL_CERTIFICATE",
        "value": row["certificate_number"],
        "source_id": source_id,
        "status": row["lifecycle_status"],
    }
    record["canonical_key"] += f"|korea_ecolabel_record:{normalize(source_id)}:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def merge_korea_ecolabel_evidence(target: dict, source_record: dict, raw: dict) -> None:
    target["specifications"].setdefault("korea_ecolabel_el611", []).append({
        "source_record_id": raw["source_record_id"],
        "certificate_number": raw["certificate_number"],
        "official_use": raw["official_use"],
        "certification_end_date": raw["certification_end_date"],
        "dataset_snapshot_date": raw["dataset_snapshot_date"],
    })
    existing_codes = {(code.get("system", key).upper(), code["value"]) for key, code in target["codes"].items()}
    for key, code in source_record["codes"].items():
        identity = (code.get("system", key).upper(), code["value"])
        if identity not in existing_codes:
            target["codes"][f"korea_ecolabel_{raw['source_record_id']}_{key}"] = code
            existing_codes.add(identity)


def green_choice_philippines_record(row: dict) -> dict:
    """Convert one historical Green Choice Philippines engine-oil licence."""
    technical = row["technical"]
    validated_sae = technical["sae_validated"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["certificate_number"],
        "brand": row["manufacturer"],
        "name": row["product_name"],
        "category": row["official_criterion"],
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": validated_sae[0] if validated_sae else "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["manufacturer"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": row["source_row_number"],
        "evidence_status": "official_government_ecolabel_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "green_choice_philippines_criterion": row["official_criterion"],
        "green_choice_philippines_criterion_code": row["official_criterion_code"],
        "green_choice_philippines_sector": row["official_sector"],
        "sae_source_reported": technical["sae_source_reported"],
        "sae_source_validated": validated_sae,
        "source_quality_flags": technical["source_quality_flags"],
        "classification_basis": row["classification_basis"],
        "source_url": row["source_url"],
        "program_url": row["program_url"],
        "criteria_url": row["criteria_url"],
    })
    record["codes"]["green_choice_philippines_licence"] = {
        "system": "GREEN_CHOICE_PHILIPPINES_LICENCE",
        "value": row["certificate_number"],
        "source_id": row["source_id"],
        "status": row["lifecycle_status"],
    }
    return record


def uae_moiat_record(row: dict) -> dict:
    """Convert one normalized UAE MOIAT product-conformity identity."""
    technical = row["technical"]
    performance = "; ".join(technical["api"] + technical["api_gl"] + technical["acea"] + technical["jaso"])
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "UAE MOIAT Product Conformity",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae"][0] if technical["sae"] else (technical["sae_gear"][0] if technical["sae_gear"] else ""),
        "api_class": performance,
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": "AE",
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_government_product_conformity_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
    })
    record["specifications"].update({
        "sae_source_reported": technical["sae"],
        "sae_gear_source_reported": technical["sae_gear"],
        "api_source_reported": technical["api"],
        "api_gl_source_reported": technical["api_gl"],
        "acea_source_reported": technical["acea"],
        "jaso_source_reported": technical["jaso"],
        "iso_vg_source_reported": technical["iso_vg"],
        "nlgi_source_reported": technical["nlgi"],
        "dot_source_reported": technical["dot"],
        "classification_basis": row["classification_basis"],
        "source_models": row["source_models"],
        "source_descriptions": row["source_descriptions"],
        "barcodes": row["barcodes"],
        "packages": row["packages"],
        "certificate_entries": row["certificate_entries"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_url": row["source_url"],
        "source_rights_url": row["source_rights_url"],
    })
    seen_certificate_values = set()
    for index, certificate in enumerate(row["certificate_entries"], 1):
        certificate_value = certificate["certificate_number"] or certificate["certificate_id"]
        if certificate_value in seen_certificate_values:
            continue
        seen_certificate_values.add(certificate_value)
        record["codes"][f"uae_moiat_certificate_{index}"] = {
            "system": "UAE_MOIAT_CERTIFICATE",
            "value": certificate_value,
            "source_id": row["source_id"],
            "status": row["lifecycle_status"],
        }
    record["canonical_key"] += f"|uae_moiat_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def eaeu_conformity_record(row: dict) -> dict:
    """Convert one conservative product identity from official EAEU open data."""
    technical = row["specifications"]
    performance = "; ".join(
        technical.get("api", [])
        + technical.get("api_gl", [])
        + technical.get("acea", [])
        + technical.get("jaso", [])
    )
    sae_engine = technical.get("sae_engine", [])
    sae_gear = technical.get("sae_gear", [])
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "EAEU conformity document product evidence",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": (sae_engine + sae_gear + [""])[0],
        "api_class": performance,
        "viscosity": (technical.get("iso_vg", []) + [""])[0],
        "grease_class": (technical.get("nlgi", []) + [""])[0],
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_government_product_conformity_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "sae_engine_source_reported": sae_engine,
        "sae_gear_source_reported": sae_gear,
        "api_source_reported": technical.get("api", []),
        "api_gl_source_reported": technical.get("api_gl", []),
        "acea_source_reported": technical.get("acea", []),
        "jaso_source_reported": technical.get("jaso", []),
        "iso_vg_source_reported": technical.get("iso_vg", []),
        "nlgi_source_reported": technical.get("nlgi", []),
        "brake_fluid_class_source_reported": technical.get("brake_fluid_class", []),
        "atf_specifications_source_reported": technical.get("atf_specifications", []),
        "brand_basis": row["brand_basis"],
        "family_basis": row["family_basis"],
        "manufacturer_candidates": row["manufacturer_candidates"],
        "technical_regulations": row["technical_regulations"],
        "standards_and_evidence_documents": row["standards_and_evidence_documents"],
        "certificate_occurrence_count": row["certificate_occurrence_count"],
        "certificate_evidence": row["certificate_evidence"],
        "source_url": row["source_url"],
        "source_registry_url": row["source_registry_url"],
        "source_rights_url": row["source_rights_url"],
    })
    for index, code in enumerate(row["tnved_codes"], 1):
        record["codes"][f"eaeu_tnved_{index}"] = {
            "system": "TNVED",
            "value": code,
            "source_id": row["source_id"],
            "status": "source_reported_official_conformity_evidence",
        }
    # The conformity register is evidence rather than a manufacturer master.
    # Keep its conservative identity independent until a cross-source review
    # proves formula/grade equivalence.
    record["canonical_key"] += f"|eaeu_conformity_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def epa_safer_choice_record(row: dict) -> dict:
    """Convert one explicitly named EPA Safer Choice lubricant product."""
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "EPA Safer Choice — special lubricant",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": "",
        "api_class": "",
        "viscosity": "",
        "grease_class": "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": "US",
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_government_program_catalog",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
    })
    record["specifications"].update({
        "epa_safer_choice_program": row["program"],
        "epa_safer_choice_categories": row["categories"],
        "epa_safer_choice_sectors": row["sectors"],
        "epa_safer_choice_partner_since": row["partner_since"],
        "epa_safer_choice_company_in_good_standing": row["company_in_good_standing"],
        "classification_basis": row["classification_basis"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_url": row["source_url"],
        "product_url": row["product_url"],
    })
    for system, values in (("UPC", row["upcs"]), ("GTIN", row["gtins"]), ("MPN", row["mpns"])):
        for index, value in enumerate(values, 1):
            record["codes"][f"epa_safer_choice_{system.lower()}_{index}"] = {
                "system": system,
                "value": value,
                "source_id": row["source_id"],
                "status": row["lifecycle_status"],
            }
    record["canonical_key"] += f"|epa_safer_choice_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def epa_chemexpo_record(row: dict) -> dict:
    """Convert one conservative EPA ChemExpo/CPDat product identity."""
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "US EPA ChemExpo/CPDat product-use evidence",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae"][0] if technical["sae"] else "",
        "api_class": "; ".join(f"API {value}" for value in technical["api"]),
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "coolant_class": technical["brake_fluid_class"][0] if technical["brake_fluid_class"] else "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": row["evidence_status"],
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
    })
    record["specifications"].update({
        "epa_chemexpo_puc_evidence": row["puc_evidence"],
        "epa_chemexpo_source_product_names": row["source_product_names"],
        "epa_chemexpo_package_and_source_name_variants": row["package_and_source_name_variants"],
        "epa_chemexpo_brand_source_reported": row["brand_source_reported"],
        "epa_chemexpo_brand_basis": row["brand_basis"],
        "epa_chemexpo_source_quality_flags": row["source_quality_flags"],
        "epa_chemexpo_source_occurrence_count": row["source_occurrence_count"],
        "sae_source_reported_in_name": technical["sae"],
        "api_source_reported_in_name": technical["api"],
        "iso_vg_source_reported_in_name": technical["iso_vg"],
        "nlgi_source_reported_in_name": technical["nlgi"],
        "brake_fluid_class_source_reported_in_name": technical["brake_fluid_class"],
        "atf_specifications_source_reported_in_name": technical["atf"],
        "source_url": row["source_page_url"],
        "source_get_data_url": row["source_get_data_url"],
        "source_puc_urls": row["source_puc_urls"],
        "source_product_urls": row["source_product_urls"],
    })
    for index, product_id in enumerate(row["source_product_ids"], 1):
        record["codes"][f"epa_chemexpo_product_id_{index}"] = {
            "system": "EPA_CHEMEXPO_PRODUCT_ID",
            "value": product_id,
            "source_id": row["source_id"],
            "status": row["lifecycle_status"],
        }
    record["canonical_key"] += f"|epa_chemexpo_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def merge_epa_chemexpo_evidence(target: dict, source_record: dict, raw: dict) -> None:
    target["specifications"].setdefault("epa_chemexpo_cpdat_evidence", []).append({
        "source_record_id": raw["source_record_id"],
        "source_product_ids": raw["source_product_ids"],
        "source_product_names": raw["source_product_names"],
        "puc_evidence": raw["puc_evidence"],
        "source_product_urls": raw["source_product_urls"],
        "dataset_snapshot_date": raw["dataset_snapshot_date"],
        "lifecycle_status": raw["lifecycle_status"],
    })
    existing_codes = {(code.get("system", key).upper(), code["value"]) for key, code in target["codes"].items()}
    for key, code in source_record["codes"].items():
        identity = (code.get("system", key).upper(), code["value"])
        if identity not in existing_codes:
            target["codes"][f"epa_chemexpo_{raw['source_record_id']}_{key}"] = code
            existing_codes.add(identity)


def psqca_engine_oil_record(row: dict) -> dict:
    """Convert one PSQCA CM licence at its conservative brand-scope grain."""
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Pakistan PSQCA CM licence — certified engine-oil brand scope",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": "",
        "api_class": "",
        "viscosity": "",
        "grease_class": "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": row["evidence_status"],
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
        "certificate_status": row["lifecycle_status"],
    })
    record["specifications"].update({
        "psqca_certified_standard": technical["certified_standard"],
        "psqca_certified_product_scope": row["certified_product_scope"],
        "psqca_search_product_scope": row["search_product_scope"],
        "psqca_product_name_basis": row["product_name_basis"],
        "psqca_source_quality_flags": row["source_quality_flags"],
        "sae_source_reported": technical["sae"],
        "api_source_reported": technical["api"],
        "source_url": row["source_url"],
        "standard_url": row["standard_url"],
        "source_facts_sha256": row["source_facts_sha256"],
    })
    record["codes"]["psqca_cm_licence"] = {
        "system": "PSQCA_CM_LICENCE",
        "value": row["licence_number"],
        "source_id": row["source_id"],
        "status": row["lifecycle_status"],
    }
    record["certificate"].update({
        "number": row["licence_number"],
        "issued_at": row["issued_at"],
        "expires_at": row["expires_at"],
        "technical_document": "; ".join(technical["certified_standard"]),
    })
    # A certified brand scope is not an individual SAE/API formulation. Keep it
    # separate even when a commercial catalog contains the same short brand name.
    record["canonical_key"] += f"|psqca_cm_licence:{normalize(row['licence_number'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def tisi_two_stroke_oil_record(row: dict) -> dict:
    """Convert one TISI licence at its conservative certified-holder grain."""
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Thailand TISI licence — certified two-stroke engine-oil holder scope",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": "",
        "api_class": "",
        "viscosity": "",
        "grease_class": "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": row["evidence_status"],
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
        "certificate_status": row["lifecycle_status"],
    })
    record["specifications"].update({
        "tisi_certified_standard": technical["certified_standard"],
        "tisi_certified_product_scope": row["certified_product_scope"],
        "tisi_certified_product_scope_source_reported": row["certified_product_scope_source_reported"],
        "tisi_product_name_basis": row["product_name_basis"],
        "tisi_permit_type": row["permit_type"],
        "tisi_permit_type_source_reported": row["permit_type_source_reported"],
        "tisi_source_quality_flags": row["source_quality_flags"],
        "sae_source_reported": technical["sae"],
        "api_source_reported": technical["api"],
        "source_url": row["source_url"],
        "dataset_url": row["dataset_url"],
        "standard_url": row["standard_url"],
        "source_facts_sha256": row["source_facts_sha256"],
    })
    record["codes"]["tisi_manufacturing_licence"] = {
        "system": "TISI_MANUFACTURING_LICENCE",
        "value": row["permit_number"],
        "source_id": row["source_id"],
        "status": row["lifecycle_status"],
    }
    record["certificate"].update({
        "number": row["permit_number"],
        "issued_at": row["issued_at"],
        "technical_document": "; ".join(technical["certified_standard"]),
    })
    # A holder-level licence scope must never collapse into a marketed product.
    record["canonical_key"] += f"|tisi_licence:{normalize(row['permit_number'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def samr_inspection_occurrence(row: dict) -> dict:
    return {
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "report_date": row["report_date"],
        "production_date_or_batch_source_reported": row["production_date_or_batch_source_reported"],
        "inspection_outcome": row["inspection_outcome"],
        "inspection_retest_confirmed_nonconforming": row.get("inspection_retest_confirmed_nonconforming", False),
        "inspection_suspected_counterfeit": row.get("inspection_suspected_counterfeit", False),
        "nonconforming_items": row["nonconforming_items"],
        "nonconforming_items_source_reported": row["nonconforming_items_source_reported"],
        "source_note": row["source_note"],
        "source_facts_sha256": row["source_facts_sha256"],
        "source_row": row.get("source_row"),
        "source_page": row.get("source_page"),
        "source_workbook": row.get("source_workbook"),
        "source_workbook_kind": row.get("source_workbook_kind"),
        "attachment_url": row.get("attachment_url"),
        "inspection_standards_scope_source_reported": row.get("inspection_standards_scope_source_reported", []),
        "model_specification_source_reported": row["model_specification_source_reported"],
    }


def samr_china_inspection_record(row: dict) -> dict:
    """Convert one historical SAMR nonconforming-product observation."""
    technical = row["technical"]
    api = technical["api_source_reported"]
    sae = technical["sae_source_reported"]
    api_gl = technical.get("api_gl_source_reported", [])
    iso_vg = technical.get("iso_vg_source_reported", [])
    china_classes = technical.get("china_lubricant_class_source_reported", [])
    brake_classes = technical["brake_fluid_dot_source_reported"] or technical["brake_fluid_hzy_source_reported"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": f"China government {row['inspection_outcome']} {row['product_kind_english']} inspection observation",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": "; ".join(sae),
        "api_class": "; ".join(
            [*(f"API {value}" for value in api), *(f"API {value}" for value in api_gl)]
        ),
        "viscosity": iso_vg[0] if iso_vg else "",
        "din_gost_class": "; ".join(china_classes),
        "grease_class": "",
        "coolant_class": brake_classes[0] if brake_classes else "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": row.get("source_row"),
        "evidence_status": row["evidence_status"],
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["snapshot_date"],
    })
    record["specifications"].update({
        "brand_basis": row["brand_basis"],
        "samr_product_name_basis": row["product_name_basis"],
        "samr_product_kind_source_reported": row["product_kind_source_reported"],
        "samr_product_kind_english": row["product_kind_english"],
        "samr_model_specification_source_reported": row["model_specification_source_reported"],
        "samr_production_date_or_batch_source_reported": row["production_date_or_batch_source_reported"],
        "samr_inspection_outcome": row["inspection_outcome"],
        "samr_nonconforming_items": row["nonconforming_items"],
        "samr_nonconforming_items_source_reported": row["nonconforming_items_source_reported"],
        "samr_source_note": row["source_note"],
        "samr_source_quality_flags": row["source_quality_flags"],
        "samr_inspection_occurrences": [samr_inspection_occurrence(row)],
        "inspection_model_specifications_source_reported": [row["model_specification_source_reported"]],
        "api_source_reported": api,
        "api_gl_source_reported": api_gl,
        "iso_vg_source_reported": iso_vg,
        "china_lubricant_class_source_reported": china_classes,
        "sae_engine_source_reported": sae,
        "brake_fluid_dot_source_reported": technical["brake_fluid_dot_source_reported"],
        "brake_fluid_hzy_source_reported": technical["brake_fluid_hzy_source_reported"],
        "brake_fluid_env_source_reported": technical.get("brake_fluid_env_source_reported", []),
        "coolant_class_source_reported": technical.get("coolant_class_source_reported", []),
        "coolant_freezing_point_source_reported": technical.get("coolant_freezing_point_source_reported", []),
        "washer_fluid_class_source_reported": technical.get("washer_fluid_class_source_reported", []),
        "washer_fluid_freezing_point_source_reported": technical.get("washer_fluid_freezing_point_source_reported", []),
        "urea_class_source_reported": technical.get("urea_class_source_reported", []),
        "acea_source_reported": technical.get("acea_source_reported", []),
        "jaso_source_reported": technical.get("jaso_source_reported", []),
        "ilsac_source_reported": technical.get("ilsac_source_reported", []),
        "oem_approval_source_reported": technical.get("oem_approval_source_reported", []),
        "source_url": row["source_url"],
        "attachment_url": row["attachment_url"],
        "rights_url": row["rights_url"],
        "report_date": row["report_date"],
        "source_facts_sha256": row["source_facts_sha256"],
        "inspection_standards_scope_source_reported": row.get("inspection_standards_scope_source_reported", []),
    })
    # A failed inspection batch is historical regulatory evidence. Keep it
    # separate from approved/current commercial identities even when names match.
    record["canonical_key"] += f"|samr_inspection_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def samr_identity(row: dict) -> tuple[str, str, str, str]:
    return (
        normalize(row["manufacturer"]),
        normalize(row["product_name"]),
        normalize(row["model_specification_source_reported"]),
        row["family_code"],
    )


def china_inspection_comparable_identity(row: dict) -> tuple[str, str, str, str]:
    """Strict cross-program identity with only generic category suffixes removed."""
    product_name = re.sub(
        r"[（(](?:车用汽油清净剂|机动车发动机冷却液|机动车辆制动液|汽车风窗玻璃清洗液)[）)]",
        "", row["product_name"],
    )
    model = row["model_specification_source_reported"].replace("毫升", "ml").replace("升", "l").replace("／", "/")
    return normalize(row["manufacturer"]), normalize(product_name), normalize(model), row["family_code"]


def china_inspection_professional_signature(row: dict) -> tuple:
    """Canonical source-reported professional signature, including DOT/HZY equivalence."""
    technical = row["technical"]
    brake_classes = set()
    for value in technical.get("brake_fluid_dot_source_reported", []):
        brake_classes.add(("DOT_OR_HZY", value.replace("DOT ", "")))
    for value in technical.get("brake_fluid_hzy_source_reported", []):
        grade = value.replace("HZY", "")
        brake_classes.add(("DOT_OR_HZY", grade) if grade in {"3", "4"} else ("HZY", grade))
    return (
        tuple(technical.get("api_source_reported", [])),
        tuple(technical.get("api_gl_source_reported", [])),
        tuple(technical.get("sae_source_reported", [])),
        tuple(technical.get("iso_vg_source_reported", [])),
        tuple(technical.get("china_lubricant_class_source_reported", [])),
        tuple(technical.get("acea_source_reported", [])),
        tuple(technical.get("jaso_source_reported", [])),
        tuple(technical.get("ilsac_source_reported", [])),
        tuple(technical.get("oem_approval_source_reported", [])),
        tuple(sorted(brake_classes)),
        tuple(technical.get("brake_fluid_env_source_reported", [])),
        tuple(technical.get("coolant_class_source_reported", [])),
        tuple(technical.get("coolant_freezing_point_source_reported", [])),
        tuple(technical.get("washer_fluid_class_source_reported", [])),
        tuple(technical.get("washer_fluid_freezing_point_source_reported", [])),
        tuple(technical.get("urea_class_source_reported", [])),
    )


def china_inspection_professional_identity(row: dict) -> tuple | None:
    """Package-independent strict identity using source-reported professional classes."""
    signature = china_inspection_professional_signature(row)
    if not any(signature):
        return None
    return (
        normalize(row["manufacturer"]), normalize(row["brand"]),
        normalize(row["product_name"]), row["family_code"], signature,
    )


def china_inspection_package_independent_identity(row: dict) -> tuple:
    """Strict formulation identity with package quantity removed but professional classes retained."""
    return (
        normalize(row["manufacturer"]), normalize(row["brand"]),
        normalize(row["product_name"]),
        normalize(row.get("model_specification_without_package", row["model_specification_source_reported"])),
        row["family_code"], china_inspection_professional_signature(row),
    )


def china_inspection_fallback_brand_match(left: dict, right: dict) -> bool:
    """Allow a graphic-mark fallback to meet one explicit brand only under strict product evidence."""
    if (
        normalize(left["manufacturer"]) != normalize(right["manufacturer"])
        or normalize(left["product_name"]) != normalize(right["product_name"])
        or left["family_code"] != right["family_code"]
    ):
        return False
    left_fallback = left.get("brand_basis", "").startswith("nominal_producer_fallback")
    right_fallback = right.get("brand_basis", "").startswith("nominal_producer_fallback")
    if not (left_fallback or right_fallback):
        return False
    left_signature = china_inspection_professional_signature(left)
    right_signature = china_inspection_professional_signature(right)
    if left_signature != right_signature:
        return False
    if any(left_signature):
        return True
    return normalize(left.get("model_specification_without_package", "")) == normalize(
        right.get("model_specification_without_package", "")
    )


def merge_china_inspection_evidence(target: dict, row: dict) -> None:
    target["specifications"]["samr_inspection_occurrences"].append(samr_inspection_occurrence(row))
    target["specifications"]["samr_source_quality_flags"] = sorted(set(
        target["specifications"]["samr_source_quality_flags"] + row["source_quality_flags"]
    ))
    target["specifications"]["inspection_model_specifications_source_reported"] = sorted(set(
        target["specifications"].get("inspection_model_specifications_source_reported", [])
        + [row["model_specification_source_reported"]]
    ))
    for field in (
        "api_source_reported", "sae_engine_source_reported", "acea_source_reported",
        "jaso_source_reported",
        "oem_approval_source_reported",
        "api_gl_source_reported", "iso_vg_source_reported", "china_lubricant_class_source_reported",
        "ilsac_source_reported", "brake_fluid_dot_source_reported",
        "brake_fluid_hzy_source_reported", "brake_fluid_env_source_reported", "coolant_class_source_reported",
        "coolant_freezing_point_source_reported", "washer_fluid_class_source_reported",
        "washer_fluid_freezing_point_source_reported",
        "urea_class_source_reported",
    ):
        technical_field = "sae_source_reported" if field == "sae_engine_source_reported" else field
        target["specifications"][field] = sorted(set(
            target["specifications"].get(field, []) + row["technical"].get(technical_field, [])
        ))
    target["specifications"]["inspection_standards_scope_source_reported"] = sorted(set(
        target["specifications"].get("inspection_standards_scope_source_reported", [])
        + row.get("inspection_standards_scope_source_reported", [])
    ))


def shenzhen_inspection_identity(row: dict) -> tuple:
    technical = row["technical"]
    return (
        normalize(row["manufacturer"]), normalize(row["brand"]), normalize(row["product_name"]),
        normalize(row["model_specification_without_package"]), row["family_code"],
        tuple(technical["api_source_reported"]), tuple(technical["sae_source_reported"]),
        tuple(technical.get("api_gl_source_reported", [])),
        tuple(technical.get("iso_vg_source_reported", [])),
        tuple(technical.get("china_lubricant_class_source_reported", [])),
        tuple(technical["acea_source_reported"]), tuple(technical["ilsac_source_reported"]),
        tuple(technical["brake_fluid_dot_source_reported"]),
        tuple(technical["brake_fluid_hzy_source_reported"]),
        tuple(technical.get("brake_fluid_env_source_reported", [])),
        tuple(technical["coolant_class_source_reported"]),
        tuple(technical.get("coolant_freezing_point_source_reported", [])),
        tuple(technical["washer_fluid_class_source_reported"]),
        tuple(technical.get("washer_fluid_freezing_point_source_reported", [])),
        tuple(technical["urea_class_source_reported"]),
    )


def philippines_bps_brake_fluid_record(row: dict) -> dict:
    """Convert one conservative Philippine BPS brake-fluid evidence identity."""
    brake_classes = row["technical"]["brake_fluid_class"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": f"Philippines BPS {row['scheme']} brake-fluid certification evidence",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": "",
        "api_class": "",
        "viscosity": "",
        "grease_class": "",
        "coolant_class": brake_classes[0] if brake_classes else "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer_or_certificate_holder"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": row["evidence_status"],
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
        "certificate_status": row["lifecycle_status"],
    })
    record["specifications"].update({
        "philippines_bps_scheme": row["scheme"],
        "philippines_bps_brand_basis": row["brand_basis"],
        "philippines_bps_certificate_entries": row["certificate_entries"],
        "philippines_bps_source_snapshot_date": row["source_snapshot_date"],
        "philippines_bps_source_quality_flags": row["source_quality_flags"],
        "brake_fluid_class_source_reported": brake_classes,
        "source_url": row["source_url"],
        "source_landing_url": row["source_landing_url"],
        "source_facts_sha256": row["source_facts_sha256"],
    })
    for key in ("source_brand_name_field", "source_model_type_field", "source_product_field", "source_occurrences", "source_occurrence_count"):
        if key in row:
            record["specifications"][f"philippines_bps_{key}"] = row[key]
    code_system = "PHILIPPINES_BPS_PS_LICENCE" if row["scheme"] == "PS" else "PHILIPPINES_BPS_ICC_CERTIFICATE"
    seen_certificate_numbers = set()
    for index, certificate in enumerate(row["certificate_entries"], 1):
        number = certificate["number"]
        if not number or number in seen_certificate_numbers:
            continue
        seen_certificate_numbers.add(number)
        record["codes"][f"philippines_bps_certificate_{index}"] = {
            "system": code_system,
            "value": number,
            "source_id": row["source_id"],
            "status": row["lifecycle_status"],
        }
    if row["certificate_entries"]:
        record["certificate"].update({
            "number": row["certificate_entries"][0]["number"],
            "technical_document": "; ".join(row["technical"].get("certified_standard", row["technical"].get("certified_standard_source_reported", []))),
        })
    # The same label/grade can have different local-production and import
    # certificates or different holders. Do not collapse those formula scopes.
    record["canonical_key"] += f"|philippines_bps_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def ghana_gsa_certified_record(row: dict) -> dict:
    """Convert one Ghana GSA product-certification licence row."""
    technical = row["technical"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["licence_number"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Ghana GSA certified lubricant or coolant",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae"][0] if technical["sae"] else "",
        "api_class": "; ".join(f"API {value}" for value in technical["api"]),
        "viscosity": "",
        "grease_class": "",
        "coolant_class": "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer_or_certificate_holder"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": row["source_item_number"],
        "evidence_status": row["evidence_status"],
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
        "certificate_status": row["lifecycle_status"],
    })
    record["specifications"].update({
        "ghana_gsa_certified_standard": technical["certified_standard"],
        "ghana_gsa_source_product_field": row["source_product_field"],
        "ghana_gsa_source_pdf_page_index": row["source_pdf_page_index"],
        "ghana_gsa_source_item_number": row["source_item_number"],
        "ghana_gsa_holder_licence_number": row["holder_licence_number"],
        "ghana_gsa_source_quality_flags": row["source_quality_flags"],
        "sae_source_reported": technical["sae"],
        "api_source_reported": technical["api"],
        "source_url": row["source_url"],
        "source_landing_url": row["source_landing_url"],
        "source_facts_sha256": row["source_facts_sha256"],
    })
    if technical.get("coolant_standard"):
        record["specifications"]["coolant_standard_source_reported"] = technical["coolant_standard"]
    record["codes"]["ghana_gsa_product_licence"] = {
        "system": "GHANA_GSA_PRODUCT_LICENCE",
        "value": row["licence_number"],
        "source_id": row["source_id"],
        "status": row["lifecycle_status"],
    }
    record["certificate"].update({
        "number": row["licence_number"],
        "issued_at": row["issued_at"],
        "expires_at": row["expires_at"],
        "technical_document": "; ".join(technical["certified_standard"]),
    })
    # Licence rows are evidence scopes. Do not merge them across holders or
    # certificates merely because a short brand/product string is identical.
    record["canonical_key"] += f"|ghana_gsa_licence:{normalize(row['licence_number'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def kebs_smark_record(row: dict) -> dict:
    """Convert one normalized product identity from the public KEBS S-Mark directory."""
    technical = row["technical"]
    sae_candidates = (
        technical["sae_gear"] + technical["sae"] + technical["sae_monograde"]
        if row["family_code"] == "T"
        else technical["sae"] + technical["sae_monograde"] + technical["sae_gear"]
    )
    sae_class = (sae_candidates + [""])[0]
    api_class = "; ".join(f"API {value}" for value in technical["api"])
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Kenya KEBS Standardization Mark",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": sae_class,
        "api_class": api_class,
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_government_product_certification_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
    })
    record["specifications"].update({
        "sae_source_reported": technical["sae"],
        "sae_monograde_source_reported": technical["sae_monograde"],
        "sae_gear_source_reported": technical["sae_gear"],
        "api_source_reported": technical["api"],
        "iso_vg_source_reported": technical["iso_vg"],
        "nlgi_source_reported": technical["nlgi"],
        "classification_basis": row["classification_basis"],
        "kebs_standards_source_reported": row["standards"],
        "kebs_smark_permits": row["permit_entries"],
        "source_permit_count": row["source_permit_count"],
        "source_url": row["source_url"],
        "source_context_url": row["source_context_url"],
    })
    for index, permit in enumerate(row["permit_entries"], 1):
        record["codes"][f"kebs_smark_permit_{index}"] = {
            "system": "KEBS_SMARK_PERMIT",
            "value": permit["permit_number"],
            "source_id": row["source_id"],
            "status": permit["status"] or row["lifecycle_status"],
        }
    record["canonical_key"] += f"|kebs_smark_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def east_africa_certified_record(row: dict) -> dict:
    """Convert one normalized UNBS or TBS public certification identity."""
    technical = row["technical"]
    sae_candidates = (
        technical["sae_gear"] + technical["sae"] + technical["sae_monograde"]
        if row["family_code"] == "T"
        else technical["sae"] + technical["sae_monograde"] + technical["sae_gear"]
    )
    iso_vg = (technical["iso_vg_explicit"] + technical["iso_vg_designation_inferred"] + [""])[0]
    performance = [f"API {value}" for value in technical["api"]] + technical["api_gl"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": f"{row['certification_authority']} certified product",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": (sae_candidates + [""])[0],
        "api_class": "; ".join(performance),
        "viscosity": iso_vg,
        "grease_class": "",
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_government_product_certification_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
    })
    if iso_vg:
        record["specifications"]["iso_vg"] = iso_vg
    record["specifications"].update({
        "sae_source_reported": technical["sae"],
        "sae_monograde_source_reported": technical["sae_monograde"],
        "sae_gear_source_reported": technical["sae_gear"],
        "api_source_reported": technical["api"],
        "api_gl_source_reported": technical["api_gl"],
        "iso_vg_explicit_source_reported": technical["iso_vg_explicit"],
        "iso_vg_product_designation_inferred": technical["iso_vg_designation_inferred"],
        "dot_source_reported": technical["dot"],
        "temperature_c_source_reported": technical["temperature_c"],
        "classification_basis": row["classification_basis"],
        "source_product_type": row["source_product_type"],
        "certification_authority": row["certification_authority"],
        "certification_standards_source_reported": row["standards"],
        "certification_permits": row["permit_entries"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_url": row["source_url"],
        "source_context_url": row["source_context_url"],
    })
    code_system = "UNBS_QMARK_PERMIT" if row["market"] == "UG" else "TBS_STANDARDS_MARK_LICENSE"
    seen_permit_values = set()
    for index, permit in enumerate(row["permit_entries"], 1):
        if not permit["permit_number"] or permit["permit_number"] in seen_permit_values:
            continue
        seen_permit_values.add(permit["permit_number"])
        record["codes"][f"east_africa_certification_{index}"] = {
            "system": code_system,
            "value": permit["permit_number"],
            "source_id": row["source_id"],
            "status": permit["status"] or row["lifecycle_status"],
        }
    record["canonical_key"] += f"|east_africa_certified_record:{normalize(row['source_id'])}:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def son_mancap_record(row: dict) -> dict:
    """Convert one normalized product from SON's public MANCAP chemical list."""
    technical = row["technical"]
    sae_candidates = (
        technical["sae_gear"] + technical["sae"] + technical["sae_monograde"]
        if row["family_code"] == "T"
        else technical["sae"] + technical["sae_monograde"] + technical["sae_gear"]
    )
    iso_vg = (technical["iso_vg_explicit"] + technical["iso_vg_designation_inferred"] + [""])[0]
    nlgi = (technical["nlgi_explicit"] + technical["nlgi_designation_inferred"] + [""])[0]
    performance = [f"API {value}" for value in technical["api"]] + technical["api_gl"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Nigeria SON MANCAP Chemical Sector certified product",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": (sae_candidates + [""])[0],
        "api_class": "; ".join(performance),
        "viscosity": iso_vg,
        "grease_class": nlgi,
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": None,
        "evidence_status": "official_government_product_certification_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
    })
    if iso_vg:
        record["specifications"]["iso_vg"] = iso_vg
    if nlgi:
        record["specifications"]["nlgi"] = nlgi
    record["specifications"].update({
        "sae_source_reported": technical["sae"],
        "sae_monograde_source_reported": technical["sae_monograde"],
        "sae_gear_source_reported": technical["sae_gear"],
        "api_source_reported": technical["api"],
        "api_gl_source_reported": technical["api_gl"],
        "acea_source_reported": technical["acea"],
        "ilsac_source_reported": technical["ilsac"],
        "jaso_source_reported": technical["jaso"],
        "iso_vg_explicit_source_reported": technical["iso_vg_explicit"],
        "iso_vg_product_designation_inferred": technical["iso_vg_designation_inferred"],
        "nlgi_explicit_source_reported": technical["nlgi_explicit"],
        "nlgi_product_designation_inferred": technical["nlgi_designation_inferred"],
        "grease_thickener_source_reported": technical["grease_thickener_source_reported"],
        "dot_source_reported": technical["dot"],
        "dexron_source_reported": technical["dexron"],
        "temperature_c_source_reported": technical["temperature_c"],
        "base_oil_grade_source_reported": technical["base_oil_grade"],
        "classification_basis": row["classification_basis"],
        "certification_authority": row["certification_authority"],
        "certification_list_period": row["source_list_period"],
        "certification_source_entries": row["source_entries"],
        "source_company_raw": row["source_company_raw"],
        "source_occurrence_count": row["source_occurrence_count"],
        "source_url": row["source_url"],
        "source_context_url": row["source_context_url"],
    })
    record["canonical_key"] += f"|son_mancap_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def rsb_smark_record(row: dict) -> dict:
    """Convert one normalized Rwanda RSB public S-Mark product."""
    technical = row["technical"]
    sae_candidates = technical["sae"] + technical["sae_monograde"]
    performance = [f"API {value}" for value in technical["api"]] + technical["api_gl"]
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": row["brand"],
        "name": row["product_name"],
        "category": "Rwanda RSB S-Mark certified product",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": (sae_candidates + [""])[0],
        "api_class": "; ".join(performance),
        "viscosity": (technical["iso_vg"] + [""])[0],
        "grease_class": (technical["nlgi"] + [""])[0],
        "source": row["source_id"],
    }
    record = canonical_record(generic)
    record.update({
        "manufacturer": row["manufacturer"],
        "brand": row["brand"],
        "market": row["market"],
        "source_id": row["source_id"],
        "source_record_id": row["source_record_id"],
        "source_row": row["source_number"],
        "evidence_status": "official_government_product_certification_registry",
        "lifecycle_status": row["lifecycle_status"],
        "snapshot_date": row["dataset_snapshot_date"],
    })
    record["specifications"].update({
        "sae_source_reported": technical["sae"],
        "sae_monograde_source_reported": technical["sae_monograde"],
        "api_source_reported": technical["api"],
        "api_gl_source_reported": technical["api_gl"],
        "classification_basis": row["classification_basis"],
        "brand_basis": row["brand_basis"],
        "certification_authority": row["certification_authority"],
        "certification_standard_source_reported": row["standard"],
        "certification_licence_source_reported": row["licence_number"],
        "certification_source_status": row["source_status"],
        "certification_expiry_date": row["expiry_date"],
        "source_url": row["source_url"],
        "source_context_url": row["source_context_url"],
    })
    record["codes"]["rsb_smark_licence"] = {
        "system": "RSB_SMARK_LICENCE",
        "value": row["licence_number"],
        "source_id": row["source_id"],
        "status": row["lifecycle_status"],
    }
    record["canonical_key"] += f"|rsb_smark_record:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    return record


def liqui_moly_identity_name(value: str) -> str:
    return re.sub(r"^liqui moly(?: gmbh)?\s+", "", normalize(value)).strip()


def brand_tokens_overlap(left: str, right: str) -> bool:
    left_tokens = set(normalize(left).split())
    right_tokens = set(normalize(right).split())
    return bool(left_tokens & right_tokens)


CHEMEXPO_OWNER_STOP_TOKENS = {
    "and", "co", "company", "corp", "corporation", "division", "group", "inc",
    "international", "intl", "limited", "llc", "ltd", "lubricant", "lubricants",
    "marketing", "of", "oil", "oils", "petroleum", "plc", "product", "products",
    "pte", "public", "subsidiary", "the", "us", "usa",
}


def chemexpo_owner_tokens_overlap(left: str, right: str) -> bool:
    """Require a meaningful owner/brand root, not a shared legal-form token."""
    left_tokens = {
        token for token in normalize(left).split()
        if len(token) >= 3 and token not in CHEMEXPO_OWNER_STOP_TOKENS
    }
    right_tokens = {
        token for token in normalize(right).split()
        if len(token) >= 3 and token not in CHEMEXPO_OWNER_STOP_TOKENS
    }
    return any(
        left_token == right_token
        or (len(left_token) >= 4 and len(right_token) >= 4 and (
            left_token in right_token or right_token in left_token
        ))
        for left_token in left_tokens
        for right_token in right_tokens
    )


def merge_man_recommendation_evidence(target: dict, raw: dict) -> None:
    target["specifications"].setdefault("man_service_recommendations", []).append({
        "application": raw["application"],
        "document_date": raw["document_date"],
        "pages": raw["source_pages"],
        "source_url": raw["source_url"],
        "specifications": raw["specifications"],
    })


def fuchs_catalog_record(row: dict, source_id: str, market_name: str) -> dict:
    technical = row["technical"]
    raw_performance = "; ".join(row["specifications"] + row["approvals"] + row["fuchs_recommendations"])
    generic = {
        "id": row["source_record_id"],
        "source_number": row["source_record_id"],
        "brand": "FUCHS",
        "name": row["product_name"],
        "category": f"Официальный каталог FUCHS {market_name}",
        "category_code": row["family_code"],
        "family": FAMILY_NAMES[row["family_code"]],
        "sae_class": technical["sae_grades"][0] if technical["sae_grades"] else "",
        "api_class": raw_performance,
        "viscosity": technical["iso_vg"][0] if technical["iso_vg"] else "",
        "grease_class": technical["nlgi"][0] if technical["nlgi"] else "",
        "source": source_id,
    }
    record = canonical_record(generic)
    record["manufacturer"] = row["manufacturer"]
    record["brand"] = "FUCHS"
    record["market"] = row["market"]
    record["source_id"] = source_id
    record["source_record_id"] = row["source_record_id"]
    record["source_row"] = None
    record["evidence_status"] = "official_manufacturer_product_catalog"
    record["lifecycle_status"] = "listed_as_of_current_product_finder"
    record["snapshot_date"] = row["snapshot_date"]
    record["specifications"].update({
        key: value for key, value in technical.items()
        if key not in {"sae_grades", "iso_vg", "nlgi"}
    })
    record["specifications"].update({
        "sae_grades_source_reported": technical["sae_grades"],
        "iso_vg_source_reported": technical["iso_vg"],
        "nlgi_source_reported": technical["nlgi"],
        "source_specifications": row["specifications"],
        "source_approvals": row["approvals"],
        "fuchs_recommendations": row["fuchs_recommendations"],
        "fuchs_brand_lines": row["brand_lines"],
        "product_group_paths": row["product_group_paths"],
        "industries": row["industries"],
        "classification_basis": row["classification_basis"],
        "source_product_dates": row["source_product_dates"],
        "source_urls": row["source_urls"],
        "grain_warning": row["grain_warning"],
    })
    source_key = {
        "FUCHS_INDIA_PRODUCT_FINDER": "fuchs_india_record",
        "FUCHS_US_PRODUCT_FINDER": "fuchs_us_record",
        "FUCHS_GERMANY_PRODUCT_FINDER": "fuchs_germany_record",
        "FUCHS_POLAND_PRODUCT_FINDER": "fuchs_poland_record",
        "FUCHS_ITALY_PRODUCT_FINDER": "fuchs_italy_record",
        "FUCHS_SWEDEN_PRODUCT_FINDER": "fuchs_sweden_record",
        "FUCHS_SPAIN_PRODUCT_FINDER": "fuchs_spain_record",
        "FUCHS_FRANCE_PRODUCT_FINDER": "fuchs_france_record",
        "FUCHS_TURKEY_PRODUCT_FINDER": "fuchs_turkey_record",
        "FUCHS_CANADA_PRODUCT_FINDER": "fuchs_canada_record",
        "FUCHS_CHINA_PRODUCT_FINDER": "fuchs_china_record",
        "FUCHS_CZECH_PRODUCT_FINDER": "fuchs_czech_record",
        "FUCHS_MEXICO_PRODUCT_FINDER": "fuchs_mexico_record",
        "FUCHS_SOUTH_AFRICA_PRODUCT_FINDER": "fuchs_south_africa_record",
        "FUCHS_BRAZIL_PRODUCT_FINDER": "fuchs_brazil_record",
        "FUCHS_NORWAY_PRODUCT_FINDER": "fuchs_norway_record",
        "FUCHS_HUNGARY_PRODUCT_FINDER": "fuchs_hungary_record",
        "FUCHS_DENMARK_PRODUCT_FINDER": "fuchs_denmark_record",
        "FUCHS_FINLAND_PRODUCT_FINDER": "fuchs_finland_record",
        "FUCHS_PORTUGAL_PRODUCT_FINDER": "fuchs_portugal_record",
        "FUCHS_ROMANIA_PRODUCT_FINDER": "fuchs_romania_record",
        "FUCHS_AUSTRIA_PRODUCT_FINDER": "fuchs_austria_record",
        "FUCHS_GREECE_PRODUCT_FINDER": "fuchs_greece_record",
        "FUCHS_SWITZERLAND_PRODUCT_FINDER": "fuchs_switzerland_record",
        "FUCHS_KOREA_PRODUCT_FINDER": "fuchs_korea_record",
        "FUCHS_UAE_PRODUCT_FINDER": "fuchs_uae_record",
        "FUCHS_ARGENTINA_PRODUCT_FINDER": "fuchs_argentina_record",
        "FUCHS_CHILE_PRODUCT_FINDER": "fuchs_chile_record",
        "FUCHS_UKRAINE_PRODUCT_FINDER": "fuchs_ukraine_record",
        "FUCHS_SLOVAKIA_PRODUCT_FINDER": "fuchs_slovakia_record",
        "FUCHS_SLOVENIA_PRODUCT_FINDER": "fuchs_slovenia_record",
        "FUCHS_CROATIA_PRODUCT_FINDER": "fuchs_croatia_record",
        "FUCHS_SAUDI_ARABIA_PRODUCT_FINDER": "fuchs_saudi_arabia_record",
        "FUCHS_MACEDONIA_PRODUCT_FINDER": "fuchs_macedonia_record",
    }[source_id]
    record["canonical_key"] += f"|{source_key}:{normalize(row['source_record_id'])}"
    record["product_id"] = "WC-" + hashlib.sha256(record["canonical_key"].encode()).hexdigest()[:20]
    for index, uid in enumerate(row["source_uids"], 1):
        record["codes"][f"fuchs_product_uid_{index}"] = {
            "system": "FUCHS_PRODUCT_UID",
            "value": str(uid),
            "source_id": source_id,
            "status": "official_manufacturer_product_catalog",
        }
    return record


def merge_fuchs_catalog_evidence(target: dict, source_record: dict, raw: dict) -> None:
    target["specifications"].setdefault("fuchs_catalog_entries", []).append({
        "source_id": raw["source_id"],
        "brand_lines": raw["brand_lines"],
        "market": raw["market"],
        "product_group_paths": raw["product_group_paths"],
        "specifications": raw["specifications"],
        "approvals": raw["approvals"],
        "recommendations": raw["fuchs_recommendations"],
        "technical": raw["technical"],
        "source_urls": raw["source_urls"],
    })
    for key, code in source_record["codes"].items():
        target["codes"][f"fuchs_{raw['source_id']}_{raw['source_record_id']}_{key}"] = code


def integrate_fuchs_market(input_records: list[dict], source_rows: list[dict], source_id: str, market_name: str, prior_market_rows: list[dict]) -> dict:
    """Attach one official FUCHS market while preserving ambiguous identities for review."""
    source_records = [fuchs_catalog_record(row, source_id, market_name) for row in source_rows]
    prior_name_families = defaultdict(set)
    for row in prior_market_rows:
        prior_name_families[normalize(row["product_name"])].add(row["family_code"])
    exact_rows = sum(row["family_code"] in prior_name_families[normalize(row["product_name"])] for row in source_rows)
    conflict_rows = sum(bool(prior_name_families[normalize(row["product_name"])]) and row["family_code"] not in prior_name_families[normalize(row["product_name"])] for row in source_rows)
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    product_key, review_keys, family_conflict_keys = {}, [], []
    added_rows = matched_rows = 0
    for raw, source_record in zip(source_rows, source_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            added_rows += 1
            if len(matches) > 1:
                review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        product_key[raw["source_record_id"]] = target["canonical_key"]
    return {"product_key": product_key, "added": added_rows, "matched": matched_rows, "review_keys": review_keys, "family_conflict_keys": family_conflict_keys, "exact": exact_rows, "conflicts": conflict_rows}


def deduplicate(records: list[dict]) -> tuple[list[dict], list[dict]]:
    by_key = defaultdict(list)
    for record in records:
        by_key[record["canonical_key"]].append(record)
    canonical = []
    candidates = []
    for group in by_key.values():
        canonical.append(group[0])
        for duplicate in group[1:]:
            candidates.append({
                "product_id_a": group[0]["product_id"],
                "product_id_b": duplicate["product_id"],
                "reason": "identical_normalized_identity_and_professional_signature",
                "score": 1.0,
                "decision": "merged",
            })
    by_name = defaultdict(list)
    for record in canonical:
        by_name[(normalize(record["brand"]), record["product_name_normalized"], record["family_code"])].append(record)
    for group in by_name.values():
        if len(group) < 2:
            continue
        by_source = defaultdict(list)
        for record in group:
            by_source[record["source_id"]].append(record)
        for source_group in by_source.values():
            for left, right in zip(source_group, source_group[1:]):
                candidates.append({
                    "product_id_a": left["product_id"],
                    "product_id_b": right["product_id"],
                    "reason": "same_brand_name_family_but_different_professional_signature_within_source",
                    "score": 0.7,
                    "decision": "keep_separate_specification_conflict",
                })
        for left, right in combinations(group, 2):
            if left["source_id"] == right["source_id"]:
                continue
            candidates.append({
                "product_id_a": left["product_id"],
                "product_id_b": right["product_id"],
                "reason": "same_company_name_family_across_sources_requires_specification_review",
                "score": 0.9,
                "decision": "review_cross_source_identity",
            })
    return canonical, candidates


def quality_issues(records: list[dict]) -> list[dict]:
    issues = []
    for row in records:
        expected = EXPECTED_ENKT_BASE.get(row["family_code"])
        for system in ["enkt", "skp"]:
            code = row["codes"].get(system, {}).get("value", "")
            if expected and code and not code.startswith(expected):
                issues.append({
                    "product_id": row["product_id"],
                    "issue_code": "classification_family_conflict",
                    "severity": "high",
                    "field": system.upper(),
                    "value": code,
                    "expected": expected,
                    "action": "Do not use for analytics; remap against current ENKT/SKP and retain legacy value as evidence.",
                })
        specs = row["specifications"]
        missing = []
        has_government_qualification = bool(specs.get("dla_qpd_qualifications"))
        if row["family_code"] == "M":
            is_jaso_two_cycle = specs.get("jaso_family_detail") == "two_cycle_gasoline_engine_oil"
            is_nmma_two_cycle = specs.get("licensed_standard") == "NMMA TC-W3"
            if not specs["sae_engine"] and not is_jaso_two_cycle and not is_nmma_two_cycle and not has_government_qualification:
                missing.append("SAE")
            if not specs["api"] and not specs["acea"] and not specs["ilsac"] and not specs.get("jaso") and not specs.get("licensed_standard") and not has_government_qualification:
                missing.append("API/ACEA/ILSAC/JASO/OEM licence")
        elif row["family_code"] in {"H", "I", "C", "U"} and not specs["iso_vg"] and not has_government_qualification:
            missing.append("ISO VG")
        if missing:
            issues.append({
                "product_id": row["product_id"],
                "issue_code": "professional_key_incomplete",
                "severity": "medium",
                "field": "professional_key",
                "value": "",
                "expected": "; ".join(missing),
                "action": "Obtain an official TDS/PDS or certificate before strict equivalence matching.",
            })
    return issues


def build_sqlite(records: list[dict], candidates: list[dict], issues: list[dict], source_links: list[dict], offers: list[dict], policies: dict, run_id: str, input_rows: int) -> None:
    if SQLITE_OUT.exists():
        SQLITE_OUT.unlink()
    db = sqlite3.connect(SQLITE_OUT)
    db.executescript("""
    PRAGMA foreign_keys=ON;
    CREATE TABLE ingest_runs(run_id TEXT PRIMARY KEY, snapshot_date TEXT NOT NULL, generated_at TEXT NOT NULL, input_rows INTEGER NOT NULL, canonical_rows INTEGER NOT NULL);
    CREATE TABLE sources(source_id TEXT PRIMARY KEY, title TEXT NOT NULL, owner TEXT, source_type TEXT, source_locator TEXT, source_sha256 TEXT, source_url TEXT, terms_url TEXT, access_status TEXT NOT NULL, bulk_ingest_allowed INTEGER NOT NULL, publication_status TEXT, observed_count INTEGER, notes TEXT);
    CREATE TABLE products(product_id TEXT PRIMARY KEY, canonical_key TEXT UNIQUE NOT NULL, manufacturer TEXT, brand TEXT NOT NULL, product_name_raw TEXT NOT NULL, product_name_normalized TEXT NOT NULL, market TEXT, family_code TEXT, family TEXT, category TEXT, candidate_technical_profile_id TEXT, profile_match_confidence REAL, profile_match_status TEXT, profile_match_basis_json TEXT NOT NULL, source_id TEXT NOT NULL REFERENCES sources(source_id), source_record_id TEXT, source_row INTEGER, evidence_status TEXT NOT NULL, lifecycle_status TEXT NOT NULL, snapshot_date TEXT NOT NULL);
    CREATE TABLE product_sources(product_id TEXT NOT NULL REFERENCES products(product_id), source_id TEXT NOT NULL REFERENCES sources(source_id), source_record_id TEXT NOT NULL, source_row INTEGER, relation TEXT NOT NULL, PRIMARY KEY(product_id, source_id, source_record_id));
    CREATE TABLE specifications(product_id TEXT NOT NULL REFERENCES products(product_id), spec_type TEXT NOT NULL, spec_value TEXT NOT NULL, is_parsed INTEGER NOT NULL, PRIMARY KEY(product_id, spec_type, spec_value));
    CREATE TABLE external_codes(product_id TEXT NOT NULL REFERENCES products(product_id), code_system TEXT NOT NULL, code_value TEXT NOT NULL, source_id TEXT NOT NULL REFERENCES sources(source_id), status TEXT NOT NULL, PRIMARY KEY(product_id, code_system, code_value));
    CREATE TABLE certificates(product_id TEXT NOT NULL REFERENCES products(product_id), certificate_number TEXT, issued_at TEXT, expires_at TEXT, local_producer_certificate TEXT, technical_document TEXT, certificate_status TEXT);
    CREATE TABLE duplicate_decisions(product_id_a TEXT NOT NULL, product_id_b TEXT NOT NULL, reason TEXT NOT NULL, score REAL NOT NULL, decision TEXT NOT NULL);
    CREATE TABLE quality_issues(product_id TEXT NOT NULL REFERENCES products(product_id), issue_code TEXT NOT NULL, severity TEXT NOT NULL, field TEXT NOT NULL, value TEXT, expected TEXT, action TEXT NOT NULL);
    CREATE TABLE product_offers(offer_id TEXT PRIMARY KEY, product_id TEXT NOT NULL REFERENCES products(product_id), market TEXT NOT NULL, package_name TEXT NOT NULL, unit TEXT NOT NULL, quantity_per_package REAL, weight_kg REAL, density_kg_per_l REAL, lifecycle_status TEXT NOT NULL, archive_type TEXT, archive_reason TEXT, source_id TEXT NOT NULL REFERENCES sources(source_id), source_record_id TEXT NOT NULL);
    CREATE INDEX products_brand_name_idx ON products(brand, product_name_normalized);
    CREATE INDEX products_family_idx ON products(family_code);
    CREATE INDEX specs_type_value_idx ON specifications(spec_type, spec_value);
    CREATE INDEX codes_system_value_idx ON external_codes(code_system, code_value);
    CREATE INDEX offers_product_idx ON product_offers(product_id);
    """)
    db.execute("INSERT INTO ingest_runs VALUES (?,?,?,?,?)", (run_id, SNAPSHOT_DATE, f"{SNAPSHOT_DATE}T00:00:00+00:00", input_rows, len(records)))
    for source in policies["sources"]:
        db.execute("INSERT INTO sources VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", (
            source["source_id"], source["title"], source.get("owner"), source.get("source_type"),
            source.get("source_locator"), source.get("source_sha256"), source.get("source_url"), source.get("terms_url"),
            source["access_status"], int(source["bulk_ingest_allowed"]), source.get("publication_status"),
            source.get("observed_count"), source.get("notes"),
        ))
    for row in records:
        db.execute("INSERT INTO products VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (
            row["product_id"], row["canonical_key"], row["manufacturer"], row["brand"],
            row["product_name_raw"], row["product_name_normalized"], row["market"], row["family_code"],
            row["family"], row["category"], row["candidate_technical_profile_id"],
            row["profile_match_confidence"], row["profile_match_status"],
            json.dumps(row["profile_match_basis"], ensure_ascii=False), row["source_id"],
            row["source_record_id"], row["source_row"], row["evidence_status"], row["lifecycle_status"],
            row["snapshot_date"],
        ))
        for spec_type, value in row["specifications"].items():
            values = value if isinstance(value, list) else [value]
            for item in values:
                if text(item):
                    db.execute("INSERT OR IGNORE INTO specifications VALUES (?,?,?,?)", (row["product_id"], spec_type, text(item), int(spec_type != "performance_raw")))
        for system, code in row["codes"].items():
            db.execute("INSERT INTO external_codes VALUES (?,?,?,?,?)", (row["product_id"], code.get("system", system).upper(), code["value"], code["source_id"], code["status"]))
        cert = row["certificate"]
        if any(cert.values()):
            db.execute("INSERT INTO certificates VALUES (?,?,?,?,?,?,?)", (
                row["product_id"], cert["number"], cert["issued_at"], cert["expires_at"],
                cert["local_producer_certificate"], cert["technical_document"], row["certificate_status"],
            ))
    db.executemany("INSERT INTO duplicate_decisions VALUES (:product_id_a,:product_id_b,:reason,:score,:decision)", candidates)
    db.executemany("INSERT INTO quality_issues VALUES (:product_id,:issue_code,:severity,:field,:value,:expected,:action)", issues)
    db.executemany("INSERT INTO product_sources VALUES (:product_id,:source_id,:source_record_id,:source_row,:relation)", source_links)
    db.executemany("INSERT INTO product_offers VALUES (:offer_id,:product_id,:market,:package_name,:unit,:quantity_per_package,:weight_kg,:density_kg_per_l,:lifecycle_status,:archive_type,:archive_reason,:source_id,:source_record_id)", offers)
    db.commit()
    db.close()


def compress_sqlite() -> None:
    """Create a deterministic repository-safe copy of the generated SQLite database."""
    with SQLITE_OUT.open("rb") as source, lzma.open(
        SQLITE_XZ_OUT, "wb", format=lzma.FORMAT_XZ, preset=9
    ) as archive:
        shutil.copyfileobj(source, archive, length=1024 * 1024)


def compress_jsonl() -> None:
    """Create a deterministic compressed copy of the generated JSONL aggregate."""
    with JSONL_OUT.open("rb") as source, JSONL_GZ_OUT.open("wb") as target:
        with gzip.GzipFile(filename="", mode="wb", fileobj=target, compresslevel=9, mtime=0) as archive:
            shutil.copyfileobj(source, archive, length=1024 * 1024)


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="17365D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        ws.column_dimensions[letter].width = min(55, max(11, max(len(text(c.value)) for c in column) + 2))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def normalize_xlsx_archive(path: Path) -> None:
    """Remove wall-clock metadata so identical workbook content is byte-stable."""
    fixed_zip_time = (2026, 7, 22, 0, 0, 0)
    fixed_w3cdtf = b"2026-07-22T00:00:00Z"
    temporary = path.with_suffix(path.suffix + ".deterministic")
    with zipfile.ZipFile(path, "r") as source:
        entries = [(info, source.read(info.filename)) for info in source.infolist()]
    with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as target:
        for source_info, payload in entries:
            if source_info.filename == "docProps/core.xml":
                payload = re.sub(
                    br"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                    lambda match: match.group(1) + fixed_w3cdtf + match.group(2),
                    payload,
                )
            info = zipfile.ZipInfo(source_info.filename, fixed_zip_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = source_info.create_system
            info.external_attr = source_info.external_attr
            info.comment = source_info.comment
            target.writestr(info, payload, compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)
    temporary.replace(path)


def build_workbook(records: list[dict], candidates: list[dict], issues: list[dict], offers: list[dict], exclusions: list[dict], policies: dict, report: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "01_Паспорт", ["Показатель", "Значение"], [
        ("Статус", "Проверенный seed; мировой каталог ещё не завершён"),
        ("Дата среза", SNAPSHOT_DATE),
        ("Входных строк", report["input_rows"]),
        ("Канонических строк", report["canonical_rows"]),
        ("Брендов", report["brands"]),
        ("Строк с проектным первичным источником", report["project_source_rows"]),
        ("Legacy-строк, ожидающих официальный TDS/PDS", report["legacy_rows_needing_evidence"]),
        ("Активных offer/SKU упаковок", report["active_offers"]),
        ("Подтверждённый мировой итог", "НЕ ОПРЕДЕЛЁН: подключён только проектный seed"),
    ])
    headers = ["Product ID", "Производитель/заявитель", "Бренд", "Исходное название", "Рынок", "Семейство", "Категория", "SAE", "SAE Gear", "ISO VG", "API", "API GL", "ACEA", "ILSAC", "JASO", "Кандидат техпрофиля", "Уверенность", "Источник", "Строка источника", "Статус доказательства"]
    add_sheet(wb, "02_Продукты", headers, [[
        r["product_id"], r["manufacturer"], r["brand"], r["product_name_raw"], r["market"], r["family"], r["category"],
        r["specifications"]["sae_engine"], r["specifications"]["sae_gear"], r["specifications"]["iso_vg"],
        "; ".join(r["specifications"]["api"]), "; ".join(r["specifications"]["api_gl"]),
        "; ".join(r["specifications"]["acea"]), "; ".join(r["specifications"]["ilsac"]), "; ".join(r["specifications"].get("jaso", [])),
        r["candidate_technical_profile_id"], r["profile_match_confidence"], r["source_id"], r["source_row"], r["evidence_status"],
    ] for r in records])
    add_sheet(wb, "03_Источники_и_права", ["Source ID", "Название", "Владелец", "URL/файл", "SHA-256", "Статус доступа", "Bulk ingest", "Публикация", "Примечание"], [[
        s["source_id"], s["title"], s.get("owner"), s.get("source_url") or s.get("source_locator"), s.get("source_sha256"), s["access_status"],
        "да" if s["bulk_ingest_allowed"] else "нет", s.get("publication_status"), s.get("notes"),
    ] for s in policies["sources"]])
    add_sheet(wb, "04_Дедупликация", ["Product A", "Product B", "Причина", "Score", "Решение"], [[
        c["product_id_a"], c["product_id_b"], c["reason"], c["score"], c["decision"]
    ] for c in candidates])
    add_sheet(wb, "05_По_брендам", ["Бренд", "Канонических строк"], sorted(Counter(r["brand"] for r in records).items(), key=lambda x: (-x[1], x[0])))
    add_sheet(wb, "06_По_семействам", ["Код", "Семейство", "Канонических строк"], [[code, FAMILY_NAMES.get(code, code), count] for code, count in sorted(Counter(r["family_code"] for r in records).items())])
    add_sheet(wb, "07_Запрос_прав", ["Поле", "Текст"], [
        ("Тема", "Request for permission / machine-readable lubricant product data for Uzbekistan classification project"),
        ("Цель", "Создание государственного профессионального классификатора смазочных материалов и сопоставления ENKT–SKP–IKPU–HS/TN VED."),
        ("Запрашиваемые данные", "Product ID, brand, product/grade name, market, lifecycle, viscosity, standards, OEM approvals, application, TDS/PDS/SDS links and update date."),
        ("Запрашиваемое право", "Непосредственная машинная выгрузка или API/feed; хранение фактических характеристик и ссылок; регулярное обновление; публикация производных классификационных соответствий."),
        ("Что не требуется", "Не требуется право переиздавать фотографии, фирменное оформление или полный текст технических документов."),
        ("Формат", "CSV/JSON/XML/API/SFTP или периодический архив; предпочтительно стабильный manufacturer product ID."),
        ("Контроль происхождения", "Для каждой строки сохраняются владелец, URL/идентификатор, рынок, дата получения, версия и хэш источника."),
        ("Контактный вопрос", "Просим подтвердить допустимый объём использования, атрибуцию, частоту обновления и ограничения на публикацию."),
    ])
    by_id = {row["product_id"]: row for row in records}
    add_sheet(wb, "08_Проблемы_качества", ["Product ID", "Бренд", "Продукт", "Код проблемы", "Уровень", "Поле", "Значение", "Ожидалось", "Действие"], [[
        issue["product_id"], by_id[issue["product_id"]]["brand"], by_id[issue["product_id"]]["product_name_raw"],
        issue["issue_code"], issue["severity"], issue["field"], issue["value"], issue["expected"], issue["action"],
    ] for issue in issues])
    add_sheet(wb, "09_Упаковки_SKU", ["Offer ID", "Product ID", "Рынок", "Упаковка", "Единица", "Количество", "Вес, кг", "Плотность, кг/л", "Статус", "Архив", "Причина", "Source record"], [[
        o["offer_id"], o["product_id"], o["market"], o["package_name"], o["unit"], o["quantity_per_package"],
        o["weight_kg"], o["density_kg_per_l"], o["lifecycle_status"], o["archive_type"], o["archive_reason"], o["source_record_id"],
    ] for o in offers])
    add_sheet(wb, "10_Исключённые_строки", ["Source record", "Название", "Причина"], [[
        e["source_record_id"], e["name"], e["reason"]
    ] for e in exclusions])
    XLSX_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUT)
    normalize_xlsx_archive(XLSX_OUT)


def main() -> None:
    source = json.loads(CATALOG.read_text(encoding="utf-8"))
    policies = json.loads(POLICY.read_text(encoding="utf-8"))
    input_records = [canonical_record(row) for row in source["products"]]
    jaso_source_rows = [json.loads(line) for line in JASO_JSONL.read_text(encoding="utf-8").splitlines() if line]
    jaso_records = [jaso_record(row) for row in jaso_source_rows]
    input_records.extend(jaso_records)
    licensed_source_rows = [json.loads(line) for line in LICENSED_JSONL.read_text(encoding="utf-8").splitlines() if line]
    licensed_records = [licensed_record(row) for row in licensed_source_rows]
    input_records.extend(licensed_records)
    blue_angel_source_rows = [json.loads(line) for line in BLUE_ANGEL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    blue_angel_records = [blue_angel_record(row) for row in blue_angel_source_rows]
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
        existing_by_name[row["product_name_normalized"]].append(row)
    blue_angel_product_key = {}
    blue_angel_added_rows = 0
    blue_angel_matched_rows = 0
    blue_angel_review_keys = []
    blue_angel_family_conflict_keys = []
    for raw, source_record in zip(blue_angel_source_rows, blue_angel_records):
        name = normalize(raw["product_name"])
        matches = [
            row for row in existing_by_name_family[(name, raw["family_code"])]
            if brand_tokens_overlap(raw["manufacturer"], row["brand"])
        ]
        if len(matches) == 1:
            target = matches[0]
            merge_blue_angel_evidence(target, source_record, raw)
            blue_angel_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [
                row for row in existing_by_name[name]
                if row["family_code"] != raw["family_code"] and brand_tokens_overlap(raw["manufacturer"], row["brand"])
            ]
            existing_by_name[name].append(target)
            blue_angel_added_rows += 1
            if len(matches) > 1:
                blue_angel_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                blue_angel_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        blue_angel_product_key[raw["source_record_id"]] = target["canonical_key"]
    austrian_uz14_source_rows = [json.loads(line) for line in AUSTRIAN_ECOLABEL_UZ14_JSONL.read_text(encoding="utf-8").splitlines() if line]
    austrian_uz14_records = [austrian_ecolabel_uz14_record(row) for row in austrian_uz14_source_rows]
    austrian_uz14_product_key = {}
    austrian_uz14_added_rows = 0
    austrian_uz14_matched_rows = 0
    austrian_uz14_cross_family_matches = 0
    austrian_uz14_review_keys = []
    for raw, source_record in zip(austrian_uz14_source_rows, austrian_uz14_records):
        name = normalize(raw["product_name"])
        identity_matches = [
            row for row in existing_by_name[name]
            if brand_tokens_overlap(raw["manufacturer"], row["brand"])
            or brand_tokens_overlap(raw["brand"], row["brand"])
            or normalize(raw["brand"]) in row["product_name_normalized"]
        ]
        family_matches = [row for row in identity_matches if row["family_code"] == raw["family_code"]]
        matches = family_matches if family_matches else identity_matches
        if len(matches) == 1:
            target = matches[0]
            merge_austrian_ecolabel_uz14_evidence(target, source_record, raw)
            austrian_uz14_matched_rows += 1
            austrian_uz14_cross_family_matches += target["family_code"] != raw["family_code"]
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            existing_by_name[name].append(target)
            austrian_uz14_added_rows += 1
            if len(matches) > 1:
                austrian_uz14_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
        austrian_uz14_product_key[raw["source_record_id"]] = target["canonical_key"]
    korea_ecolabel_source_rows = [json.loads(line) for line in KOREA_ECOLABEL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    korea_ecolabel_records = [korea_ecolabel_record(row) for row in korea_ecolabel_source_rows]
    korea_ecolabel_product_key = {}
    korea_ecolabel_added_rows = 0
    korea_ecolabel_matched_rows = 0
    korea_ecolabel_review_keys = []
    korea_ecolabel_family_conflict_keys = []
    for raw, source_record in zip(korea_ecolabel_source_rows, korea_ecolabel_records):
        name = normalize(raw["product_name"])
        matches = [
            row for row in existing_by_name_family[(name, raw["family_code"])]
            if brand_tokens_overlap(raw["manufacturer"], row["brand"])
        ]
        if len(matches) == 1:
            target = matches[0]
            merge_korea_ecolabel_evidence(target, source_record, raw)
            korea_ecolabel_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [
                row for row in existing_by_name[name]
                if row["family_code"] != raw["family_code"] and brand_tokens_overlap(raw["manufacturer"], row["brand"])
            ]
            existing_by_name[name].append(target)
            korea_ecolabel_added_rows += 1
            if len(matches) > 1:
                korea_ecolabel_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                korea_ecolabel_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        korea_ecolabel_product_key[raw["source_record_id"]] = target["canonical_key"]
    korea_el509_source_rows = [json.loads(line) for line in KOREA_ECOLABEL_EL509_JSONL.read_text(encoding="utf-8").splitlines() if line]
    korea_el509_records = [korea_ecolabel_record(row) for row in korea_el509_source_rows]
    korea_el509_product_key = {}
    korea_el509_added_rows = 0
    korea_el509_matched_rows = 0
    for raw, source_record in zip(korea_el509_source_rows, korea_el509_records):
        name = normalize(raw["product_name"])
        matches = [
            row for row in existing_by_name_family[(name, raw["family_code"])]
            if brand_tokens_overlap(raw["manufacturer"], row["brand"])
        ]
        if len(matches) == 1:
            target = matches[0]
            merge_korea_ecolabel_evidence(target, source_record, raw)
            korea_el509_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            existing_by_name[name].append(target)
            korea_el509_added_rows += 1
        korea_el509_product_key[raw["source_record_id"]] = target["canonical_key"]
    green_choice_philippines_source_rows = [json.loads(line) for line in GREEN_CHOICE_PHILIPPINES_JSONL.read_text(encoding="utf-8").splitlines() if line]
    green_choice_philippines_records = [green_choice_philippines_record(row) for row in green_choice_philippines_source_rows]
    input_records.extend(green_choice_philippines_records)
    uae_moiat_source_rows = [json.loads(line) for line in UAE_MOIAT_JSONL.read_text(encoding="utf-8").splitlines() if line]
    uae_moiat_records = [uae_moiat_record(row) for row in uae_moiat_source_rows]
    input_records.extend(uae_moiat_records)
    eaeu_conformity_source_rows = [json.loads(line) for line in EAEU_CONFORMITY_JSONL.read_text(encoding="utf-8").splitlines() if line]
    eaeu_conformity_records = [eaeu_conformity_record(row) for row in eaeu_conformity_source_rows]
    input_records.extend(eaeu_conformity_records)
    epa_safer_choice_source_rows = [json.loads(line) for line in EPA_SAFER_CHOICE_JSONL.read_text(encoding="utf-8").splitlines() if line]
    epa_safer_choice_records = [epa_safer_choice_record(row) for row in epa_safer_choice_source_rows]
    input_records.extend(epa_safer_choice_records)
    epa_chemexpo_source_rows = [json.loads(line) for line in EPA_CHEMEXPO_JSONL.read_text(encoding="utf-8").splitlines() if line]
    epa_chemexpo_records = [epa_chemexpo_record(row) for row in epa_chemexpo_source_rows]
    psqca_engine_oil_source_rows = [json.loads(line) for line in PSQCA_ENGINE_OIL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    psqca_engine_oil_records = [psqca_engine_oil_record(row) for row in psqca_engine_oil_source_rows]
    input_records.extend(psqca_engine_oil_records)
    tisi_two_stroke_oil_source_rows = [json.loads(line) for line in TISI_TWO_STROKE_OIL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    tisi_two_stroke_oil_records = [tisi_two_stroke_oil_record(row) for row in tisi_two_stroke_oil_source_rows]
    input_records.extend(tisi_two_stroke_oil_records)
    samr_china_2025_source_rows = [json.loads(line) for line in SAMR_CHINA_2025_JSONL.read_text(encoding="utf-8").splitlines() if line]
    samr_china_2025_records = [samr_china_inspection_record(row) for row in samr_china_2025_source_rows]
    input_records.extend(samr_china_2025_records)
    china_inspection_target_by_record_id = {
        raw["source_record_id"]: record
        for raw, record in zip(samr_china_2025_source_rows, samr_china_2025_records)
    }
    samr_china_2025_by_identity = {
        samr_identity(raw): record
        for raw, record in zip(samr_china_2025_source_rows, samr_china_2025_records)
    }
    samr_china_2024_source_rows = [json.loads(line) for line in SAMR_CHINA_2024_JSONL.read_text(encoding="utf-8").splitlines() if line]
    samr_china_2024_records = [samr_china_inspection_record(row) for row in samr_china_2024_source_rows]
    input_records.extend(samr_china_2024_records)
    china_inspection_target_by_record_id.update({
        raw["source_record_id"]: record
        for raw, record in zip(samr_china_2024_source_rows, samr_china_2024_records)
    })
    samr_china_2023_source_rows = [json.loads(line) for line in SAMR_CHINA_2023_JSONL.read_text(encoding="utf-8").splitlines() if line]
    samr_china_2023_product_key = {}
    samr_china_2023_products_matched_to_existing = 0
    samr_china_2023_products_added = 0
    for raw in samr_china_2023_source_rows:
        target = samr_china_2025_by_identity.get(samr_identity(raw))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            samr_china_2023_products_matched_to_existing += 1
        else:
            target = samr_china_inspection_record(raw)
            input_records.append(target)
            samr_china_2023_products_added += 1
        china_inspection_target_by_record_id[raw["source_record_id"]] = target
        samr_china_2023_product_key[raw["source_record_id"]] = target["canonical_key"]
    china_inspection_by_comparable_identity = {
        china_inspection_comparable_identity(raw): china_inspection_target_by_record_id[raw["source_record_id"]]
        for raw in samr_china_2025_source_rows + samr_china_2024_source_rows + samr_china_2023_source_rows
    }
    shenzhen_china_2021_source_rows = [json.loads(line) for line in SHENZHEN_CHINA_2021_JSONL.read_text(encoding="utf-8").splitlines() if line]
    shenzhen_china_2021_product_key = {}
    shenzhen_china_2021_target_by_record_id = {}
    shenzhen_china_2021_products_matched_to_existing = 0
    shenzhen_china_2021_products_added = 0
    for raw in shenzhen_china_2021_source_rows:
        target = china_inspection_by_comparable_identity.get(china_inspection_comparable_identity(raw))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            shenzhen_china_2021_products_matched_to_existing += 1
        else:
            target = samr_china_inspection_record(raw)
            input_records.append(target)
            shenzhen_china_2021_products_added += 1
        shenzhen_china_2021_target_by_record_id[raw["source_record_id"]] = target
        shenzhen_china_2021_product_key[raw["source_record_id"]] = target["canonical_key"]
    shenzhen_china_2025_source_rows = [json.loads(line) for line in SHENZHEN_CHINA_2025_JSONL.read_text(encoding="utf-8").splitlines() if line]
    shenzhen_china_2025_product_key = {}
    shenzhen_china_2025_target_by_record_id = {}
    shenzhen_china_2025_by_identity = {}
    shenzhen_china_2025_products_matched_to_existing = 0
    shenzhen_china_2025_duplicate_occurrences_merged = 0
    shenzhen_china_2025_products_added = 0
    for raw in shenzhen_china_2025_source_rows:
        target = china_inspection_by_comparable_identity.get(china_inspection_comparable_identity(raw))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            shenzhen_china_2025_products_matched_to_existing += 1
        else:
            identity = shenzhen_inspection_identity(raw)
            target = shenzhen_china_2025_by_identity.get(identity)
            if target is not None:
                merge_china_inspection_evidence(target, raw)
                shenzhen_china_2025_duplicate_occurrences_merged += 1
            else:
                target = samr_china_inspection_record(raw)
                input_records.append(target)
                shenzhen_china_2025_by_identity[identity] = target
                shenzhen_china_2025_products_added += 1
        shenzhen_china_2025_target_by_record_id[raw["source_record_id"]] = target
        shenzhen_china_2025_product_key[raw["source_record_id"]] = target["canonical_key"]
    china_inspection_complete_by_comparable_identity = dict(china_inspection_by_comparable_identity)
    china_inspection_complete_by_comparable_identity.update({
        china_inspection_comparable_identity(raw): shenzhen_china_2021_target_by_record_id[raw["source_record_id"]]
        for raw in shenzhen_china_2021_source_rows
    })
    china_inspection_complete_by_comparable_identity.update({
        china_inspection_comparable_identity(raw): shenzhen_china_2025_target_by_record_id[raw["source_record_id"]]
        for raw in shenzhen_china_2025_source_rows
    })
    china_inspection_by_professional_identity = {}
    ambiguous_china_inspection_professional_identities = set()
    for raw, target in [
        *[(row, china_inspection_target_by_record_id[row["source_record_id"]]) for row in samr_china_2025_source_rows + samr_china_2024_source_rows + samr_china_2023_source_rows],
        *[(row, shenzhen_china_2021_target_by_record_id[row["source_record_id"]]) for row in shenzhen_china_2021_source_rows],
        *[(row, shenzhen_china_2025_target_by_record_id[row["source_record_id"]]) for row in shenzhen_china_2025_source_rows],
    ]:
        identity = china_inspection_professional_identity(raw)
        if identity is None or identity in ambiguous_china_inspection_professional_identities:
            continue
        existing_target = china_inspection_by_professional_identity.get(identity)
        if existing_target is not None and existing_target is not target:
            china_inspection_by_professional_identity.pop(identity)
            ambiguous_china_inspection_professional_identities.add(identity)
        else:
            china_inspection_by_professional_identity[identity] = target
    shenzhen_china_2020_source_rows = [json.loads(line) for line in SHENZHEN_CHINA_2020_JSONL.read_text(encoding="utf-8").splitlines() if line]
    shenzhen_china_2020_product_key = {}
    shenzhen_china_2020_target_by_record_id = {}
    shenzhen_china_2020_by_identity = {}
    shenzhen_china_2020_products_matched_to_existing = 0
    shenzhen_china_2020_duplicate_occurrences_merged = 0
    shenzhen_china_2020_products_added = 0
    for raw in shenzhen_china_2020_source_rows:
        target = china_inspection_complete_by_comparable_identity.get(china_inspection_comparable_identity(raw))
        if target is None:
            professional_identity = china_inspection_professional_identity(raw)
            if professional_identity is not None:
                target = china_inspection_by_professional_identity.get(professional_identity)
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            shenzhen_china_2020_products_matched_to_existing += 1
        else:
            identity = shenzhen_inspection_identity(raw)
            target = shenzhen_china_2020_by_identity.get(identity)
            if target is not None:
                merge_china_inspection_evidence(target, raw)
                shenzhen_china_2020_duplicate_occurrences_merged += 1
            else:
                target = samr_china_inspection_record(raw)
                input_records.append(target)
                shenzhen_china_2020_by_identity[identity] = target
                shenzhen_china_2020_products_added += 1
        shenzhen_china_2020_target_by_record_id[raw["source_record_id"]] = target
        shenzhen_china_2020_product_key[raw["source_record_id"]] = target["canonical_key"]
    china_inspection_existing_entries = [
        *[(row, china_inspection_target_by_record_id[row["source_record_id"]]) for row in samr_china_2025_source_rows + samr_china_2024_source_rows + samr_china_2023_source_rows],
        *[(row, shenzhen_china_2021_target_by_record_id[row["source_record_id"]]) for row in shenzhen_china_2021_source_rows],
        *[(row, shenzhen_china_2025_target_by_record_id[row["source_record_id"]]) for row in shenzhen_china_2025_source_rows],
        *[(row, shenzhen_china_2020_target_by_record_id[row["source_record_id"]]) for row in shenzhen_china_2020_source_rows],
    ]

    def unique_inspection_target_map(identity_function) -> dict:
        result = {}
        ambiguous = set()
        for source_row, source_target in china_inspection_existing_entries:
            identity = identity_function(source_row)
            if identity is None or identity in ambiguous:
                continue
            existing_target = result.get(identity)
            if existing_target is not None and existing_target is not source_target:
                result.pop(identity)
                ambiguous.add(identity)
            else:
                result[identity] = source_target
        return result

    china_inspection_all_by_comparable_identity = unique_inspection_target_map(china_inspection_comparable_identity)
    china_inspection_all_by_professional_identity = unique_inspection_target_map(china_inspection_professional_identity)
    china_inspection_all_by_package_independent_identity = unique_inspection_target_map(china_inspection_package_independent_identity)
    shenzhen_china_2019_source_rows = [json.loads(line) for line in SHENZHEN_CHINA_2019_JSONL.read_text(encoding="utf-8").splitlines() if line]
    shenzhen_china_2019_product_key = {}
    shenzhen_china_2019_target_by_record_id = {}
    shenzhen_china_2019_by_identity = {}
    shenzhen_china_2019_products_matched_to_existing = 0
    shenzhen_china_2019_duplicate_occurrences_merged = 0
    shenzhen_china_2019_products_added = 0
    for raw in shenzhen_china_2019_source_rows:
        target = china_inspection_all_by_comparable_identity.get(china_inspection_comparable_identity(raw))
        if target is None:
            professional_identity = china_inspection_professional_identity(raw)
            if professional_identity is not None:
                target = china_inspection_all_by_professional_identity.get(professional_identity)
        if target is None:
            target = china_inspection_all_by_package_independent_identity.get(
                china_inspection_package_independent_identity(raw)
            )
        if target is None:
            fallback_candidates = {
                candidate_target["canonical_key"]: candidate_target
                for candidate_raw, candidate_target in china_inspection_existing_entries
                if china_inspection_fallback_brand_match(raw, candidate_raw)
            }
            if len(fallback_candidates) == 1:
                target = next(iter(fallback_candidates.values()))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            shenzhen_china_2019_products_matched_to_existing += 1
        else:
            identity = shenzhen_inspection_identity(raw)
            target = shenzhen_china_2019_by_identity.get(identity)
            if target is not None:
                merge_china_inspection_evidence(target, raw)
                shenzhen_china_2019_duplicate_occurrences_merged += 1
            else:
                target = samr_china_inspection_record(raw)
                input_records.append(target)
                shenzhen_china_2019_by_identity[identity] = target
                shenzhen_china_2019_products_added += 1
        shenzhen_china_2019_product_key[raw["source_record_id"]] = target["canonical_key"]
        shenzhen_china_2019_target_by_record_id[raw["source_record_id"]] = target
    china_inspection_existing_entries.extend(
        (raw, shenzhen_china_2019_target_by_record_id[raw["source_record_id"]])
        for raw in shenzhen_china_2019_source_rows
    )
    shanghai_china_source_rows = [
        json.loads(line)
        for line in SHANGHAI_CHINA_2023_2025_JSONL.read_text(encoding="utf-8").splitlines()
        if line
    ]
    shanghai_china_product_key = {}
    shanghai_china_products_matched_to_existing = 0
    shanghai_china_products_added = 0
    for raw in shanghai_china_source_rows:
        comparable_map = unique_inspection_target_map(china_inspection_comparable_identity)
        professional_map = unique_inspection_target_map(china_inspection_professional_identity)
        package_independent_map = unique_inspection_target_map(china_inspection_package_independent_identity)
        target = comparable_map.get(china_inspection_comparable_identity(raw))
        if target is None:
            professional_identity = china_inspection_professional_identity(raw)
            if professional_identity is not None:
                target = professional_map.get(professional_identity)
        if target is None:
            target = package_independent_map.get(china_inspection_package_independent_identity(raw))
        if target is None:
            fallback_candidates = {
                candidate_target["canonical_key"]: candidate_target
                for candidate_raw, candidate_target in china_inspection_existing_entries
                if china_inspection_fallback_brand_match(raw, candidate_raw)
            }
            if len(fallback_candidates) == 1:
                target = next(iter(fallback_candidates.values()))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            shanghai_china_products_matched_to_existing += 1
        else:
            target = samr_china_inspection_record(raw)
            input_records.append(target)
            shanghai_china_products_added += 1
        china_inspection_existing_entries.append((raw, target))
        shanghai_china_product_key[raw["source_record_id"]] = target["canonical_key"]
    beijing_china_source_rows = [
        json.loads(line)
        for line in BEIJING_CHINA_2018_JSONL.read_text(encoding="utf-8").splitlines()
        if line
    ]
    beijing_china_product_key = {}
    beijing_china_products_matched_to_existing = 0
    beijing_china_products_added = 0
    for raw in beijing_china_source_rows:
        comparable_map = unique_inspection_target_map(china_inspection_comparable_identity)
        professional_map = unique_inspection_target_map(china_inspection_professional_identity)
        package_independent_map = unique_inspection_target_map(china_inspection_package_independent_identity)
        target = comparable_map.get(china_inspection_comparable_identity(raw))
        if target is None:
            professional_identity = china_inspection_professional_identity(raw)
            if professional_identity is not None:
                target = professional_map.get(professional_identity)
        if target is None:
            target = package_independent_map.get(china_inspection_package_independent_identity(raw))
        if target is None:
            fallback_candidates = {
                candidate_target["canonical_key"]: candidate_target
                for candidate_raw, candidate_target in china_inspection_existing_entries
                if china_inspection_fallback_brand_match(raw, candidate_raw)
            }
            if len(fallback_candidates) == 1:
                target = next(iter(fallback_candidates.values()))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            beijing_china_products_matched_to_existing += 1
        else:
            target = samr_china_inspection_record(raw)
            input_records.append(target)
            beijing_china_products_added += 1
        china_inspection_existing_entries.append((raw, target))
        beijing_china_product_key[raw["source_record_id"]] = target["canonical_key"]
    shenzhen_china_2016_2017_source_rows = [
        json.loads(line)
        for line in SHENZHEN_CHINA_2016_2017_JSONL.read_text(encoding="utf-8").splitlines()
        if line
    ]
    shenzhen_china_2016_2017_product_key = {}
    shenzhen_china_2016_2017_products_matched_to_existing = 0
    shenzhen_china_2016_2017_products_added = 0
    for raw in shenzhen_china_2016_2017_source_rows:
        comparable_map = unique_inspection_target_map(china_inspection_comparable_identity)
        professional_map = unique_inspection_target_map(china_inspection_professional_identity)
        package_independent_map = unique_inspection_target_map(china_inspection_package_independent_identity)
        target = comparable_map.get(china_inspection_comparable_identity(raw))
        if target is None:
            professional_identity = china_inspection_professional_identity(raw)
            if professional_identity is not None:
                target = professional_map.get(professional_identity)
        if target is None:
            target = package_independent_map.get(china_inspection_package_independent_identity(raw))
        if target is None:
            fallback_candidates = {
                candidate_target["canonical_key"]: candidate_target
                for candidate_raw, candidate_target in china_inspection_existing_entries
                if china_inspection_fallback_brand_match(raw, candidate_raw)
            }
            if len(fallback_candidates) == 1:
                target = next(iter(fallback_candidates.values()))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            shenzhen_china_2016_2017_products_matched_to_existing += 1
        else:
            target = samr_china_inspection_record(raw)
            input_records.append(target)
            shenzhen_china_2016_2017_products_added += 1
        china_inspection_existing_entries.append((raw, target))
        shenzhen_china_2016_2017_product_key[raw["source_record_id"]] = target["canonical_key"]
    qingdao_china_source_rows = [
        json.loads(line)
        for line in QINGDAO_CHINA_2021_2025_JSONL.read_text(encoding="utf-8").splitlines()
        if line
    ]
    qingdao_china_product_key = {}
    qingdao_china_products_matched_to_existing = 0
    qingdao_china_duplicate_occurrences_merged = 0
    qingdao_china_products_added = 0
    qingdao_created_keys = set()
    for raw in qingdao_china_source_rows:
        comparable_map = unique_inspection_target_map(china_inspection_comparable_identity)
        professional_map = unique_inspection_target_map(china_inspection_professional_identity)
        package_independent_map = unique_inspection_target_map(china_inspection_package_independent_identity)
        target = comparable_map.get(china_inspection_comparable_identity(raw))
        if target is None:
            professional_identity = china_inspection_professional_identity(raw)
            if professional_identity is not None:
                target = professional_map.get(professional_identity)
        if target is None:
            target = package_independent_map.get(china_inspection_package_independent_identity(raw))
        if target is None:
            fallback_candidates = {
                candidate_target["canonical_key"]: candidate_target
                for candidate_raw, candidate_target in china_inspection_existing_entries
                if china_inspection_fallback_brand_match(raw, candidate_raw)
            }
            if len(fallback_candidates) == 1:
                target = next(iter(fallback_candidates.values()))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            if target["canonical_key"] in qingdao_created_keys:
                qingdao_china_duplicate_occurrences_merged += 1
            else:
                qingdao_china_products_matched_to_existing += 1
        else:
            target = samr_china_inspection_record(raw)
            input_records.append(target)
            qingdao_created_keys.add(target["canonical_key"])
            qingdao_china_products_added += 1
        china_inspection_existing_entries.append((raw, target))
        qingdao_china_product_key[raw["source_record_id"]] = target["canonical_key"]
    jilin_china_source_rows = [
        json.loads(line)
        for line in JILIN_CHINA_2024_2025_JSONL.read_text(encoding="utf-8").splitlines()
        if line
    ]
    jilin_china_product_key = {}
    jilin_china_products_matched_to_existing = 0
    jilin_china_duplicate_occurrences_merged = 0
    jilin_china_products_added = 0
    jilin_created_keys = set()
    for raw in jilin_china_source_rows:
        comparable_map = unique_inspection_target_map(china_inspection_comparable_identity)
        professional_map = unique_inspection_target_map(china_inspection_professional_identity)
        package_independent_map = unique_inspection_target_map(china_inspection_package_independent_identity)
        target = comparable_map.get(china_inspection_comparable_identity(raw))
        if target is None:
            professional_identity = china_inspection_professional_identity(raw)
            if professional_identity is not None:
                target = professional_map.get(professional_identity)
        if target is None:
            target = package_independent_map.get(china_inspection_package_independent_identity(raw))
        if target is None:
            fallback_candidates = {
                candidate_target["canonical_key"]: candidate_target
                for candidate_raw, candidate_target in china_inspection_existing_entries
                if china_inspection_fallback_brand_match(raw, candidate_raw)
            }
            if len(fallback_candidates) == 1:
                target = next(iter(fallback_candidates.values()))
        if target is not None:
            merge_china_inspection_evidence(target, raw)
            if target["canonical_key"] in jilin_created_keys:
                jilin_china_duplicate_occurrences_merged += 1
            else:
                jilin_china_products_matched_to_existing += 1
        else:
            target = samr_china_inspection_record(raw)
            input_records.append(target)
            jilin_created_keys.add(target["canonical_key"])
            jilin_china_products_added += 1
        china_inspection_existing_entries.append((raw, target))
        jilin_china_product_key[raw["source_record_id"]] = target["canonical_key"]
    philippines_bps_brake_fluid_source_rows = [json.loads(line) for line in PHILIPPINES_BPS_BRAKE_FLUID_JSONL.read_text(encoding="utf-8").splitlines() if line]
    philippines_bps_brake_fluid_records = [philippines_bps_brake_fluid_record(row) for row in philippines_bps_brake_fluid_source_rows]
    input_records.extend(philippines_bps_brake_fluid_records)
    ghana_gsa_source_rows = [json.loads(line) for line in GHANA_GSA_CERTIFIED_JSONL.read_text(encoding="utf-8").splitlines() if line]
    ghana_gsa_records = [ghana_gsa_certified_record(row) for row in ghana_gsa_source_rows]
    input_records.extend(ghana_gsa_records)
    kebs_smark_source_rows = [json.loads(line) for line in KEBS_SMARK_JSONL.read_text(encoding="utf-8").splitlines() if line]
    kebs_smark_records = [kebs_smark_record(row) for row in kebs_smark_source_rows]
    input_records.extend(kebs_smark_records)
    east_africa_certified_source_rows = [json.loads(line) for line in EAST_AFRICA_CERTIFIED_JSONL.read_text(encoding="utf-8").splitlines() if line]
    east_africa_certified_records = [east_africa_certified_record(row) for row in east_africa_certified_source_rows]
    input_records.extend(east_africa_certified_records)
    son_mancap_source_rows = [json.loads(line) for line in SON_MANCAP_JSONL.read_text(encoding="utf-8").splitlines() if line]
    son_mancap_records = [son_mancap_record(row) for row in son_mancap_source_rows]
    input_records.extend(son_mancap_records)
    rsb_smark_source_rows = [json.loads(line) for line in RSB_SMARK_JSONL.read_text(encoding="utf-8").splitlines() if line]
    rsb_smark_records = [rsb_smark_record(row) for row in rsb_smark_source_rows]
    input_records.extend(rsb_smark_records)
    biopreferred_source_rows = [json.loads(line) for line in USDA_BIOPREFERRED_JSONL.read_text(encoding="utf-8").splitlines() if line]
    biopreferred_records = [biopreferred_record(row) for row in biopreferred_source_rows]
    input_records.extend(biopreferred_records)
    zf_source_rows = [json.loads(line) for line in ZF_TE_ML_JSONL.read_text(encoding="utf-8").splitlines() if line]
    zf_records = [zf_te_ml_record(row) for row in zf_source_rows]
    input_records.extend(zf_records)
    allison_source_rows = [json.loads(line) for line in ALLISON_JSONL.read_text(encoding="utf-8").splitlines() if line]
    allison_records = [allison_record(row) for row in allison_source_rows]
    input_records.extend(allison_records)
    driventic_source_rows = [json.loads(line) for line in DRIVENTIC_DIWA_JSONL.read_text(encoding="utf-8").splitlines() if line]
    driventic_records = [driventic_diwa_record(row) for row in driventic_source_rows]
    input_records.extend(driventic_records)
    mercedes_dtfr_source_rows = [json.loads(line) for line in MERCEDES_DTFR_JSONL.read_text(encoding="utf-8").splitlines() if line]
    mercedes_dtfr_records = [mercedes_dtfr_record(row) for row in mercedes_dtfr_source_rows]
    input_records.extend(mercedes_dtfr_records)
    mercedes_bevo_source_rows = [json.loads(line) for line in MERCEDES_BEVO_JSONL.read_text(encoding="utf-8").splitlines() if line]
    mercedes_bevo_records = [mercedes_bevo_record(row) for row in mercedes_bevo_source_rows]
    existing_by_name = defaultdict(list)
    for row in input_records:
        existing_by_name[(normalize(row["brand"]), row["product_name_normalized"], row["family_code"])].append(row)
    mercedes_bevo_product_key = {}
    mercedes_bevo_added_rows = 0
    mercedes_bevo_matched_rows = 0
    for raw, source_record in zip(mercedes_bevo_source_rows, mercedes_bevo_records):
        key = (normalize(raw["company"]), normalize(raw["product_name"]), raw["family_code"])
        matches = existing_by_name.get(key, [])
        if matches:
            target = matches[0]
            merge_mercedes_bevo_evidence(target, source_record, raw)
            mercedes_bevo_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name[key].append(target)
            mercedes_bevo_added_rows += 1
        mercedes_bevo_product_key[raw["source_record_id"]] = target["canonical_key"]
    volvo_genuine_source_rows = [json.loads(line) for line in VOLVO_GENUINE_JSONL.read_text(encoding="utf-8").splitlines() if line]
    volvo_genuine_records = [volvo_genuine_record(row) for row in volvo_genuine_source_rows]
    input_records.extend(volvo_genuine_records)
    mack_genuine_source_rows = [json.loads(line) for line in MACK_GENUINE_JSONL.read_text(encoding="utf-8").splitlines() if line]
    mack_genuine_records = [mack_genuine_record(row) for row in mack_genuine_source_rows]
    input_records.extend(mack_genuine_records)
    scania_genuine_source_rows = [json.loads(line) for line in SCANIA_GENUINE_JSONL.read_text(encoding="utf-8").splitlines() if line]
    scania_genuine_records = [scania_genuine_record(row) for row in scania_genuine_source_rows]
    input_records.extend(scania_genuine_records)
    brava_official_source_rows = [json.loads(line) for line in BRAVA_OFFICIAL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    brava_official_records = [brava_official_record(row) for row in brava_official_source_rows]
    input_records.extend(brava_official_records)
    taiwan_cpc_source_rows = [json.loads(line) for line in TAIWAN_CPC_JSONL.read_text(encoding="utf-8").splitlines() if line]
    taiwan_cpc_records = [taiwan_cpc_record(row) for row in taiwan_cpc_source_rows]
    input_records.extend(taiwan_cpc_records)
    volvo_websds_source_rows = [json.loads(line) for line in VOLVO_WEBSDS_JSONL.read_text(encoding="utf-8").splitlines() if line]
    volvo_websds_records = [volvo_websds_record(row) for row in volvo_websds_source_rows]
    input_records.extend(volvo_websds_records)
    parker_denison_source_rows = [json.loads(line) for line in PARKER_DENISON_JSONL.read_text(encoding="utf-8").splitlines() if line]
    parker_denison_records = [parker_denison_record(row) for row in parker_denison_source_rows]
    input_records.extend(parker_denison_records)
    ceypetco_source_rows = [json.loads(line) for line in CEYPETCO_JSONL.read_text(encoding="utf-8").splitlines() if line]
    ceypetco_records = [ceypetco_record(row) for row in ceypetco_source_rows]
    input_records.extend(ceypetco_records)
    pso_source_rows = [json.loads(line) for line in PSO_OFFICIAL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    pso_records = [pso_official_record(row) for row in pso_source_rows]
    pso_product_key = {}
    pso_added_rows = 0
    pso_matched_rows = 0
    for raw, source_record in zip(pso_source_rows, pso_records):
        matches = [
            row for row in input_records
            if row["family_code"] == raw["family_code"]
            and row["product_name_normalized"] == normalize(raw["source_series"])
            and row["specifications"].get("sae_engine", "") == raw["specifications"].get("sae_engine", "")
            and "pakistan state oil" in normalize(row["manufacturer"])
        ]
        if len(matches) == 1:
            target = matches[0]
            merge_pso_catalog_evidence(target, source_record)
            pso_matched_rows += 1
        else:
            assert not matches, (raw["source_record_id"], len(matches))
            target = source_record
            input_records.append(target)
            pso_added_rows += 1
        pso_product_key[raw["source_record_id"]] = target["canonical_key"]
    man_service_source_rows = [json.loads(line) for line in MAN_SERVICE_JSONL.read_text(encoding="utf-8").splitlines() if line]
    man_service_records = [man_service_record(row) for row in man_service_source_rows]
    existing_by_name_family = defaultdict(list)
    for row in input_records:
        existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
    man_service_product_key = {}
    man_service_added_rows = 0
    man_service_matched_rows = 0
    man_service_review_keys = []
    for raw, source_record in zip(man_service_source_rows, man_service_records):
        matches = [
            row for row in existing_by_name_family[(normalize(raw["product_name"]), raw["family_code"])]
            if brand_tokens_overlap(raw["brand"], row["brand"])
        ]
        if len(matches) == 1:
            target = matches[0]
            merge_man_recommendation_evidence(target, raw)
            man_service_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(target["product_name_normalized"], target["family_code"])].append(target)
            man_service_added_rows += 1
            if len(matches) > 1:
                man_service_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
        man_service_product_key[raw["source_record_id"]] = target["canonical_key"]
    liqui_moly_source_rows = [json.loads(line) for line in LIQUI_MOLY_2020_JSONL.read_text(encoding="utf-8").splitlines() if line]
    liqui_moly_records = [liqui_moly_catalog_record(row) for row in liqui_moly_source_rows]
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("LIQUI MOLY", row["brand"]):
            identity_name = liqui_moly_identity_name(row["product_name_raw"])
            existing_by_name_family[(identity_name, row["family_code"])].append(row)
            existing_by_name[identity_name].append(row)
    liqui_moly_product_key = {}
    liqui_moly_added_rows = liqui_moly_matched_rows = 0
    liqui_moly_review_keys, liqui_moly_family_conflict_keys = [], []
    for raw, source_record in zip(liqui_moly_source_rows, liqui_moly_records):
        name = liqui_moly_identity_name(raw["product_name"])
        matches = [row for row in existing_by_name_family[(name, raw["family_code"])] if brand_tokens_overlap("LIQUI MOLY", row["brand"])]
        if len(matches) == 1:
            target = matches[0]
            merge_liqui_moly_evidence(target, source_record, raw)
            liqui_moly_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if brand_tokens_overlap("LIQUI MOLY", row["brand"]) and row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            liqui_moly_added_rows += 1
            if len(matches) > 1:
                liqui_moly_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                liqui_moly_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        liqui_moly_product_key[raw["source_record_id"]] = target["canonical_key"]
    liqui_moly_current_source_rows = [json.loads(line) for line in LIQUI_MOLY_CURRENT_JSONL.read_text(encoding="utf-8").splitlines() if line]
    liqui_moly_current_records = [liqui_moly_current_record(row) for row in liqui_moly_current_source_rows]
    input_by_key = {row["canonical_key"]: row for row in input_records}
    liqui_moly_current_product_key = {}
    liqui_moly_current_added_rows = 0
    liqui_moly_current_matched_rows = 0
    liqui_moly_current_review_keys = []
    for raw, source_record in zip(liqui_moly_current_source_rows, liqui_moly_current_records):
        candidate_keys = sorted({
            liqui_moly_product_key[candidate["source_record_id"]]
            for candidate in raw["historical_candidates"]
            if candidate["source_record_id"] in liqui_moly_product_key
        })
        if len(candidate_keys) == 1 and input_by_key[candidate_keys[0]]["family_code"] == raw["family_code"]:
            target = input_by_key[candidate_keys[0]]
            merge_liqui_moly_current_evidence(target, source_record, raw)
            liqui_moly_current_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            input_by_key[target["canonical_key"]] = target
            liqui_moly_current_added_rows += 1
            if candidate_keys:
                liqui_moly_current_review_keys.append((target["canonical_key"], candidate_keys))
        liqui_moly_current_product_key[raw["source_record_id"]] = target["canonical_key"]
    anp_brazil_source_rows = [json.loads(line) for line in ANP_BRAZIL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    anp_brazil_records = [anp_brazil_record(row) for row in anp_brazil_source_rows]
    input_records.extend(anp_brazil_records)
    anp_monitoring_xlsx_source_rows = [json.loads(line) for line in ANP_BRAZIL_MONITORING_JSONL.read_text(encoding="utf-8").splitlines() if line]
    anp_monitoring_pdf_source_rows = [json.loads(line) for line in ANP_BRAZIL_MONITORING_PDF_JSONL.read_text(encoding="utf-8").splitlines() if line]
    anp_monitoring_pdf_exception_source_rows = [json.loads(line) for line in ANP_BRAZIL_MONITORING_PDF_EXCEPTIONS_JSONL.read_text(encoding="utf-8").splitlines() if line]
    anp_monitoring_complete_source_rows = anp_monitoring_xlsx_source_rows + anp_monitoring_pdf_source_rows
    anp_monitoring_source_rows = anp_monitoring_complete_source_rows + anp_monitoring_pdf_exception_source_rows
    anp_monitoring_groups = defaultdict(list)
    complete_identities_by_base = defaultdict(set)
    for raw in anp_monitoring_complete_source_rows:
        identity = anp_monitoring_identity_key(raw)
        anp_monitoring_groups[identity].append(raw)
        base = identity[:3] + (identity[4:] if len(identity) > 4 else ())
        complete_identities_by_base[base].add(identity)
    exception_raw_identity_keys = set()
    exception_semantic_identity_keys = set()
    exception_semantically_remapped_observations = 0
    exception_semantically_remapped_identity_keys = set()
    for raw in anp_monitoring_pdf_exception_source_rows:
        raw_identity = anp_monitoring_identity_key(raw)
        identity = raw_identity
        base = identity[:3] + (identity[4:] if len(identity) > 4 else ())
        # Exception appendices often omit API while a complete ANP table has
        # exactly one grade for the same registration + name + SAE (+ holder
        # when unregistered).  Only that one-candidate case is safe to merge.
        if not identity[3] and len(complete_identities_by_base[base]) == 1:
            identity = next(iter(complete_identities_by_base[base]))
        if identity != raw_identity:
            exception_semantically_remapped_observations += 1
            exception_semantically_remapped_identity_keys.add(raw_identity)
        exception_raw_identity_keys.add(raw_identity)
        exception_semantic_identity_keys.add(identity)
        anp_monitoring_groups[identity].append(raw)
    current_anp_index = defaultdict(list)
    for raw, record in zip(anp_brazil_source_rows, anp_brazil_records):
        current_anp_index[(
            normalize(raw["registration_number"]),
            normalize(raw["product_name"]),
            normalize(raw["sae"]),
        )].append((raw, record))
    anp_monitoring_product_key = {}
    anp_monitoring_matched_identities = 0
    anp_monitoring_ambiguous_match_identities = 0
    anp_monitoring_added_identities = 0
    anp_monitoring_added_registered_identities = 0
    anp_monitoring_added_unregistered_identities = 0
    anp_monitoring_added_identities_by_primary_source = Counter()
    for identity, occurrences in sorted(anp_monitoring_groups.items()):
        row = occurrences[0]
        base_key = (
            normalize(row["registration_number"]),
            normalize(row["product_name"]),
            normalize(row["sae"]),
        )
        base_candidates = current_anp_index[base_key]
        performance_key = anp_monitoring_performance_key(row["performance_level"])
        performance_candidates = [
            pair for pair in base_candidates
            if not performance_key or performance_key in anp_monitoring_current_performance_keys(pair[0]["performance_level"])
        ]
        candidates = performance_candidates if performance_candidates else []
        if candidates:
            # Holder renames occur through the archive.  Prefer the exact
            # current holder when available, then use a deterministic record.
            holder_matches = [
                pair for pair in candidates
                if normalize(pair[0]["registration_holder"]) == normalize(row["registration_holder"])
            ]
            ranked = holder_matches or candidates
            raw_target, target = min(ranked, key=lambda pair: pair[1]["source_record_id"])
            match_basis = "registration_name_sae_performance"
            if holder_matches:
                match_basis += "_holder"
            merge_anp_monitoring_evidence(target, occurrences, match_basis, len(candidates))
            anp_monitoring_matched_identities += 1
            if len(candidates) > 1:
                anp_monitoring_ambiguous_match_identities += 1
        else:
            target = anp_monitoring_history_record(identity, occurrences)
            merge_anp_monitoring_evidence(target, occurrences, "historical_grade_absent_from_current_anp_snapshot", 0)
            input_records.append(target)
            anp_monitoring_added_identities += 1
            anp_monitoring_added_identities_by_primary_source[row["source_id"]] += 1
            if row["registration_number"]:
                anp_monitoring_added_registered_identities += 1
            else:
                anp_monitoring_added_unregistered_identities += 1
        for raw in occurrences:
            anp_monitoring_product_key[raw["source_record_id"]] = target["canonical_key"]
    indonesia_npt_source_rows = [json.loads(line) for line in INDONESIA_NPT_JSONL.read_text(encoding="utf-8").splitlines() if line]
    indonesia_npt_records = [indonesia_npt_record(row) for row in indonesia_npt_source_rows]
    thailand_doeb_source_rows = [json.loads(line) for line in THAILAND_DOEB_JSONL.read_text(encoding="utf-8").splitlines() if line]
    thailand_doeb_records = [thailand_doeb_record(row) for row in thailand_doeb_source_rows]
    dla_qpd_source_rows = [json.loads(line) for line in DLA_QPD_JSONL.read_text(encoding="utf-8").splitlines() if line]
    dla_qpd_records = [dla_qpd_record(row) for row in dla_qpd_source_rows]
    fuchs_india_source_rows = [json.loads(line) for line in FUCHS_INDIA_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_india_records = [fuchs_catalog_record(row, "FUCHS_INDIA_PRODUCT_FINDER", "India") for row in fuchs_india_source_rows]
    existing_by_name_family = defaultdict(list)
    for row in input_records:
        existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
    fuchs_india_product_key = {}
    fuchs_india_added_rows = 0
    fuchs_india_matched_rows = 0
    fuchs_india_review_keys = []
    for raw, source_record in zip(fuchs_india_source_rows, fuchs_india_records):
        matches = [
            row for row in existing_by_name_family[(normalize(raw["product_name"]), raw["family_code"])]
            if brand_tokens_overlap("FUCHS", row["brand"])
        ]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_india_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(target["product_name_normalized"], target["family_code"])].append(target)
            fuchs_india_added_rows += 1
            if len(matches) > 1:
                fuchs_india_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
        fuchs_india_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_us_source_rows = [json.loads(line) for line in FUCHS_US_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_us_records = [fuchs_catalog_record(row, "FUCHS_US_PRODUCT_FINDER", "USA") for row in fuchs_us_source_rows]
    india_name_families = defaultdict(set)
    for row in fuchs_india_source_rows:
        india_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_cross_market_exact_name_family_rows = sum(
        row["family_code"] in india_name_families[normalize(row["product_name"])]
        for row in fuchs_us_source_rows
    )
    fuchs_cross_market_family_conflict_rows = sum(
        bool(india_name_families[normalize(row["product_name"])])
        and row["family_code"] not in india_name_families[normalize(row["product_name"])]
        for row in fuchs_us_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_us_product_key = {}
    fuchs_us_added_rows = 0
    fuchs_us_matched_rows = 0
    fuchs_us_review_keys = []
    fuchs_us_family_conflict_keys = []
    for raw, source_record in zip(fuchs_us_source_rows, fuchs_us_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_us_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_us_added_rows += 1
            if len(matches) > 1:
                fuchs_us_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_us_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_us_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_germany_source_rows = [json.loads(line) for line in FUCHS_GERMANY_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_germany_records = [fuchs_catalog_record(row, "FUCHS_GERMANY_PRODUCT_FINDER", "Germany") for row in fuchs_germany_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_germany_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_germany_source_rows
    )
    fuchs_germany_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_germany_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_germany_product_key = {}
    fuchs_germany_added_rows = 0
    fuchs_germany_matched_rows = 0
    fuchs_germany_review_keys = []
    fuchs_germany_family_conflict_keys = []
    for raw, source_record in zip(fuchs_germany_source_rows, fuchs_germany_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_germany_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_germany_added_rows += 1
            if len(matches) > 1:
                fuchs_germany_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_germany_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_germany_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_poland_source_rows = [json.loads(line) for line in FUCHS_POLAND_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_poland_records = [fuchs_catalog_record(row, "FUCHS_POLAND_PRODUCT_FINDER", "Poland") for row in fuchs_poland_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_poland_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_poland_source_rows
    )
    fuchs_poland_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_poland_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_poland_product_key = {}
    fuchs_poland_added_rows = 0
    fuchs_poland_matched_rows = 0
    fuchs_poland_review_keys = []
    fuchs_poland_family_conflict_keys = []
    for raw, source_record in zip(fuchs_poland_source_rows, fuchs_poland_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_poland_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_poland_added_rows += 1
            if len(matches) > 1:
                fuchs_poland_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_poland_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_poland_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_italy_source_rows = [json.loads(line) for line in FUCHS_ITALY_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_italy_records = [fuchs_catalog_record(row, "FUCHS_ITALY_PRODUCT_FINDER", "Italy") for row in fuchs_italy_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_italy_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_italy_source_rows
    )
    fuchs_italy_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_italy_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_italy_product_key = {}
    fuchs_italy_added_rows = 0
    fuchs_italy_matched_rows = 0
    fuchs_italy_review_keys = []
    fuchs_italy_family_conflict_keys = []
    for raw, source_record in zip(fuchs_italy_source_rows, fuchs_italy_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_italy_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_italy_added_rows += 1
            if len(matches) > 1:
                fuchs_italy_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_italy_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_italy_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_sweden_source_rows = [json.loads(line) for line in FUCHS_SWEDEN_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_sweden_records = [fuchs_catalog_record(row, "FUCHS_SWEDEN_PRODUCT_FINDER", "Sweden") for row in fuchs_sweden_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows + fuchs_italy_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_sweden_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])] for row in fuchs_sweden_source_rows
    )
    fuchs_sweden_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_sweden_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_sweden_product_key = {}
    fuchs_sweden_added_rows = 0
    fuchs_sweden_matched_rows = 0
    fuchs_sweden_review_keys = []
    fuchs_sweden_family_conflict_keys = []
    for raw, source_record in zip(fuchs_sweden_source_rows, fuchs_sweden_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_sweden_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_sweden_added_rows += 1
            if len(matches) > 1:
                fuchs_sweden_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_sweden_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_sweden_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_spain_source_rows = [json.loads(line) for line in FUCHS_SPAIN_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_spain_records = [fuchs_catalog_record(row, "FUCHS_SPAIN_PRODUCT_FINDER", "Spain") for row in fuchs_spain_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows + fuchs_italy_source_rows + fuchs_sweden_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_spain_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])] for row in fuchs_spain_source_rows
    )
    fuchs_spain_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_spain_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_spain_product_key = {}
    fuchs_spain_added_rows = 0
    fuchs_spain_matched_rows = 0
    fuchs_spain_review_keys = []
    fuchs_spain_family_conflict_keys = []
    for raw, source_record in zip(fuchs_spain_source_rows, fuchs_spain_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_spain_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_spain_added_rows += 1
            if len(matches) > 1:
                fuchs_spain_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_spain_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_spain_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_france_source_rows = [json.loads(line) for line in FUCHS_FRANCE_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_france_records = [fuchs_catalog_record(row, "FUCHS_FRANCE_PRODUCT_FINDER", "France") for row in fuchs_france_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows + fuchs_italy_source_rows + fuchs_sweden_source_rows + fuchs_spain_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_france_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])] for row in fuchs_france_source_rows
    )
    fuchs_france_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_france_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_france_product_key = {}
    fuchs_france_added_rows = 0
    fuchs_france_matched_rows = 0
    fuchs_france_review_keys = []
    fuchs_france_family_conflict_keys = []
    for raw, source_record in zip(fuchs_france_source_rows, fuchs_france_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_france_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_france_added_rows += 1
            if len(matches) > 1:
                fuchs_france_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_france_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_france_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_turkey_source_rows = [json.loads(line) for line in FUCHS_TURKEY_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_turkey_records = [fuchs_catalog_record(row, "FUCHS_TURKEY_PRODUCT_FINDER", "Turkey") for row in fuchs_turkey_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows + fuchs_italy_source_rows + fuchs_sweden_source_rows + fuchs_spain_source_rows + fuchs_france_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_turkey_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])] for row in fuchs_turkey_source_rows
    )
    fuchs_turkey_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_turkey_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_turkey_product_key = {}
    fuchs_turkey_added_rows = 0
    fuchs_turkey_matched_rows = 0
    fuchs_turkey_review_keys = []
    fuchs_turkey_family_conflict_keys = []
    for raw, source_record in zip(fuchs_turkey_source_rows, fuchs_turkey_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_turkey_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_turkey_added_rows += 1
            if len(matches) > 1:
                fuchs_turkey_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_turkey_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_turkey_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_canada_source_rows = [json.loads(line) for line in FUCHS_CANADA_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_canada_records = [fuchs_catalog_record(row, "FUCHS_CANADA_PRODUCT_FINDER", "Canada") for row in fuchs_canada_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows + fuchs_italy_source_rows + fuchs_sweden_source_rows + fuchs_spain_source_rows + fuchs_france_source_rows + fuchs_turkey_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_canada_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])] for row in fuchs_canada_source_rows
    )
    fuchs_canada_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_canada_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_canada_product_key = {}
    fuchs_canada_added_rows = 0
    fuchs_canada_matched_rows = 0
    fuchs_canada_review_keys = []
    fuchs_canada_family_conflict_keys = []
    for raw, source_record in zip(fuchs_canada_source_rows, fuchs_canada_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_canada_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_canada_added_rows += 1
            if len(matches) > 1:
                fuchs_canada_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_canada_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_canada_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_china_source_rows = [json.loads(line) for line in FUCHS_CHINA_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_china_records = [fuchs_catalog_record(row, "FUCHS_CHINA_PRODUCT_FINDER", "China") for row in fuchs_china_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows + fuchs_italy_source_rows + fuchs_sweden_source_rows + fuchs_spain_source_rows + fuchs_france_source_rows + fuchs_turkey_source_rows + fuchs_canada_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_china_cross_market_exact_name_family_rows = sum(
        row["family_code"] in prior_market_name_families[normalize(row["product_name"])] for row in fuchs_china_source_rows
    )
    fuchs_china_cross_market_family_conflict_rows = sum(
        bool(prior_market_name_families[normalize(row["product_name"])])
        and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])]
        for row in fuchs_china_source_rows
    )
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_china_product_key = {}
    fuchs_china_added_rows = 0
    fuchs_china_matched_rows = 0
    fuchs_china_review_keys = []
    fuchs_china_family_conflict_keys = []
    for raw, source_record in zip(fuchs_china_source_rows, fuchs_china_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_china_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_china_added_rows += 1
            if len(matches) > 1:
                fuchs_china_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_china_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_china_product_key[raw["source_record_id"]] = target["canonical_key"]
    fuchs_czech_source_rows = [json.loads(line) for line in FUCHS_CZECH_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_czech_records = [fuchs_catalog_record(row, "FUCHS_CZECH_PRODUCT_FINDER", "Czech Republic") for row in fuchs_czech_source_rows]
    prior_market_name_families = defaultdict(set)
    for row in fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows + fuchs_italy_source_rows + fuchs_sweden_source_rows + fuchs_spain_source_rows + fuchs_france_source_rows + fuchs_turkey_source_rows + fuchs_canada_source_rows + fuchs_china_source_rows:
        prior_market_name_families[normalize(row["product_name"])].add(row["family_code"])
    fuchs_czech_cross_market_exact_name_family_rows = sum(row["family_code"] in prior_market_name_families[normalize(row["product_name"])] for row in fuchs_czech_source_rows)
    fuchs_czech_cross_market_family_conflict_rows = sum(bool(prior_market_name_families[normalize(row["product_name"])]) and row["family_code"] not in prior_market_name_families[normalize(row["product_name"])] for row in fuchs_czech_source_rows)
    existing_by_name_family = defaultdict(list)
    existing_by_name = defaultdict(list)
    for row in input_records:
        if brand_tokens_overlap("FUCHS", row["brand"]):
            existing_by_name_family[(row["product_name_normalized"], row["family_code"])].append(row)
            existing_by_name[row["product_name_normalized"]].append(row)
    fuchs_czech_product_key = {}
    fuchs_czech_added_rows = fuchs_czech_matched_rows = 0
    fuchs_czech_review_keys, fuchs_czech_family_conflict_keys = [], []
    for raw, source_record in zip(fuchs_czech_source_rows, fuchs_czech_records):
        name = normalize(raw["product_name"])
        matches = existing_by_name_family[(name, raw["family_code"])]
        if len(matches) == 1:
            target = matches[0]
            merge_fuchs_catalog_evidence(target, source_record, raw)
            fuchs_czech_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            existing_by_name_family[(name, raw["family_code"])].append(target)
            other_family_matches = [row for row in existing_by_name[name] if row["family_code"] != raw["family_code"]]
            existing_by_name[name].append(target)
            fuchs_czech_added_rows += 1
            if len(matches) > 1:
                fuchs_czech_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in matches]))
            if other_family_matches:
                fuchs_czech_family_conflict_keys.append((target["canonical_key"], [row["canonical_key"] for row in other_family_matches]))
        fuchs_czech_product_key[raw["source_record_id"]] = target["canonical_key"]
    prior_fuchs_rows = fuchs_india_source_rows + fuchs_us_source_rows + fuchs_germany_source_rows + fuchs_poland_source_rows + fuchs_italy_source_rows + fuchs_sweden_source_rows + fuchs_spain_source_rows + fuchs_france_source_rows + fuchs_turkey_source_rows + fuchs_canada_source_rows + fuchs_china_source_rows + fuchs_czech_source_rows
    fuchs_mexico_source_rows = [json.loads(line) for line in FUCHS_MEXICO_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_mexico = integrate_fuchs_market(input_records, fuchs_mexico_source_rows, "FUCHS_MEXICO_PRODUCT_FINDER", "Mexico", prior_fuchs_rows)
    fuchs_mexico_product_key = fuchs_mexico["product_key"]
    fuchs_mexico_added_rows, fuchs_mexico_matched_rows = fuchs_mexico["added"], fuchs_mexico["matched"]
    fuchs_mexico_review_keys, fuchs_mexico_family_conflict_keys = fuchs_mexico["review_keys"], fuchs_mexico["family_conflict_keys"]
    fuchs_mexico_cross_market_exact_name_family_rows, fuchs_mexico_cross_market_family_conflict_rows = fuchs_mexico["exact"], fuchs_mexico["conflicts"]
    fuchs_south_africa_source_rows = [json.loads(line) for line in FUCHS_SOUTH_AFRICA_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_south_africa = integrate_fuchs_market(input_records, fuchs_south_africa_source_rows, "FUCHS_SOUTH_AFRICA_PRODUCT_FINDER", "South Africa", prior_fuchs_rows + fuchs_mexico_source_rows)
    fuchs_south_africa_product_key = fuchs_south_africa["product_key"]
    fuchs_south_africa_added_rows, fuchs_south_africa_matched_rows = fuchs_south_africa["added"], fuchs_south_africa["matched"]
    fuchs_south_africa_review_keys, fuchs_south_africa_family_conflict_keys = fuchs_south_africa["review_keys"], fuchs_south_africa["family_conflict_keys"]
    fuchs_south_africa_cross_market_exact_name_family_rows, fuchs_south_africa_cross_market_family_conflict_rows = fuchs_south_africa["exact"], fuchs_south_africa["conflicts"]
    fuchs_brazil_source_rows = [json.loads(line) for line in FUCHS_BRAZIL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_brazil = integrate_fuchs_market(input_records, fuchs_brazil_source_rows, "FUCHS_BRAZIL_PRODUCT_FINDER", "Brazil", prior_fuchs_rows + fuchs_mexico_source_rows + fuchs_south_africa_source_rows)
    fuchs_brazil_product_key = fuchs_brazil["product_key"]
    fuchs_brazil_added_rows, fuchs_brazil_matched_rows = fuchs_brazil["added"], fuchs_brazil["matched"]
    fuchs_brazil_review_keys, fuchs_brazil_family_conflict_keys = fuchs_brazil["review_keys"], fuchs_brazil["family_conflict_keys"]
    fuchs_brazil_cross_market_exact_name_family_rows, fuchs_brazil_cross_market_family_conflict_rows = fuchs_brazil["exact"], fuchs_brazil["conflicts"]
    fuchs_norway_source_rows = [json.loads(line) for line in FUCHS_NORWAY_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_norway = integrate_fuchs_market(input_records, fuchs_norway_source_rows, "FUCHS_NORWAY_PRODUCT_FINDER", "Norway", prior_fuchs_rows + fuchs_mexico_source_rows + fuchs_south_africa_source_rows + fuchs_brazil_source_rows)
    fuchs_norway_product_key = fuchs_norway["product_key"]
    fuchs_norway_added_rows, fuchs_norway_matched_rows = fuchs_norway["added"], fuchs_norway["matched"]
    fuchs_norway_review_keys, fuchs_norway_family_conflict_keys = fuchs_norway["review_keys"], fuchs_norway["family_conflict_keys"]
    fuchs_norway_cross_market_exact_name_family_rows, fuchs_norway_cross_market_family_conflict_rows = fuchs_norway["exact"], fuchs_norway["conflicts"]
    fuchs_hungary_source_rows = [json.loads(line) for line in FUCHS_HUNGARY_JSONL.read_text(encoding="utf-8").splitlines() if line]
    fuchs_hungary = integrate_fuchs_market(input_records, fuchs_hungary_source_rows, "FUCHS_HUNGARY_PRODUCT_FINDER", "Hungary", prior_fuchs_rows + fuchs_mexico_source_rows + fuchs_south_africa_source_rows + fuchs_brazil_source_rows + fuchs_norway_source_rows)
    fuchs_hungary_product_key = fuchs_hungary["product_key"]
    fuchs_hungary_added_rows, fuchs_hungary_matched_rows = fuchs_hungary["added"], fuchs_hungary["matched"]
    fuchs_hungary_review_keys, fuchs_hungary_family_conflict_keys = fuchs_hungary["review_keys"], fuchs_hungary["family_conflict_keys"]
    fuchs_hungary_cross_market_exact_name_family_rows, fuchs_hungary_cross_market_family_conflict_rows = fuchs_hungary["exact"], fuchs_hungary["conflicts"]
    additional_fuchs = {}
    additional_prior_rows = prior_fuchs_rows + fuchs_mexico_source_rows + fuchs_south_africa_source_rows + fuchs_brazil_source_rows + fuchs_norway_source_rows + fuchs_hungary_source_rows
    for slug, source_path, source_id, market_name in FUCHS_ADDITIONAL_MARKETS:
        source_rows = [json.loads(line) for line in source_path.read_text(encoding="utf-8").splitlines() if line]
        integration = integrate_fuchs_market(input_records, source_rows, source_id, market_name, additional_prior_rows)
        additional_fuchs[slug] = {"source_path": source_path, "source_id": source_id, "source_rows": source_rows, **integration}
        additional_prior_rows += source_rows
    chemexpo_by_name_family = defaultdict(list)
    for existing in input_records:
        chemexpo_by_name_family[(existing["product_name_normalized"], existing["family_code"])].append(existing)
    epa_chemexpo_product_key = {}
    epa_chemexpo_added_rows = 0
    epa_chemexpo_matched_rows = 0
    epa_chemexpo_review_keys = []
    for raw, source_record in zip(epa_chemexpo_source_rows, epa_chemexpo_records):
        name = normalize(raw["product_name"])
        candidates_for_name = chemexpo_by_name_family[(name, raw["family_code"])]
        raw_owner = " ".join(value for value in [raw["brand_source_reported"], raw["manufacturer"]] if value)
        strong_matches = []
        for candidate in candidates_for_name:
            if candidate["evidence_status"] in {
                "official_government_product_certification_brand_scope",
                "official_government_product_certification_holder_scope",
            }:
                continue
            candidate_owner = " ".join(value for value in [candidate["brand"], candidate["manufacturer"]] if value)
            candidate_brand = normalize(candidate["brand"])
            name_supports_candidate_brand = bool(candidate_brand and len(candidate_brand) >= 4 and candidate_brand in name)
            if chemexpo_owner_tokens_overlap(raw_owner, candidate_owner) or name_supports_candidate_brand:
                strong_matches.append(candidate)
        if len(strong_matches) == 1:
            target = strong_matches[0]
            merge_epa_chemexpo_evidence(target, source_record, raw)
            epa_chemexpo_matched_rows += 1
        else:
            target = source_record
            input_records.append(target)
            chemexpo_by_name_family[(name, raw["family_code"])].append(target)
            epa_chemexpo_added_rows += 1
            if len(strong_matches) > 1:
                epa_chemexpo_review_keys.append((target["canonical_key"], [row["canonical_key"] for row in strong_matches]))
        epa_chemexpo_product_key[raw["source_record_id"]] = target["canonical_key"]
    aichilon_products, aichilon_packages, exclusions = aichilon_seed()
    existing_by_name = defaultdict(list)
    for row in input_records:
        existing_by_name[(normalize(row["brand"]), row["product_name_normalized"])].append(row)
    aichilon_product_key = {}
    aichilon_new_rows = 0
    aichilon_matched_rows = 0
    for raw in aichilon_products:
        key = (normalize(raw["brand"]), normalize(raw["name"]))
        matches = existing_by_name.get(key, [])
        if matches:
            target = matches[0]
            aichilon_matched_rows += 1
        else:
            target = canonical_record(raw)
            target["lifecycle_status"] = "active" if raw["is_active"] else "archived"
            input_records.append(target)
            existing_by_name[key].append(target)
            aichilon_new_rows += 1
        aichilon_product_key[int(raw["source_number"])] = target["canonical_key"]
    # Keep regulatory registry rows independent from the sequential FUCHS market
    # consolidation. Cross-source identities remain review candidates and cannot
    # perturb already established manufacturer-market deduplication decisions.
    input_records.extend(indonesia_npt_records)
    pertamina_source_rows = [json.loads(line) for line in PERTAMINA_OFFICIAL_JSONL.read_text(encoding="utf-8").splitlines() if line]
    pertamina_records = [pertamina_official_record(row) for row in pertamina_source_rows]
    pertamina_product_key = {}
    pertamina_added_rows = 0
    pertamina_exact_matched_rows = 0
    pertamina_blank_npt_fallback_matched_rows = 0
    pertamina_ambiguous_rows_added = 0
    pertamina_cross_family_corrections = []
    for raw, source_record in zip(pertamina_source_rows, pertamina_records):
        affiliated = [
            row for row in input_records
            if row["source_id"] != "PERTAMINA_LUBRICANTS_OFFICIAL_CATALOG"
            and is_pertamina_affiliated(row)
        ]
        core = pertamina_identity_core(source_record)
        signature = pertamina_grade_signature(source_record)
        core_matches = [row for row in affiliated if pertamina_identity_core(row) == core]
        exact_matches = [row for row in core_matches if pertamina_grade_signature(row) == signature]
        exact_npt = [row for row in exact_matches if row["source_id"] == "INDONESIA_NPT_LUBRICANT_REGISTRY"]
        match_basis = ""
        if len(exact_matches) == 1:
            target = exact_matches[0]
            match_basis = "exact_affiliated_name_core_and_professional_grade_signature"
        elif len(exact_npt) == 1:
            target = exact_npt[0]
            match_basis = "exact_signature_unique_indonesia_npt_target_among_repeated_approval_evidence"
        else:
            blank_npt = [
                row for row in core_matches
                if row["source_id"] == "INDONESIA_NPT_LUBRICANT_REGISTRY"
                and not any(pertamina_grade_signature(row))
            ]
            graded_npt = [
                row for row in core_matches
                if row["source_id"] == "INDONESIA_NPT_LUBRICANT_REGISTRY"
                and any(pertamina_grade_signature(row))
            ]
            if len(blank_npt) == 1 and not graded_npt:
                target = blank_npt[0]
                match_basis = "unique_blank_spec_indonesia_npt_identity_enriched_by_current_official_pds"
            else:
                target = source_record
                input_records.append(target)
                pertamina_added_rows += 1
                pertamina_ambiguous_rows_added += bool(exact_matches or len(blank_npt) > 1)
        if match_basis:
            old_family = target["family_code"]
            merge_pertamina_catalog_evidence(target, source_record, match_basis)
            if match_basis.startswith("unique_blank_spec"):
                pertamina_blank_npt_fallback_matched_rows += 1
            else:
                pertamina_exact_matched_rows += 1
            if old_family != raw["family_code"]:
                target["family_code"] = raw["family_code"]
                target["family"] = FAMILY_NAMES[raw["family_code"]]
                target["category"] = source_record["category"]
                target["specifications"]["pertamina_official_family_correction"] = {
                    "previous_family_code": old_family,
                    "current_official_family_code": raw["family_code"],
                }
                pertamina_cross_family_corrections.append({
                    "source_record_id": raw["source_record_id"],
                    "product_name": raw["product_name"],
                    "previous_family_code": old_family,
                    "current_official_family_code": raw["family_code"],
                    "target_source_id": target["source_id"],
                })
        pertamina_product_key[raw["source_record_id"]] = target["canonical_key"]
    input_records.extend(thailand_doeb_records)
    input_records.extend(dla_qpd_records)
    # Historical Mack approvals are appended only after every current-catalog
    # matching pass, so they cannot become accidental lifecycle-merge targets.
    mack_2014_source_rows = [json.loads(line) for line in MACK_2014_APPROVED_JSONL.read_text(encoding="utf-8").splitlines() if line]
    mack_2014_records = [mack_2014_approved_record(row) for row in mack_2014_source_rows]
    input_records.extend(mack_2014_records)
    # This is a dated corporate catalog, not the login-gated current Cummins
    # registration list. Append it after all current matching passes so its
    # historical lifecycle can never overwrite a current product identity.
    cummins_valvoline_2022_source_rows = [json.loads(line) for line in CUMMINS_VALVOLINE_2022_JSONL.read_text(encoding="utf-8").splitlines() if line]
    cummins_valvoline_2022_records = [cummins_valvoline_2022_record(row) for row in cummins_valvoline_2022_source_rows]
    input_records.extend(cummins_valvoline_2022_records)
    records, candidates = deduplicate(input_records)
    canonical_by_key = {row["canonical_key"]: row for row in records}
    for source_key, match_keys in blue_angel_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "blue_angel_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_blue_angel_multi_registry_identity",
            })
    for source_key, match_keys in blue_angel_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "same_blue_angel_product_name_but_conflicting_professional_family_across_registries",
                "score": 0.75, "decision": "keep_separate_blue_angel_family_conflict",
            })
    for source_key, match_keys in austrian_uz14_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "austrian_uz14_exact_product_name_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_austrian_uz14_multi_registry_identity",
            })
    for source_key, match_keys in korea_ecolabel_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "korea_ecolabel_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_korea_ecolabel_multi_registry_identity",
            })
    for source_key, match_keys in korea_ecolabel_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "same_korea_ecolabel_product_name_but_conflicting_professional_family_across_registries",
                "score": 0.75, "decision": "keep_separate_korea_ecolabel_family_conflict",
            })
    for source_key, match_keys in man_service_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "exact_product_name_and_family_with_probable_brand_owner_alias",
                "score": 0.99,
                "decision": "review_brand_alias_identity",
            })
    for source_key, match_keys in liqui_moly_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({"product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"], "reason": "liqui_moly_2020_exact_product_name_and_family_with_multiple_existing_registry_records", "score": 0.995, "decision": "review_liqui_moly_multi_registry_identity"})
    for source_key, match_keys in liqui_moly_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({"product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"], "reason": "same_liqui_moly_product_name_but_conflicting_professional_family", "score": 0.70, "decision": "keep_separate_liqui_moly_family_conflict"})
    for source_key, match_keys in liqui_moly_current_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "current_liqui_moly_product_links_to_multiple_2020_catalog_identities_by_name_or_article_sku",
                "score": 0.98,
                "decision": "review_liqui_moly_current_multiple_historical_candidates",
            })
    for source_key, match_keys in fuchs_india_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "fuchs_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995,
                "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_us_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "fuchs_us_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995,
                "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_us_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70,
                "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_germany_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "fuchs_germany_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995,
                "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_germany_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70,
                "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_poland_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "fuchs_poland_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995,
                "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_poland_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70,
                "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_italy_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "fuchs_italy_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995,
                "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_italy_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70,
                "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_sweden_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "fuchs_sweden_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_sweden_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70, "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_spain_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "fuchs_spain_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_spain_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70, "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_france_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "fuchs_france_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_france_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70, "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_turkey_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "fuchs_turkey_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_turkey_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70, "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_canada_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "fuchs_canada_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_canada_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70, "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_china_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "fuchs_china_exact_product_name_and_family_with_multiple_existing_registry_records",
                "score": 0.995, "decision": "review_fuchs_multi_registry_identity",
            })
    for source_key, match_keys in fuchs_china_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"],
                "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets",
                "score": 0.70, "decision": "keep_separate_fuchs_market_family_conflict",
            })
    for source_key, match_keys in fuchs_czech_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({"product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"], "reason": "fuchs_czech_exact_product_name_and_family_with_multiple_existing_registry_records", "score": 0.995, "decision": "review_fuchs_multi_registry_identity"})
    for source_key, match_keys in fuchs_czech_family_conflict_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({"product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"], "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets", "score": 0.70, "decision": "keep_separate_fuchs_market_family_conflict"})
    for market_slug, review_keys, conflict_keys in [
        ("mexico", fuchs_mexico_review_keys, fuchs_mexico_family_conflict_keys),
        ("south_africa", fuchs_south_africa_review_keys, fuchs_south_africa_family_conflict_keys),
        ("brazil", fuchs_brazil_review_keys, fuchs_brazil_family_conflict_keys),
        ("norway", fuchs_norway_review_keys, fuchs_norway_family_conflict_keys),
        ("hungary", fuchs_hungary_review_keys, fuchs_hungary_family_conflict_keys),
    ] + [(slug, data["review_keys"], data["family_conflict_keys"]) for slug, data in additional_fuchs.items()]:
        for source_key, match_keys in review_keys:
            source_product = canonical_by_key[source_key]
            for match_key in match_keys:
                match_product = canonical_by_key[match_key]
                candidates.append({"product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"], "reason": f"fuchs_{market_slug}_exact_product_name_and_family_with_multiple_existing_registry_records", "score": 0.995, "decision": "review_fuchs_multi_registry_identity"})
        for source_key, match_keys in conflict_keys:
            source_product = canonical_by_key[source_key]
            for match_key in match_keys:
                match_product = canonical_by_key[match_key]
                candidates.append({"product_id_a": match_product["product_id"], "product_id_b": source_product["product_id"], "reason": "same_fuchs_product_name_but_conflicting_professional_family_across_markets", "score": 0.70, "decision": "keep_separate_fuchs_market_family_conflict"})
    for source_key, match_keys in epa_chemexpo_review_keys:
        source_product = canonical_by_key[source_key]
        for match_key in match_keys:
            match_product = canonical_by_key[match_key]
            candidates.append({
                "product_id_a": match_product["product_id"],
                "product_id_b": source_product["product_id"],
                "reason": "epa_chemexpo_exact_product_name_and_family_with_multiple_owner_supported_existing_records",
                "score": 0.99,
                "decision": "review_chemexpo_multi_registry_identity",
            })
    source_links = [{
        "product_id": row["product_id"], "source_id": row["source_id"], "source_record_id": row["source_record_id"],
        "source_row": row["source_row"], "relation": "primary_seed_record",
    } for row in records]
    source_link_keys = {(link["product_id"], link["source_id"], link["source_record_id"]) for link in source_links}
    for raw in anp_monitoring_source_rows:
        target = canonical_by_key[anp_monitoring_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"],
            "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"],
            "source_row": raw.get("source_row", raw.get("source_page")),
            "relation": "official_government_historical_market_monitoring_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in shenzhen_china_2016_2017_source_rows:
        target = canonical_by_key[shenzhen_china_2016_2017_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in qingdao_china_source_rows:
        target = canonical_by_key[qingdao_china_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in jilin_china_source_rows:
        target = canonical_by_key[jilin_china_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in beijing_china_source_rows:
        target = canonical_by_key[beijing_china_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in samr_china_2023_source_rows:
        target = canonical_by_key[samr_china_2023_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_government_nonconforming_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in shenzhen_china_2019_source_rows:
        target = canonical_by_key[shenzhen_china_2019_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in shenzhen_china_2020_source_rows:
        target = canonical_by_key[shenzhen_china_2020_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in shenzhen_china_2025_source_rows:
        target = canonical_by_key[shenzhen_china_2025_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in shenzhen_china_2021_source_rows:
        target = canonical_by_key[shenzhen_china_2021_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_government_nonconforming_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in shanghai_china_source_rows:
        target = canonical_by_key[shanghai_china_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_product_inspection_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in aichilon_products:
        target = canonical_by_key[aichilon_product_key[int(raw["source_number"])]]
        link = {
            "product_id": target["product_id"], "source_id": "aichilon-internal", "source_record_id": text(raw["source_number"]),
            "source_row": None, "relation": "internal_product_master",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(jaso_source_rows, jaso_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": text(raw["source_row_number"]), "source_row": raw["source_row_number"],
            "relation": "official_filed_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(licensed_source_rows, licensed_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": raw["source_id"],
            "source_record_id": text(raw["source_record_id"]), "source_row": raw["source_row_number"],
            "relation": "official_licensed_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in blue_angel_source_rows:
        target = canonical_by_key[blue_angel_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "BLUE_ANGEL_DE_UZ_178",
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row_numbers"][0],
            "relation": "official_ecolabel_product_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in austrian_uz14_source_rows:
        target = canonical_by_key[austrian_uz14_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "AUSTRIAN_ECOLABEL_UZ14_LUBRICANTS",
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row"],
            "relation": "official_government_ecolabel_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in korea_ecolabel_source_rows:
        target = canonical_by_key[korea_ecolabel_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "KOREA_ECOLABEL_EL611",
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row_numbers"][0],
            "relation": "official_government_ecolabel_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in korea_el509_source_rows:
        target = canonical_by_key[korea_el509_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "KOREA_ECOLABEL_EL509",
            "source_record_id": raw["source_record_id"], "source_row": raw["source_row_numbers"][0],
            "relation": "official_government_ecolabel_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(green_choice_philippines_source_rows, green_choice_philippines_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"],
            "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"],
            "source_row": raw["source_row_number"],
            "relation": "official_government_ecolabel_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in epa_chemexpo_source_rows:
        target = canonical_by_key[epa_chemexpo_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"],
            "source_id": raw["source_id"],
            "source_record_id": raw["source_record_id"],
            "source_row": None,
            "relation": "official_government_compiled_product_database",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(biopreferred_source_rows, biopreferred_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "usda-biopreferred",
            "source_record_id": text(raw["product_id"]), "source_row": None,
            "relation": "official_government_program_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(zf_source_rows, zf_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "ZF_TE_ML",
            "source_record_id": text(raw["approval_number"]), "source_row": None,
            "relation": "official_oem_approval_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(allison_source_rows, allison_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "ALLISON_APPROVED_FLUIDS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_oem_approval_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(driventic_source_rows, driventic_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "DRIVENTIC_DIWA_APPROVED_OILS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_oem_approval_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(mercedes_dtfr_source_rows, mercedes_dtfr_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "MERCEDES_DTFR_APPROVED_FLUIDS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_oem_approval_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in mercedes_bevo_source_rows:
        target = canonical_by_key[mercedes_bevo_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "MERCEDES_BENZ_BEVO_APPROVED_FLUIDS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_oem_approval_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(volvo_genuine_source_rows, volvo_genuine_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "VOLVO_GENUINE_FLUIDS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(mack_genuine_source_rows, mack_genuine_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "MACK_GENUINE_FLUIDS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(mack_2014_source_rows, mack_2014_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "MACK_2014_APPROVED_OILS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "historical_official_oem_approval_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(cummins_valvoline_2022_source_rows, cummins_valvoline_2022_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "CUMMINS_VALVOLINE_EU_2022_CATALOG",
            "source_record_id": raw["source_record_id"], "source_row": raw["source_page"],
            "relation": "historical_official_manufacturer_oem_corporate_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(scania_genuine_source_rows, scania_genuine_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "SCANIA_GENUINE_ENGINE_OILS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(brava_official_source_rows, brava_official_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "BRAVA_LUBRICANTS_OFFICIAL_CATALOG",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(volvo_websds_source_rows, volvo_websds_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"],
            "source_id": "VOLVO_GROUP_WEBSDS_CHANGE_ARCHIVE",
            "source_record_id": raw["source_record_id"],
            "source_row": None,
            "relation": "official_manufacturer_sds_change_observation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(parker_denison_source_rows, parker_denison_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"],
            "source_id": "PARKER_DENISON_CURRENT_FLUID_RATINGS",
            "source_record_id": raw["source_record_id"],
            "source_row": raw["source_table_row"],
            "relation": "official_oem_approval_registry",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw, normalized_row in zip(ceypetco_source_rows, ceypetco_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        link = {
            "product_id": target["product_id"], "source_id": "CEYPETCO_OFFICIAL_LUBRICANT_CATALOG",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_state_owned_supplier_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in pso_source_rows:
        target = canonical_by_key[pso_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"],
            "source_id": "PAKISTAN_STATE_OIL_OFFICIAL_CATALOG",
            "source_record_id": raw["source_record_id"],
            "source_row": None,
            "relation": "official_state_owned_manufacturer_product_catalog_noncommercial_attribution",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in pertamina_source_rows:
        target = canonical_by_key[pertamina_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"],
            "source_id": "PERTAMINA_LUBRICANTS_OFFICIAL_CATALOG",
            "source_record_id": raw["source_record_id"],
            "source_row": None,
            "relation": "official_state_owned_manufacturer_current_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in man_service_source_rows:
        target = canonical_by_key[man_service_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "MAN_CURRENT_SERVICE_PRODUCTS",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_oem_service_recommendation",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in liqui_moly_source_rows:
        target = canonical_by_key[liqui_moly_product_key[raw["source_record_id"]]]
        link = {"product_id": target["product_id"], "source_id": "LIQUI_MOLY_2020_PRODUCT_CATALOG", "source_record_id": raw["source_record_id"], "source_row": None, "relation": "historical_official_manufacturer_product_catalog"}
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in liqui_moly_current_source_rows:
        target = canonical_by_key[liqui_moly_current_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"],
            "source_id": "LIQUI_MOLY_CURRENT_OPENAPI",
            "source_record_id": raw["source_record_id"],
            "source_row": None,
            "relation": "current_official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_india_source_rows:
        target = canonical_by_key[fuchs_india_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_INDIA_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_us_source_rows:
        target = canonical_by_key[fuchs_us_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_US_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_germany_source_rows:
        target = canonical_by_key[fuchs_germany_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_GERMANY_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_poland_source_rows:
        target = canonical_by_key[fuchs_poland_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_POLAND_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_italy_source_rows:
        target = canonical_by_key[fuchs_italy_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_ITALY_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_sweden_source_rows:
        target = canonical_by_key[fuchs_sweden_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_SWEDEN_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_spain_source_rows:
        target = canonical_by_key[fuchs_spain_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_SPAIN_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_france_source_rows:
        target = canonical_by_key[fuchs_france_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_FRANCE_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_turkey_source_rows:
        target = canonical_by_key[fuchs_turkey_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_TURKEY_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_canada_source_rows:
        target = canonical_by_key[fuchs_canada_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_CANADA_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_china_source_rows:
        target = canonical_by_key[fuchs_china_product_key[raw["source_record_id"]]]
        link = {
            "product_id": target["product_id"], "source_id": "FUCHS_CHINA_PRODUCT_FINDER",
            "source_record_id": raw["source_record_id"], "source_row": None,
            "relation": "official_manufacturer_product_catalog",
        }
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for raw in fuchs_czech_source_rows:
        target = canonical_by_key[fuchs_czech_product_key[raw["source_record_id"]]]
        link = {"product_id": target["product_id"], "source_id": "FUCHS_CZECH_PRODUCT_FINDER", "source_record_id": raw["source_record_id"], "source_row": None, "relation": "official_manufacturer_product_catalog"}
        link_key = (link["product_id"], link["source_id"], link["source_record_id"])
        if link_key not in source_link_keys:
            source_links.append(link)
            source_link_keys.add(link_key)
    for source_rows, product_key, source_id in [
        (fuchs_mexico_source_rows, fuchs_mexico_product_key, "FUCHS_MEXICO_PRODUCT_FINDER"),
        (fuchs_south_africa_source_rows, fuchs_south_africa_product_key, "FUCHS_SOUTH_AFRICA_PRODUCT_FINDER"),
        (fuchs_brazil_source_rows, fuchs_brazil_product_key, "FUCHS_BRAZIL_PRODUCT_FINDER"),
        (fuchs_norway_source_rows, fuchs_norway_product_key, "FUCHS_NORWAY_PRODUCT_FINDER"),
        (fuchs_hungary_source_rows, fuchs_hungary_product_key, "FUCHS_HUNGARY_PRODUCT_FINDER"),
    ] + [(data["source_rows"], data["product_key"], data["source_id"]) for data in additional_fuchs.values()]:
        for raw in source_rows:
            target = canonical_by_key[product_key[raw["source_record_id"]]]
            link = {"product_id": target["product_id"], "source_id": source_id, "source_record_id": raw["source_record_id"], "source_row": None, "relation": "official_manufacturer_product_catalog"}
            link_key = (link["product_id"], link["source_id"], link["source_record_id"])
            if link_key not in source_link_keys:
                source_links.append(link)
                source_link_keys.add(link_key)
    offers = []
    for package in aichilon_packages:
        canonical_key = aichilon_product_key.get(int(package["source_product_id"]))
        if not canonical_key:
            continue
        target = canonical_by_key[canonical_key]
        offers.append({
            "offer_id": f"AICHILON-PACKAGE-{package['package_id']}",
            "product_id": target["product_id"],
            "market": "UZ",
            "package_name": text(package["package_name"]),
            "unit": text(package["unit"]),
            "quantity_per_package": package["quantity_per_package"],
            "weight_kg": package["weight_kg"],
            "density_kg_per_l": package["density_kg_per_l"],
            "lifecycle_status": "active" if package["is_active"] else "archived",
            "archive_type": text(package["archive_type"]),
            "archive_reason": text(package["archive_reason"]),
            "source_id": "aichilon-internal",
            "source_record_id": text(package["package_id"]),
        })
    for raw in liqui_moly_current_source_rows:
        target = canonical_by_key[liqui_moly_current_product_key[raw["source_record_id"]]]
        for article in raw["articles"]:
            package_match = re.fullmatch(r"([0-9]+(?:[.,][0-9]+)?)\s*([A-Za-z]+)", article["content"])
            quantity = float(package_match.group(1).replace(",", ".")) if package_match else None
            unit = package_match.group(2).lower() if package_match else article["content"]
            weight_kg = None
            if quantity is not None and unit == "kg":
                weight_kg = quantity
            elif quantity is not None and unit == "g":
                weight_kg = quantity / 1000
            offers.append({
                "offer_id": f"LIQUI-MOLY-GB-{article['sku']}",
                "product_id": target["product_id"],
                "market": "GB",
                "package_name": " / ".join(value for value in [article["content"], article["container_type"]] if value),
                "unit": unit,
                "quantity_per_package": quantity,
                "weight_kg": weight_kg,
                "density_kg_per_l": None,
                "lifecycle_status": "listed_current_catalog",
                "archive_type": "",
                "archive_reason": "",
                "source_id": "LIQUI_MOLY_CURRENT_OPENAPI",
                "source_record_id": article["sku"],
            })
    for raw, normalized_row in zip(brava_official_source_rows, brava_official_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        for package in raw["packages"]:
            package_name = package["package_name"]
            quantity = None
            unit = "package"
            if package_name == "Quart":
                quantity, unit = 1.0, "qt"
            else:
                package_match = re.match(r"([0-9]+(?:[.,][0-9]+)?)\s*(L|GL|oz)\b", package_name, flags=re.I)
                if package_match:
                    quantity = float(package_match.group(1).replace(",", "."))
                    unit = {"l": "l", "gl": "gal", "oz": "fl oz"}[package_match.group(2).lower()]
            offers.append({
                "offer_id": f"BRAVA-{raw['source_record_id']}-{package['part_number']}",
                "product_id": target["product_id"],
                "market": "PR_US",
                "package_name": package_name,
                "unit": unit,
                "quantity_per_package": quantity,
                "weight_kg": None,
                "density_kg_per_l": None,
                "lifecycle_status": "listed_current_catalog",
                "archive_type": "",
                "archive_reason": "",
                "source_id": "BRAVA_LUBRICANTS_OFFICIAL_CATALOG",
                "source_record_id": package["part_number"],
            })
    for raw, normalized_row in zip(mack_genuine_source_rows, mack_genuine_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        for package in raw["packages"]:
            package_name = package["package_name"]
            quantity = None
            unit = "package"
            package_match = re.match(r"([0-9]+(?:[.,][0-9]+)?)\s*(Gallon|Oz|kg)\b", package_name, flags=re.I)
            if package_match:
                quantity = float(package_match.group(1).replace(",", "."))
                unit = {"gallon": "gal", "oz": "oz", "kg": "kg"}[package_match.group(2).lower()]
            offers.append({
                "offer_id": f"MACK-{raw['source_record_id']}-{hashlib.sha256(package['part_number'].encode()).hexdigest()[:16]}",
                "product_id": target["product_id"],
                "market": "BAHAMAS_EN_MACK",
                "package_name": package_name,
                "unit": unit,
                "quantity_per_package": quantity,
                "weight_kg": quantity if unit == "kg" else None,
                "density_kg_per_l": None,
                "lifecycle_status": "listed_current_catalog",
                "archive_type": "",
                "archive_reason": "",
                "source_id": "MACK_GENUINE_FLUIDS",
                "source_record_id": package["part_number"],
            })
    for raw, normalized_row in zip(cummins_valvoline_2022_source_rows, cummins_valvoline_2022_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        for package in raw["packages"]:
            offer_identity = "|".join([
                raw["source_record_id"], package["article_number"], package["variant"], package["package_name"]
            ])
            offers.append({
                "offer_id": "CUMMINS-VALVOLINE-2022-" + hashlib.sha256(offer_identity.encode()).hexdigest()[:20],
                "product_id": target["product_id"],
                "market": raw["market"],
                "package_name": " / ".join([package["package_name"], package["variant"]]),
                "unit": "catalog_package",
                "quantity_per_package": None,
                "weight_kg": None,
                "density_kg_per_l": None,
                "lifecycle_status": "historical_catalog_current_status_unverified",
                "archive_type": "dated_catalog_snapshot",
                "archive_reason": "Current availability is not established by the April 2022 catalog.",
                "source_id": "CUMMINS_VALVOLINE_EU_2022_CATALOG",
                "source_record_id": package["article_number"],
            })
    for raw, normalized_row in zip(taiwan_cpc_source_rows, taiwan_cpc_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        for package_index, package in enumerate(raw["packages"], 1):
            package_name = package["package_name_source"]
            if package.get("units_per_case"):
                package_name += f" ({package['units_per_case']} units/case)"
            offer_identity = f"{raw['source_record_id']}|{package_index}|{package_name}"
            quantity = package.get("quantity")
            unit = package.get("unit") or "catalog_package"
            offers.append({
                "offer_id": "TAIWAN-CPC-" + hashlib.sha256(offer_identity.encode()).hexdigest()[:20],
                "product_id": target["product_id"],
                "market": raw["market"],
                "package_name": package_name,
                "unit": unit,
                "quantity_per_package": quantity,
                "weight_kg": quantity if unit == "kg" else None,
                "density_kg_per_l": None,
                "lifecycle_status": "listed_current_catalog",
                "archive_type": "",
                "archive_reason": "",
                "source_id": "TAIWAN_CPC_CURRENT_LUBRICANT_CATALOG",
                "source_record_id": f"{raw['source_record_id']}:package:{package_index}",
            })
    issues = quality_issues(records)
    cpc_issue_rules = {
        "source_multigrade_table_not_safely_aligned_to_listing_title": (
            "medium", "ISO_VG", "One unambiguous source grade mapped to this exact CPC product code",
            "Retain the product without an ISO VG assertion until CPC publishes an unambiguous product-specific mapping.",
        ),
        "source_package_text_present_but_no_product_specific_structured_offer": (
            "medium", "PACKAGE", "One or more package facts attributable to this exact CPC product code",
            "Retain the source package statement but exclude it from structured offers until attribution is unambiguous.",
        ),
    }
    for raw, normalized_row in zip(taiwan_cpc_source_rows, taiwan_cpc_records):
        target = canonical_by_key[normalized_row["canonical_key"]]
        for flag in raw["source_quality_flags"]:
            severity, field, expected, action = cpc_issue_rules[flag]
            issues.append({
                "product_id": target["product_id"],
                "issue_code": flag,
                "severity": severity,
                "field": field,
                "value": raw["product_name_source"],
                "expected": expected,
                "action": action,
            })
    issues.extend({
        "product_id": row["product_id"],
        "issue_code": "source_registration_number_missing",
        "severity": "high",
        "field": "INDONESIA_NPT_REGISTRATION_NUMBER",
        "value": row["specifications"]["indonesia_npt_registration_number_raw"],
        "expected": "Published NPT registration number",
        "action": "Retain as product-name evidence, but do not treat it as a verified registered product until the authority corrects the source row.",
    } for row in records if row["evidence_status"] == "official_government_registry_source_data_issue")
    issues.extend({
        "product_id": row["product_id"],
        "issue_code": "thailand_doeb_registration_number_collision",
        "severity": "high",
        "field": "THAILAND_DOEB_REGISTRATION_NUMBER",
        "value": "; ".join(code["value"] for code in row["codes"].values() if code.get("system") == "THAILAND_DOEB_REGISTRATION_NUMBER"),
        "expected": "One registration number per product and SAE identity",
        "action": "Keep conflicting rows separate and request source-authority clarification before treating the registration number as a unique identity.",
    } for row in records if row["source_id"] == "THAILAND_DOEB_LUBRICANT_REGISTRY" and "registration_number_collision" in row["specifications"]["source_quality_flags"])
    issues.extend({
        "product_id": row["product_id"],
        "issue_code": "thailand_doeb_nonstandard_sae_notation",
        "severity": "medium",
        "field": "SAE",
        "value": "; ".join(row["specifications"]["sae_source_raw_values"]),
        "expected": "Recognized SAE J300 viscosity notation",
        "action": "Retain the source value but exclude it from strict SAE equivalence until the authority or an official TDS confirms the grade.",
    } for row in records if row["source_id"] == "THAILAND_DOEB_LUBRICANT_REGISTRY" and "nonstandard_sae_notation" in row["specifications"]["source_quality_flags"])
    ceypetco_issue_meta = {
        "conflicting_sae_within_current_tds": ("high", "SAE", "One unambiguous SAE grade in the current TDS", "Retain both source values but keep SAE empty in strict equivalence until Ceypetco publishes a corrected TDS."),
        "tds_color_table_conflicts_with_product_variant": ("medium", "COOLANT_COLOR", "Color consistent with the TDS title and product variant", "Retain the Red product identity and flag the contradictory Green property cell for supplier clarification."),
        "nonstandard_oem_notation_retained_verbatim": ("medium", "OEM_SPECIFICATION", "Recognized OEM specification notation", "Retain the printed value verbatim and do not normalize it to a guessed Ford code without corrected source evidence."),
    }
    for row in records:
        if row["source_id"] != "CEYPETCO_OFFICIAL_LUBRICANT_CATALOG":
            continue
        for flag in row["specifications"].get("source_quality_flags", []):
            severity, field, expected, action = ceypetco_issue_meta[flag]
            issues.append({
                "product_id": row["product_id"],
                "issue_code": f"ceypetco_{flag}",
                "severity": severity,
                "field": field,
                "value": row["specifications"].get("source_conflict_note", flag),
                "expected": expected,
                "action": action,
            })
    pso_issue_meta = {
        "current_page_grade_not_observed_in_linked_pds": ("medium", "GRADE", "A grade confirmed by the linked current PDS", "Retain the current page grade with lower evidence and exclude it from PDS-confirmed claims until PSO updates the document."),
        "linked_pds_grade_not_listed_in_current_page_grade_summary": ("low", "GRADE", "Consistent grade lists on the current page and linked PDS", "Retain the PDS-confirmed grade and record the page-summary omission."),
        "current_page_ow_20_typo_resolved_by_linked_pds_0w_20": ("medium", "SAE", "Recognized SAE 0W-20 notation", "Use 0W-20 from the linked PDS and retain the webpage typo only as a provenance flag."),
        "current_page_title_break_typo_resolved_by_linked_pds_brake": ("low", "PRODUCT_NAME", "Hydraulic Brake Fluid", "Use the product name proven by the linked PDS and retain the webpage title typo as a provenance flag."),
        "dot_4_table_grade_without_separate_dot_4_performance_bullet": ("high", "BRAKE_FLUID_CLASS", "A product-specific DOT 4 performance statement", "Retain the DOT 4 table grade, but do not infer FMVSS 116 DOT 4 compliance from the separate DOT 3 performance bullet."),
    }
    for raw in pso_source_rows:
        target = canonical_by_key[pso_product_key[raw["source_record_id"]]]
        for flag in raw["source_quality_flags"]:
            severity, field, expected, action = pso_issue_meta[flag]
            issues.append({
                "product_id": target["product_id"],
                "issue_code": f"pso_{flag}",
                "severity": severity,
                "field": field,
                "value": flag,
                "expected": expected,
                "action": action,
            })
    for raw in pertamina_source_rows:
        target = canonical_by_key[pertamina_product_key[raw["source_record_id"]]]
        if "linked_pds_title_conflicts_with_current_product_card" in raw["source_quality_flags"]:
            issues.append({
                "product_id": target["product_id"],
                "issue_code": "pertamina_linked_pds_title_conflicts_with_current_product_card",
                "severity": "high",
                "field": "TECHNICAL_DOCUMENT",
                "value": "Rusguard Lube 10 card links to a PDS titled Rusguard Lube X",
                "expected": "A linked PDS identifying the same current catalog product",
                "action": "Retain the official product card and link as provenance, but exclude the mismatched PDF from technical evidence for Rusguard Lube 10.",
            })
    for correction in pertamina_cross_family_corrections:
        target = canonical_by_key[pertamina_product_key[correction["source_record_id"]]]
        issues.append({
            "product_id": target["product_id"],
            "issue_code": "pertamina_current_official_family_corrects_prior_registry_family",
            "severity": "medium",
            "field": "FAMILY_CODE",
            "value": correction["previous_family_code"],
            "expected": correction["current_official_family_code"],
            "action": "Use the current manufacturer catalog family for professional analytics and retain the previous registry family as auditable source history.",
        })
    brava_issue_meta = {
        "nonstandard_acea_class_source_reported_verbatim": ("medium", "ACEA", "A recognized published ACEA class", "Retain A4/B4 verbatim but exclude it from strict equivalence until the manufacturer publishes a corrected class."),
        "source_part_number_placeholder_excluded": ("low", "PART_NUMBER", "A concrete manufacturer part number", "Keep the product-grade row but do not create an offer/code from the source placeholder."),
        "source_part_number_cross_product_collision": ("high", "PART_NUMBER", "One product-grade identity per current manufacturer part number", "Retain both source assignments as separate offers and request manufacturer clarification before treating the part number as a unique identity."),
        "vague_additional_specifications_not_enumerated": ("medium", "PERFORMANCE", "An explicitly enumerated specification", "Exclude the words 'and more' from equivalence and retain only the named Allison C-4 claim."),
    }
    for row in records:
        if row["source_id"] != "BRAVA_LUBRICANTS_OFFICIAL_CATALOG":
            continue
        for flag in row["specifications"].get("source_quality_flags", []):
            severity, field, expected, action = brava_issue_meta[flag]
            issues.append({
                "product_id": row["product_id"],
                "issue_code": f"brava_{flag}",
                "severity": severity,
                "field": field,
                "value": flag,
                "expected": expected,
                "action": action,
            })
    mack_issue_meta = {
        "source_part_number_contains_embedded_formatting_spaces_retained_verbatim": ("medium", "PART_NUMBER", "A consistently formatted manufacturer part number", "Retain the exact catalog value, but do not silently remove embedded spaces or use it as a cross-market identity without confirmation."),
        "source_sds_link_target_name_mismatch_not_used_as_technical_evidence": ("high", "SDS_URL", "An SDS whose target document identifies the same product", "Keep the official catalog product and part numbers, but exclude the mismatched SDS link from technical-composition evidence."),
        "official_regional_part_number_variants_preserved_separately": ("low", "PART_NUMBER", "Market-specific manufacturer part-number evidence", "Preserve both official regional variants and do not collapse them into one code without a manufacturer cross-reference."),
        "source_mack_standard_37319_retained_verbatim_not_silently_corrected": ("medium", "OEM_SPECIFICATION", "A confirmed current Mack transmission-fluid standard", "Retain 37319 exactly as printed on the official page, but exclude it from asserted equivalence until Mack confirms or corrects the notation."),
    }
    for row in records:
        if row["source_id"] != "MACK_GENUINE_FLUIDS":
            continue
        for flag in row["specifications"].get("source_quality_flags", []):
            severity, field, expected, action = mack_issue_meta[flag]
            issues.append({
                "product_id": row["product_id"],
                "issue_code": f"mack_{flag}",
                "severity": severity,
                "field": field,
                "value": flag,
                "expected": expected,
                "action": action,
            })
    mack_2014_issue_meta = {
        "source_duplicate_approval_row_merged": ("low", "SOURCE_OCCURRENCE", "One source occurrence per approval identity", "Preserve both page/row occurrences while exposing one normalized historical product."),
        "source_multi_grade_approval_row_split": ("low", "SAE", "One product-grade identity per viscosity", "Split the source row into separate grade identities and retain the shared source occurrence."),
        "viscosity_inferred_from_product_name_due_empty_table_cell": ("medium", "SAE", "A populated source viscosity cell", "Use the explicit 75W-90 printed in the product name, retain the empty-cell anomaly, and require current evidence before present-day equivalence."),
        "cross_section_same_product_identity_merged": ("low", "OEM_APPROVAL", "One identity carrying every source approval section", "Merge the exact supplier-product-grade identity and retain both historical approval sections."),
        "source_name_grade_notation_variants_merged": ("low", "PRODUCT_NAME", "Consistent grade punctuation", "Normalize the SAE notation while retaining both source spellings and occurrences."),
    }
    for row in records:
        if row["source_id"] != "MACK_2014_APPROVED_OILS":
            continue
        for flag in row["specifications"].get("source_quality_flags", []):
            severity, field, expected, action = mack_2014_issue_meta[flag]
            issues.append({
                "product_id": row["product_id"],
                "issue_code": f"mack_2014_{flag}",
                "severity": severity,
                "field": field,
                "value": flag,
                "expected": expected,
                "action": action,
            })
    cummins_valvoline_issue_meta = {
        "source_product_title_missing_viscosity_grade": ("medium", "ISO_VG", "A product title carrying an explicit viscosity grade", "Retain the source title without guessing a grade; exclude it from strict ISO VG equivalence until another official document confirms the value."),
        "source_product_title_hlvp_notation_retained_verbatim": ("medium", "PRODUCT_NAME", "Consistent HVLP notation", "Retain HLVP exactly as printed in the catalog title; do not silently normalize it to HVLP for product identity."),
        "source_article_number_cross_product_collision_retained_verbatim": ("high", "ARTICLE_NUMBER", "One product identity per manufacturer article number", "Preserve every printed assignment, but do not use the colliding article number alone as a cross-product identity without current manufacturer confirmation."),
    }
    for row in records:
        if row["source_id"] != "CUMMINS_VALVOLINE_EU_2022_CATALOG":
            continue
        for flag in row["specifications"].get("source_quality_flags", []):
            severity, field, expected, action = cummins_valvoline_issue_meta[flag]
            issues.append({
                "product_id": row["product_id"],
                "issue_code": f"cummins_valvoline_2022_{flag}",
                "severity": severity,
                "field": field,
                "value": flag,
                "expected": expected,
                "action": action,
            })
    issues.extend({
        "product_id": row["product_id"],
        "issue_code": "dla_qpd_lifecycle_restriction",
        "severity": "high" if row["lifecycle_status"] in {"stop_ship", "sam_inactive_source_review"} else "medium",
        "field": "DLA_QPD_QUALIFICATION_STATUS",
        "value": row["lifecycle_status"],
        "expected": "qualified_source_certified",
        "action": "Retain as official qualification history, but review the live QPD status before sourcing or declaring current equivalence.",
    } for row in records if row["evidence_status"] == "official_government_qualified_product_registry" and row["lifecycle_status"] != "qualified_source_certified")
    issues.extend({
        "product_id": row["product_id"],
        "issue_code": "parker_denison_source_invalid_validity_month",
        "severity": "high",
        "field": "RATING_VALIDITY_UNTIL",
        "value": row["specifications"]["rating_validity_until_source_reported"],
        "expected": "YYYY-MM with a calendar month from 01 to 12",
        "action": "Retain the official row, but do not assert current validity until Parker publishes a corrected value.",
    } for row in records if row["source_id"] == "PARKER_DENISON_CURRENT_FLUID_RATINGS" and row["lifecycle_status"] == "source_validity_value_invalid_review_required")
    run_id = f"seed-{SNAPSHOT_DATE}"
    JSONL_OUT.write_text("".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records), encoding="utf-8")
    compress_jsonl()
    report = {
        "schema_version": SCHEMA_VERSION,
        "run_id": run_id,
        "snapshot_date": SNAPSHOT_DATE,
        "status": "seed_only_world_catalog_incomplete",
        "input_rows": len(input_records),
        "normalized_input_sha256": hashlib.sha256(CATALOG.read_bytes()).hexdigest(),
        "jaso_normalized_input_sha256": hashlib.sha256(JASO_JSONL.read_bytes()).hexdigest(),
        "official_licensed_input_sha256": hashlib.sha256(LICENSED_JSONL.read_bytes()).hexdigest(),
        "blue_angel_input_sha256": hashlib.sha256(BLUE_ANGEL_JSONL.read_bytes()).hexdigest(),
        "austrian_ecolabel_uz14_input_sha256": hashlib.sha256(AUSTRIAN_ECOLABEL_UZ14_JSONL.read_bytes()).hexdigest(),
        "korea_ecolabel_input_sha256": hashlib.sha256(KOREA_ECOLABEL_JSONL.read_bytes()).hexdigest(),
        "korea_ecolabel_el509_input_sha256": hashlib.sha256(KOREA_ECOLABEL_EL509_JSONL.read_bytes()).hexdigest(),
        "green_choice_philippines_input_sha256": hashlib.sha256(GREEN_CHOICE_PHILIPPINES_JSONL.read_bytes()).hexdigest(),
        "uae_moiat_input_sha256": hashlib.sha256(UAE_MOIAT_JSONL.read_bytes()).hexdigest(),
        "eaeu_conformity_input_sha256": hashlib.sha256(EAEU_CONFORMITY_JSONL.read_bytes()).hexdigest(),
        "epa_safer_choice_input_sha256": hashlib.sha256(EPA_SAFER_CHOICE_JSONL.read_bytes()).hexdigest(),
        "epa_chemexpo_input_sha256": hashlib.sha256(EPA_CHEMEXPO_JSONL.read_bytes()).hexdigest(),
        "psqca_engine_oil_input_sha256": hashlib.sha256(PSQCA_ENGINE_OIL_JSONL.read_bytes()).hexdigest(),
        "tisi_two_stroke_oil_input_sha256": hashlib.sha256(TISI_TWO_STROKE_OIL_JSONL.read_bytes()).hexdigest(),
        "samr_china_2025_input_sha256": hashlib.sha256(SAMR_CHINA_2025_JSONL.read_bytes()).hexdigest(),
        "samr_china_2023_input_sha256": hashlib.sha256(SAMR_CHINA_2023_JSONL.read_bytes()).hexdigest(),
        "samr_china_2024_input_sha256": hashlib.sha256(SAMR_CHINA_2024_JSONL.read_bytes()).hexdigest(),
        "shenzhen_china_2021_input_sha256": hashlib.sha256(SHENZHEN_CHINA_2021_JSONL.read_bytes()).hexdigest(),
        "shenzhen_china_2020_input_sha256": hashlib.sha256(SHENZHEN_CHINA_2020_JSONL.read_bytes()).hexdigest(),
        "shenzhen_china_2019_input_sha256": hashlib.sha256(SHENZHEN_CHINA_2019_JSONL.read_bytes()).hexdigest(),
        "shenzhen_china_2025_input_sha256": hashlib.sha256(SHENZHEN_CHINA_2025_JSONL.read_bytes()).hexdigest(),
        "shanghai_china_2023_2025_input_sha256": hashlib.sha256(SHANGHAI_CHINA_2023_2025_JSONL.read_bytes()).hexdigest(),
        "beijing_china_2018_input_sha256": hashlib.sha256(BEIJING_CHINA_2018_JSONL.read_bytes()).hexdigest(),
        "shenzhen_china_2016_2017_input_sha256": hashlib.sha256(SHENZHEN_CHINA_2016_2017_JSONL.read_bytes()).hexdigest(),
        "qingdao_china_2021_2025_input_sha256": hashlib.sha256(QINGDAO_CHINA_2021_2025_JSONL.read_bytes()).hexdigest(),
        "jilin_china_2024_2025_input_sha256": hashlib.sha256(JILIN_CHINA_2024_2025_JSONL.read_bytes()).hexdigest(),
        "philippines_bps_brake_fluid_input_sha256": hashlib.sha256(PHILIPPINES_BPS_BRAKE_FLUID_JSONL.read_bytes()).hexdigest(),
        "ghana_gsa_certified_input_sha256": hashlib.sha256(GHANA_GSA_CERTIFIED_JSONL.read_bytes()).hexdigest(),
        "kebs_smark_input_sha256": hashlib.sha256(KEBS_SMARK_JSONL.read_bytes()).hexdigest(),
        "east_africa_certified_input_sha256": hashlib.sha256(EAST_AFRICA_CERTIFIED_JSONL.read_bytes()).hexdigest(),
        "son_mancap_input_sha256": hashlib.sha256(SON_MANCAP_JSONL.read_bytes()).hexdigest(),
        "rsb_smark_input_sha256": hashlib.sha256(RSB_SMARK_JSONL.read_bytes()).hexdigest(),
        "usda_biopreferred_input_sha256": hashlib.sha256(USDA_BIOPREFERRED_JSONL.read_bytes()).hexdigest(),
        "zf_te_ml_input_sha256": hashlib.sha256(ZF_TE_ML_JSONL.read_bytes()).hexdigest(),
        "allison_input_sha256": hashlib.sha256(ALLISON_JSONL.read_bytes()).hexdigest(),
        "driventic_diwa_input_sha256": hashlib.sha256(DRIVENTIC_DIWA_JSONL.read_bytes()).hexdigest(),
        "mercedes_dtfr_input_sha256": hashlib.sha256(MERCEDES_DTFR_JSONL.read_bytes()).hexdigest(),
        "mercedes_bevo_input_sha256": hashlib.sha256(MERCEDES_BEVO_JSONL.read_bytes()).hexdigest(),
        "volvo_genuine_input_sha256": hashlib.sha256(VOLVO_GENUINE_JSONL.read_bytes()).hexdigest(),
        "mack_genuine_input_sha256": hashlib.sha256(MACK_GENUINE_JSONL.read_bytes()).hexdigest(),
        "mack_2014_approved_input_sha256": hashlib.sha256(MACK_2014_APPROVED_JSONL.read_bytes()).hexdigest(),
        "cummins_valvoline_2022_input_sha256": hashlib.sha256(CUMMINS_VALVOLINE_2022_JSONL.read_bytes()).hexdigest(),
        "taiwan_cpc_input_sha256": hashlib.sha256(TAIWAN_CPC_JSONL.read_bytes()).hexdigest(),
        "volvo_websds_input_sha256": hashlib.sha256(VOLVO_WEBSDS_JSONL.read_bytes()).hexdigest(),
        "parker_denison_input_sha256": hashlib.sha256(PARKER_DENISON_JSONL.read_bytes()).hexdigest(),
        "scania_genuine_input_sha256": hashlib.sha256(SCANIA_GENUINE_JSONL.read_bytes()).hexdigest(),
        "brava_official_input_sha256": hashlib.sha256(BRAVA_OFFICIAL_JSONL.read_bytes()).hexdigest(),
        "ceypetco_input_sha256": hashlib.sha256(CEYPETCO_JSONL.read_bytes()).hexdigest(),
        "pso_official_input_sha256": hashlib.sha256(PSO_OFFICIAL_JSONL.read_bytes()).hexdigest(),
        "pertamina_official_input_sha256": hashlib.sha256(PERTAMINA_OFFICIAL_JSONL.read_bytes()).hexdigest(),
        "man_service_input_sha256": hashlib.sha256(MAN_SERVICE_JSONL.read_bytes()).hexdigest(),
        "liqui_moly_2020_input_sha256": hashlib.sha256(LIQUI_MOLY_2020_JSONL.read_bytes()).hexdigest(),
        "liqui_moly_current_input_sha256": hashlib.sha256(LIQUI_MOLY_CURRENT_JSONL.read_bytes()).hexdigest(),
        "anp_brazil_input_sha256": hashlib.sha256(ANP_BRAZIL_JSONL.read_bytes()).hexdigest(),
        "anp_brazil_monitoring_input_sha256": hashlib.sha256(ANP_BRAZIL_MONITORING_JSONL.read_bytes()).hexdigest(),
        "anp_brazil_monitoring_pdf_input_sha256": hashlib.sha256(ANP_BRAZIL_MONITORING_PDF_JSONL.read_bytes()).hexdigest(),
        "anp_brazil_monitoring_pdf_exceptions_input_sha256": hashlib.sha256(ANP_BRAZIL_MONITORING_PDF_EXCEPTIONS_JSONL.read_bytes()).hexdigest(),
        "indonesia_npt_input_sha256": hashlib.sha256(INDONESIA_NPT_JSONL.read_bytes()).hexdigest(),
        "thailand_doeb_input_sha256": hashlib.sha256(THAILAND_DOEB_JSONL.read_bytes()).hexdigest(),
        "dla_qpd_input_sha256": hashlib.sha256(DLA_QPD_JSONL.read_bytes()).hexdigest(),
        "fuchs_india_input_sha256": hashlib.sha256(FUCHS_INDIA_JSONL.read_bytes()).hexdigest(),
        "fuchs_us_input_sha256": hashlib.sha256(FUCHS_US_JSONL.read_bytes()).hexdigest(),
        "fuchs_germany_input_sha256": hashlib.sha256(FUCHS_GERMANY_JSONL.read_bytes()).hexdigest(),
        "fuchs_poland_input_sha256": hashlib.sha256(FUCHS_POLAND_JSONL.read_bytes()).hexdigest(),
        "fuchs_italy_input_sha256": hashlib.sha256(FUCHS_ITALY_JSONL.read_bytes()).hexdigest(),
        "fuchs_sweden_input_sha256": hashlib.sha256(FUCHS_SWEDEN_JSONL.read_bytes()).hexdigest(),
        "fuchs_spain_input_sha256": hashlib.sha256(FUCHS_SPAIN_JSONL.read_bytes()).hexdigest(),
        "fuchs_france_input_sha256": hashlib.sha256(FUCHS_FRANCE_JSONL.read_bytes()).hexdigest(),
        "fuchs_turkey_input_sha256": hashlib.sha256(FUCHS_TURKEY_JSONL.read_bytes()).hexdigest(),
        "fuchs_canada_input_sha256": hashlib.sha256(FUCHS_CANADA_JSONL.read_bytes()).hexdigest(),
        "fuchs_china_input_sha256": hashlib.sha256(FUCHS_CHINA_JSONL.read_bytes()).hexdigest(),
        "fuchs_czech_input_sha256": hashlib.sha256(FUCHS_CZECH_JSONL.read_bytes()).hexdigest(),
        "fuchs_mexico_input_sha256": hashlib.sha256(FUCHS_MEXICO_JSONL.read_bytes()).hexdigest(),
        "fuchs_south_africa_input_sha256": hashlib.sha256(FUCHS_SOUTH_AFRICA_JSONL.read_bytes()).hexdigest(),
        "fuchs_brazil_input_sha256": hashlib.sha256(FUCHS_BRAZIL_JSONL.read_bytes()).hexdigest(),
        "fuchs_norway_input_sha256": hashlib.sha256(FUCHS_NORWAY_JSONL.read_bytes()).hexdigest(),
        "fuchs_hungary_input_sha256": hashlib.sha256(FUCHS_HUNGARY_JSONL.read_bytes()).hexdigest(),
        "canonical_rows": len(records),
        "brands": len({r["brand"] for r in records}),
        "families": dict(sorted(Counter(r["family_code"] for r in records).items())),
        "project_source_rows": sum(r["evidence_status"] == "project_source_row" for r in records),
        "legacy_rows_needing_evidence": sum(r["evidence_status"] == "legacy_needs_official_tds" for r in records),
        "internal_master_rows": sum(r["evidence_status"] == "internal_product_master" for r in records),
        "official_filed_registry_rows": sum(r["evidence_status"] == "official_filed_registry" for r in records),
        "official_licensed_registry_rows": sum(r["evidence_status"] == "official_licensed_registry" for r in records),
        "official_licensed_source_rows": len(licensed_source_rows),
        "blue_angel_source_rows": len(blue_angel_source_rows),
        "blue_angel_products_matched_to_existing": blue_angel_matched_rows,
        "blue_angel_products_added": blue_angel_added_rows,
        "austrian_ecolabel_uz14_source_rows": len(austrian_uz14_source_rows),
        "austrian_ecolabel_uz14_products_matched_to_existing": austrian_uz14_matched_rows,
        "austrian_ecolabel_uz14_cross_family_evidence_matches": austrian_uz14_cross_family_matches,
        "austrian_ecolabel_uz14_products_added": austrian_uz14_added_rows,
        "official_ecolabel_product_registry_rows": sum(r["evidence_status"] == "official_ecolabel_product_registry" for r in records),
        "korea_ecolabel_source_rows": len(korea_ecolabel_source_rows),
        "korea_ecolabel_products_matched_to_existing": korea_ecolabel_matched_rows,
        "korea_ecolabel_products_added": korea_ecolabel_added_rows,
        "official_government_ecolabel_registry_rows": sum(r["evidence_status"] == "official_government_ecolabel_registry" for r in records),
        "korea_ecolabel_el509_source_rows": len(korea_el509_source_rows),
        "korea_ecolabel_el509_products_matched_to_existing": korea_el509_matched_rows,
        "korea_ecolabel_el509_products_added": korea_el509_added_rows,
        "green_choice_philippines_source_rows": len(green_choice_philippines_source_rows),
        "green_choice_philippines_expired_rows": sum(row["lifecycle_status"] == "ecolabel_certificate_expired" for row in green_choice_philippines_source_rows),
        "uae_moiat_source_rows": len(uae_moiat_source_rows),
        "eaeu_conformity_source_rows": len(eaeu_conformity_source_rows),
        "eaeu_conformity_explicit_brand_rows": sum(row["brand_basis"] == "explicit_source_trademark" for row in eaeu_conformity_source_rows),
        "eaeu_conformity_manufacturer_holder_fallback_rows": sum(row["brand_basis"] == "manufacturer_holder_fallback" for row in eaeu_conformity_source_rows),
        "epa_safer_choice_source_rows": len(epa_safer_choice_source_rows),
        "epa_chemexpo_source_rows": len(epa_chemexpo_source_rows),
        "epa_chemexpo_source_product_occurrences": sum(row["source_occurrence_count"] for row in epa_chemexpo_source_rows),
        "epa_chemexpo_products_matched_to_existing": epa_chemexpo_matched_rows,
        "epa_chemexpo_products_added": epa_chemexpo_added_rows,
        "psqca_engine_oil_source_rows": len(psqca_engine_oil_source_rows),
        "official_government_product_certification_brand_scope_rows": sum(r["evidence_status"] == "official_government_product_certification_brand_scope" for r in records),
        "tisi_two_stroke_oil_source_rows": len(tisi_two_stroke_oil_source_rows),
        "official_government_product_certification_holder_scope_rows": sum(r["evidence_status"] == "official_government_product_certification_holder_scope" for r in records),
        "samr_china_2025_source_rows": len(samr_china_2025_source_rows),
        "samr_china_2023_source_rows": len(samr_china_2023_source_rows),
        "samr_china_2023_products_matched_to_existing": samr_china_2023_products_matched_to_existing,
        "samr_china_2023_products_added": samr_china_2023_products_added,
        "samr_china_2024_source_rows": len(samr_china_2024_source_rows),
        "shenzhen_china_2021_source_rows": len(shenzhen_china_2021_source_rows),
        "shenzhen_china_2021_products_matched_to_existing": shenzhen_china_2021_products_matched_to_existing,
        "shenzhen_china_2021_products_added": shenzhen_china_2021_products_added,
        "shenzhen_china_2020_source_rows": len(shenzhen_china_2020_source_rows),
        "shenzhen_china_2020_products_matched_to_existing": shenzhen_china_2020_products_matched_to_existing,
        "shenzhen_china_2020_duplicate_occurrences_merged": shenzhen_china_2020_duplicate_occurrences_merged,
        "shenzhen_china_2020_products_added": shenzhen_china_2020_products_added,
        "shenzhen_china_2019_source_rows": len(shenzhen_china_2019_source_rows),
        "shenzhen_china_2019_products_matched_to_existing": shenzhen_china_2019_products_matched_to_existing,
        "shenzhen_china_2019_duplicate_occurrences_merged": shenzhen_china_2019_duplicate_occurrences_merged,
        "shenzhen_china_2019_products_added": shenzhen_china_2019_products_added,
        "shenzhen_china_2025_source_rows": len(shenzhen_china_2025_source_rows),
        "shenzhen_china_2025_products_matched_to_existing": shenzhen_china_2025_products_matched_to_existing,
        "shenzhen_china_2025_duplicate_occurrences_merged": shenzhen_china_2025_duplicate_occurrences_merged,
        "shenzhen_china_2025_products_added": shenzhen_china_2025_products_added,
        "shanghai_china_2023_2025_source_rows": len(shanghai_china_source_rows),
        "shanghai_china_products_matched_to_existing": shanghai_china_products_matched_to_existing,
        "shanghai_china_products_added": shanghai_china_products_added,
        "beijing_china_2018_source_rows": len(beijing_china_source_rows),
        "beijing_china_products_matched_to_existing": beijing_china_products_matched_to_existing,
        "beijing_china_products_added": beijing_china_products_added,
        "shenzhen_china_2016_2017_source_rows": len(shenzhen_china_2016_2017_source_rows),
        "shenzhen_china_2016_2017_products_matched_to_existing": shenzhen_china_2016_2017_products_matched_to_existing,
        "shenzhen_china_2016_2017_products_added": shenzhen_china_2016_2017_products_added,
        "qingdao_china_2021_2025_source_rows": len(qingdao_china_source_rows),
        "qingdao_china_products_matched_to_existing": qingdao_china_products_matched_to_existing,
        "qingdao_china_duplicate_occurrences_merged": qingdao_china_duplicate_occurrences_merged,
        "qingdao_china_products_added": qingdao_china_products_added,
        "jilin_china_2024_2025_source_rows": len(jilin_china_source_rows),
        "jilin_china_products_matched_to_existing": jilin_china_products_matched_to_existing,
        "jilin_china_duplicate_occurrences_merged": jilin_china_duplicate_occurrences_merged,
        "jilin_china_products_added": jilin_china_products_added,
        "china_government_inspection_source_observations": sum(map(len, (
            samr_china_2025_source_rows, samr_china_2024_source_rows,
            samr_china_2023_source_rows, shenzhen_china_2021_source_rows,
            shenzhen_china_2025_source_rows, shenzhen_china_2020_source_rows,
            shenzhen_china_2019_source_rows,
            shanghai_china_source_rows,
            beijing_china_source_rows,
            shenzhen_china_2016_2017_source_rows,
            qingdao_china_source_rows,
            jilin_china_source_rows,
        ))),
        "samr_china_source_observations": len(samr_china_2025_source_rows) + len(samr_china_2024_source_rows) + len(samr_china_2023_source_rows),
        "official_government_nonconforming_product_inspection_observation_rows": sum(r["evidence_status"] == "official_government_nonconforming_product_inspection_observation" for r in records),
        "official_government_conforming_product_inspection_observation_rows": sum(r["evidence_status"] == "official_government_conforming_product_inspection_observation" for r in records),
        "philippines_bps_brake_fluid_source_rows": len(philippines_bps_brake_fluid_source_rows),
        "philippines_bps_ps_brake_fluid_rows": sum(r["source_id"] == "PHILIPPINES_BPS_PS_BRAKE_FLUID_LICENCES" for r in records),
        "philippines_bps_icc_brake_fluid_rows": sum(r["source_id"] == "PHILIPPINES_BPS_ICC_BRAKE_FLUID_CERTIFICATES" for r in records),
        "ghana_gsa_certified_source_rows": len(ghana_gsa_source_rows),
        "kebs_smark_source_rows": len(kebs_smark_source_rows),
        "east_africa_certified_source_rows": len(east_africa_certified_source_rows),
        "east_africa_certified_source_rows_by_source": dict(sorted(Counter(row["source_id"] for row in east_africa_certified_source_rows).items())),
        "son_mancap_source_rows": len(son_mancap_source_rows),
        "rsb_smark_source_rows": len(rsb_smark_source_rows),
        "official_government_product_conformity_registry_rows": sum(r["evidence_status"] == "official_government_product_conformity_registry" for r in records),
        "official_government_product_certification_registry_rows": sum(r["evidence_status"] == "official_government_product_certification_registry" for r in records),
        "official_government_program_rows": sum(r["evidence_status"] == "official_government_program_catalog" for r in records),
        "official_government_compiled_product_database_rows": sum(r["evidence_status"] == "official_government_compiled_product_database" for r in records),
        "official_government_regulatory_registry_rows": sum(r["evidence_status"] == "official_government_regulatory_registry" for r in records),
        "official_government_qualified_product_registry_rows": sum(r["evidence_status"] == "official_government_qualified_product_registry" for r in records),
        "usda_biopreferred_source_rows": len(biopreferred_source_rows),
        "anp_brazil_source_rows": len(anp_brazil_source_rows),
        "anp_brazil_monitoring_source_observations": len(anp_monitoring_source_rows),
        "anp_brazil_monitoring_xlsx_source_observations": len(anp_monitoring_xlsx_source_rows),
        "anp_brazil_monitoring_pdf_source_observations": len(anp_monitoring_pdf_source_rows),
        "anp_brazil_monitoring_pdf_exception_source_observations": len(anp_monitoring_pdf_exception_source_rows),
        "anp_brazil_monitoring_pdf_exception_raw_identities": len(exception_raw_identity_keys),
        "anp_brazil_monitoring_pdf_exception_semantic_target_identities": len(exception_semantic_identity_keys),
        "anp_brazil_monitoring_pdf_exception_semantically_remapped_observations": exception_semantically_remapped_observations,
        "anp_brazil_monitoring_pdf_exception_semantically_remapped_identities": len(exception_semantically_remapped_identity_keys),
        "anp_brazil_monitoring_product_grade_identities": len(anp_monitoring_groups),
        "anp_brazil_monitoring_identities_matched_to_current_registry": anp_monitoring_matched_identities,
        "anp_brazil_monitoring_ambiguous_current_registry_matches": anp_monitoring_ambiguous_match_identities,
        "anp_brazil_monitoring_historical_identities_added": anp_monitoring_added_identities,
        "anp_brazil_monitoring_registered_historical_identities_added": anp_monitoring_added_registered_identities,
        "anp_brazil_monitoring_unregistered_historical_identities_added": anp_monitoring_added_unregistered_identities,
        "anp_brazil_monitoring_historical_identities_added_by_primary_source": dict(sorted(anp_monitoring_added_identities_by_primary_source.items())),
        "official_government_historical_market_monitoring_rows": sum(r["evidence_status"] == "official_government_historical_market_monitoring" for r in records),
        "indonesia_npt_source_rows": len(indonesia_npt_source_rows),
        "indonesia_npt_rows_with_registration_value": sum(bool(row["registration_number"]) for row in indonesia_npt_source_rows),
        "indonesia_npt_rows_with_source_data_issue": sum(not row["registration_number"] for row in indonesia_npt_source_rows),
        "thailand_doeb_normalized_products": len(thailand_doeb_source_rows),
        "thailand_doeb_source_occurrences": sum(row["source_occurrence_count"] for row in thailand_doeb_source_rows),
        "thailand_doeb_unique_registration_numbers": len({number for row in thailand_doeb_source_rows for number in row["registration_numbers"]}),
        "thailand_doeb_registration_collision_products": sum("registration_number_collision" in row["source_quality_flags"] for row in thailand_doeb_source_rows),
        "thailand_doeb_published_end_date_not_expired_products": sum(row["lifecycle_status"] == "not_expired_by_published_end_date_as_of_catalog_snapshot" for row in thailand_doeb_source_rows),
        "dla_qpd_source_rows": len(dla_qpd_source_rows),
        "dla_qpd_source_rows_by_source": dict(sorted(Counter(row["source_id"] for row in dla_qpd_source_rows).items())),
        "dla_qpd_lifecycle_statuses": dict(sorted(Counter(row["lifecycle_status"] for row in dla_qpd_source_rows).items())),
        "official_government_registry_source_data_issue_rows": sum(r["evidence_status"] == "official_government_registry_source_data_issue" for r in records),
        "official_oem_approval_rows": sum(r["evidence_status"] == "official_oem_approval_registry" for r in records),
        "official_manufacturer_catalog_rows": sum(r["evidence_status"] == "official_manufacturer_product_catalog" for r in records),
        "official_manufacturer_sds_change_observation_rows": sum(r["evidence_status"] == "official_manufacturer_sds_change_observation" for r in records),
        "official_oem_service_recommendation_rows": sum(r["evidence_status"] == "official_oem_service_recommendation" for r in records),
        "zf_te_ml_source_rows": len(zf_source_rows),
        "allison_source_rows": len(allison_source_rows),
        "driventic_diwa_source_rows": len(driventic_source_rows),
        "mercedes_dtfr_source_rows": len(mercedes_dtfr_source_rows),
        "mercedes_bevo_source_rows": len(mercedes_bevo_source_rows),
        "mercedes_bevo_products_matched_to_existing": mercedes_bevo_matched_rows,
        "mercedes_bevo_products_added": mercedes_bevo_added_rows,
        "volvo_genuine_source_rows": len(volvo_genuine_source_rows),
        "mack_genuine_source_rows": len(mack_genuine_source_rows),
        "mack_2014_approved_source_rows": len(mack_2014_source_rows),
        "cummins_valvoline_2022_source_rows": len(cummins_valvoline_2022_source_rows),
        "taiwan_cpc_source_rows": len(taiwan_cpc_source_rows),
        "volvo_websds_source_rows": len(volvo_websds_source_rows),
        "volvo_websds_unique_part_numbers": len({part for row in volvo_websds_source_rows for part in row["part_numbers"]}),
        "parker_denison_source_rows": len(parker_denison_source_rows),
        "parker_denison_lifecycle_statuses": dict(sorted(Counter(row["lifecycle_status"] for row in parker_denison_source_rows).items())),
        "taiwan_cpc_structured_package_offers": sum(len(row["packages"]) for row in taiwan_cpc_source_rows),
        "scania_genuine_source_rows": len(scania_genuine_source_rows),
        "brava_official_source_rows": len(brava_official_source_rows),
        "ceypetco_source_rows": len(ceypetco_source_rows),
        "pso_official_source_rows": len(pso_source_rows),
        "pso_official_products_matched_to_existing": pso_matched_rows,
        "pso_official_products_added": pso_added_rows,
        "pertamina_official_source_rows": len(pertamina_source_rows),
        "pertamina_official_products_exact_matched_to_existing": pertamina_exact_matched_rows,
        "pertamina_official_products_blank_npt_fallback_matched": pertamina_blank_npt_fallback_matched_rows,
        "pertamina_official_products_added": pertamina_added_rows,
        "pertamina_official_ambiguous_rows_added_separately": pertamina_ambiguous_rows_added,
        "pertamina_official_cross_family_corrections": pertamina_cross_family_corrections,
        "man_service_source_rows": len(man_service_source_rows),
        "man_service_products_matched_to_existing": man_service_matched_rows,
        "man_service_products_added": man_service_added_rows,
        "liqui_moly_2020_source_rows": len(liqui_moly_source_rows),
        "liqui_moly_2020_products_matched_to_existing": liqui_moly_matched_rows,
        "liqui_moly_2020_products_added": liqui_moly_added_rows,
        "liqui_moly_current_source_rows": len(liqui_moly_current_source_rows),
        "liqui_moly_current_products_matched_to_2020": liqui_moly_current_matched_rows,
        "liqui_moly_current_products_added": liqui_moly_current_added_rows,
        "liqui_moly_current_article_skus": len({article["sku"] for row in liqui_moly_current_source_rows for article in row["articles"]}),
        "liqui_moly_current_lifecycle_assessments": dict(Counter(row["lifecycle_assessment"] for row in liqui_moly_current_source_rows)),
        "fuchs_india_source_rows": len(fuchs_india_source_rows),
        "fuchs_india_products_matched_to_existing": fuchs_india_matched_rows,
        "fuchs_india_products_added": fuchs_india_added_rows,
        "fuchs_us_source_rows": len(fuchs_us_source_rows),
        "fuchs_us_products_matched_to_existing": fuchs_us_matched_rows,
        "fuchs_us_products_added": fuchs_us_added_rows,
        "fuchs_cross_market_exact_name_family_rows": fuchs_cross_market_exact_name_family_rows,
        "fuchs_cross_market_family_conflict_rows": fuchs_cross_market_family_conflict_rows,
        "fuchs_germany_source_rows": len(fuchs_germany_source_rows),
        "fuchs_germany_products_matched_to_existing": fuchs_germany_matched_rows,
        "fuchs_germany_products_added": fuchs_germany_added_rows,
        "fuchs_germany_cross_market_exact_name_family_rows": fuchs_germany_cross_market_exact_name_family_rows,
        "fuchs_germany_cross_market_family_conflict_rows": fuchs_germany_cross_market_family_conflict_rows,
        "fuchs_poland_source_rows": len(fuchs_poland_source_rows),
        "fuchs_poland_products_matched_to_existing": fuchs_poland_matched_rows,
        "fuchs_poland_products_added": fuchs_poland_added_rows,
        "fuchs_poland_cross_market_exact_name_family_rows": fuchs_poland_cross_market_exact_name_family_rows,
        "fuchs_poland_cross_market_family_conflict_rows": fuchs_poland_cross_market_family_conflict_rows,
        "fuchs_italy_source_rows": len(fuchs_italy_source_rows),
        "fuchs_italy_products_matched_to_existing": fuchs_italy_matched_rows,
        "fuchs_italy_products_added": fuchs_italy_added_rows,
        "fuchs_italy_cross_market_exact_name_family_rows": fuchs_italy_cross_market_exact_name_family_rows,
        "fuchs_italy_cross_market_family_conflict_rows": fuchs_italy_cross_market_family_conflict_rows,
        "fuchs_sweden_source_rows": len(fuchs_sweden_source_rows),
        "fuchs_sweden_products_matched_to_existing": fuchs_sweden_matched_rows,
        "fuchs_sweden_products_added": fuchs_sweden_added_rows,
        "fuchs_sweden_cross_market_exact_name_family_rows": fuchs_sweden_cross_market_exact_name_family_rows,
        "fuchs_sweden_cross_market_family_conflict_rows": fuchs_sweden_cross_market_family_conflict_rows,
        "fuchs_spain_source_rows": len(fuchs_spain_source_rows),
        "fuchs_spain_products_matched_to_existing": fuchs_spain_matched_rows,
        "fuchs_spain_products_added": fuchs_spain_added_rows,
        "fuchs_spain_cross_market_exact_name_family_rows": fuchs_spain_cross_market_exact_name_family_rows,
        "fuchs_spain_cross_market_family_conflict_rows": fuchs_spain_cross_market_family_conflict_rows,
        "fuchs_france_source_rows": len(fuchs_france_source_rows),
        "fuchs_france_products_matched_to_existing": fuchs_france_matched_rows,
        "fuchs_france_products_added": fuchs_france_added_rows,
        "fuchs_france_cross_market_exact_name_family_rows": fuchs_france_cross_market_exact_name_family_rows,
        "fuchs_france_cross_market_family_conflict_rows": fuchs_france_cross_market_family_conflict_rows,
        "fuchs_turkey_source_rows": len(fuchs_turkey_source_rows),
        "fuchs_turkey_products_matched_to_existing": fuchs_turkey_matched_rows,
        "fuchs_turkey_products_added": fuchs_turkey_added_rows,
        "fuchs_turkey_cross_market_exact_name_family_rows": fuchs_turkey_cross_market_exact_name_family_rows,
        "fuchs_turkey_cross_market_family_conflict_rows": fuchs_turkey_cross_market_family_conflict_rows,
        "fuchs_canada_source_rows": len(fuchs_canada_source_rows),
        "fuchs_canada_products_matched_to_existing": fuchs_canada_matched_rows,
        "fuchs_canada_products_added": fuchs_canada_added_rows,
        "fuchs_canada_cross_market_exact_name_family_rows": fuchs_canada_cross_market_exact_name_family_rows,
        "fuchs_canada_cross_market_family_conflict_rows": fuchs_canada_cross_market_family_conflict_rows,
        "fuchs_china_source_rows": len(fuchs_china_source_rows),
        "fuchs_china_products_matched_to_existing": fuchs_china_matched_rows,
        "fuchs_china_products_added": fuchs_china_added_rows,
        "fuchs_china_cross_market_exact_name_family_rows": fuchs_china_cross_market_exact_name_family_rows,
        "fuchs_china_cross_market_family_conflict_rows": fuchs_china_cross_market_family_conflict_rows,
        "fuchs_czech_source_rows": len(fuchs_czech_source_rows),
        "fuchs_czech_products_matched_to_existing": fuchs_czech_matched_rows,
        "fuchs_czech_products_added": fuchs_czech_added_rows,
        "fuchs_czech_cross_market_exact_name_family_rows": fuchs_czech_cross_market_exact_name_family_rows,
        "fuchs_czech_cross_market_family_conflict_rows": fuchs_czech_cross_market_family_conflict_rows,
        "fuchs_mexico_source_rows": len(fuchs_mexico_source_rows),
        "fuchs_mexico_products_matched_to_existing": fuchs_mexico_matched_rows,
        "fuchs_mexico_products_added": fuchs_mexico_added_rows,
        "fuchs_mexico_cross_market_exact_name_family_rows": fuchs_mexico_cross_market_exact_name_family_rows,
        "fuchs_mexico_cross_market_family_conflict_rows": fuchs_mexico_cross_market_family_conflict_rows,
        "fuchs_south_africa_source_rows": len(fuchs_south_africa_source_rows),
        "fuchs_south_africa_products_matched_to_existing": fuchs_south_africa_matched_rows,
        "fuchs_south_africa_products_added": fuchs_south_africa_added_rows,
        "fuchs_south_africa_cross_market_exact_name_family_rows": fuchs_south_africa_cross_market_exact_name_family_rows,
        "fuchs_south_africa_cross_market_family_conflict_rows": fuchs_south_africa_cross_market_family_conflict_rows,
        "fuchs_brazil_source_rows": len(fuchs_brazil_source_rows),
        "fuchs_brazil_products_matched_to_existing": fuchs_brazil_matched_rows,
        "fuchs_brazil_products_added": fuchs_brazil_added_rows,
        "fuchs_brazil_cross_market_exact_name_family_rows": fuchs_brazil_cross_market_exact_name_family_rows,
        "fuchs_brazil_cross_market_family_conflict_rows": fuchs_brazil_cross_market_family_conflict_rows,
        "fuchs_norway_source_rows": len(fuchs_norway_source_rows),
        "fuchs_norway_products_matched_to_existing": fuchs_norway_matched_rows,
        "fuchs_norway_products_added": fuchs_norway_added_rows,
        "fuchs_norway_cross_market_exact_name_family_rows": fuchs_norway_cross_market_exact_name_family_rows,
        "fuchs_norway_cross_market_family_conflict_rows": fuchs_norway_cross_market_family_conflict_rows,
        "fuchs_hungary_source_rows": len(fuchs_hungary_source_rows),
        "fuchs_hungary_products_matched_to_existing": fuchs_hungary_matched_rows,
        "fuchs_hungary_products_added": fuchs_hungary_added_rows,
        "fuchs_hungary_cross_market_exact_name_family_rows": fuchs_hungary_cross_market_exact_name_family_rows,
        "fuchs_hungary_cross_market_family_conflict_rows": fuchs_hungary_cross_market_family_conflict_rows,
        "jaso_source_rows": len(jaso_source_rows),
        "jaso_unique_oil_codes": len({r["oil_code"] for r in jaso_source_rows}),
        "aichilon_source_products": len(aichilon_products) + len(exclusions),
        "aichilon_products_matched_to_existing": aichilon_matched_rows,
        "aichilon_products_added": aichilon_new_rows,
        "aichilon_rows_excluded": len(exclusions),
        "offers": len(offers),
        "active_offers": sum(o["lifecycle_status"] in {"active", "listed_current_catalog"} for o in offers),
        "current_catalog_listed_offers": sum(o["lifecycle_status"] == "listed_current_catalog" for o in offers),
        "archived_offers": sum(o["lifecycle_status"] == "archived" for o in offers),
        "duplicate_decisions": dict(Counter(c["decision"] for c in candidates)),
        "quality_issues": dict(Counter(i["issue_code"] for i in issues)),
        "bulk_sources_allowed": [s["source_id"] for s in policies["sources"] if s["bulk_ingest_allowed"]],
        "bulk_sources_blocked": [s["source_id"] for s in policies["sources"] if not s["bulk_ingest_allowed"]],
        "confirmed_world_total": None,
        "completion_note": "The seed proves the pipeline, not worldwide coverage. Exact worldwide count remains pending licensed/authorized source ingestion.",
    }
    for slug, data in additional_fuchs.items():
        report.update({
            f"fuchs_{slug}_input_sha256": hashlib.sha256(data["source_path"].read_bytes()).hexdigest(),
            f"fuchs_{slug}_source_rows": len(data["source_rows"]),
            f"fuchs_{slug}_products_matched_to_existing": data["matched"],
            f"fuchs_{slug}_products_added": data["added"],
            f"fuchs_{slug}_cross_market_exact_name_family_rows": data["exact"],
            f"fuchs_{slug}_cross_market_family_conflict_rows": data["conflicts"],
        })
    REPORT_OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    build_sqlite(records, candidates, issues, source_links, offers, policies, run_id, len(input_records))
    compress_sqlite()
    build_workbook(records, candidates, issues, offers, exclusions, policies, report)
    print(json.dumps(report, ensure_ascii=False))


if __name__ == "__main__":
    main()
