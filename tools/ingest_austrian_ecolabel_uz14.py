#!/usr/bin/env python3
"""Ingest factual product identities from Austria's official UZ 14 directory."""

from __future__ import annotations

import hashlib
import html
import io
import json
import re
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "austrian-ecolabel-uz14-products.jsonl"
REPORT = ROOT / "data" / "austrian-ecolabel-uz14-products-report.json"
SOURCE_ID = "AUSTRIAN_ECOLABEL_UZ14_LUBRICANTS"
SNAPSHOT_DATE = "2026-07-21"
BASE = "https://www.umweltzeichen.at"
XLSX_URL = BASE + "/uz-export/umweltzeichen-lizenznehmerinnen-UZ%2014.xlsx"
CATEGORY_URL = BASE + "/de/produkte/garten-gr%C3%BCnraum"
IMPRINT_URL = BASE + "/de/zertifizierung/impressum"


def fetch(url: str) -> bytes:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "MFClassifier public-sector catalog research/1.0"},
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


def clean_markup(value: str) -> str:
    return " ".join(html.unescape(re.sub(r"<[^>]+>", " ", value)).split())


def page_products(page: bytes) -> list[str]:
    source = page.decode("utf-8", errors="replace")
    return [clean_markup(value) for value in re.findall(r'<h5 class="card-title">(.*?)</h5>', source, re.S)]


def family_for(product_name: str) -> tuple[str, str]:
    normalized = product_name.casefold()
    if any(token in normalized for token in ("hydraul", "synthohyd")) or re.search(r"\bHE\s+(?:32|46)\b", product_name, re.I):
        return "H", "explicit hydraulic product designation on official UZ 14 product card"
    if any(token in normalized for token in ("kettenöl", "kettenoel", "chain oil", "medialub")):
        return "I", "explicit chain-oil product designation on official UZ 14 product card"
    return "S", "official UZ 14 lubricant scope; detailed family not explicit in product designation"


def brand_for(product_name: str, holder: str) -> tuple[str, str]:
    normalized = product_name.casefold()
    for token, brand in (
        ("biostar", "BIOSTAR"),
        ("oregon", "OREGON"),
        ("kettlitz", "KETTLITZ"),
        ("kox", "KOX"),
        ("divinol", "Divinol"),
    ):
        if normalized.startswith(token):
            return brand, "explicit product-name prefix"
    return holder, "licence-holder fallback; no separate brand stated"


def main() -> None:
    xlsx = fetch(XLSX_URL)
    category = fetch(CATEGORY_URL)
    imprint = fetch(IMPRINT_URL)
    workbook = load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    sheet = workbook.active
    holders = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        licence = str(values[0] or "").strip()
        holder = str(values[1] or "").strip()
        if not licence or not holder:
            continue
        holders.append({"licence": licence, "holder": holder, "country": str(values[2] or "").strip(), "source_row": row_number})

    rows = []
    page_hashes = {}
    for holder in holders:
        query = urllib.parse.urlencode({"cert_number": holder["licence"]})
        page_url = CATEGORY_URL + "?" + query
        page = fetch(page_url)
        page_hashes[holder["licence"]] = hashlib.sha256(page).hexdigest()
        products = page_products(page)
        if not products:
            raise RuntimeError(f"No products found for {holder['licence']}")
        for product_index, product_name in enumerate(products, start=1):
            family_code, classification_basis = family_for(product_name)
            brand, brand_basis = brand_for(product_name, holder["holder"])
            source_record_id = f"{holder['licence']}:{product_index:02d}"
            rows.append({
                "source_id": SOURCE_ID,
                "source_record_id": source_record_id,
                "source_row": holder["source_row"],
                "licence_number": holder["licence"],
                "licence_holder": holder["holder"],
                "holder_country_source": holder["country"],
                "manufacturer": holder["holder"],
                "brand": brand,
                "brand_basis": brand_basis,
                "product_name": product_name,
                "family_code": family_code,
                "classification_basis": classification_basis,
                "certification_standard": "Austrian Ecolabel UZ 14 — Schmierstoffe",
                "lifecycle_status": "listed_in_current_uz14_directory_status_not_individually_dated",
                "market": "Austria",
                "snapshot_date": SNAPSHOT_DATE,
                "source_url": page_url,
                "directory_url": CATEGORY_URL,
                "licensee_export_url": XLSX_URL,
                "technical": {"sae": [], "iso_vg": [], "nlgi": []},
            })

    rows.sort(key=lambda row: (row["licence_number"], row["source_record_id"]))
    normalized = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in rows)
    OUT.write_text(normalized, encoding="utf-8")
    report = {
        "schema_version": 1,
        "source_id": SOURCE_ID,
        "snapshot_date": SNAPSHOT_DATE,
        "source_url": CATEGORY_URL,
        "licensee_export_url": XLSX_URL,
        "imprint_url": IMPRINT_URL,
        "rights_note": "BMLUK copyright permits reproduction with attribution except for commercial purposes; only factual product identity and certification fields are retained.",
        "licensees": len(holders),
        "normalized_products": len(rows),
        "products_by_family": dict(sorted({code: sum(row["family_code"] == code for row in rows) for code in {row["family_code"] for row in rows}}.items())),
        "licensee_export_sha256": hashlib.sha256(xlsx).hexdigest(),
        "category_page_sha256": hashlib.sha256(category).hexdigest(),
        "imprint_page_sha256": hashlib.sha256(imprint).hexdigest(),
        "product_page_sha256": page_hashes,
        "normalized_output_sha256": hashlib.sha256(normalized.encode("utf-8")).hexdigest(),
        "excluded_fields": ["postal address", "telephone", "email", "marketing descriptions", "images", "logo artwork"],
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(f"licensees={len(holders)} products={len(rows)} output={OUT}")


if __name__ == "__main__":
    main()
