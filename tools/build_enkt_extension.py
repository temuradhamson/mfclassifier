#!/usr/bin/env python3
"""Build a pilot ENKT extension for professional lubricant classes."""

from __future__ import annotations

import json
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CATALOG = ROOT / "data" / "catalog-v3.json"
JSON_OUTPUT = ROOT / "data" / "enkt-extension-pilot.json"
XLSX_OUTPUT = ROOT / "deliverables" / "ENKT_GSM_extension_pilot.xlsx"
VERSION = "0.1-pilot"

FAMILY_RULES = {
    "M": ("19.20.29.110", "Масла моторные", "SAE + API/ACEA + тип базового масла"),
    "T": ("19.20.29.120", "Масла трансмиссионные", "SAE J306 + API GL/OEM + тип жидкости"),
    "H": ("19.20.29.130", "Масла гидравлические", "ISO VG + DIN/ISO эксплуатационный класс"),
    "I": ("19.20.29.140", "Масла индустриальные", "ISO VG + DIN/ISO/ГОСТ + назначение"),
    "C": ("19.20.29.150", "Масла компрессорные", "ISO VG + DIN 51506/тип компрессора"),
    "U": ("19.20.29.160", "Масла турбинные", "ISO VG/ГОСТ + тип турбины"),
    "E": ("19.20.29.172", "Масла электроизоляционные", "IEC/ГОСТ/ASTM + тип оборудования"),
    "G": ("19.20.29.210", "Смазки пластичные", "загуститель + NLGI + DIN 51502 + назначение"),
}

CURRENT_ENKT = [
    ("19.20.29.100-00001", "Масло нефтяное смазочное", 459991),
    ("19.20.29.100-00002", "Масло смазочное отработанное", 580852),
    ("19.20.29.100-00003", "Масло осевое", 653217),
    ("19.20.29.100-00004", "Масло шпиндельное", 649542),
    ("19.20.29.110-00001", "Масло моторное", 665526),
    ("19.20.29.111-00001", "Масло универсальное", 640923),
    ("19.20.29.111-00002", "Жидкость для быстрого запуска двигателя", 540374),
    ("19.20.29.111-00003", "Калибровочная жидкость", 661128),
    ("19.20.29.111-00004", "Масло белое минеральное", 683508),
    ("19.20.29.113-00001", "Масло авиационное", 540497),
    ("19.20.29.114-00001", "Масло моторное авиационное", 666674),
    ("19.20.29.119-00001", "Масло двухтактное для бензопилы", 469955),
    ("19.20.29.120-00001", "Масло трансмиссионное", 668672),
    ("19.20.29.120-00002", "Жидкость гидроусилителя руля", 562900),
    ("19.20.29.130-00001", "Масло гидравлическое", 668120),
    ("19.20.29.130-00002", "Масло для вакуумных насосов", 640924),
    ("19.20.29.140-00001", "Масло индустриальное", 668236),
    ("19.20.29.140-00002", "Масло трансформаторное", 659371),
    ("19.20.29.150-00001", "Масло компрессорное", 668528),
    ("19.20.29.150-00002", "Нигрол", 519038),
    ("19.20.29.160-00001", "Масло турбинное", 665449),
    ("19.20.29.171-00001", "Масло антикоррозионное", 554106),
    ("19.20.29.172-00001", "Масло электроизоляционное трансформаторное", 667542),
    ("19.20.29.180-00001", "Солидол", 476575),
    ("19.20.29.180-00002", "Фаза масляная", 657178),
    ("19.20.29.190-00001", "Автол", 458138),
    ("19.20.29.190-00002", "Масло-теплоноситель", 640920),
    ("19.20.29.190-00003", "Масло формовочное", 640911),
    ("19.20.29.190-00004", "Замазка защитная консервационная", 655413),
    ("19.20.29.210-00001", "Смазка пластичная универсальная", 665652),
    ("19.20.29.210-00002", "Литол", 462257),
    ("19.20.29.210-00003", "Смазка Графитная", 497766),
    ("19.20.29.210-00004", "Консталин", 490888),
    ("19.20.29.210-00005", "Цепная высоковязкая специальная смазка", 484302),
    ("19.20.29.210-00006", "Смазка железнодорожная", 628872),
    ("19.20.29.210-00007", "Смазка антивибрационная", 670973),
    ("19.20.29.210-00008", "Смазка противозадирочная для отбора кернов", 670977),
    ("19.20.29.230-00001", "Эмульсол", 471826),
    ("19.20.29.230-00002", "Смазочно-охлаждающая жидкость", 682785),
]


def clean(value) -> str:
    return str(value or "").strip()


def class_specs(row: dict) -> list[str]:
    values = []
    for label, key in [
        ("ISO VG", "iso_vg"), ("SAE", "sae_engine"), ("SAE Gear", "sae_gear"),
        ("API", "api"), ("API GL", "api_gl"), ("ACEA", "acea"),
        ("DIN", "din"), ("ISO", "iso_class"), ("ГОСТ", "gost"),
        ("NLGI", "nlgi"), ("ASTM", "astm_product"),
    ]:
        value = clean(row.get(key))
        if value and value != "-":
            values.append(f"{label} {value}" if not value.upper().startswith(label.upper()) else value)
    return values


def standardized_title(row: dict) -> str:
    family = FAMILY_RULES[row["category_code"]][1]
    specs = class_specs(row)
    base = clean(row.get("base_oil"))
    thickener = clean(row.get("thickener"))
    additions = specs[:4]
    if row["category_code"] == "G" and thickener:
        additions.insert(0, f"загуститель: {thickener}")
    elif base:
        additions.append(f"основа: {base}")
    application = clean(row.get("application"))
    if application:
        additions.append(f"назначение: {application}")
    return family + (", " + ", ".join(additions) if additions else "")


def build_rows(classes: list[dict]) -> list[dict]:
    counters = Counter()
    rows = []
    for item in sorted((row for row in classes if row.get("category_code") in FAMILY_RULES), key=lambda row: (row["category_code"], row["id"])):
        family = item["category_code"]
        counters[family] += 1
        skp, skp_name, key_rule = FAMILY_RULES[family]
        suffix = 10000 + counters[family]
        rows.append({
            "proposed_enkt_code": f"{skp}-{suffix:05d}",
            "skp_code": skp,
            "skp_name": skp_name,
            "technical_profile_id": item["id"],
            "standardized_name_ru": standardized_title(item),
            "professional_key_rule": key_rule,
            "category_code": family,
            "product_form": item.get("product_form"),
            "application": item.get("application"),
            "industry": item.get("industry"),
            "specifications": class_specs(item),
            "tnved_chapter_hint": item.get("tnved_chapter"),
            "tnved_code_hint": item.get("tnved_hint"),
            "status": "PILOT_PROPOSED_NOT_OFFICIAL",
        })
    return rows


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min(55, max(11, max(len(str(cell.value or "")) for cell in column) + 2))
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_table(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def build_workbook(rows: list[dict]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    summary = [
        ["Статус", "Пилотное предложение; не является официальным справочником"],
        ["Цель", "Расширить ЕНКТ профессиональными классами ГСМ без кодирования бренда и упаковки"],
        ["Формат", "СКП (родитель) + пятизначный идентификатор технического профиля"],
        ["Пилотных профилей", len(rows)],
        ["Семейств", len(FAMILY_RULES)],
        ["Текущих укрупнённых ЕНКТ-позиций из архива", len(CURRENT_ENKT)],
        ["Версия", VERSION],
        ["Дата формирования", date.today().isoformat()],
        ["Главный принцип", "Один технический класс = один ЕНКТ; бренды и SKU связываются отдельно many-to-one"],
    ]
    add_table(wb, "01_Предложение", ["Показатель", "Содержание"], summary)

    pilot_headers = [
        "Проектный код ЕНКТ", "Код СКП", "Название СКП", "ID техпрофиля",
        "Стандартизированное название", "Правило профессионального ключа", "Семейство",
        "Форма", "Применение", "Отрасль", "Спецификации", "Глава ТН ВЭД (подсказка)",
        "ТН ВЭД (подсказка)", "Статус",
    ]
    add_table(wb, "02_Пилот_171", pilot_headers, [[
        row["proposed_enkt_code"], row["skp_code"], row["skp_name"], row["technical_profile_id"],
        row["standardized_name_ru"], row["professional_key_rule"], row["category_code"],
        row["product_form"], row["application"], row["industry"], "; ".join(row["specifications"]),
        row["tnved_chapter_hint"], row["tnved_code_hint"], row["status"],
    ] for row in rows])

    add_table(wb, "03_Текущий_ЕНКТ_39", ["Код ЕНКТ", "Наименование", "XARID product ID", "Проверено", "Источник", "Диагноз"], [[
        code, title, product_id, "2026-07-20",
        "https://xarid-api-trade.uzex.uz/Lib/GetProducts/113821",
        "Укрупнённая позиция: недостаточно SAE/API/ISO VG/DIN/NLGI для точного ценового сравнения",
    ] for code, title, product_id in CURRENT_ENKT])

    blocks = [
        ["00001–09999", "Сохранённые действующие/legacy позиции", "Не перенумеровывать"],
        ["10000–69999", "Профессиональные технические профили", "Основной рабочий диапазон"],
        ["70000–79999", "Экстремальные и специальные применения", "Отдельная экспертная валидация"],
        ["80000–89999", "Резерв новых международных стандартов", "Не выдавать до утверждения"],
        ["90000–99998", "Временные коды на экспертной проверке", "Запрещены для ценового бенчмарка"],
        ["99999", "Неопределённый прочий продукт", "Обязательная последующая реклассификация"],
    ]
    add_table(wb, "04_Правила_5_цифр", ["Диапазон", "Назначение", "Правило"], blocks)

    crosswalk = [
        ["СКП → ЕНКТ", "1:N", "Наследуется из первых цифр ЕНКТ", "Автоматически"],
        ["ЕНКТ ↔ ИКПУ", "N:M", "Таблица связей с датой и источником", "Налоговая валидация"],
        ["ЕНКТ ↔ ТН ВЭД", "N:M", "Состав, доля нефтепродукта, форма и назначение", "Таможенная/экспертная валидация"],
        ["Бренд/SKU → ЕНКТ", "N:1", "Точный профессиональный ключ", "Сертификат/TDS/PDS"],
        ["ЕНКТ → единица", "1:N", "Базовая аналитическая единица кг; торговые единицы конвертируются", "Коэффициент и источник"],
    ]
    add_table(wb, "05_Crosswalk", ["Связь", "Кардинальность", "Механизм", "Контроль"], crosswalk)

    rollout = [
        [1, "Утвердить модель данных и 8 профессиональных ключей", "Минэкономфин + Статкомитет + Налоговый комитет + таможня", "Согласованная схема"],
        [2, "Пилот 171 класса на исторических закупках", "Рабочая группа ГСМ", "Метрики качества и ценового эффекта"],
        [3, "Экспертная проверка спорных ТН ВЭД/ИКПУ связей", "Таможня + Налоговый комитет", "Версионированный crosswalk"],
        [4, "API ЕНКТ с обязательными атрибутами по семейству", "Оператор ЕНКТ", "Машиночитаемая валидация"],
        [5, "Миграция лотов и ЭСФ без потери исходного текста", "UZEX/XARID/ЭСФ", "Параллельное хранение raw + normalized"],
        [6, "Расширение на технические жидкости и спецпродукты", "Межведомственная группа", "Следующая редакция"],
    ]
    add_table(wb, "06_План_внедрения", ["Этап", "Действие", "Ответственные", "Результат"], rollout)

    archive = [
        ["1057", "08.12.2025", "Стандартизация 114 086 строк ЭСФ; 998 целевых названий; SAE/API/ISO/DIN/ГОСТ"],
        ["1109–1120", "19–20.12.2025", "Сбор ОКЭД, ШЭК, СКП, ТН ВЭД, ИКПУ и функциональной модели ГСМ"],
        ["1124", "21.12.2025", "ЕНКТ = код СКП + пятизначный суффикс; 22 513 продуктов; 548 033 связей"],
        ["1127–1128", "21.12.2025", "Crosswalk ИКПУ → ЕНКТ → СКП → ТН ВЭД; 70,3% детального покрытия"],
        ["1130", "22.12.2025", "39 позиций ЕНКТ 19.20.29; 7 без ИКПУ; детализация характеристик"],
        ["1136–1137", "23.12.2025", "ASTM и DIN 51502 как профессиональные признаки"],
        ["1241", "30.01.2026", "Продолжение каталога моторных масел"],
    ]
    add_table(wb, "07_Архивная_основа", ["Сессии", "Дата", "Что восстановлено"], archive)

    XLSX_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUTPUT)


def main() -> None:
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    rows = build_rows(catalog["classes"])
    payload = {
        "schema_version": 1,
        "version": VERSION,
        "status": "pilot_proposed_not_official",
        "principle": "one technical profile equals one ENKT code; brands and SKUs map many-to-one",
        "current_enkt_reference_count": len(CURRENT_ENKT),
        "current_enkt_source": "https://xarid-api-trade.uzex.uz/Lib/GetProducts/113821",
        "current_enkt_verified_at": "2026-07-20",
        "current_enkt_reference": [
            {"code": code, "name_ru": title, "xarid_product_id": product_id}
            for code, title, product_id in CURRENT_ENKT
        ],
        "pilot_profiles": len(rows),
        "families": FAMILY_RULES,
        "rows": rows,
    }
    JSON_OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    build_workbook(rows)
    print(json.dumps({"pilot_profiles": len(rows), "families": len(FAMILY_RULES), "xlsx": str(XLSX_OUTPUT)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
