#!/usr/bin/env python3
"""Normalize all 106 automotive-fluid rows in Shenzhen's 2020 inspection workbooks."""

from __future__ import annotations

import hashlib
import html
import io
import json
import re
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data/shenzhen-2020-automotive-fluid-inspection.jsonl"
REPORT = ROOT / "data/shenzhen-2020-automotive-fluid-inspection-report.json"
SOURCE_ID = "SHENZHEN_CHINA_2020_AUTOMOTIVE_FLUID_INSPECTION"
SOURCE_URL = "https://www.sz.gov.cn/cn/xxgk/zfxxgj/tzgg/content/post_8537379.html"
CONFORMING_URL = "http://www.sz.gov.cn/attachment/0/748/748883/8537379.xlsx"
NONCONFORMING_URL = "http://www.sz.gov.cn/attachment/0/757/757058/8537379.xlsx"
RIGHTS_URL = "https://www.sz.gov.cn/cn/qt/gywm/"
EXPECTED_SHA256 = {
    "conforming": "6eb62911fd9b0e6a8c4e89cd33f2b5e0463ec174812277b04219967a5d2d94eb",
    "nonconforming": "9dda9aa6d8b4b295e559b63a7dc6b3e2a1a3c47484bd40d1e8a9ea9c81eb6a83",
}
SNAPSHOT_DATE = "2026-07-21"
REPORT_DATE = "2021-02-01"


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def empty_marker(value: object) -> str:
    value = clean(value)
    return "" if value in {"/", "-", "—"} else value


def fetch(url: str, expected_sha256: str) -> bytes:
    request = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 MFClassifierResearch/1.0",
        "Referer": SOURCE_URL,
    })
    with urllib.request.urlopen(request, timeout=120) as response:
        payload = response.read()
    digest = hashlib.sha256(payload).hexdigest()
    if digest != expected_sha256:
        raise RuntimeError(f"Shenzhen 2020 XLSX changed: expected {expected_sha256}, received {digest}")
    return payload


def manufacturer(value: object) -> tuple[str, str]:
    source = clean(value)
    normalized = re.sub(r"^(?:进口商|委托方|分装|销售商|出品)[：:]\s*", "", source)
    normalized = re.sub(r"^广东[·•]\s*", "", normalized)
    return normalized, source


def classify(name: str, model: str) -> tuple[str, str, str]:
    joined = name + " " + model
    if "制动液" in joined or "刹车油" in joined or "刹车液" in joined:
        return "TF", "机动车制动液", "motor-vehicle brake fluid"
    if "尿素" in joined or "尾气净化液" in joined or "尾气处理液" in joined or "氮氧化物还原剂" in joined:
        return "TF", "车用尿素溶液", "automotive aqueous urea solution"
    if "冷却液" in joined or "防冻液" in joined or "水箱宝" in joined:
        return "TF", "机动车发动机冷却液", "motor-vehicle engine coolant"
    if any(token in joined for token in ("燃油", "汽油添加剂", "汽油复合剂", "喷油嘴清洗剂")):
        return "S", "车用汽油清净剂", "automotive gasoline detergent additive"
    if "机油" in joined or "发动机油" in joined or "润滑油" in joined:
        return "M", "机动车发动机润滑油", "motor-vehicle engine lubricating oil"
    if re.search(r"(?:0W|5W|10W|15W|20W|25W)\s*[-/]?\s*\d{2}", joined, re.I):
        return "M", "机动车发动机润滑油", "motor-vehicle engine lubricating oil"
    raise RuntimeError(f"Unclassified Shenzhen 2020 product: {name!r} / {model!r}")


def technical(name: str, model: str) -> tuple[dict, list[str]]:
    joined = f"{name} {model}".upper().replace("Ⅱ", "II").replace("－", "-").replace("／", "/")
    flags = []
    if re.search(r"(?<![A-Z0-9])1OW\s*[-/]?\s*\d{2}", joined):
        joined = re.sub(r"(?<![A-Z0-9])1OW(?=\s*[-/]?\s*\d{2})", "10W", joined)
        flags.append("source_1ow_typo_normalized_to_sae_10w")
    if re.search(r"(?<![A-Z0-9])OW\s*[-/]?\s*\d{2}", joined):
        joined = re.sub(r"(?<![A-Z0-9])OW(?=\s*[-/]?\s*\d{2})", "0W", joined)
        flags.append("source_ow_typo_normalized_to_sae_0w")
    api_order = r"CF-4|CH-4|CI-4|CJ-4|CK-4|CD|CF|SG|SJ|SL|SM|SN|SP|SQ"
    api = re.findall(rf"(?<![A-Z])({api_order})(?![A-Z])", joined)
    sae = [
        f"{winter}-{summer}"
        for winter, summer in re.findall(r"(?<![0-9])(0W|5W|10W|15W|20W|25W)\s*[-/]?\s*([0-9]{2})(?![0-9])", joined)
    ]
    acea = re.findall(r"(?<![A-Z0-9])(A3/B4|A5/B5|C[235])(?![A-Z0-9])", joined)
    ilsac = re.findall(r"GF-[456](?:A|B)?", joined)
    if re.search(r"D0T\s*-?\s*[345]", joined):
        flags.append("source_dot_zero_typo_normalized_to_dot")
    dot = [f"DOT {value}" for value in re.findall(r"D[O0]T\s*-?\s*([345])", joined)]
    hzy = [f"HZY{value}" for value in re.findall(r"HZY\s*([3456])", joined)]
    compact = joined.replace(" ", "")
    coolant = re.findall(r"(?:LEC|HEC|LOC)-(?:II|11)(?:-(?:15|25|26|35|36|40|45|50))?", compact)
    freezing_point = [f"{value} °C" for value in re.findall(r"冰点[：:]?\s*(-?\d+)\s*℃", joined)]
    urea = ["AUS 32"] if re.search(r"AUS\s*32", joined) else []
    return {
        "api_source_reported": sorted(set(api)),
        "sae_source_reported": sorted(set(sae)),
        "acea_source_reported": sorted(set(acea)),
        "ilsac_source_reported": sorted(set(ilsac)),
        "brake_fluid_dot_source_reported": sorted(set(dot)),
        "brake_fluid_hzy_source_reported": sorted(set(hzy)),
        "coolant_class_source_reported": sorted(set(coolant)),
        "coolant_freezing_point_source_reported": sorted(set(freezing_point)),
        "washer_fluid_class_source_reported": [],
        "urea_class_source_reported": urea,
    }, flags


def brand(source_brand: str, name: str, producer: str) -> tuple[str, str]:
    source_brand = empty_marker(source_brand)
    if source_brand and source_brand not in {"图形", "图案", "图案商标", "图形商标"}:
        return source_brand, "source_reported_text_brand"
    prefixes = (
        ("上汽大众", "Volkswagen"), ("奥迪", "Audi"), ("雷克萨斯", "Lexus"),
        ("丰田", "Toyota"), ("壳牌", "Shell"), ("本田", "Honda"),
    )
    for prefix, normalized in prefixes:
        if name.startswith(prefix):
            return normalized, "brand_in_explicit_oem_product_name"
    return producer, "nominal_producer_fallback_no_usable_text_brand"


def package_removed_model(value: str) -> str:
    value = value.replace("毫升", "ml").replace("千克", "kg").replace("公斤", "kg").replace("升", "l")
    value = re.sub(r"\d+(?:\.\d+)?\s*(?:ml|mL|L|l|kg|KG|g)(?:\s*[×xX]\s*\d+)?\s*(?:/\s*(?:瓶|桶|罐|盒))?", "", value)
    return clean(value).strip("，,、 ：:;；")


def split_failures(value: str) -> list[str]:
    value = clean(value)
    if not value:
        return []
    value = re.sub(r"(?:^|[；;])\s*\d+[、.]\s*", ";", value)
    return [clean(part) for part in value.strip(";").split(";") if clean(part)]


def workbook_rows(payload: bytes, outcome: str) -> list[dict]:
    sheet = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)["Sheet1"]
    header = tuple(clean(value) for value in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    expected = ("序号", "受检单位名称", "样品名称", "标称商标")
    if header[:4] != expected:
        raise RuntimeError(f"Unexpected Shenzhen 2020 columns: {header!r}")
    rows = []
    for values in sheet.iter_rows(min_row=4, values_only=True):
        if not isinstance(values[0], (int, float)):
            continue
        rows.append({
            "source_row": int(values[0]), "product_name": clean(values[2]),
            "brand_source_reported": empty_marker(values[3]), "model": clean(values[4]),
            "production_date": empty_marker(values[5]), "producer": clean(values[6]),
            "outcome_source_reported": clean(values[7]),
            "nonconforming_items_source_reported": clean(values[8]) if outcome == "nonconforming" else "",
            "source_note": empty_marker(values[9]) if outcome == "nonconforming" else empty_marker(values[8]),
            "source_workbook": outcome,
        })
    expected_count = 95 if outcome == "conforming" else 11
    if len(rows) != expected_count:
        raise RuntimeError(f"Expected {expected_count} {outcome} rows, received {len(rows)}")
    return rows


def main() -> None:
    payloads = {
        "conforming": fetch(CONFORMING_URL, EXPECTED_SHA256["conforming"]),
        "nonconforming": fetch(NONCONFORMING_URL, EXPECTED_SHA256["nonconforming"]),
    }
    raw_rows = []
    for outcome, payload in payloads.items():
        raw_rows.extend(workbook_rows(payload, outcome))

    records = []
    for row in raw_rows:
        outcome = row["source_workbook"]
        family, kind_zh, kind_en = classify(row["product_name"], row["model"])
        extracted, technical_flags = technical(row["product_name"], row["model"])
        producer, producer_source = manufacturer(row["producer"])
        product_brand, brand_basis = brand(row["brand_source_reported"], row["product_name"], producer)
        record_marker = "C" if outcome == "conforming" else "NC"
        flags = [
            "official_2020_municipal_inspection_observation_not_current_catalog_offer",
            *technical_flags,
        ]
        if outcome == "nonconforming":
            flags.append("nonconforming_product_do_not_treat_as_approved_or_recommended")
        if "假冒" in row["source_note"]:
            flags.append("source_reported_counterfeit_trademark_product")
        source_facts = {key: row[key] for key in sorted(row)}
        records.append({
            "source_id": SOURCE_ID,
            "source_record_id": f"SZ-CN-2020-{record_marker}-{row['source_row']:03d}",
            "source_row": row["source_row"], "source_workbook": outcome,
            "source_url": SOURCE_URL,
            "attachment_url": CONFORMING_URL if outcome == "conforming" else NONCONFORMING_URL,
            "rights_url": RIGHTS_URL, "snapshot_date": SNAPSHOT_DATE, "report_date": REPORT_DATE,
            "market": "China / Shenzhen", "manufacturer": producer,
            "manufacturer_source_reported": producer_source,
            "brand": product_brand, "brand_source_reported": row["brand_source_reported"],
            "brand_basis": brand_basis, "product_name": row["product_name"],
            "product_name_basis": "source_reported_sample_name_from_municipal_quality_inspection",
            "model_specification_source_reported": row["model"],
            "model_specification_without_package": package_removed_model(row["model"]),
            "product_kind_source_reported": kind_zh, "product_kind_english": kind_en,
            "family_code": family, "technical": extracted,
            "production_date_or_batch_source_reported": row["production_date"],
            "inspection_outcome": outcome,
            "inspection_retest_confirmed_nonconforming": False,
            "nonconforming_items": split_failures(row["nonconforming_items_source_reported"]),
            "nonconforming_items_source_reported": row["nonconforming_items_source_reported"],
            "source_note": row["source_note"],
            "lifecycle_status": f"official_2020_shenzhen_inspection_{outcome}_at_test_current_market_status_unverified",
            "source_quality_flags": flags,
            "source_facts_sha256": hashlib.sha256(json.dumps(source_facts, ensure_ascii=False, sort_keys=True).encode()).hexdigest(),
            "evidence_status": f"official_government_{outcome}_product_inspection_observation",
        })

    output_text = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records)
    OUTPUT.write_text(output_text, encoding="utf-8")
    report = {
        "status": "official_shenzhen_2020_complete_automotive_fluid_inspection_normalized",
        "source_id": SOURCE_ID, "source_url": SOURCE_URL,
        "attachment_urls": {"conforming": CONFORMING_URL, "nonconforming": NONCONFORMING_URL},
        "snapshot_date": SNAPSHOT_DATE, "report_date": REPORT_DATE,
        "source_xlsx_sha256": {key: hashlib.sha256(value).hexdigest() for key, value in payloads.items()},
        "source_rows": len(records),
        "outcomes": dict(sorted(Counter(row["inspection_outcome"] for row in records).items())),
        "source_product_types": dict(sorted(Counter(row["product_kind_source_reported"] for row in records).items())),
        "families_before_identity_merging": dict(sorted(Counter(row["family_code"] for row in records).items())),
        "rows_with_api": sum(bool(row["technical"]["api_source_reported"]) for row in records),
        "rows_with_sae": sum(bool(row["technical"]["sae_source_reported"]) for row in records),
        "rows_with_acea": sum(bool(row["technical"]["acea_source_reported"]) for row in records),
        "rows_with_ilsac": sum(bool(row["technical"]["ilsac_source_reported"]) for row in records),
        "rows_with_brake_class": sum(bool(row["technical"]["brake_fluid_dot_source_reported"] or row["technical"]["brake_fluid_hzy_source_reported"]) for row in records),
        "rows_with_coolant_class_or_freezing_point": sum(bool(row["technical"]["coolant_class_source_reported"] or row["technical"]["coolant_freezing_point_source_reported"]) for row in records),
        "rows_with_aus32": sum(bool(row["technical"]["urea_class_source_reported"]) for row in records),
        "counterfeit_source_notes": sum("source_reported_counterfeit_trademark_product" in row["source_quality_flags"] for row in records),
        "normalized_output_sha256": hashlib.sha256(output_text.encode()).hexdigest(),
        "grain_note": "One output row is one inspected batch observation. Product identities may merge later only under strict producer, product and package-independent technical identity while every occurrence remains linked.",
        "scope_note": "Both official workbooks are complete for this program: all 95 conforming and all 11 nonconforming automotive-product rows are retained.",
        "lifecycle_note": "Conforming means only that the sampled batch passed this inspection. Neither outcome is treated as a current commercial offer or blanket approval.",
        "privacy_note": "Inspected retailers are excluded; nominal producer, product, brand, model, batch date, outcome and failure facts are retained.",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
