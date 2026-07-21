#!/usr/bin/env python3
"""Ingest lubricant-scope product/part observations from Volvo Group WebSDS news."""

from __future__ import annotations

import hashlib
import io
import json
import re
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "volvo-group-websds-lubricant-products.jsonl"
REPORT = ROOT / "data" / "volvo-group-websds-lubricant-products-report.json"
SOURCE_ID = "VOLVO_GROUP_WEBSDS_CHANGE_ARCHIVE"
SNAPSHOT_DATE = "2026-07-21"
BASE_URL = "https://websds.volvo.com/websds/"
NEWS_URL = BASE_URL + "sdsNews"
USER_AGENT = "MFClassifier research catalog/1.0 (+government classification research)"

SCOPE_RE = re.compile(
    r"(?:\boil\b|\bgrease\b|\blubricant\b|\bcoolant\b|\bantifreeze\b|"
    r"\bbrake fluid\b|\btransmission fluid\b|\bhydraulic\b|\bretarder\b|"
    r"\bgearbox\b|\baxle\b|\bcompressor\b|\bpower steering\b|\bATF\b|"
    r"\bAdBlue\b|\burea\b|breaker paste)",
    re.I,
)
PART_RE = re.compile(r"(?:VOE)?\d+[A-Z]?", re.I)


def fetch(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def identity_name(value: str) -> str:
    value = unicodedata.normalize("NFKC", clean(value)).casefold()
    value = value.replace("biodegradeble", "biodegradable").replace("g0102", "go102")
    value = value.replace("40-60", "40/60").replace("at 102", "at102")
    value = re.sub(r"\bvolvo\b", " ", value)
    value = re.sub(r"automatic transmission fluid\s+97342\s+at102", "automatic transmission fluid 97342 at102", value)
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9./-]+", " ", value)).strip()


GENERIC_IDENTITIES = {
    identity_name(value)
    for value in (
        "Brake fluid", "Compressor Oil", "Concentrated coolant", "Coolant concentrated",
        "Engine oil", "Grease", "Hydraulic Oil", "Power Steering oil", "Retarder oil",
        "Silicon grease", "Silicone oil", "Synthetic Compressor oil", "Transmission oil",
    )
}


def organization(title: str, company: str) -> str:
    if re.search(r"\bRT\b", title):
        return "Renault Trucks"
    if "VCC" in title or company == "VBC":
        return "Volvo Cars"
    return "Volvo Group"


def parse_table(table: list[list], document_id: str, document_title: str, page: int) -> list[dict]:
    header_index = next(
        (index for index, row in enumerate(table) if any("Part No" in clean(cell) for cell in row)),
        None,
    )
    if header_index is None:
        return []
    header = [clean(cell) for cell in table[header_index]]
    try:
        part_index = next(index for index, value in enumerate(header) if "Part No" in value)
        name_index = next(index for index, value in enumerate(header) if "Product name" in value)
        date_index = next(index for index, value in enumerate(header) if "SDS Month" in value)
        change_index = next(index for index, value in enumerate(header) if "Type of change" in value)
    except StopIteration:
        return []
    company_index = next((index for index, value in enumerate(header) if "Company" in value), None)
    rows = []
    for source_row, row in enumerate(table[header_index + 1:], header_index + 2):
        if len(row) <= max(part_index, name_index, date_index, change_index):
            continue
        part_cell = clean(row[part_index])
        product_name = clean(row[name_index])
        date_cell = clean(row[date_index])
        change_type = clean(row[change_index])
        company = clean(row[company_index]) if company_index is not None and len(row) > company_index else ""
        if not part_cell or not product_name or not date_cell.startswith("SDS "):
            continue
        part_numbers = []
        for value in PART_RE.findall(part_cell):
            value = value.upper()
            if value not in part_numbers:
                part_numbers.append(value)
        if not part_numbers:
            continue
        rows.append({
            "document_id": document_id,
            "document_title": document_title,
            "document_url": BASE_URL + f"streamSdsDoc?pDocumentId={document_id}",
            "source_page": page,
            "source_table_row": source_row,
            "sds_date": date_cell.removeprefix("SDS "),
            "organization": organization(document_title, company),
            "source_company_code": company,
            "part_number_cell_source": part_cell,
            "part_numbers": part_numbers,
            "product_name_source": product_name,
            "change_type_source": change_type,
        })
    return rows


def family_for(name: str) -> str:
    key = name.casefold()
    if "engine oil" in key:
        return "M"
    if any(value in key for value in ("transmission", "gearbox", "axle", "retarder", "wet brake", "atf")):
        return "T"
    if "hydraulic" in key:
        return "H"
    if "compressor" in key:
        return "C"
    if "grease" in key or "breaker paste" in key:
        return "G"
    if any(value in key for value in ("coolant", "antifreeze", "brake fluid", "adblue", "urea")):
        return "TF"
    return "I"


def specifications(names: list[str], family: str) -> dict:
    source_text = " | ".join(names)
    result = {}
    multigrades = sorted(set(
        f"{match.group(1).upper()}-{match.group(2)}"
        for match in re.finditer(r"(?<!\d)(\d{1,2}W)[- ]?(\d{2,3})(?!\d)", source_text, re.I)
    ))
    monogrades = sorted(set(re.findall(r"\bSAE\s*(\d{2,3})\b", source_text, re.I)))
    if multigrades or monogrades:
        grades = multigrades + [f"SAE {value}" for value in monogrades]
        result["sae_engine" if family == "M" else "sae_gear"] = grades[0] if len(grades) == 1 else grades
    iso_values = sorted(set(re.findall(r"(?:ISO\s*VG|\bVG)\s*(\d{2,3})\b", source_text, re.I)), key=int)
    if iso_values:
        result["iso_vg"] = iso_values[0] if len(iso_values) == 1 else iso_values
    volvo_standards = sorted(set(re.findall(r"\b(?:97|98)\d{3}\b", source_text)))
    if volvo_standards:
        result["volvo_standards_source_reported"] = volvo_standards
    renault_standards = sorted(set(re.findall(r"\b(?:RLD-\d|RN\d{2,4}|TF\s*\d{4}|AO\s*\d{4}|HF\s*\d{2})\b", source_text, re.I)))
    if renault_standards:
        result["renault_trucks_standards_source_reported"] = renault_standards
    mix_values = sorted(set(re.findall(r"\b(?:40/60|50/50)\b", source_text.replace("40-60", "40/60"))))
    if mix_values:
        result["coolant_mix_source_reported"] = mix_values
    return result


def preferred_name(names: list[str]) -> str:
    return max(
        names,
        key=lambda value: (
            bool(re.search(r"\b(?:Volvo|Renault Trucks)\b", value, re.I)),
            len(re.findall(r"\d", value)),
            len(value),
            value,
        ),
    )


def main() -> None:
    news_payload = fetch(NEWS_URL)
    news_html = news_payload.decode("utf-8", "replace")
    links = [
        (document_id, clean(re.sub(r"<[^>]+>", "", title)))
        for document_id, title in re.findall(
            r'href="streamSdsDoc\?pDocumentId=(\d+)"[^>]*>.*?</a></td>\s*<td[^>]*>(.*?)</td>',
            news_html,
            re.S | re.I,
        )
    ]
    assert len(links) == 50
    assert len({document_id for document_id, _ in links}) == len(links)

    document_payloads = {}
    all_rows = []
    document_kinds = Counter()
    for document_id, document_title in links:
        payload = fetch(BASE_URL + f"streamSdsDoc?pDocumentId={document_id}")
        document_payloads[document_id] = payload
        if payload.startswith(b"%PDF"):
            document_kinds["pdf"] += 1
            with pdfplumber.open(io.BytesIO(payload)) as pdf:
                for page_number, page in enumerate(pdf.pages, 1):
                    for table in page.extract_tables():
                        all_rows.extend(parse_table(table, document_id, document_title, page_number))
        elif payload.startswith(b"PK"):
            document_kinds["xlsx"] += 1
            workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
            for worksheet in workbook.worksheets:
                all_rows.extend(parse_table(list(worksheet.iter_rows(values_only=True)), document_id, document_title, 1))
        else:
            raise AssertionError((document_id, payload[:16]))

    assert document_kinds == {"pdf": 49, "xlsx": 1}
    assert len(all_rows) == 833
    scoped_rows = [row for row in all_rows if SCOPE_RE.search(row["product_name_source"])]
    assert len(scoped_rows) == 537

    parent = list(range(len(scoped_rows)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        left, right = find(left), find(right)
        if left != right:
            parent[right] = left

    by_part_number = defaultdict(list)
    for index, row in enumerate(scoped_rows):
        for part_number in row["part_numbers"]:
            by_part_number[part_number].append(index)
    for indices in by_part_number.values():
        for index in indices[1:]:
            union(indices[0], index)

    by_informative_name = defaultdict(list)
    for index, row in enumerate(scoped_rows):
        key = identity_name(row["product_name_source"])
        if key not in GENERIC_IDENTITIES:
            by_informative_name[key].append(index)
    for indices in by_informative_name.values():
        for index in indices[1:]:
            union(indices[0], index)

    components = defaultdict(list)
    for index, row in enumerate(scoped_rows):
        components[find(index)].append(row)

    records = []
    for component in components.values():
        # Include the original spelling as a tie-breaker: a set can otherwise
        # emit case-only aliases in process-random order and change JSONL bytes.
        names = sorted(
            {row["product_name_source"] for row in component},
            key=lambda value: (value.casefold(), value),
        )
        part_numbers = sorted({value for row in component for value in row["part_numbers"]})
        organizations = sorted({row["organization"] for row in component})
        product_name = preferred_name(names)
        brand = "Renault Trucks" if organizations == ["Renault Trucks"] else "Volvo"
        manufacturer = "Renault Trucks" if brand == "Renault Trucks" else "Volvo Group"
        family = family_for(product_name)
        record_id_basis = "|".join([brand, identity_name(product_name), family, *part_numbers])
        source_record_id = "VOLVO-WEBSDS-" + hashlib.sha256(record_id_basis.encode()).hexdigest()[:20].upper()
        quality_flags = []
        if any("G0102" in name for name in names):
            quality_flags.append("source_g0102_notation_normalized_to_go102_for_identity_only")
        if identity_name(product_name) in GENERIC_IDENTITIES:
            quality_flags.append("generic_source_product_name_kept_part_number_specific")
        records.append({
            "source_id": SOURCE_ID,
            "source_record_id": source_record_id,
            "manufacturer": manufacturer,
            "brand": brand,
            "product_name": product_name,
            "product_name_source_aliases": names,
            "family_code": family,
            "market": "MULTI_MARKET_WEBSDS",
            "supplier_organizations": organizations,
            "part_numbers": part_numbers,
            "specifications": specifications(names, family),
            "source_occurrences": sorted(component, key=lambda row: (
                row["sds_date"], row["document_id"], row["source_page"], row["source_table_row"]
            )),
            "first_observed_sds_date": min(row["sds_date"] for row in component),
            "last_observed_sds_date": max(row["sds_date"] for row in component),
            "change_types_source_reported": sorted({row["change_type_source"] for row in component}),
            "lifecycle_status": "official_sds_change_observation_2023_09_to_2026_06_current_availability_unverified",
            "source_url": NEWS_URL,
            "snapshot_date": SNAPSHOT_DATE,
            "source_quality_flags": quality_flags,
        })

    records.sort(key=lambda row: (row["brand"], row["family_code"], identity_name(row["product_name"]), row["source_record_id"]))
    assert len({row["source_record_id"] for row in records}) == len(records)
    assert sum(len(row["source_occurrences"]) for row in records) == len(scoped_rows)
    OUT.write_text("".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records), encoding="utf-8")
    report = {
        "schema_version": 1,
        "snapshot_date": SNAPSHOT_DATE,
        "source_id": SOURCE_ID,
        "source_url": NEWS_URL,
        "source_documents": len(links),
        "source_document_formats": dict(sorted(document_kinds.items())),
        "source_table_rows": len(all_rows),
        "lubricant_scope_occurrences": len(scoped_rows),
        "normalized_product_part_identities": len(records),
        "unique_part_numbers": len({value for row in records for value in row["part_numbers"]}),
        "brands": dict(sorted(Counter(row["brand"] for row in records).items())),
        "families": dict(sorted(Counter(row["family_code"] for row in records).items())),
        "source_quality_flags": dict(sorted(Counter(flag for row in records for flag in row["source_quality_flags"]).items())),
        "news_page_sha256": hashlib.sha256(news_payload).hexdigest(),
        "source_documents_aggregate_sha256": hashlib.sha256("".join(
            hashlib.sha256(document_payloads[document_id]).hexdigest()
            for document_id, _ in sorted(links)
        ).encode()).hexdigest(),
        "normalized_output_sha256": hashlib.sha256(OUT.read_bytes()).hexdigest(),
        "publication_scope": "Attributed non-expressive SDS change-list facts only: product names, part numbers, dates, change types and derived technical grades. Full SDS text, hazard content, images and document layout are not republished.",
        "lifecycle_note": "A WebSDS change-list observation proves that an official SDS/transport record was created or updated on the stated date; it does not by itself prove current market availability.",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))


if __name__ == "__main__":
    main()
