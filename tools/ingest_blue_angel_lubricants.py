#!/usr/bin/env python3
"""Download and normalize the official Blue Angel DE-UZ 178 product export."""

from __future__ import annotations

import hashlib
import html
import io
import json
import re
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "blue-angel-de-uz-178-products.jsonl"
REPORT = ROOT / "data" / "blue-angel-de-uz-178-products-report.json"
SOURCE_URL = "https://www.blauer-engel.de/de/produktwelt/schmierstoffe-und-hydraulikfluessigkeiten"
IMPRINT_URL = "https://www.blauer-engel.de/de/impressum"
CRITERIA_URL = "https://produktinfo.blauer-engel.de/uploads/criteriafile/de/202/DE-UZ%20178-202201-de%20Kriterien-V4.pdf"
USER_AGENT = "MFClassifierResearch/1.0 (official-ecolabel-product-export research)"

CATEGORIES = {
    "getriebeschmierstoffe": ("Getriebeschmierstoffe", "T"),
    "hydraulikfluessigkeiten": ("Hydraulikflüssigkeiten", "H"),
    "kettenschmierstoffe-fuer-motorsaegen": ("Kettenschmierstoffe für Motorsägen", "I"),
    "schaloele-und-betontrennmittel": ("Schalöle und Betontrennmittel", "S"),
    "schmierfette": ("Schmierfette", "G"),
    "schmierstoffe": ("Schmierstoffe", "I"),
    "verlustschmierstoffe": ("Verlustschmierstoffe", "I"),
}
FAMILY_PRIORITY = {"G": 0, "H": 1, "T": 2, "S": 3, "I": 4}


def clean(value) -> str:
    return re.sub(r"\s+", " ", html.unescape(str(value or ""))).strip()


def plain(fragment: str) -> str:
    return clean(re.sub(r"<[^>]+>", " ", fragment or ""))


def normalize(value) -> str:
    value = unicodedata.normalize("NFKC", clean(value)).casefold().replace("ß", "ss")
    return re.sub(r"[^0-9a-z]+", " ", value).strip()


def download(url: str, data: bytes | None = None) -> bytes:
    request = urllib.request.Request(url, data=data, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def export_xlsx(page: bytes) -> bytes:
    text = page.decode("utf-8", errors="replace")
    form_match = re.search(r'<form class="bep-export-form.*?</form>', text, re.I | re.S)
    assert form_match, "official export form missing"
    form = form_match.group(0)
    fields = {}
    for name in ["type", "id", "export_type", "slug", "fileslug", "form_build_id", "form_id"]:
        tag = re.search(rf'<input[^>]+name="{re.escape(name)}"[^>]*>', form, re.I)
        assert tag, name
        value = re.search(r'value="([^"]*)"', tag.group(0), re.I)
        fields[name] = html.unescape(value.group(1)) if value else ""
    assert fields["id"] == "202" and fields["form_id"] == "bep_export_form"
    fields["xlsx"] = "XLS-Datei"
    content = download(SOURCE_URL, urllib.parse.urlencode(fields).encode())
    assert content.startswith(b"PK\x03\x04"), "export is not XLSX"
    return content


def cards(page: bytes) -> list[dict]:
    text = page.decode("utf-8", errors="replace")
    result = []
    for href, body in re.findall(
        r'<a\s+href="([^"]+)"[^>]*class="m-bep_raluz__productslink"[^>]*>(.*?)</a>', text, re.I | re.S
    ):
        title = re.search(r'<h2[^>]*>(.*?)</h2>', body, re.I | re.S)
        company = re.search(r'<div class="m-bep_productcard__company">(.*?)</div>', body, re.I | re.S)
        assert title and company
        result.append({
            "product_page_path": html.unescape(href),
            "product_name": plain(title.group(1)),
            "manufacturer": plain(company.group(1)),
        })
    return result


def technical(name: str, family: str) -> dict:
    upper = name.upper()
    common_iso = {"10", "15", "22", "32", "46", "68", "100", "150", "220", "320", "460", "680", "1000"}
    explicit_iso = set(re.findall(r"\bISO\s*(?:VG\s*)?(\d{1,4})\b", upper))
    contextual = set()
    if family in {"H", "T", "I"}:
        for value in re.findall(r"(?<!\d)(\d{2,4})(?!\d)", upper):
            if value in common_iso:
                contextual.add(value)
    iso_vg = sorted(explicit_iso | contextual, key=int)
    nlgi = []
    if family == "G":
        match = re.search(r"\b(?:NLGI\s*)?(00|0|[1-6])(?:\s*[A-Z])?\b", upper)
        if match:
            nlgi = [match.group(1)]
    sae = sorted(set(re.findall(r"\b(?:SAE\s*)?((?:0W|5W|10W|15W|20W|25W)(?:[- ]?\d{2,3})?|(?:70W|75W|80W|85W)(?:[- ]?\d{2,3})?)\b", upper)))
    return {"iso_vg": iso_vg, "nlgi": nlgi, "sae": [value.replace(" ", "-") for value in sae]}


def main() -> None:
    main_page = download(SOURCE_URL)
    xlsx = export_xlsx(main_page)

    main_cards = cards(main_page)
    assert len(main_cards) == 149
    page_facts_text = json.dumps(sorted(main_cards, key=lambda row: row["product_page_path"]), ensure_ascii=False, sort_keys=True)
    page_facts_hash = hashlib.sha256(page_facts_text.encode()).hexdigest()
    cards_by_key: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for card in main_cards:
        cards_by_key[(normalize(card["product_name"]), normalize(card["manufacturer"]))].append(card)

    category_membership: dict[str, set[tuple[str, str]]] = defaultdict(set)
    category_page_hashes = {}
    category_hits = {}
    base = SOURCE_URL + "/"
    for slug, (label, _) in CATEGORIES.items():
        url = base + slug
        page = download(url)
        category_cards = cards(page)
        category_facts_text = json.dumps(sorted(category_cards, key=lambda row: row["product_page_path"]), ensure_ascii=False, sort_keys=True)
        category_page_hashes[url] = hashlib.sha256(category_facts_text.encode()).hexdigest()
        category_hits[label] = len(category_cards)
        for card in category_cards:
            category_membership[card["product_page_path"]].add((slug, label))
    assert sum(category_hits.values()) == 159

    workbook = load_workbook(io.BytesIO(xlsx), data_only=True, read_only=True)
    assert workbook.sheetnames == ["Blauer Engel"]
    sheet = workbook["Blauer Engel"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Produktname", "Herstellername", "Adresse", "Telefon", "E-Mail", "Link zur Firmenwebsite", "Link zur Produktwebsite")
    source_rows = rows[1:]
    assert len(source_rows) == 149
    export_facts_text = json.dumps([[clean(value) for value in row] for row in source_rows], ensure_ascii=False, sort_keys=True)
    export_facts_hash = hashlib.sha256(export_facts_text.encode()).hexdigest()

    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for source_row_number, row in enumerate(source_rows, 2):
        product_name, manufacturer = clean(row[0]), clean(row[1])
        key = (normalize(product_name), normalize(manufacturer))
        matching_cards = cards_by_key.get(key, [])
        assert matching_cards, (product_name, manufacturer)
        grouped[key].append({
            "source_row_number": source_row_number,
            "product_name": product_name,
            "manufacturer": manufacturer,
            "product_external_url": clean(row[6]),
            "cards": matching_cards,
        })

    records = []
    for key, occurrences in sorted(grouped.items()):
        first = occurrences[0]
        paths = sorted({card["product_page_path"] for occurrence in occurrences for card in occurrence["cards"]})
        categories = sorted({label for path in paths for _, label in category_membership[path]})
        category_slugs = sorted({slug for path in paths for slug, _ in category_membership[path]})
        assert categories, first
        families = {CATEGORIES[slug][1] for slug in category_slugs}
        family = min(families, key=FAMILY_PRIORITY.__getitem__)
        fingerprint = hashlib.sha256("|".join(key).encode()).hexdigest()[:16]
        records.append({
            "source_id": "BLUE_ANGEL_DE_UZ_178",
            "source_record_id": f"BLUE-ANGEL-178-{fingerprint}",
            "source_url": SOURCE_URL,
            "source_export_facts_sha256": export_facts_hash,
            "source_page_facts_sha256": page_facts_hash,
            "category_page_sha256": {base + slug: category_page_hashes[base + slug] for slug in category_slugs},
            "criteria_url": CRITERIA_URL,
            "snapshot_date": date.today().isoformat(),
            "market": "Germany / Blue Angel certification",
            "product_name": first["product_name"],
            "manufacturer": first["manufacturer"],
            "product_page_urls": ["https://www.blauer-engel.de" + path for path in paths],
            "product_external_urls": sorted({occurrence["product_external_url"] for occurrence in occurrences if occurrence["product_external_url"]}),
            "certification_standard": "DE-UZ 178",
            "lifecycle_status": "listed_in_current_official_blue_angel_product_export",
            "family_code": family,
            "classification_basis": "official_blue_angel_product_categories",
            "official_categories": categories,
            "technical": technical(first["product_name"], family),
            "source_occurrence_count": len(occurrences),
            "source_row_numbers": sorted(occurrence["source_row_number"] for occurrence in occurrences),
        })

    output_text = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records)
    OUTPUT.write_text(output_text, encoding="utf-8")
    report = {
        "schema_version": 1,
        "status": "official_current_ecolabel_product_export_normalized",
        "snapshot_date": date.today().isoformat(),
        "source_url": SOURCE_URL,
        "imprint_url": IMPRINT_URL,
        "criteria_url": CRITERIA_URL,
        "source_page_facts_sha256": page_facts_hash,
        "source_export_facts_sha256": export_facts_hash,
        "category_page_facts_sha256": category_page_hashes,
        "normalized_output_sha256": hashlib.sha256(output_text.encode()).hexdigest(),
        "export_rows": len(source_rows),
        "official_product_cards": len(main_cards),
        "normalized_products": len(records),
        "duplicate_export_occurrences_merged": len(source_rows) - len(records),
        "manufacturers": len({normalize(row["manufacturer"]) for row in records}),
        "category_occurrences": sum(category_hits.values()),
        "category_hits": category_hits,
        "families": dict(sorted(Counter(row["family_code"] for row in records).items())),
        "rights_note": "The official Blue Angel product page provides an explicit XLS export. Only factual product, manufacturer, category, certification and source-link fields are republished with attribution; addresses, phone numbers, email, images and marketing descriptions are omitted.",
        "lifecycle_note": "A row means the product is present in the current official DE-UZ 178 export on the snapshot date. It is not a claim about availability in every market.",
        "grain_note": "One row is normalized manufacturer + product name. The one exact duplicate export identity is merged while both official product-page paths and source row numbers are retained.",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({key: report[key] for key in ["normalized_products", "manufacturers", "category_occurrences", "families"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
