#!/usr/bin/env python3
"""Normalize relevant 2023 China SAMR national-inspection product rows.

The original SAMR report and its complete XLSX are preserved on a Chinese
local-government portal.  Rows explicitly marked as suspected counterfeit are
kept in the audit report but excluded from canonical product identities.
"""

from __future__ import annotations

import hashlib
import html
import io
import json
import re
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "samr-china-2023-nonconforming-fluids.jsonl"
REPORT = ROOT / "data" / "samr-china-2023-nonconforming-fluids-report.json"
SOURCE_ID = "SAMR_CHINA_2023_NONCONFORMING_FLUIDS"
SOURCE_URL = "https://www.shandan.gov.cn/zfxxgk/fdzdgknr/ssjygkhjdjc_5883/202310/t20231020_1128402.html"
ATTACHMENT_URL = "https://www.shandan.gov.cn/zfxxgk/fdzdgknr/ssjygkhjdjc_5883/202310/P020231020329120054720.xlsx"
RIGHTS_URL = "https://www.shandan.gov.cn/"
EXPECTED_XLSX_SHA256 = "6eff7f988cba8b51e04ac3e7735affa790431cb9e51aec21eb9087928d38d45b"
SNAPSHOT_DATE = "2026-07-21"
REPORT_DATE = "2023-10-20"
USER_AGENT = "Mozilla/5.0 MFClassifierResearch/1.0 (public-government-inspection-data)"

SCOPE = {
    "机动车辆制动液": ("TF", "motor-vehicle brake fluid"),
    "车用尿素水溶液": ("TF", "automotive aqueous urea solution"),
    "机动车发动机冷却液": ("TF", "motor-vehicle engine coolant"),
    "机动车发动机润滑油": ("M", "motor-vehicle engine lubricating oil"),
    "汽车风窗玻璃清洗液": ("TF", "automotive windshield washer fluid"),
    "车用汽油清净剂": ("S", "automotive gasoline detergent additive"),
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def empty_marker(value: object) -> str:
    value = clean(value)
    return "" if value in {"/", "-", "—"} else value


def fetch() -> bytes:
    request = urllib.request.Request(ATTACHMENT_URL, headers={"User-Agent": USER_AGENT, "Referer": SOURCE_URL})
    with urllib.request.urlopen(request, timeout=120) as response:
        payload = response.read()
    digest = hashlib.sha256(payload).hexdigest()
    if digest != EXPECTED_XLSX_SHA256:
        raise RuntimeError(f"SAMR 2023 XLSX changed: expected {EXPECTED_XLSX_SHA256}, received {digest}")
    return payload


def split_failures(value: str) -> list[str]:
    return [clean(item).rstrip("*") for item in re.split(r"[,，]", value) if clean(item)]


def extract_engine_specs(value: str) -> tuple[list[str], list[str]]:
    upper = value.upper().replace("／", "/").replace("－", "-")
    api = re.findall(r"(?<![A-Z0-9])(CD|CF|CF-4|CH-4|CI-4|CJ-4|CK-4|SJ|SL|SM|SN|SP)(?![A-Z0-9])", upper)
    sae = []
    for winter, summer in re.findall(r"(?<![0-9])(0W|5W|10W|15W|20W|25W)\s*[-/]?\s*([0-9]{2})(?![0-9])", upper):
        sae.append(f"{winter}-{summer}")
    if re.search(r"(?<![0-9])SAE\s*40(?![0-9])", upper) or re.search(r"\bCD\s+40\b", upper):
        sae.append("40")
    return sorted(set(api)), sorted(set(sae))


def extract_brake_classes(value: str) -> tuple[list[str], list[str]]:
    upper = value.upper().replace("－", "-")
    dot = [f"DOT {item}" for item in re.findall(r"DOT\s*-?\s*([345])", upper)]
    hzy = [f"HZY{item}" for item in re.findall(r"HZY\s*([345])", upper)]
    return sorted(set(dot)), sorted(set(hzy))


def extract_coolant_classes(value: str) -> list[str]:
    upper = value.upper().replace("Ⅱ", "II").replace("－", "-")
    return sorted(set(re.findall(r"(?:LEC|HEC)-II-(?:15|25|35|40|45|50)", upper)))


def rows(payload: bytes) -> tuple[list[dict], list[dict]]:
    sheet = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)["Sheet1"]
    header = tuple(clean(value) for value in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[:11])
    expected = ("序号", "产品种类", "受检单位", "标称生产单位", "标称生产单位所在地", "产品名称", "规格型号", "生产日期/批号", "主要不合格项目", "承检机构", "备注")
    if header != expected:
        raise RuntimeError(f"Unexpected SAMR 2023 columns: {header!r}")

    included, excluded = [], []
    for values in sheet.iter_rows(min_row=4, values_only=True):
        if not isinstance(values[0], (int, float)) or clean(values[1]) not in SCOPE:
            continue
        raw = {
            "source_row": int(values[0]),
            "product_kind_source_reported": clean(values[1]),
            "producer": empty_marker(values[3]),
            "product_name_source_reported": clean(values[5]),
            "model_specification_source_reported": clean(values[6]),
            "production_date_or_batch_source_reported": empty_marker(values[7]),
            "nonconforming_items_source_reported": clean(values[8]),
            "source_note": clean(values[10]),
        }
        if "涉嫌假冒" in raw["source_note"]:
            excluded.append({**raw, "exclusion_reason": "suspected_counterfeit_nominal_producer_attribution_unreliable"})
            continue
        if not raw["producer"] or not raw["product_name_source_reported"] or not raw["model_specification_source_reported"]:
            raise RuntimeError(f"Incomplete relevant SAMR 2023 row: {raw!r}")
        included.append(raw)
    return included, excluded


def main() -> None:
    payload = fetch()
    included, excluded = rows(payload)
    if len(included) != 126 or len(excluded) != 10:
        raise RuntimeError(f"Expected 126 included and 10 counterfeit exclusions, received {len(included)} and {len(excluded)}")

    records = []
    identities: set[tuple[str, str, str, str]] = set()
    for row in included:
        family_code, kind_en = SCOPE[row["product_kind_source_reported"]]
        model = row["model_specification_source_reported"]
        api, sae = extract_engine_specs(model)
        dot, hzy = extract_brake_classes(model)
        coolant = extract_coolant_classes(model)
        washer = []
        if "水基型" in model:
            washer = ["水基型低温型" if "低温型" in model else "水基型普通型"]
        urea = ["AUS 32"] if re.search(r"AUS\s*32", row["product_name_source_reported"], re.I) else []
        identity = (row["producer"], row["product_name_source_reported"], model, family_code)
        if identity in identities:
            raise RuntimeError(f"Duplicate SAMR 2023 product identity requires occurrence merging: {identity!r}")
        identities.add(identity)
        flags = [
            "historical_regulatory_observation_not_current_catalog_listing",
            "nonconforming_product_do_not_treat_as_approved_or_recommended",
            "brand_falls_back_to_nominal_producer_because_source_has_no_brand_column",
        ]
        if row["product_kind_source_reported"] == "汽车风窗玻璃清洗液":
            flags.append("source_narrative_reports_one_unidentified_counterfeit_in_washer_scope")
        source_facts = {key: row[key] for key in sorted(row)}
        records.append({
            "source_id": SOURCE_ID,
            "source_record_id": f"SAMR-CN-2023-{row['source_row']:03d}",
            "source_url": SOURCE_URL,
            "attachment_url": ATTACHMENT_URL,
            "rights_url": RIGHTS_URL,
            "snapshot_date": SNAPSHOT_DATE,
            "report_date": REPORT_DATE,
            "market": "China",
            "manufacturer": row["producer"],
            "brand": row["producer"],
            "brand_basis": "source_reported_nominal_producer_fallback_no_separate_brand_column",
            "product_name": row["product_name_source_reported"],
            "product_name_basis": "source_reported_product_name_from_national_quality_inspection",
            "model_specification_source_reported": model,
            "product_kind_source_reported": row["product_kind_source_reported"],
            "product_kind_english": kind_en,
            "family_code": family_code,
            "technical": {
                "api_source_reported": api,
                "sae_source_reported": sae,
                "brake_fluid_dot_source_reported": dot,
                "brake_fluid_hzy_source_reported": hzy,
                "coolant_class_source_reported": coolant,
                "washer_fluid_class_source_reported": washer,
                "urea_class_source_reported": urea,
            },
            "production_date_or_batch_source_reported": row["production_date_or_batch_source_reported"],
            "inspection_outcome": "nonconforming",
            "inspection_retest_confirmed_nonconforming": "复检仍不合格" in row["source_note"],
            "nonconforming_items": split_failures(row["nonconforming_items_source_reported"]),
            "nonconforming_items_source_reported": row["nonconforming_items_source_reported"],
            "source_note": row["source_note"],
            "lifecycle_status": "official_2023_national_inspection_nonconforming_current_market_status_unverified",
            "source_quality_flags": flags,
            "source_facts_sha256": hashlib.sha256(json.dumps(source_facts, ensure_ascii=False, sort_keys=True).encode()).hexdigest(),
            "evidence_status": "official_government_nonconforming_product_inspection_observation",
        })

    records.sort(key=lambda row: row["source_record_id"])
    output_text = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records)
    OUTPUT.write_text(output_text, encoding="utf-8")
    report = {
        "status": "official_china_2023_nonconforming_fluid_product_observations_normalized",
        "source_id": SOURCE_ID,
        "source_url": SOURCE_URL,
        "attachment_url": ATTACHMENT_URL,
        "snapshot_date": SNAPSHOT_DATE,
        "report_date": REPORT_DATE,
        "source_xlsx_sha256": hashlib.sha256(payload).hexdigest(),
        "source_relevant_rows": len(included) + len(excluded),
        "normalized_product_observations": len(records),
        "excluded_suspected_counterfeit_rows": len(excluded),
        "excluded_source_rows": excluded,
        "rows_by_source_product_kind": dict(sorted(Counter(row["product_kind_source_reported"] for row in records).items())),
        "families": dict(sorted(Counter(row["family_code"] for row in records).items())),
        "rows_with_api": sum(bool(row["technical"]["api_source_reported"]) for row in records),
        "rows_with_sae": sum(bool(row["technical"]["sae_source_reported"]) for row in records),
        "rows_with_dot": sum(bool(row["technical"]["brake_fluid_dot_source_reported"]) for row in records),
        "rows_with_hzy": sum(bool(row["technical"]["brake_fluid_hzy_source_reported"]) for row in records),
        "rows_with_coolant_class": sum(bool(row["technical"]["coolant_class_source_reported"]) for row in records),
        "rows_with_washer_class": sum(bool(row["technical"]["washer_fluid_class_source_reported"]) for row in records),
        "rows_with_aus32": sum(bool(row["technical"]["urea_class_source_reported"]) for row in records),
        "retest_confirmed_nonconforming_rows": sum(row["inspection_retest_confirmed_nonconforming"] for row in records),
        "normalized_output_sha256": hashlib.sha256(output_text.encode()).hexdigest(),
        "grain_note": "One row is one exact nominal producer + reported product name + model/specification identity observed as nonconforming in the 2023 national inspection.",
        "lifecycle_note": "These are historical regulatory observations, never approval, recommendation or evidence of current sale.",
        "counterfeit_note": "Ten attachment rows explicitly marked suspected counterfeit are audit-only. The narrative also reports one unidentified counterfeit washer-fluid sample; because the attachment does not identify it, all eight washer rows remain with an explicit uncertainty flag.",
        "provenance_note": "Complete SAMR national-report XLSX mirrored verbatim by the Shandan County government portal.",
        "privacy_note": "Retailers, testing laboratories and producer locations are excluded; only company-level nominal producer and product facts required for identity and regulatory provenance are retained.",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
