#!/usr/bin/env python3
"""Normalize Shenzhen's 2021 nonconforming automotive-fluid inspection rows."""

from __future__ import annotations

import hashlib
import html
import io
import json
import re
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data/shenzhen-2021-nonconforming-automotive-fluids.jsonl"
REPORT = ROOT / "data/shenzhen-2021-nonconforming-automotive-fluids-report.json"
SOURCE_ID = "SHENZHEN_CHINA_2021_NONCONFORMING_AUTOMOTIVE_FLUIDS"
SOURCE_URL = "https://www.sz.gov.cn/cn/xxgk/zfxxgj/scjg/cpzljy/content/mpost_9934070.html"
ATTACHMENT_URL = "http://www.sz.gov.cn/attachment/0/997/997590/9934070.xlsx"
EXPECTED_XLSX_SHA256 = "73a011b46ee94611f487c89791bf9cf01c7a2a978c59c495f46bd95dd7f414bd"
SNAPSHOT_DATE = "2026-07-21"
REPORT_DATE = "2022-07-06"


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def empty_marker(value: object) -> str:
    value = clean(value)
    return "" if value in {"/", "-", "—"} else value


def fetch() -> bytes:
    request = urllib.request.Request(ATTACHMENT_URL, headers={
        "User-Agent": "Mozilla/5.0 MFClassifierResearch/1.0",
        "Referer": SOURCE_URL,
    })
    with urllib.request.urlopen(request, timeout=120) as response:
        payload = response.read()
    digest = hashlib.sha256(payload).hexdigest()
    if digest != EXPECTED_XLSX_SHA256:
        raise RuntimeError(f"Shenzhen XLSX changed: expected {EXPECTED_XLSX_SHA256}, received {digest}")
    return payload


def manufacturer(value: object) -> str:
    value = empty_marker(value)
    value = re.sub(r"^出品[：:]\s*", "", value)
    value = re.sub(r"^广东[·•]\s*", "", value)
    return value


def classify(name: str, model: str) -> tuple[str, str, str]:
    joined = name + " " + model
    if "机油" in joined or "发动机油" in joined or "润滑油" in joined:
        return "M", "机动车发动机润滑油", "motor-vehicle engine lubricating oil"
    if "玻璃" in joined:
        return "TF", "汽车风窗玻璃清洗液", "automotive windshield washer fluid"
    if "冷却液" in joined or "防冻液" in joined:
        return "TF", "机动车发动机冷却液", "motor-vehicle engine coolant"
    return "S", "车用汽油清净剂", "automotive gasoline detergent additive"


def extract_engine_specs(value: str) -> tuple[list[str], list[str]]:
    upper = value.upper().replace("：", ":").replace("－", "-")
    api = re.findall(r"(?<![A-Z0-9])(CF-4|CH-4|CI-4|CJ-4|CK-4|SG|SJ|SL|SM|SN|SP|CF)(?![A-Z0-9])", upper)
    sae = [f"{winter}-{summer}" for winter, summer in re.findall(r"(?<![0-9])(0W|5W|10W|15W|20W|25W)\s*-?\s*([0-9]{2})(?![0-9])", upper)]
    return sorted(set(api)), sorted(set(sae))


def split_failures(value: str) -> list[str]:
    value = re.sub(r"(?:^|[；;])\s*\d+[、.]\s*", ";", value)
    return [clean(item) for item in value.strip(";").split(";") if clean(item)]


def main() -> None:
    payload = fetch()
    sheet = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)["Sheet1"]
    header = tuple(clean(value) for value in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    if header[:4] != ("序号", "产品类型", "受检单位名称", "样品名称"):
        raise RuntimeError(f"Unexpected Shenzhen columns: {header!r}")

    records = []
    for values in sheet.iter_rows(min_row=4, values_only=True):
        if not isinstance(values[0], (int, float)) or values[0] == 1:
            continue  # Exclude the single diesel-fuel row.
        row_no = int(values[0])
        product_name = clean(values[3])
        model = clean(values[5])
        producer = manufacturer(values[7])
        if not product_name or not model or not producer:
            raise RuntimeError(f"Incomplete Shenzhen relevant row {row_no}")
        family, kind_zh, kind_en = classify(product_name, model)
        api, sae = extract_engine_specs(model)
        coolant = []
        normalized_model = model.upper().replace("Ⅱ", "II").replace(" ", "")
        coolant.extend(re.findall(r"(?:LEC|HEC)-II-(?:15|25|35|40|45|50)", normalized_model))
        washer = []
        if "水基型" in model:
            washer = ["水基型低温型" if "低温型" in model else "水基型普通型"]
        source_facts = {
            "source_row": row_no, "product_name": product_name, "brand_source_reported": clean(values[4]),
            "model": model, "production_date": empty_marker(values[6]), "producer_source_reported": clean(values[7]),
            "outcome": clean(values[8]), "failures": clean(values[9]), "note": clean(values[10]),
        }
        records.append({
            "source_id": SOURCE_ID,
            "source_record_id": f"SZ-CN-2021-{row_no:03d}",
            "source_url": SOURCE_URL,
            "attachment_url": ATTACHMENT_URL,
            "rights_url": "https://www.sz.gov.cn/cn/qt/gywm/",
            "snapshot_date": SNAPSHOT_DATE,
            "report_date": REPORT_DATE,
            "market": "China / Shenzhen",
            "manufacturer": producer,
            "manufacturer_source_reported": clean(values[7]),
            "brand": producer,
            "brand_source_reported": clean(values[4]),
            "brand_basis": "nominal_producer_fallback_because_source_brand_is_only_generic_graphic_mark",
            "product_name": product_name,
            "product_name_basis": "source_reported_sample_name_from_municipal_quality_inspection",
            "model_specification_source_reported": model,
            "product_kind_source_reported": kind_zh,
            "product_kind_english": kind_en,
            "family_code": family,
            "technical": {
                "api_source_reported": api, "sae_source_reported": sae,
                "brake_fluid_dot_source_reported": [], "brake_fluid_hzy_source_reported": [],
                "coolant_class_source_reported": sorted(set(coolant)),
                "washer_fluid_class_source_reported": washer,
                "urea_class_source_reported": [],
            },
            "production_date_or_batch_source_reported": empty_marker(values[6]),
            "inspection_outcome": "nonconforming",
            "inspection_retest_confirmed_nonconforming": "复检仍不合格" in clean(values[10]),
            "nonconforming_items": split_failures(clean(values[9])),
            "nonconforming_items_source_reported": clean(values[9]),
            "source_note": empty_marker(values[10]),
            "lifecycle_status": "official_2021_shenzhen_inspection_nonconforming_current_market_status_unverified",
            "source_quality_flags": [
                "historical_regulatory_observation_not_current_catalog_listing",
                "nonconforming_product_do_not_treat_as_approved_or_recommended",
                "brand_falls_back_to_nominal_producer_due_generic_graphic_mark_only",
            ],
            "source_facts_sha256": hashlib.sha256(json.dumps(source_facts, ensure_ascii=False, sort_keys=True).encode()).hexdigest(),
            "evidence_status": "official_government_nonconforming_product_inspection_observation",
        })
    if len(records) != 12:
        raise RuntimeError(f"Expected 12 automotive-fluid rows after diesel exclusion, received {len(records)}")

    output_text = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records)
    OUTPUT.write_text(output_text, encoding="utf-8")
    report = {
        "status": "official_shenzhen_2021_nonconforming_automotive_fluid_observations_normalized",
        "source_id": SOURCE_ID, "source_url": SOURCE_URL, "attachment_url": ATTACHMENT_URL,
        "snapshot_date": SNAPSHOT_DATE, "report_date": REPORT_DATE,
        "source_xlsx_sha256": hashlib.sha256(payload).hexdigest(),
        "source_rows": 13, "excluded_diesel_fuel_rows": 1,
        "normalized_product_observations": len(records),
        "families": dict(sorted(Counter(row["family_code"] for row in records).items())),
        "rows_by_source_product_kind": dict(sorted(Counter(row["product_kind_source_reported"] for row in records).items())),
        "rows_with_api": sum(bool(row["technical"]["api_source_reported"]) for row in records),
        "rows_with_sae": sum(bool(row["technical"]["sae_source_reported"]) for row in records),
        "rows_with_coolant_class": sum(bool(row["technical"]["coolant_class_source_reported"]) for row in records),
        "rows_with_washer_class": sum(bool(row["technical"]["washer_fluid_class_source_reported"]) for row in records),
        "retest_confirmed_nonconforming_rows": sum(row["inspection_retest_confirmed_nonconforming"] for row in records),
        "normalized_output_sha256": hashlib.sha256(output_text.encode()).hexdigest(),
        "grain_note": "One row is one nominal producer + sample name + exact source model/specification identity observed as nonconforming.",
        "lifecycle_note": "Historical municipal inspection evidence only; never approval, recommendation or evidence of current sale.",
        "privacy_note": "Inspected retailers are excluded. Only producer-level and product facts required for identity and regulatory provenance are retained.",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
