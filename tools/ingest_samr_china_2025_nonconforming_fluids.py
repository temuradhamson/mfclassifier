#!/usr/bin/env python3
"""Normalize relevant 2025 China SAMR national-inspection product rows.

SAMR publishes an official XLSX of products found nonconforming in the 2025
national product-quality inspection.  This importer retains only engine oils,
motor-vehicle brake fluids and automotive aqueous urea solution.  A row is a
historical regulatory product observation, not evidence of current sale.
Counterfeit rows without a reported producer are counted in the audit report
but excluded from the canonical product dataset.
"""

from __future__ import annotations

import hashlib
import html
import io
import json
import re
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "samr-china-2025-nonconforming-fluids.jsonl"
REPORT = ROOT / "data" / "samr-china-2025-nonconforming-fluids-report.json"
SOURCE_ID = "SAMR_CHINA_2025_NONCONFORMING_FLUIDS"
SOURCE_URL = "https://www.samr.gov.cn/zw/zfxxgk/fdzdgknr/zljds/art/2026/art_2edd988c928c47ff92cdaa2941bbe013.html"
ATTACHMENT_URL = "https://www.samr.gov.cn/cms_files/filemanager/1647978232/attach/20261/5ced9b6ecebe48d1a6842dad0bbc6873.xlsx"
RIGHTS_URL = "https://www.samr.gov.cn/jg/wzsm/art/2021/art_c30de52ec3264bd29886479e5471dc72.html"
EXPECTED_XLSX_SHA256 = "0042e67b52973d78660f2c137f912cac2ceb785ff56132467f5eb24248c0f801"
SNAPSHOT_DATE = "2026-07-21"
REPORT_DATE = "2026-01-13"
USER_AGENT = "Mozilla/5.0 MFClassifierResearch/1.0 (public-government-inspection-data)"

SCOPE = {
    "机动车辆制动液": ("TF", "motor-vehicle brake fluid"),
    "车用尿素水溶液": ("TF", "automotive aqueous urea solution"),
    "发动机润滑油": ("M", "engine lubricating oil"),
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def empty_marker(value: object) -> str:
    value = clean(value)
    return "" if value in {"/", "-", "—"} else value


def fetch() -> bytes:
    request = urllib.request.Request(
        ATTACHMENT_URL,
        headers={"User-Agent": USER_AGENT, "Referer": SOURCE_URL},
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        payload = response.read()
    digest = hashlib.sha256(payload).hexdigest()
    if digest != EXPECTED_XLSX_SHA256:
        raise RuntimeError(f"SAMR XLSX changed: expected {EXPECTED_XLSX_SHA256}, received {digest}")
    return payload


def split_failures(value: str) -> list[str]:
    return [clean(item).rstrip("*") for item in re.split(r"[,，]", value) if clean(item)]


def extract_engine_specs(value: str) -> tuple[list[str], list[str]]:
    upper = value.upper().replace("／", "/")
    api = re.findall(r"(?<![A-Z0-9])(SJ|SL|CF-4|CH-4|CI-4)(?![A-Z0-9])", upper)
    sae = []
    for winter, summer in re.findall(r"(?<![0-9])(0W|5W|10W|15W|20W|25W)[-/]([0-9]{2})(?![0-9])", upper):
        sae.append(f"{winter}-{summer}")
    return sorted(set(api)), sorted(set(sae))


def extract_brake_classes(value: str) -> tuple[list[str], list[str]]:
    upper = value.upper()
    dot = [f"DOT {item}" for item in re.findall(r"DOT\s*([345])", upper)]
    hzy = [f"HZY{item}" for item in re.findall(r"HZY\s*([345])", upper)]
    return sorted(set(dot)), sorted(set(hzy))


def rows(payload: bytes) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    header = tuple(clean(value) for value in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[:12])
    expected = (
        "序号", "产品种类", "受检单位", "标称生产单位", "标称生产 单位所在地",
        "产品名称", "规格型号", "生产日期/批号", "主要不合格项目", "承检机构", "电商平台", "备注",
    )
    if header != expected:
        raise RuntimeError(f"Unexpected SAMR columns: {header!r}")

    included = []
    excluded = []
    for values in sheet.iter_rows(min_row=4, values_only=True):
        if not isinstance(values[0], (int, float)):
            continue
        source_row = int(values[0])
        product_kind = clean(values[1])
        if product_kind not in SCOPE:
            continue
        producer = empty_marker(values[3])
        product_name = clean(values[5])
        model = clean(values[6])
        note = clean(values[11])
        raw = {
            "source_row": source_row,
            "product_kind_source_reported": product_kind,
            "producer": producer,
            "product_name_source_reported": product_name,
            "model_specification_source_reported": model,
            "production_date_or_batch_source_reported": empty_marker(values[7]),
            "nonconforming_items_source_reported": clean(values[8]),
            "source_note": note,
        }
        if note == "假冒" and not producer:
            excluded.append({**raw, "exclusion_reason": "counterfeit_row_without_reported_producer"})
        else:
            if not producer or not product_name or not model:
                raise RuntimeError(f"Incomplete relevant SAMR row: {raw!r}")
            included.append(raw)
    return included, excluded


def main() -> None:
    payload = fetch()
    included, excluded = rows(payload)
    if len(included) != 25 or len(excluded) != 3:
        raise RuntimeError(f"Expected 25 included and 3 counterfeit exclusions, received {len(included)} and {len(excluded)}")

    records = []
    identities: set[tuple[str, str, str, str]] = set()
    for row in included:
        family_code, product_kind_en = SCOPE[row["product_kind_source_reported"]]
        api, sae = extract_engine_specs(row["model_specification_source_reported"])
        dot, hzy = extract_brake_classes(row["model_specification_source_reported"])
        identity = (
            row["producer"], row["product_name_source_reported"],
            row["model_specification_source_reported"], family_code,
        )
        if identity in identities:
            raise RuntimeError(f"Duplicate SAMR product identity requires occurrence merging: {identity!r}")
        identities.add(identity)
        source_facts = {key: row[key] for key in sorted(row)}
        records.append({
            "source_id": SOURCE_ID,
            "source_record_id": f"SAMR-CN-2025-{row['source_row']:03d}",
            "source_url": SOURCE_URL,
            "attachment_url": ATTACHMENT_URL,
            "rights_url": RIGHTS_URL,
            "snapshot_date": SNAPSHOT_DATE,
            "report_date": REPORT_DATE,
            "market": "China",
            "manufacturer": row["producer"],
            "brand": row["producer"],
            "brand_basis": "source_reported_nominal_producer_fallback_no_separate_brand_column",
            "product_name": row["product_name_source_reported"],
            "product_name_basis": "source_reported_product_name_from_national_quality_inspection",
            "model_specification_source_reported": row["model_specification_source_reported"],
            "product_kind_source_reported": row["product_kind_source_reported"],
            "product_kind_english": product_kind_en,
            "family_code": family_code,
            "technical": {
                "api_source_reported": api,
                "sae_source_reported": sae,
                "brake_fluid_dot_source_reported": dot,
                "brake_fluid_hzy_source_reported": hzy,
            },
            "production_date_or_batch_source_reported": row["production_date_or_batch_source_reported"],
            "inspection_outcome": "nonconforming",
            "nonconforming_items": split_failures(row["nonconforming_items_source_reported"]),
            "nonconforming_items_source_reported": row["nonconforming_items_source_reported"],
            "source_note": row["source_note"],
            "lifecycle_status": "official_2025_national_inspection_nonconforming_current_market_status_unverified",
            "source_quality_flags": [
                "historical_regulatory_observation_not_current_catalog_listing",
                "nonconforming_product_do_not_treat_as_approved_or_recommended",
                "brand_falls_back_to_nominal_producer_because_source_has_no_brand_column",
            ],
            "source_facts_sha256": hashlib.sha256(
                json.dumps(source_facts, ensure_ascii=False, sort_keys=True).encode("utf-8")
            ).hexdigest(),
            "evidence_status": "official_government_nonconforming_product_inspection_observation",
        })

    records.sort(key=lambda row: row["source_record_id"])
    output_text = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records)
    OUTPUT.write_text(output_text, encoding="utf-8")
    report = {
        "status": "official_china_2025_nonconforming_fluid_product_observations_normalized",
        "source_id": SOURCE_ID,
        "source_url": SOURCE_URL,
        "attachment_url": ATTACHMENT_URL,
        "rights_url": RIGHTS_URL,
        "snapshot_date": SNAPSHOT_DATE,
        "report_date": REPORT_DATE,
        "source_xlsx_sha256": hashlib.sha256(payload).hexdigest(),
        "source_relevant_rows": len(included) + len(excluded),
        "normalized_product_observations": len(records),
        "excluded_counterfeit_rows_without_producer": len(excluded),
        "excluded_source_rows": excluded,
        "rows_by_source_product_kind": dict(sorted(Counter(row["product_kind_source_reported"] for row in records).items())),
        "families": dict(sorted(Counter(row["family_code"] for row in records).items())),
        "rows_with_api": sum(bool(row["technical"]["api_source_reported"]) for row in records),
        "rows_with_sae": sum(bool(row["technical"]["sae_source_reported"]) for row in records),
        "rows_with_dot": sum(bool(row["technical"]["brake_fluid_dot_source_reported"]) for row in records),
        "rows_with_hzy": sum(bool(row["technical"]["brake_fluid_hzy_source_reported"]) for row in records),
        "normalized_output_sha256": hashlib.sha256(output_text.encode("utf-8")).hexdigest(),
        "grain_note": "One row is one exact nominal producer + reported product name + model/specification identity observed as nonconforming in the 2025 national inspection.",
        "lifecycle_note": "All retained rows failed one or more inspection items. They are historical regulatory observations and are never presented as approved, recommended or currently marketed products.",
        "counterfeit_note": "Three counterfeit rows without a reported producer are retained only in this audit report and do not inflate the canonical product total.",
        "rights_note": "SAMR reserves website copyright and prohibits commercial verbatim republication. The source XLSX is not redistributed; only attributed, non-expressive factual product and inspection fields are normalized.",
        "privacy_note": "Retailers, platform names, testing laboratories and producer locations are excluded. Only company-level nominal producer and product facts needed for product identity and regulatory provenance are retained.",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
