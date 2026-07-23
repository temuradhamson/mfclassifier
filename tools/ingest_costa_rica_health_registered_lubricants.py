#!/usr/bin/env python3
"""Normalize lubricant-scope registrations from Costa Rica Ministry of Health."""

from __future__ import annotations

import hashlib
import io
import json
import re
import urllib.request
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "costa-rica-health-registered-lubricants.jsonl"
REPORT = ROOT / "data" / "costa-rica-health-registered-lubricants-report.json"
SOURCE_ID = "COSTA_RICA_HEALTH_HISTORICAL_CHEMICAL_PRODUCT_REGISTRATIONS"
LANDING_URL = "https://www.ministeriodesalud.go.cr/index.php/biblioteca-de-archivos-left/documentos-ministerio-de-salud/tramites/registros/productos-registrados"
SNAPSHOT_DATE = "2026-07-23"
SOURCES = (
    {
        "key": "after_2007",
        "url": "https://web.archive.org/web/20220808151132id_/https://www.ministeriodesalud.go.cr/index.php/biblioteca-de-archivos-left/documentos-ministerio-de-salud/tramites/registros/productos-registrados/2574-quimicos-peligrosos-despues-del-2007/file",
        "official_url": "https://www.ministeriodesalud.go.cr/index.php/biblioteca-de-archivos-left/documentos-ministerio-de-salud/tramites/registros/productos-registrados/2574-quimicos-peligrosos-despues-del-2007/file",
        "sha256": "ab5d966547dc81c6d485090411bf6ee7156ccb4414565a4305f880246f3c758e",
        "rows": 62532,
    },
    {
        "key": "before_2007",
        "url": "https://web.archive.org/web/20220808151129id_/https://www.ministeriodesalud.go.cr/index.php/biblioteca-de-archivos-left/documentos-ministerio-de-salud/tramites/registros/productos-registrados/2580-quimicos-peligrosos-antes-del-2007/file",
        "official_url": "https://www.ministeriodesalud.go.cr/index.php/biblioteca-de-archivos-left/documentos-ministerio-de-salud/tramites/registros/productos-registrados/2580-quimicos-peligrosos-antes-del-2007/file",
        "sha256": "5cc6ca26fa8baaebc416e324ac91aef146cb7837d0d25d0df4ec3385d74f7b6f",
        "rows": 43840,
    },
)


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def upper(value: object) -> str:
    return clean(value).upper()


def fetch(source: dict) -> bytes:
    request = urllib.request.Request(source["url"], headers={"User-Agent": "MFClassifierResearch/1.0 (official-open-data-archive)"})
    with urllib.request.urlopen(request, timeout=300) as response:
        payload = response.read()
    digest = hashlib.sha256(payload).hexdigest()
    if digest != source["sha256"]:
        raise RuntimeError(f"{source['key']} workbook changed: {digest}")
    return payload


def included(row: dict) -> bool:
    generic = upper(row["NOM_GENERICO"])
    commercial = upper(row["NOM_COMERCIAL"])
    combined = f"{generic} {commercial}"
    excluded = (
        "ACEITE ESENCIAL", "PINTURA", "COSMET", "ACEITE DE LINAZA",
        "ACEITE DE RICINO", "ACEITE DE PINO", "ACEITE VEGETAL",
        "ACEITE AROMATIC", "REMOVEDOR DE GRASA", "DESENGRAS",
        "LIMPIADOR DE FRENOS", "GAS REFRIGERANTE",
        "HIDROCARBURO REFRIGERANTE", "REFRIGERANTE R-",
        "REFRIGERANTE 404", "REFRIGERANTE 407", "REFRIGERANTE 410",
        "REFRIGERANTE 507", "TETRAFLUOR", "ISOBUTANO",
    )
    if any(token in combined for token in excluded):
        return False
    if any(token in generic for token in ("LUBRIC", "LIBRIC", "LUBRICNATE", "PARA LUBRICAR")):
        return True
    if any(token in generic for token in (
        "LIQUIDO DE FRENOS", "LÍQUIDO DE FRENOS", "LIQUIDO PARA FRENOS",
        "LÍQUIDO PARA FRENOS", "FLUIDO HIDRAUL", "FLUIDO HIDRÁUL",
        "LIQUIDO HIDRAUL", "LÍQUIDO HIDRÁUL", "FLUIDO DE TRANSM",
        "FLUIDO PARA TRANSM", "LIQUIDO DE TRANSM", "LÍQUIDO DE TRANSM",
        "ANTICONGELANTE", "COOLANT", "REFRIGERANTE PARA RADIADOR",
        "REFRIGERANTE DE MOTOR", "ANTICONGELANTE Y REFRIGERANTE",
    )):
        return True
    if "GRASA" in generic and any(token in combined for token in (
        "LITIO", "SILICON", "SINTET", "PETROLE", "RODAM", "ENGRANA",
        "MULTIPROP", "AUTOMOT", "INDUSTR", "MAQUIN", "CHASIS",
        "COJIN", "GRADO ALIMENTICIO", "MARINA", "DIELECTR",
        "ALTA TEMPERATURA", "EP ",
    )):
        return True
    if "ACEITE" in generic and any(token in combined for token in (
        "MOTOR", "TRANSM", "HIDRAUL", "HIDRÁUL", "ENGRANA", "COMPRES",
        "TURBINA", "DIELECTR", "DIELÉCTR", "CORTE", "INDUSTR",
        "AUTOMOT", "CAJA DE CAMBIO", "DIFERENCIAL", "CADENA", "MAQUIN",
        "REFRIGERACION", "REFRIGERACIÓN", "PENETRANTE", "PROCESO",
        "AISLANTE", "CIGÜEÑAL", "NAFTEN", "NAFTÉN", "ATF", "GEAR",
        "HYDRAULIC", "ENGINE OIL", "MOTOR OIL",
    )):
        return True
    if generic in (
        "REFRIGERANTE", "REFRIGERANTE.", "REFRIGERANTE.-",
        "REFRIGERANTE.--", "LIQUIDO REFRIGERANTE", "LÍQUIDO REFRIGERANTE",
    ) and any(token in commercial for token in ("COOLANT", "ANTIFREEZE", "ANTICONGEL", "RADIADOR", "RADIATOR", "MOTOR")):
        return True
    return False


def family(generic: str, commercial: str) -> str:
    value = upper(f"{generic} {commercial}")
    if "FRENO" in value:
        return "TF"
    if any(token in value for token in ("ANTICONGEL", "COOLANT", "RADIADOR", "RADIATOR", "REFRIGERANTE DE MOTOR")):
        return "C"
    if "GRASA" in upper(generic):
        return "G"
    if any(token in value for token in ("TRANSM", "ENGRANA", "DIFERENCIAL", "CAJA DE CAMBIO", "GEAR OIL", " ATF")):
        return "T"
    if any(token in value for token in ("HIDRAUL", "HIDRÁUL", "HYDRAULIC")):
        return "H"
    if any(token in value for token in ("MOTOR", "ENGINE OIL", "MARINO", "2T", "4T")):
        return "M"
    if any(token in value for token in ("COMPRES", "TURBINA", "DIELECTR", "DIELÉCTR", "CORTE", "INDUSTR", "MAQUIN", "PROCESO")):
        return "I"
    return "S"


def technical(text: str) -> dict:
    value = upper(text).replace("OW-", "0W-")
    sae = []
    for item in re.findall(r"(?<![A-Z0-9])(?:0W|5W|10W|15W|20W|25W|70W|75W|80W|85W)[- ]?\d{2,3}(?![A-Z0-9])", value):
        item = item.replace(" ", "")
        item = re.sub(r"^(\d+W)(\d+)$", r"\1-\2", item)
        if item not in sae:
            sae.append(item)
    api = sorted(set(re.findall(r"\bAPI\s+(SP|SN(?:\s+PLUS)?|SM|SL|SJ|SH|SG|SF|CK-4|CJ-4|CI-4|CH-4|CF-4|CF|CE|CD)\b", value)))
    api_gl = [f"GL-{item}" for item in sorted(set(re.findall(r"\bGL[- ]?([1-6])\b", value)))]
    iso_vg = sorted(set(re.findall(r"\bISO(?:\s+VG)?\s*[-:]?\s*(22|32|46|68|100|150|220|320|460|680)\b", value)), key=int)
    dot = [f"DOT {item}" for item in sorted(set(re.findall(r"\bDOT\s*([345](?:\.1)?)\b", value)))]
    nlgi = sorted(set(re.findall(r"\bNLGI\s*([0-6])\b", value)))
    return {"sae": sae, "api": api, "api_gl": api_gl, "iso_vg": iso_vg, "dot": dot, "nlgi": nlgi}


def date_value(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return clean(value)


def main() -> None:
    records = []
    audited = {}
    retained = {}
    for source in SOURCES:
        workbook = openpyxl.load_workbook(io.BytesIO(fetch(source)), read_only=True, data_only=True)
        sheet = workbook.active
        header = [clean(value) for value in next(sheet.iter_rows(values_only=True))]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) != source["rows"]:
            raise RuntimeError(f"{source['key']} row count changed: {len(rows)}")
        audited[source["key"]] = len(rows)
        selected = 0
        for source_row, values in enumerate(rows, 2):
            row = dict(zip(header, values))
            if not included(row):
                continue
            selected += 1
            commercial = clean(row["NOM_COMERCIAL"])
            generic = clean(row["NOM_GENERICO"])
            manufacturer = clean(row["FABRICANTE"])
            registration = clean(row["CONSECUTIVO"])
            tech = technical(f"{commercial} {generic}")
            facts = {
                "registration": registration, "date": date_value(row["FECHA_REGISTRO"]),
                "commercial": commercial, "generic": generic, "manufacturer": manufacturer,
                "country": clean(row["NOM_PAIS"]), "class": clean(row["CLASE"]),
            }
            records.append({
                "source_id": SOURCE_ID,
                "source_record_id": f"CR-MS-{source['key']}-{source_row}",
                "source_url": source["official_url"],
                "source_archive_url": source["url"],
                "source_dataset_part": source["key"],
                "source_row": source_row,
                "registration_number": registration,
                "registration_date": date_value(row["FECHA_REGISTRO"]),
                "source_indicator": clean(row["INDICADOR"]),
                "source_precursor": clean(row["PRECURSOR"]),
                "source_country": clean(row["NOM_PAIS"]),
                "source_other_countries": clean(row["OTROS_PAISES"]),
                "source_class": clean(row["CLASE"]),
                "source_generic_name": generic,
                "manufacturer": manufacturer,
                "registration_holder": clean(row["REGISTRANTE"]),
                "brand": f"Manufacturer fallback: {manufacturer}" if manufacturer else "Brand not stated (Costa Rica source)",
                "product_name": commercial,
                "family_code": family(generic, commercial),
                "technical": tech,
                "market": "Costa Rica",
                "dataset_snapshot_date": SNAPSHOT_DATE,
                "lifecycle_status": "historical_registration_observation_current_status_unverified",
                "evidence_status": "official_government_historical_chemical_product_registration",
                "source_facts_sha256": hashlib.sha256(json.dumps(facts, ensure_ascii=False, sort_keys=True).encode()).hexdigest(),
                "source_quality_flags": [
                    "registration_proves_historical_regulatory_identity_not_current_market_availability",
                    "manufacturer_used_as_explicit_brand_fallback_not_asserted_trademark",
                    "source_personal_representative_fields_excluded",
                    "archived_official_workbook_snapshot_due_current_site_access_denial",
                ],
            })
        retained[source["key"]] = selected
    OUTPUT.write_text("".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in records), encoding="utf-8")
    report = {
        "status": "official_costa_rica_historical_chemical_registrations_normalized",
        "landing_url": LANDING_URL,
        "source_workbook_sha256": {source["key"]: source["sha256"] for source in SOURCES},
        "audited_source_rows": audited,
        "audited_total_rows": sum(audited.values()),
        "retained_source_rows": retained,
        "normalized_products": len(records),
        "families": dict(sorted(Counter(row["family_code"] for row in records).items())),
        "rows_with_sae": sum(bool(row["technical"]["sae"]) for row in records),
        "rows_with_api": sum(bool(row["technical"]["api"]) for row in records),
        "rows_with_api_gl": sum(bool(row["technical"]["api_gl"]) for row in records),
        "rows_with_iso_vg": sum(bool(row["technical"]["iso_vg"]) for row in records),
        "rows_with_dot": sum(bool(row["technical"]["dot"]) for row in records),
        "rows_with_nlgi": sum(bool(row["technical"]["nlgi"]) for row in records),
        "normalized_output_sha256": hashlib.sha256(OUTPUT.read_bytes()).hexdigest(),
        "method": "all 106,372 rows in two official archived workbooks audited; strict Spanish generic-name inclusion with explicit cosmetic/food/paint/cleaner/HVAC-gas exclusions",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
