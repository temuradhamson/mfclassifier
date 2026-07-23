#!/usr/bin/env python3
"""Build the current Venezuelan PDV lubricant layer from official downloads.

The current distributor page links a digital catalogue and five evidence
bundles.  The CPE workbook is the product/package denominator: 39 package
certificates collapse to 23 product-grade identities.  Linked technical sheets
enrich only compatible identities; filename/body and grade conflicts remain
explicit quality flags.
"""

from __future__ import annotations

import hashlib
import html
import io
import json
import re
import urllib.request
import zipfile
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data/venezuela-pdv-current-lubricants.jsonl"
REPORT = ROOT / "data/venezuela-pdv-current-lubricants-report.json"
SOURCE_ID = "VENEZUELA_PDV_CURRENT_CPE_LUBRICANT_CATALOG"
HOME = "https://lubricantespdv.com/"
DOWNLOAD_BASE = "https://t6g.cb7.myftpupload.com/wp-content/documentos/"
SNAPSHOT_DATE = "2026-07-23"
UA = "MFClassifier evidence catalog/1.0"
BUNDLES = {
    "catalog": "CATALOGO.pdf",
    "cpe": "CPE.zip",
    "technical": "FICHA_TECNICA.zip",
    "sencamer": "SENCAMER.zip",
    "quality": "cCALIDAD.zip",
    "safety": "cSEGURIDAD.zip",
}
EXPECTED_BUNDLE_SHA256 = {
    "catalog": "dabc454acaba307dbafbdc39d7bbd7e0fbe96f946bb24eccfb171b18fb8202bb",
    "cpe": "cb476ea2deb5a42d3547c02edaae4f8815552e2c6244c023d203a1f93e4e2bb9",
    "technical": "958881705b218a13f08f133b4ace1ed1b04adab16e7a343a39c85a2eb3c7c639",
    "sencamer": "6b7cc4ddbf946f8989ed8dfbe7ecc5898d130517a6a16dbf6c7d01b57500048b",
    "quality": "e145dcafeb49461c3ec8befe1d27e52a10e21093beb2727d597d2342d9378278",
    "safety": "225884a222945a3a3ade30f9fd7d16028b8b5c74c1b027078a22b7977132d3d7",
}


def P(cpe_name, name, family, *, sae_engine="", sae_gear="", iso="", source_grade="",
      api=(), api_gl=(), acea=(), ilsac=(), performance=(), tds="", flags=()):
    return {
        "cpe_name": cpe_name, "name": name, "family": family,
        "sae_engine": sae_engine, "sae_gear": sae_gear, "iso": iso,
        "source_grade": source_grade, "api": api, "api_gl": api_gl,
        "acea": acea, "ilsac": ilsac, "performance": performance,
        "tds": tds, "flags": flags,
    }


PRODUCTS = [
    P("ULTRA DIESEL CJ-4 SN 15W40", "PDV Ultra Diesel CJ-4/SN 15W-40", "M",
      sae_engine="15W-40", api=("CJ-4", "SN"),
      tds="ULTRADIESEL 15W40 CJ-4.pdf",
      flags=("current_catalog_and_cpe_support_cj_4_sn",
             "linked_cj_4_tds_payload_is_identical_to_ci_4_tds_and_describes_ci_4_sl_not_used_for_cj_4_claim")),
    P("PDV TRANSLUB EP GL-5 SAE 85W140", "PDV Translub EP GL-5 85W-140", "T",
      sae_gear="85W-140", api_gl=("GL-5", "MT-1"),
      performance=("Mack GO-J", "MIL-PRF-2105E", "ZF TE-ML 07A", "ZF TE-ML 08", "ZF TE-ML 24A"),
      tds="TRANSLUB EP GL-5.pdf"),
    P("PDV TRANSLUB EP GL-5 SAE 80W90", "PDV Translub EP GL-5 80W-90", "T",
      sae_gear="80W-90", api_gl=("GL-5", "MT-1"),
      performance=("Mack GO-J", "MIL-PRF-2105E", "ZF TE-ML 07A", "ZF TE-ML 08", "ZF TE-ML 24A"),
      tds="TRANSLUB EP GL-5.pdf"),
    P("PDV TRANSFLUIDO DIII", "PDV Transfluido D III", "TF",
      performance=("GM Dexron III G", "Ford Mercon", "Allison C-4", "Caterpillar TO-2"),
      tds="TRANSFLUIDO D III.pdf"),
    P("PDV SUPRA SEMI SINTETICO MX SAE SP 20W-50", "PDV Supra MX Semi-Synthetic SP 20W-50", "M",
      sae_engine="20W-50", api=("SP",), tds="SUPRA SP MX.pdf"),
    P("PDV SUPRA SEMI SINTETICO MX SAE SP 15W-40", "PDV Supra MX Semi-Synthetic SP 15W-40", "M",
      sae_engine="15W-40", api=("SP",), tds="SUPRA SP MX.pdf"),
    P("PDV SUPRA SEMI SINTETICO MX SAE SP 10W-30", "PDV Supra MX Semi-Synthetic SP 10W-30", "M",
      sae_engine="10W-30", api=("SP",), tds="SUPRA SP MX.pdf"),
    P("PDV SUPRA PREMIUM SAE SN 20W-50", "PDV Supra Premium SN 20W-50", "M",
      sae_engine="20W-50", api=("SN",), tds="SUPRA PREMIUM-API SN 2.pdf"),
    P("PDV SUPRA PREMIUM SAE SN 15W-40", "PDV Supra Premium SN 15W-40", "M",
      sae_engine="15W-40", api=("SN",), tds="SUPRA PREMIUM-API SN 2.pdf"),
    P("PDV SUPRA PREMIUM SAE SL 15W-40", "PDV Supra Premium SL 15W-40", "M",
      sae_engine="15W-40", api=("SL",), tds="SUPRA PREMIUM-API SL.pdf"),
    P("SAE 50 MAXIDIESEL PLUS", "PDV Maxidiesel Plus SAE 50", "M",
      sae_engine="50", api=("CF",), tds="MAXIDIESEL PLUS CF .pdf"),
    P("SAE 15W40 ULTRADIESEL CI-4/SL", "PDV Ultradiesel CI-4/SL 15W-40", "M",
      sae_engine="15W-40", api=("CI-4", "CH-4", "CG-4", "CF-4", "CF", "SL"),
      acea=("E7-12", "E7-16", "A3/B4-12"),
      performance=("MAN M 3275", "Cummins CES 20071", "Cummins CES 20072",
                   "Cummins CES 20076", "Cummins CES 20077", "Cummins CES 20078",
                   "MTU Type 2", "CAT ECF-2", "CAT ECF-1-a", "Volvo VDS-3",
                   "Mack EO-M Plus", "Detroit Diesel DDC 93K215", "MB 228.3",
                   "MB 229.1", "Renault RLD", "Renault RLD-2", "Global DHD-1"),
      tds="ULTRADIESEL 15W40 CI-4.pdf"),
    P("MOTOR GAS W40", "PDV MotorGas W SAE 40", "M",
      sae_engine="40", api=("CF",), tds="MotorGas W.pdf",
      flags=("technical_sheet_is_image_only_manually_audited",)),
    P("PDV MOTO PREMIUM 4T SAE 20W50", "PDV Moto Premium 4T 20W-50", "M",
      sae_engine="20W-50", api=("SL",),
      performance=("JASO MA", "JASO MA2", "JASO MB with booster (as printed)"),
      tds="MOTO PREMIUM 4T.pdf"),
    P("MAXITREN EMD 40", "PDV Maxitren EMD SAE 40", "M",
      sae_engine="40", api=("CF", "CD"),
      performance=("LMOA Generation IV Long Life", "GE Generation IV Long Life"),
      tds="MAXITREN EMD.pdf"),
    P("MAXI DIESEL CF-4 CF 20W50", "PDV Maxidiesel Plus CF-4/CF 20W-50", "M",
      sae_engine="20W-50", api=("CF-4", "CF", "CF-2", "SF"),
      acea=("E2-96 Issue 5 (2007)",),
      performance=("MB 228.1", "MB 228.0", "MAN 270", "MAN 271"),
      tds="MAXIDIESEL PLUS CF -4.pdf"),
    P("ISO 68 HIDRALUB", "PDV Hidralub ISO VG 68", "H",
      iso="68", performance=("Denison HF-1", "Denison HF-2", "Cincinnati P-69", "Vickers I-286-S"),
      tds="HIDRALUB.pdf"),
    P("HIDRALUB AW68", "PDV Hidralub AW ISO VG 68", "H",
      iso="68", performance=("DIN 51524 Parts 1/2/3", "ISO 11158 HL/HM/HV",
                             "Parker Denison HF-0/HF-1/HF-2", "Bosch Rexroth RDE-90235",
                             "ZF TE-ML 07H", "ZF TE-ML 21M"),
      tds="HIDRALUB AW.pdf"),
    P("PDV FUERA DE BORDA TC-W3", "PDV Fuera de Borda TC-W3", "M",
      performance=("NMMA TC-W3",), tds="FUERA DE BORDA TC-W3.pdf"),
    P("PDV DOS TIEMPOS", "PDV Dos Tiempos", "M",
      api=("TC",), performance=("JASO FB", "ISO-L-EGB"), tds="DOS TIEMPOS.pdf"),
    P("DIESEL MAR 40", "PDV Dieselmar MR 40", "M",
      source_grade="40", api=("CF",), tds="Dieselmar MR.pdf",
      flags=("current_cpe_grade_40_not_silently_mapped_to_historical_tds_proprietary_3030_4020_4030_4040",
             "technical_sheet_is_image_only_manually_audited")),
    P("CRUMAR SAE 30", "PDV Crumar SAE 30", "M",
      sae_engine="30", tds="CRUMAR.pdf",
      flags=("technical_sheet_is_image_only_manually_audited",)),
    P("AGROFLUIDO 100", "PDV Agrofluido 100", "TF",
      source_grade="100", api_gl=("GL-4",),
      performance=("John Deere J20C", "John Deere J20D", "Massey Ferguson M1145",
                   "Ford ESN-M2C134-D", "Allison C-3"),
      tds="AGROFLUIDO.pdf",
      flags=("current_cpe_and_catalog_grade_100_conflicts_with_linked_tds_sae_10w30",
             "tds_grade_not_assigned_to_current_cpe_identity")),
]

EXPECTED_TDS_SHA256 = {
    "AGROFLUIDO.pdf": "1daaed47adb41b6c2e46d0db472c3eb4748fdd0ea6d32280be06c30325a84adc",
    "CRUMAR.pdf": "767cbbbc6d9fef7ec9a48a8964cf8859e545ceef27a0c0e7e6933aa464ff4162",
    "DOS TIEMPOS.pdf": "70662446af7d1256570a8f4b79e75576a9a668b4461dfeb50a8ae6a507a1ce60",
    "Dieselmar MR.pdf": "60d63b187f8948e4651c60f6e00e9b00b8d092b13877413f6dfb75a26432e856",
    "FUERA DE BORDA TC-W3.pdf": "cd3544c60a4148d37d47e477bf6f9495cae65bcaa1c7830d5d0add10e3bb1db3",
    "HIDRALUB AW.pdf": "d4af3832c99ca1e6dcca9a187f2eaf7899f6ad3403e63ae5339cc79d4cd71732",
    "HIDRALUB.pdf": "0ad2b15e80d64bb04563bcf4bbc2e505a890bff1db430fcf017d0208980da8cb",
    "MAXIDIESEL PLUS CF -4.pdf": "c2d56d4ad54d01adc8b45113856f38c3f254686f865382d889441cf9e3662b8e",
    "MAXIDIESEL PLUS CF .pdf": "5170ed1b21d6a0d4b1017b2902af48674253bd79da312e7c2592098d138c60e1",
    "MAXITREN EMD.pdf": "9227c89a8eea2662c5f49cea737246018c5b6321f34bc9eb1bd736555de857b4",
    "MOTO PREMIUM 4T.pdf": "af1dfb3d3f85a73ee5fe8c9a55a54b39b6ccede7fb3f3e94d3ab6c6b21bfee89",
    "MotorGas W.pdf": "72c6d594d7bcc284c0d8724c0d95413d87fc8565a040970c82f35f9c9430944b",
    "SUPRA PREMIUM-API SL.pdf": "b2783f2f08ae3f32bb6bd7fd6fafa952ed1a42701b00c4e3a84e885c6dec11f3",
    "SUPRA PREMIUM-API SN 2.pdf": "2401ee090b0d7b485525a7b9078ca1bdef4b9593d3c4a5c8984f4138433a91c6",
    "SUPRA SP MX.pdf": "47c736bbeb591dc5af87d854eae423a62c7a68a708fbc1d2ac7bd9759c8d272d",
    "TRANSFLUIDO D III.pdf": "fa2988d799987cb2f840c2f8d8e995e33ea03367591431aa8d0450247fe43e35",
    "TRANSLUB EP GL-5.pdf": "035a70394b5f5a3c0ff3a0fedb293c7c5fe223578b64338012cb295a5688d9b4",
    "ULTRADIESEL 15W40 CI-4.pdf": "4adaf4515ef34daf3b80a3a39948315d1c87a980044923d0631cb2577e05cc45",
    "ULTRADIESEL 15W40 CJ-4.pdf": "4adaf4515ef34daf3b80a3a39948315d1c87a980044923d0631cb2577e05cc45",
}


class TextParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.skip = 0
        self.parts = []

    def handle_starttag(self, tag, attrs):
        if tag in {"script", "style", "noscript"}:
            self.skip += 1

    def handle_endtag(self, tag):
        if tag in {"script", "style", "noscript"} and self.skip:
            self.skip -= 1

    def handle_data(self, data):
        if not self.skip and data.strip():
            self.parts.append(data.strip())


def get(url):
    request = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(request, timeout=180) as response:
        return response.read()


def sha(value):
    return hashlib.sha256(value).hexdigest()


def normalized_page_text(value):
    parser = TextParser()
    parser.feed(value.decode("utf-8", "replace"))
    return re.sub(r"\s+", " ", html.unescape(" ".join(parser.parts))).strip()


def real_members(payload):
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        return {
            item.filename: archive.read(item)
            for item in archive.infolist()
            if not item.is_dir() and not item.filename.startswith("__MACOSX")
            and "/._" not in item.filename
        }


def main():
    page = get(HOME)
    page_text = normalized_page_text(page)
    if not all(filename in page.decode("utf-8", "replace") for filename in BUNDLES.values()):
        raise RuntimeError("PDV official download set changed")

    def fetch(item):
        label, filename = item
        payload = get(DOWNLOAD_BASE + filename)
        if sha(payload) != EXPECTED_BUNDLE_SHA256[label]:
            raise RuntimeError(f"PDV bundle payload changed: {label}")
        return label, payload

    with ThreadPoolExecutor(max_workers=6) as pool:
        bundles = dict(pool.map(fetch, BUNDLES.items()))

    cpe_members = real_members(bundles["cpe"])
    workbook_name = next(name for name in cpe_members if name.endswith("Listado de CPE para PDV.xlsx"))
    sheet = load_workbook(io.BytesIO(cpe_members[workbook_name]), data_only=True, read_only=True).active
    cpe_rows = [
        {"description": str(description).strip(), "litres": float(litres), "cpe": str(cpe).strip()}
        for description, litres, cpe in sheet.iter_rows(min_row=2, values_only=True)
        if description and str(cpe or "").startswith("CPE")
    ]
    if len(cpe_rows) != 39 or len({row["cpe"] for row in cpe_rows}) != 39:
        raise RuntimeError("PDV CPE denominator changed")
    if {row["description"] for row in cpe_rows} != {product["cpe_name"] for product in PRODUCTS}:
        raise RuntimeError("PDV CPE product membership changed")

    technical_members = {
        name.rsplit("/", 1)[-1]: payload
        for name, payload in real_members(bundles["technical"]).items()
    }
    technical_pdf_names = {name for name in technical_members if name.lower().endswith(".pdf")}
    if len(technical_pdf_names) != 21:
        raise RuntimeError(f"PDV technical bundle changed: {len(technical_pdf_names)} PDFs")
    for name, expected in EXPECTED_TDS_SHA256.items():
        if sha(technical_members[name]) != expected:
            raise RuntimeError(f"PDV linked technical sheet changed: {name}")

    rows = []
    for index, product in enumerate(PRODUCTS, 1):
        package_rows = [row for row in cpe_rows if row["description"] == product["cpe_name"]]
        technical = {
            "sae_engine": product["sae_engine"], "sae_gear": product["sae_gear"],
            "api": list(product["api"]), "api_gl": list(product["api_gl"]),
            "acea": list(product["acea"]), "ilsac": list(product["ilsac"]),
            "iso_vg": product["iso"], "nlgi": "", "source_grade": product["source_grade"],
            "performance": list(product["performance"]),
        }
        rows.append({
            "source_id": SOURCE_ID,
            "source_record_id": f"PDV-VE-CPE-{index:02d}",
            "market": "Venezuela",
            "manufacturer": "PDVSA Ecuador",
            "importer_and_distributor": "Lubricantes PDV Venezuela",
            "brand": "PDV",
            "product_name": product["name"],
            "source_product_description": product["cpe_name"],
            "family_code": product["family"],
            "technical": technical,
            "packages": package_rows,
            "lifecycle_status": "listed_in_current_official_venezuela_cpe_download",
            "evidence_status": "official_current_distributor_cpe_and_linked_manufacturer_technical_evidence",
            "snapshot_date": SNAPSHOT_DATE,
            "source_url": HOME,
            "source_catalog_url": DOWNLOAD_BASE + BUNDLES["catalog"],
            "source_cpe_url": DOWNLOAD_BASE + BUNDLES["cpe"],
            "source_technical_bundle_url": DOWNLOAD_BASE + BUNDLES["technical"],
            "source_technical_document": product["tds"],
            "source_technical_document_sha256": EXPECTED_TDS_SHA256[product["tds"]],
            "source_quality_flags": [
                "current_official_distributor_cpe_product_and_package_evidence",
                "linked_pdvsa_ecuador_technical_sheet",
                "source_reported_performance_claims_not_independent_approvals",
                "marketing_prose_excluded",
                *product["flags"],
            ],
        })

    facts = [{
        "source_record_id": row["source_record_id"],
        "source_product_description": row["source_product_description"],
        "technical": row["technical"],
        "packages": row["packages"],
        "source_technical_document_sha256": row["source_technical_document_sha256"],
    } for row in rows]
    facts_sha = sha(json.dumps(facts, ensure_ascii=False, sort_keys=True).encode())
    for row in rows:
        row["source_facts_sha256"] = facts_sha
        row["source_page_text_sha256"] = sha(page_text.encode())
        row["source_bundle_sha256"] = EXPECTED_BUNDLE_SHA256
    OUT.write_text(
        "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in rows),
        encoding="utf-8",
    )
    report = {
        "source_id": SOURCE_ID,
        "snapshot_date": SNAPSHOT_DATE,
        "normalized_products": len(rows),
        "cpe_package_rows": len(cpe_rows),
        "unique_cpe_codes": len({row["cpe"] for row in cpe_rows}),
        "technical_pdfs_audited": len(technical_pdf_names),
        "families": dict(sorted(Counter(row["family_code"] for row in rows).items())),
        "rows_with_sae": sum(bool(row["technical"]["sae_engine"] or row["technical"]["sae_gear"]) for row in rows),
        "rows_with_api_or_api_gl": sum(bool(row["technical"]["api"] or row["technical"]["api_gl"]) for row in rows),
        "rows_with_iso_vg": sum(bool(row["technical"]["iso_vg"]) for row in rows),
        "bundle_sha256": EXPECTED_BUNDLE_SHA256,
        "source_page_text_sha256": sha(page_text.encode()),
        "source_facts_sha256": facts_sha,
        "normalized_output_sha256": sha(OUT.read_bytes()),
        "audited_not_ingested_as_current_products": sorted(
            technical_pdf_names - {product["tds"] for product in PRODUCTS}
        ),
        "critical_source_conflicts": [
            "ULTRADIESEL CJ-4 PDF is byte-identical to CI-4 PDF and its body describes CI-4/SL; "
            "CJ-4/SN is retained only from current catalogue/CPE evidence.",
            "Current Agrofluido 100 CPE/catalog designation conflicts with linked historical SAE 10W-30 TDS; "
            "100 is retained as source_grade and 10W-30 is not assigned.",
        ],
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({k: v for k, v in report.items() if k != "bundle_sha256"}, ensure_ascii=False))


if __name__ == "__main__":
    main()
