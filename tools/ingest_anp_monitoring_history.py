#!/usr/bin/env python3
"""Normalize every machine-readable ANP lubricant-monitoring annex (2017+).

The public ANP page is the manifest of record.  Workbooks are downloaded to a
temporary directory, hashed, parsed, and discarded.  The published output is a
row-level factual observation layer; collection locations, CNPJ values and
other unnecessary personal/business contact fields are deliberately omitted.
"""

from __future__ import annotations

import hashlib
import json
import re
import tempfile
import unicodedata
import urllib.request
from collections import Counter, defaultdict
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "anp-brazil-monitoring-observations.jsonl"
REPORT = ROOT / "data" / "anp-brazil-monitoring-report.json"
SOURCE_PAGE = (
    "https://www.gov.br/anp/pt-br/centrais-de-conteudo/publicacoes/"
    "boletins-anp/boletins/boletim-de-monitoramento-de-lubrificantes"
)
TERMS_URL = "https://www.gov.br/pt-br/termos-de-uso"
USER_AGENT = "MFClassifierResearch/1.0 (open-government-data research)"


HEADER_ALIASES = {
    "sample_id": {"AMOSTRA", "NUMERO AMOSTRA", "ID CPT"},
    "product_name": {"MARCA", "MARCA COMERCIAL"},
    "registration_holder": {"DETENTOR"},
    "registration_number": {"REGISTRO", "REGISTRO ANP", "REG ANP"},
    "sae": {"GRAU SAE", "SAE"},
    "performance_level": {"NIVEL DE DESEMPENHO", "ND"},
    "lot": {"LOTE"},
    "manufactured_at": {"DATA DE FABRICACAO", "DATA FABRICACAO"},
    "collected_at": {"DATA DA COLETA", "DATA COLETA"},
}

MISSING_VALUES = {"", "N A", "N.A", "N.A.", "NA", "NAO IDENTIFICADO", "NAO INFORMADO"}


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = text.encode("ascii", "ignore").decode().upper()
    return re.sub(r"[^A-Z0-9+.-]+", " ", text).strip()


def factual_value(value: object) -> str:
    cleaned = clean(value)
    return "" if normalize(cleaned) in MISSING_VALUES else cleaned


def download(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


class AnnexLinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.urls: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag != "a":
            return
        href = dict(attrs).get("href") or ""
        url = urljoin(SOURCE_PAGE, href)
        path = urlparse(url).path.lower()
        if path.endswith(".xlsx") and "boletim-monitoramento-lubrificantes" in path:
            self.urls.append(url)


def header_mapping(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    candidates = []
    for row_index, row in enumerate(rows[:12]):
        normalized = [normalize(value) for value in row]
        mapping = {}
        for field, aliases in HEADER_ALIASES.items():
            for column_index, value in enumerate(normalized):
                if value in aliases:
                    mapping[field] = column_index
                    break
        score = sum(field in mapping for field in ("product_name", "registration_number", "sae", "performance_level"))
        candidates.append((score, row_index, mapping))
    score, row_index, mapping = max(candidates, default=(0, 0, {}))
    if score < 3 or "product_name" not in mapping:
        raise ValueError(f"No supported product header found: best score={score}")
    return row_index, mapping


def parse_sheet(path: Path, sheet, source_url: str, source_sha256: str) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index, mapping = header_mapping(rows)
    preamble = " ".join(normalize(value) for row in rows[:header_index] for value in row if value)
    source_reports_conforming_scope = "AMOSTRAS CONFORMES" in preamble
    year_match = re.search(r"/(20\d{2})/", source_url)
    issue_year = int(year_match.group(1)) if year_match else None
    records = []
    for source_row, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = {
            field: clean(row[column]) if column < len(row) else ""
            for field, column in mapping.items()
        }
        product_name = factual_value(values.get("product_name"))
        if not product_name or normalize(product_name) in HEADER_ALIASES["product_name"]:
            continue
        if not any(factual_value(values.get(field)) for field in ("registration_number", "sae", "performance_level")):
            continue
        sample_id = factual_value(values.get("sample_id"))
        locator = f"{source_sha256[:12]}:{sheet.title}:{source_row}"
        records.append({
            "source_id": "ANP_BRAZIL_LUBRICANT_MONITORING_HISTORY",
            "source_record_id": f"ANP-PML-{hashlib.sha256((locator + '|' + sample_id).encode()).hexdigest()[:18]}",
            "source_page_url": SOURCE_PAGE,
            "source_url": source_url,
            "source_file_sha256": source_sha256,
            "source_file_name": path.name,
            "source_sheet": sheet.title,
            "source_row": source_row,
            "snapshot_date": date.today().isoformat(),
            "issue_year": issue_year,
            "market": "Brazil",
            "sample_id": sample_id,
            "product_name": product_name,
            "registration_holder": factual_value(values.get("registration_holder")),
            "registration_number": factual_value(values.get("registration_number")),
            "sae": factual_value(values.get("sae")),
            "performance_level": factual_value(values.get("performance_level")),
            "lot": factual_value(values.get("lot")),
            "manufactured_at": factual_value(values.get("manufactured_at")),
            "collected_at": factual_value(values.get("collected_at")),
            "source_reports_conforming_scope": source_reports_conforming_scope,
            "quality_flags": [],
            "lifecycle_status": "historical_market_sample_observation",
        })
    return records


def join_supplementary_flags(main_rows: list[dict], supplementary_rows: list[dict]) -> int:
    by_sample = defaultdict(list)
    for row in main_rows:
        if row["sample_id"]:
            by_sample[normalize(row["sample_id"])].append(row)
    joined = 0
    for row in supplementary_rows:
        title = normalize(row["source_sheet"])
        flag = None
        if "ADITIVACAO AUSENTE" in title:
            flag = "source_reported_additive_absence"
        elif "SEM REGISTRO" in title:
            flag = "source_reported_without_registration"
        if not flag:
            continue
        candidates = by_sample.get(normalize(row["sample_id"]), []) if row["sample_id"] else []
        if not candidates:
            identity = tuple(normalize(row[field]) for field in ("product_name", "registration_number", "sae", "performance_level", "lot"))
            candidates = [
                candidate for candidate in main_rows
                if candidate["source_file_sha256"] == row["source_file_sha256"]
                and tuple(normalize(candidate[field]) for field in ("product_name", "registration_number", "sae", "performance_level", "lot")) == identity
            ]
        for candidate in candidates:
            if flag not in candidate["quality_flags"]:
                candidate["quality_flags"].append(flag)
                joined += 1
    return joined


def main() -> None:
    page = download(SOURCE_PAGE)
    parser = AnnexLinkParser()
    parser.feed(page.decode("utf-8", errors="replace"))
    urls = list(dict.fromkeys(parser.urls))
    if len(urls) < 28:
        raise RuntimeError(f"Expected at least 28 official XLSX annexes, found {len(urls)}")

    all_rows = []
    files = []
    supplementary_rows = []
    with tempfile.TemporaryDirectory(prefix="anp-pml-") as temporary:
        directory = Path(temporary)
        for index, url in enumerate(urls, start=1):
            payload = download(url)
            source_hash = hashlib.sha256(payload).hexdigest()
            year = re.search(r"/(20\d{2})/", url)
            file_name = f"{year.group(1) if year else 'unknown'}-{index:02d}-{Path(urlparse(url).path).name}"
            path = directory / file_name
            path.write_bytes(payload)
            workbook = load_workbook(path, read_only=True, data_only=True)
            parsed_count = 0
            for sheet in workbook:
                title = normalize(sheet.title)
                if "DICION" in title:
                    continue
                try:
                    parsed = parse_sheet(path, sheet, url, source_hash)
                except ValueError:
                    continue
                parsed_count += len(parsed)
                if "ADITIVACAO AUSENTE" in title or "SEM REGISTRO" in title:
                    supplementary_rows.extend(parsed)
                else:
                    all_rows.extend(parsed)
            files.append({
                "source_url": url,
                "source_file_name": Path(urlparse(url).path).name,
                "source_sha256": source_hash,
                "parsed_rows_including_supplementary": parsed_count,
            })

    joined_flags = join_supplementary_flags(all_rows, supplementary_rows)
    for row in all_rows:
        row["quality_flags"].sort()
    all_rows.sort(key=lambda row: (row["issue_year"] or 0, row["source_url"], row["source_sheet"], row["source_row"]))
    output_text = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in all_rows)
    OUTPUT.write_text(output_text, encoding="utf-8")

    identity_fields = ("registration_number", "product_name", "sae", "performance_level", "registration_holder")
    identities = {
        tuple(normalize(row[field]) for field in identity_fields)
        for row in all_rows
    }
    report = {
        "schema_version": 1,
        "status": "official_open_government_monitoring_archive_normalized",
        "snapshot_date": date.today().isoformat(),
        "source_page_url": SOURCE_PAGE,
        "terms_url": TERMS_URL,
        "source_page_sha256": hashlib.sha256(page).hexdigest(),
        "normalized_output_sha256": hashlib.sha256(output_text.encode()).hexdigest(),
        "official_xlsx_files": len(files),
        "files": files,
        "source_observations": len(all_rows),
        "unique_sample_ids": len({normalize(row["sample_id"]) for row in all_rows if row["sample_id"]}),
        "normalized_product_grade_holder_identities": len(identities),
        "observations_by_issue_year": dict(sorted(Counter(str(row["issue_year"]) for row in all_rows).items())),
        "supplementary_rows": len(supplementary_rows),
        "supplementary_quality_flags_joined": joined_flags,
        "quality_flags": dict(sorted(Counter(flag for row in all_rows for flag in row["quality_flags"]).items())),
        "rights_note": "ANP publishes the archive on gov.br, whose page states CC BY-ND 3.0. Only normalized factual fields and provenance are republished; workbook layout and narrative text are omitted.",
        "scope_note": "Complete machine-readable XLSX annex scope linked by the official PML page at snapshot time. The page has no XLSX annexes for 2024 and only PDF bulletins before late 2017; those PDF-only years remain a separate extraction backlog.",
        "lifecycle_note": "A monitoring observation proves that a labelled sample existed and was analysed. It is historical evidence, not proof of current registration or market availability.",
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({key: value for key, value in report.items() if key != "files"}, ensure_ascii=False))


if __name__ == "__main__":
    main()
