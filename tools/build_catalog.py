#!/usr/bin/env python3
"""Build the MF Classifier v3 data model from the supplied workbooks and legacy data."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_AZMOL = Path("/workspace/agentbox-uploads/228bb9727e01/Каталог Азмол.PDF")


def scalar(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    value = re.sub(r"\s+", " ", str(value)).strip()
    return value or None


def normalized(value: Any) -> str:
    text = str(value or "").lower().replace("ё", "е").replace("с", "c")
    return re.sub(r"[^a-zа-я0-9]+", "", text)


def normalized_grade(value: Any) -> str:
    return normalized(str(value or "").replace("SAE", "").replace("ISO VG", ""))


def split_semicolon(value: Any) -> list[str]:
    return [x.strip() for x in str(value or "").split(";") if x.strip()]


def rows_from_sheet(workbook, sheet: str, header_row: int = 1) -> list[dict]:
    rows = list(workbook[sheet].iter_rows(values_only=True))
    headers = [clean_text(value) for value in rows[header_row - 1]]
    output = []
    for raw in rows[header_row:]:
        if not any(value not in (None, "") for value in raw):
            continue
        output.append({header: scalar(value) for header, value in zip(headers, raw) if header})
    return output


def load_classes(path: Path) -> tuple[list[dict], dict[str, list[dict]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    lubricant_rows = rows_from_sheet(workbook, "DATA_PRODUCTS")
    classes = []
    for row in lubricant_rows:
        if not row.get("InternalCode"):
            continue
        classes.append({
            "id": row["InternalCode"],
            "kind": "lubricant",
            "category_code": row.get("CategoryCode"),
            "category": row.get("CategoryName"),
            "product_form": row.get("ProductForm"),
            "aggregate_state": row.get("AggregateState"),
            "base_composition": row.get("BaseComposition"),
            "base_oil": row.get("BaseOil"),
            "thickener": row.get("Thickener"),
            "nlgi": row.get("NLGI"),
            "iso_vg": row.get("ISO_VG"),
            "sae_engine": row.get("SAE_Engine"),
            "sae_gear": row.get("SAE_Gear"),
            "api": row.get("API"),
            "api_gl": row.get("API_GL"),
            "acea": row.get("ACEA"),
            "gost": row.get("GOST"),
            "astm_product": row.get("ASTM_Product"),
            "astm_tests": split_semicolon(row.get("ASTM_Tests")),
            "astm_notes": row.get("ASTM_Notes"),
            "temp_min": row.get("TempMin_C"),
            "temp_max": row.get("TempMax_C"),
            "application": row.get("Application"),
            "examples": row.get("Examples"),
            "analogues": row.get("Analogues"),
            "notes": row.get("Notes"),
        })

    for row in rows_from_sheet(workbook, "TECH_FLUIDS"):
        if not row.get("TechFluidCode"):
            continue
        standards = [row.get(key) for key in (
            "ISO_Standard", "DIN_Standard", "ASTM_SAE", "OEM_Standards"
        ) if row.get(key)]
        classes.append({
            "id": row["TechFluidCode"],
            "kind": "technical_fluid",
            "category_code": "TF",
            "category": row.get("Класс"),
            "product_form": row.get("Тип продукта"),
            "aggregate_state": "Жидкое",
            "base_composition": row.get("Основа"),
            "tnved_chapter": row.get("TNVED_Chapter"),
            "tnved_hint": row.get("TNVED_Code"),
            "standards": standards,
            "application": row.get("Применение"),
            "is_concentrate": row.get("IsConcentrate"),
        })

    references = {
        "standards": rows_from_sheet(workbook, "Стандарты", 3),
        "astm_methods": rows_from_sheet(workbook, "ASTM_Справочник"),
        "temperature_zones": rows_from_sheet(workbook, "Температурные зоны", 3),
        "industries": rows_from_sheet(workbook, "Отрасли применения", 3),
        "keywords": rows_from_sheet(workbook, "KEYWORDS_DICTIONARY"),
        "tnved_rules": rows_from_sheet(workbook, "TNVED_RULES"),
        "coding_system": rows_from_sheet(workbook, "Система кодирования", 4),
    }
    return classes, references


PRODUCT_COLUMNS = {
    "№": "source_number",
    "Столбец1": "brand_source",
    "Наименование": "name",
    "продукт": "category",
    "вязкость": "viscosity",
    "компрессорные,гидравлика,редукторное DIN(ГОСТ)": "din_gost_class",
    "моторка,трансмиссия SAE": "sae_class",
    "моторка,трансмисия API": "api_class",
    "название по ГОСТ": "gost_name",
    "антифриз": "coolant_class",
    "смазки": "grease_class",
    "сертификат соотвествия": "certificate_number",
    "сертификат соотвествия выдан(ДАТА)": "certificate_issued_at",
    "сертификат соотвествия действует до(ДАТА)": "certificate_expires_at",
    "сертификат локального производителя2": "local_producer_certificate",
    "ГОСТР/ТС/УЗТР продукция Общий": "technical_document",
    "код ТН ВЭД, указываемый на сертификате": "tnved_code",
}


def product_brand(value: Any, name: str) -> str:
    source = str(value or "").upper()
    if "CHILON" in source or "CHILON" in name.upper():
        return "CHILON"
    if "UNO" in source or "UNO" in name.upper():
        return "UNO"
    return clean_text(value) or "НЕ УКАЗАН"


def category_code(category: Any) -> str | None:
    value = normalized(category)
    routes = (
        (("мотор",), "M"), (("трансмис",), "T"), (("гидрав",), "H"),
        (("компресс", "холодиль"), "C"), (("турбин",), "U"),
        (("трансформ", "электроизоля"), "E"), (("смазк",), "G"),
        (("редукт", "индустри", "шпиндель", "направляющ", "прокат"), "I"),
        (("охлажда", "антифриз", "сож"), "TF"),
    )
    for needles, code in routes:
        if any(needle in value for needle in needles):
            return code
    return "S"


def load_products(path: Path) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = rows_from_sheet(workbook, "продукция")
    products_by_key: dict[str, dict] = {}
    anomalies = []
    for source_row, row in enumerate(rows, 2):
        name = clean_text(row.get("Наименование"))
        if not name:
            continue
        product = {target: scalar(row.get(source)) for source, target in PRODUCT_COLUMNS.items()}
        product["name"] = name
        product["brand"] = product_brand(row.get("Столбец1"), name)
        product["family"] = family_for_category(product.get("category"))
        product["category_code"] = category_code(product.get("category"))
        product["source"] = "products_classified_2026"
        product["source_row"] = source_row
        key = normalized(name)
        if key in products_by_key:
            previous = products_by_key[key]
            anomalies.append({
                "type": "duplicate_name",
                "name": name,
                "kept_row": previous["source_row"],
                "duplicate_row": source_row,
                "kept_brand": previous["brand_source"],
                "duplicate_brand": product["brand_source"],
                "note": "Имя содержит CHILON, поэтому строка с брендом UNO помечена как конфликт источника.",
            })
            continue
        products_by_key[key] = product
    return list(products_by_key.values()), anomalies


def family_for_category(category: Any) -> str:
    value = normalized(category)
    if "смазк" in value:
        return "Пластичные смазки"
    if "охлажда" in value or "антифриз" in value:
        return "Охлаждающие жидкости"
    if "сож" in value:
        return "СОЖ"
    return "Масла"


def legacy_name_tokens(value: str) -> list[str]:
    value = re.sub(r"\([^)]*\)", " ", value.lower())
    return [token for token in re.findall(r"[a-zа-яё0-9]+", value) if len(token) > 1]


def legacy_match(row: dict, products: list[dict]) -> dict | None:
    if str(row.get("brand", "")).lower() != "chilon":
        return None
    tokens = legacy_name_tokens(row["name"])
    grade = normalized_grade(row.get("viscosity_class"))
    candidates = []
    for product in products:
        if product["brand"] != "CHILON" or normalized_grade(product.get("sae_class")) != grade:
            continue
        haystack = normalized(product["name"])
        if tokens and all(normalized(token) in haystack for token in tokens):
            candidates.append(product)
    return candidates[0] if len(candidates) == 1 else None


def legacy_product(row: dict) -> dict:
    viscosity = clean_text(str(row.get("viscosity_class") or "").replace("SAE", ""))
    return {
        "name": " ".join(filter(None, [row.get("brand"), row.get("name"), row.get("viscosity_class")])),
        "brand": "CHILON" if str(row.get("brand", "")).lower() == "chilon" else row.get("brand"),
        "brand_source": row.get("brand"),
        "family": "Масла",
        "category": "Масла моторные",
        "category_code": "M",
        "viscosity": None,
        "din_gost_class": row.get("viscosity_class_gost"),
        "sae_class": viscosity,
        "api_class": row.get("specification"),
        "gost_name": None,
        "coolant_class": None,
        "grease_class": None,
        "certificate_number": None,
        "certificate_issued_at": None,
        "certificate_expires_at": None,
        "local_producer_certificate": None,
        "technical_document": row.get("standard"),
        "tnved_code": row.get("tnved"),
        "unit": row.get("unit"),
        "packaging": row.get("packaging"),
        "container": row.get("container"),
        "ikpu": row.get("ikpu"),
        "enkt": row.get("enkt"),
        "skp": row.get("skp"),
        "source": "legacy_mfclassifier",
        "legacy_id": row.get("id"),
    }


def merge_legacy(products: list[dict], legacy_path: Path) -> tuple[int, int]:
    legacy = json.loads(legacy_path.read_text(encoding="utf-8"))["motor_oils"]
    enriched = 0
    appended = 0
    for row in legacy:
        match = legacy_match(row, products)
        if match:
            for source, target in (
                ("unit", "unit"), ("packaging", "packaging"), ("container", "container"),
                ("ikpu", "ikpu"), ("enkt", "enkt"), ("skp", "skp"),
            ):
                match[target] = row.get(source)
            match["legacy_id"] = row.get("id")
            match["source"] = "products_classified_2026+legacy_mfclassifier"
            enriched += 1
        else:
            products.append(legacy_product(row))
            appended += 1
    return enriched, appended


def class_match(product: dict, classes: list[dict]) -> dict | None:
    code = product.get("category_code")
    candidates = []
    for item in classes:
        if code == "TF":
            if item["kind"] != "technical_fluid":
                continue
        elif item.get("category_code") != code:
            continue
        score = 35
        basis = ["категория"]
        for product_field, class_field, label, weight in (
            ("sae_class", "sae_engine", "SAE", 35),
            ("sae_class", "sae_gear", "SAE Gear", 35),
            ("viscosity", "iso_vg", "ISO VG", 35),
        ):
            left, right = normalized_grade(product.get(product_field)), normalized_grade(item.get(class_field))
            if left and right and left == right:
                score += weight
                basis.append(label)
        api = normalized(product.get("api_class"))
        class_api = normalized(item.get("api") or item.get("api_gl"))
        if api and class_api and (api in class_api or class_api in api):
            score += 15
            basis.append("API")
        details = normalized(" ".join(str(product.get(key) or "") for key in (
            "din_gost_class", "grease_class", "coolant_class", "name"
        )))
        marker = normalized(" ".join(str(item.get(key) or "") for key in (
            "id", "nlgi", "thickener", "product_form"
        )))
        shared = [token for token in ("hlp", "clp", "nlgi0", "nlgi1", "nlgi2", "nlgi3", "dot4", "dot5")
                  if token in details and token in marker]
        if shared:
            score += 20
            basis.append("класс/тип")
        candidates.append((score, item["id"], basis))
    if not candidates:
        return None
    candidates.sort(key=lambda x: (-x[0], x[1]))
    best = candidates[0]
    if best[0] <= 35:
        return None
    ties = sum(score == best[0] for score, _, _ in candidates)
    confidence = min(98, best[0] - (8 if ties > 1 else 0))
    return {"class_id": best[1], "confidence": confidence, "basis": best[2], "status": "suggested"}


def parse_azmol_index(pdf_path: Path) -> list[dict]:
    if not pdf_path.exists():
        return []
    try:
        from pypdf import PdfReader
    except ImportError:
        return []
    reader = PdfReader(pdf_path)
    output = []
    for pdf_page in range(107, 111):
        lines = [clean_text(x) for x in (reader.pages[pdf_page - 1].extract_text() or "").splitlines()]
        lines = [x for x in lines if x]
        headers = [i for i, line in enumerate(lines) if line == "№н/п Наименование продукции стр."]
        headers.append(len(lines))
        for start, end in zip(headers, headers[1:]):
            block = lines[start + 1:end]
            numbers = []
            while block and re.fullmatch(r"\d+", block[0]):
                numbers.append(int(block.pop(0)))
            if not numbers:
                continue
            reference_pattern = re.compile(r"^\d+(?:\s*,\s*\d+)*$")
            split_at = next((i for i, line in enumerate(block) if i >= len(numbers) - 2 and reference_pattern.fullmatch(line)), len(block))
            names, references = block[:split_at], block[split_at:]
            while len(names) > len(numbers):
                merge_index = next((i for i, name in enumerate(names) if name in {"®"} or name.lower().startswith("и рулевого")), None)
                if merge_index is None:
                    merge_index = min(range(1, len(names)), key=lambda i: len(names[i]))
                names[merge_index - 1] = f"{names[merge_index - 1]} {names[merge_index]}"
                names.pop(merge_index)
            for index, number in enumerate(numbers):
                if index >= len(names):
                    break
                output.append({
                    "id": f"AZMOL-{number:03d}",
                    "name": names[index],
                    "brand": "АЗМОЛ",
                    "catalog_pages": references[index] if index < len(references) else None,
                    "pdf_page": pdf_page,
                    "source": "Каталог АЗМОЛ, 2008",
                })
    return sorted(output, key=lambda row: row["id"])


def certificate_status(product: dict, today: date) -> str:
    if not product.get("certificate_number"):
        return "missing"
    expires = product.get("certificate_expires_at")
    if not expires:
        return "no_expiry"
    return "expired" if str(expires) < today.isoformat() else "valid"


def build(args: argparse.Namespace) -> dict:
    classes, references = load_classes(args.classifier)
    products, anomalies = load_products(args.products)
    enriched, appended = merge_legacy(products, args.legacy)
    for index, product in enumerate(products, 1):
        product["id"] = f"P-{index:04d}"
        product["class_match"] = class_match(product, classes)
        product["certificate_status"] = certificate_status(product, date.today())

    azmol = parse_azmol_index(args.azmol_pdf)
    category_counts = Counter(product["category"] for product in products)
    brand_counts = Counter(product["brand"] for product in products)
    linked = sum(product["class_match"] is not None for product in products)
    payload = {
        "schema_version": 3,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "title": "Единый классификатор смазочных материалов и технических жидкостей",
        "metrics": {
            "products": len(products),
            "brands": len(brand_counts),
            "product_categories": len(category_counts),
            "classes": len(classes),
            "class_links": linked,
            "standards": len(references["standards"]),
            "astm_methods": len(references["astm_methods"]),
            "reference_products": len(azmol),
        },
        "products": products,
        "classes": classes,
        "references": {**references, "azmol_products": azmol},
        "facets": {
            "brands": dict(sorted(brand_counts.items())),
            "categories": dict(sorted(category_counts.items())),
            "class_categories": dict(sorted(Counter(item["category"] for item in classes).items())),
        },
        "quality": {
            "anomalies": anomalies,
            "legacy_enriched": enriched,
            "legacy_appended": appended,
            "unlinked_products": len(products) - linked,
            "tnved_notice": "Код ТН ВЭД зависит от состава и формы продукта; подсказка не заменяет юридическую классификацию.",
        },
        "sources": [
            {"id": "products-2026", "title": "Продукция с классификаторами", "type": "XLSX", "records": 444, "role": "коммерческая номенклатура и сертификаты"},
            {"id": "classifier-v2.1", "title": "Классификатор смазочных материалов ASTM v2.1", "type": "XLSX", "records": 205, "role": "отраслевые классы, стандарты и правила"},
            {"id": "api-1509-23", "title": "API 1509, 23rd Edition", "type": "PDF", "edition": "February 2025", "pages": 248, "role": "лицензирование и сертификация моторных масел"},
            {"id": "atc-2024", "title": "Petroleum Additives Product Approval Code of Practice", "type": "PDF", "edition": "August 2024", "pages": 135, "role": "испытания и подтверждение эксплуатационных характеристик"},
            {"id": "azmol-2008", "title": "Каталог продукции АЗМОЛ", "type": "PDF", "edition": "2008", "pages": 112, "records": len(azmol), "role": "исторические марки, аналоги и примеры"},
            {"id": "legacy-mfclassifier", "title": "MF Classifier legacy", "type": "JSON", "records": 107, "role": "мультибрендовые товары и коды ИКПУ/ЕНКТ/СКП"},
        ],
    }
    return payload


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--classifier", type=Path, default=ROOT / "sources/classifier_astm_v2_1.xlsx")
    parser.add_argument("--products", type=Path, default=ROOT / "sources/products_classified_2026.xlsx")
    parser.add_argument("--legacy", type=Path, default=ROOT / "motor_oils.json")
    parser.add_argument("--azmol-pdf", type=Path, default=DEFAULT_AZMOL)
    parser.add_argument("--output", type=Path, default=ROOT / "data/catalog-v3.json")
    args = parser.parse_args()
    payload = build(args)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    compact_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    args.output.write_text(compact_json + "\n", encoding="utf-8")
    script_path = args.output.with_suffix(".js")
    script_path.write_text("window.MF_CLASSIFIER_DATA=" + compact_json + ";\n", encoding="utf-8")
    print(json.dumps(payload["metrics"], ensure_ascii=False))
    print(json.dumps(payload["quality"], ensure_ascii=False))


if __name__ == "__main__":
    main()
